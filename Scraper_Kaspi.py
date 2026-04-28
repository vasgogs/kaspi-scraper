from playwright.sync_api import sync_playwright, TimeoutError
import math
import os
import pandas as pd
import re
import tempfile
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from collections.abc import Sequence
from datetime import datetime, timedelta
import threading
import json
from pathlib import Path
from urllib.parse import urlparse, quote_plus, urljoin, urlencode, parse_qs
from concurrent.futures import ProcessPoolExecutor, as_completed
from concurrent.futures.process import BrokenProcessPool
import sys
import random
import time
import signal
import contextlib
import fcntl
import csv
import sqlite3
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill
from PIL import Image, ImageDraw, ImageFont
import textwrap
import smtplib
from email.message import EmailMessage
from zoneinfo import ZoneInfo
from typing import Any

DEFAULT_SHAREPOINT_SITE_URL = "https://stadaarz-my.sharepoint.com/personal/vasily_gogolev_stada_kz"
DEFAULT_SHAREPOINT_FILE_URL = "/personal/vasily_gogolev_stada_kz/Documents/Book 1.xlsx"
DEFAULT_SHAREPOINT_USERNAME = "vasily.gogolev@stada.kz"
DEFAULT_SHAREPOINT_PASSWORD = os.environ.get("SHAREPOINT_PASSWORD", "")
DEFAULT_TENANT_ID = os.environ.get("AZURE_TENANT_ID", "organizations")
DEFAULT_GRAPH_UPLOAD_PATH = "Documents/Book 1.xlsx"
TELEGRAM_SAFE_CHUNK = 3000  # запас до лимита 4096 и учёт смайлов/UTF-8
PRICE_TOLERANCE_KZT = int(os.environ.get("PRICE_TOLERANCE_KZT", "5"))
TELEGRAM_FILE_ONLY = os.environ.get("TELEGRAM_FILE_ONLY", "0").strip().lower() in {"1", "true", "yes"}
MISSION_SUPPRESS_TELEGRAM = os.environ.get("MISSION_SUPPRESS_TELEGRAM", "0").strip().lower() in {"1", "true", "yes"}
ALMATY_TZ = ZoneInfo("Asia/Almaty")
CITY_CODES = {
    "Алматы": "750000000",
    "Астана": "710000000",
    "Шымкент": "511010000",
}

BUNDLE_COLUMNS = [
    "bundle_of",
    "bundle_qty",
    "bundle_1_sku_price_kzt",
    "bundle_price_kzt",
    "bundle_unit_price_kzt",
    "bundle_same_sku_total_kzt",
    "bundle_savings_vs_same_sku_total_kzt",
    "bundle_vs_single_kzt",
    "bundle_best_seller",
    "1_sku_best_price_kzt",
    "bundle_best_price_kzt",
    "bundle_savings_vs_best_total_kzt",
    "bundle_vs_best_kzt",
    "bundle_discount_pct",
    "bundle_note",
]

# Быстрые таймауты/паузы, можно повысить при блокировках.
FAST_NAV_TIMEOUT = int(os.environ.get("SCRAPER_FAST_NAV_TIMEOUT_MS", "18000"))  # первая попытка загрузки, мс
SLOW_NAV_TIMEOUT = int(os.environ.get("SCRAPER_SLOW_NAV_TIMEOUT_MS", "32000"))  # вторая попытка загрузки, мс
POST_LOAD_DELAY_FAST = int(os.environ.get("SCRAPER_POST_LOAD_DELAY_FAST_MS", "1600"))  # задержка после goto, мс
POST_LOAD_DELAY_SLOW = int(os.environ.get("SCRAPER_POST_LOAD_DELAY_SLOW_MS", "3200"))  # задержка после goto (повтор), мс
POST_PAGINATION_DELAY = int(os.environ.get("SCRAPER_POST_PAGINATION_DELAY_MS", "1200"))  # задержка после клика Следующая, мс
SELLER_WAIT_TIMEOUT = int(os.environ.get("SCRAPER_SELLER_WAIT_TIMEOUT_MS", "24000"))  # ожидание блока продавцов, мс
OVERLAY_WAIT_TIMEOUT = int(os.environ.get("SCRAPER_OVERLAY_WAIT_TIMEOUT_MS", "9000"))
MAX_PAGINATION_RETRIES = int(os.environ.get("SCRAPER_PAGINATION_RETRIES", "8"))
MAX_BROWSER_RETRIES = int(os.environ.get("SCRAPER_BROWSER_RETRIES", "2"))
WORKER_JITTER_MIN = float(os.environ.get("SCRAPER_WORKER_JITTER_MIN", "0.2"))
WORKER_JITTER_MAX = float(os.environ.get("SCRAPER_WORKER_JITTER_MAX", "0.8"))
JOB_TIMEOUT_SECONDS = int(os.environ.get("SCRAPER_JOB_TIMEOUT", "420"))
SCRAPER_CITY_RETRIES = int(os.environ.get("SCRAPER_CITY_RETRIES", "1"))
SCRAPER_CITY_RETRY_BACKOFF_SEC = int(os.environ.get("SCRAPER_CITY_RETRY_BACKOFF_SEC", "20"))
SELLER_DISCOVERY_ATTEMPTS = int(os.environ.get("SCRAPER_SELLER_DISCOVERY_ATTEMPTS", "14"))
SELLER_DISCOVERY_RELOAD_EVERY = int(os.environ.get("SCRAPER_SELLER_DISCOVERY_RELOAD_EVERY", "3"))
SELLER_DISCOVERY_PAUSE_MS = int(os.environ.get("SCRAPER_SELLER_DISCOVERY_PAUSE_MS", "1200"))
NO_OFFERS_TEXT_FRAGMENT = "нет предложений по данному товару"
STATE_DIR = Path(__file__).resolve().parent / "state"
PROGRESS_PATH = STATE_DIR / "progress.json"
REVIEW_PROGRESS_PATH = STATE_DIR / "review_progress.json"
BUNDLE_CACHE_PATH = STATE_DIR / "bundle_price_cache.json"
BUNDLE_CACHE_VERSION = 1
REVIEWS_DB_PATH = Path(os.environ.get("REVIEWS_DB_PATH", str(STATE_DIR / "reviews.db")))
SCRAPER_LOCK_FILE = Path(os.environ.get("SCRAPER_LOCK_FILE", "/tmp/kaspi_scraper.lock"))
SCRAPER_PRIMARY_FLAG = Path(os.environ.get("SCRAPER_PRIMARY_FLAG", "/tmp/kaspi_scraper.primary"))
SCRAPER_PRIMARY_TTL_SEC = int(os.environ.get("SCRAPER_PRIMARY_TTL_SEC", "21600"))
SCRAPER_METRICS_PATH = Path(
    os.environ.get(
        "SCRAPER_METRICS_PATH",
        str(Path(__file__).resolve().parent / "logs" / "scrape_metrics.csv"),
    )
)
SNAPSHOT_TTL_HOURS = 12
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
OPENAI_MODEL = os.environ.get("OPENAI_MODEL", "gpt-4o-mini")
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/129.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 13_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Safari/605.1.15",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36",
]

REQUEST_SESSION = requests.Session()
_retry = Retry(
    total=3,
    backoff_factor=0.4,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["GET", "HEAD", "OPTIONS"],
)
_adapter = HTTPAdapter(max_retries=_retry)
REQUEST_SESSION.mount("http://", _adapter)
REQUEST_SESSION.mount("https://", _adapter)


def _random_user_agent() -> str:
    return random.choice(USER_AGENTS)


def _requests_get(url: str, **kwargs):
    headers = kwargs.pop("headers", {})
    headers.setdefault("User-Agent", _random_user_agent())
    return REQUEST_SESSION.get(url, headers=headers, **kwargs)


def _jitter_sleep(min_s: float = WORKER_JITTER_MIN, max_s: float = WORKER_JITTER_MAX):
    if max_s <= 0:
        return
    try:
        time.sleep(random.uniform(max(0.01, min_s), max(max_s, min_s)))
    except Exception:
        pass


def _wait_for_overlay(page, timeout: int = OVERLAY_WAIT_TIMEOUT):
    """Ждём, пока пропадут перекрывающие затемнения (curtain)."""
    try:
        page.wait_for_function(
            """() => {
                const overlays = Array.from(document.querySelectorAll('div.curtain'));
                if (!overlays.length) {
                    return true;
                }
                return overlays.every((node) => {
                    const styles = window.getComputedStyle(node);
                    const hidden = styles.visibility === 'hidden' || styles.display === 'none';
                    const notVisible = !node.classList.contains('_visible');
                    return hidden || notVisible;
                });
            }""",
            timeout=timeout,
        )
    except Exception:
        pass


def _append_scrape_metric(row: dict):
    """Append one metrics row to CSV with a file lock for multi-process safety."""
    try:
        path = SCRAPER_METRICS_PATH
        path.parent.mkdir(parents=True, exist_ok=True)
        header = ["ts", "city", "product", "url", "attempt", "status", "elapsed_s", "total_s", "error"]
        with open(path, "a+", encoding="utf-8", newline="") as fh:
            fcntl.flock(fh, fcntl.LOCK_EX)
            fh.seek(0, os.SEEK_END)
            if fh.tell() == 0:
                writer = csv.DictWriter(fh, fieldnames=header)
                writer.writeheader()
            writer = csv.DictWriter(fh, fieldnames=header)
            writer.writerow({k: row.get(k, "") for k in header})
            fcntl.flock(fh, fcntl.LOCK_UN)
    except Exception:
        pass


def _is_transient_worker_error(exc: Exception) -> bool:
    msg = str(exc).lower()
    if isinstance(exc, BrokenProcessPool):
        return True
    return "epipe" in msg or "brokenprocesspool" in msg or "pipe" in msg and "broken" in msg


def _safe_click(locator, page, description: str, retries: int = 3) -> bool:
    """Кликает по элементу с доп. ожиданиями и скроллом."""
    for attempt in range(retries):
        try:
            _wait_for_overlay(page)
            locator.scroll_into_view_if_needed(timeout=2000)
            locator.click(timeout=6000)
            return True
        except Exception as exc:
            exc_text = str(exc).lower()
            if "intercepts pointer events" in exc_text or "curtain" in exc_text:
                # Overlay sometimes stays visible even when the tab is technically active.
                # For such cases force a click / JS click and continue scraping instead of hanging.
                try:
                    locator.click(timeout=1500, force=True)
                    return True
                except Exception:
                    pass
                try:
                    locator.evaluate("el => el.click()")
                    return True
                except Exception:
                    pass
                try:
                    page.keyboard.press("Escape")
                except Exception:
                    pass
            if attempt + 1 == retries:
                print(f"⚠️ Could not click {description}: {exc}")
                return False
            try:
                page.mouse.wheel(0, 600)
            except Exception:
                pass
            page.wait_for_timeout(400 + attempt * 150)
    return False


def _load_progress_state(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return {}


def _save_progress_state(path: Path, data: dict):
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _load_review_progress_state(path: Path) -> dict:
    try:
        if path.exists():
            with open(path, "r", encoding="utf-8") as fh:
                return json.load(fh) or {}
    except Exception:
        return {}
    return {}


def _save_review_progress_state(path: Path, data: dict):
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _primary_flag_active() -> bool:
    if not SCRAPER_PRIMARY_FLAG.exists():
        return False
    try:
        raw = SCRAPER_PRIMARY_FLAG.read_text(encoding="utf-8").strip()
        ts = float(raw) if raw else SCRAPER_PRIMARY_FLAG.stat().st_mtime
        if time.time() - ts > SCRAPER_PRIMARY_TTL_SEC:
            SCRAPER_PRIMARY_FLAG.unlink(missing_ok=True)
            return False
        return True
    except Exception:
        return True


def _set_primary_flag():
    try:
        SCRAPER_PRIMARY_FLAG.parent.mkdir(parents=True, exist_ok=True)
        SCRAPER_PRIMARY_FLAG.write_text(str(time.time()), encoding="utf-8")
    except Exception:
        pass


def _clear_primary_flag():
    try:
        SCRAPER_PRIMARY_FLAG.unlink(missing_ok=True)
    except Exception:
        pass


@contextlib.contextmanager
def _exclusive_scrape_lock(priority: str = "primary"):
    """Глобальная блокировка, чтобы сборы не перекрывались."""
    if priority == "secondary":
        while _primary_flag_active():
            print(
                f"[{datetime.now(ALMATY_TZ):%Y-%m-%d %H:%M:%S}] ⏳ Ждём основной скрейп (приоритет)..."
            )
            time.sleep(5)

    flag_set = False
    if priority == "primary":
        _set_primary_flag()
        flag_set = True

    lock_path = SCRAPER_LOCK_FILE
    lock_path.parent.mkdir(parents=True, exist_ok=True)
    fd = os.open(lock_path, os.O_CREAT | os.O_RDWR)
    waited = False
    try:
        try:
            fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
        except BlockingIOError:
            waited = True
            print(
                f"[{datetime.now(ALMATY_TZ):%Y-%m-%d %H:%M:%S}] ⏳ Уже идёт другой запуск скрейпера, ждём завершения..."
            )
            fcntl.flock(fd, fcntl.LOCK_EX)
        yield
    finally:
        fcntl.flock(fd, fcntl.LOCK_UN)
        os.close(fd)
        if flag_set:
            _clear_primary_flag()
        if waited:
            print(f"[{datetime.now(ALMATY_TZ):%Y-%m-%d %H:%M:%S}] ✅ Очередь скрейпера освобождена.")


def _update_last_processed(payload: dict):
    try:
        STATE_DIR.mkdir(parents=True, exist_ok=True)
        snapshot = {
            **payload,
            "timestamp": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
        }
        target = STATE_DIR / "last_product.json"
        with open(target, "w", encoding="utf-8") as fh:
            json.dump(snapshot, fh, ensure_ascii=False, indent=2)
    except Exception:
        pass


def _wait_after_navigation(page, base_delay_ms: int):
    jitter = random.randint(120, 380)
    page.wait_for_timeout(max(base_delay_ms, 0) + jitter)
    _wait_for_overlay(page)


def _has_no_offers_text(page) -> bool:
    selectors = [
        "text=К сожалению, в настоящее время нет предложений по данному товару.",
        "text=К сожалению, в настоящее время нет предложений по данному товару",
        "text=нет предложений по данному товару",
    ]
    def _visible_marker_present() -> bool:
        for selector in selectors:
            try:
                locator = page.locator(selector)
                count = min(locator.count(), 5)
                for idx in range(count):
                    try:
                        if locator.nth(idx).is_visible():
                            return True
                    except Exception:
                        continue
            except Exception:
                continue
        return False

    if not _visible_marker_present():
        return False
    try:
        page.wait_for_timeout(700)
        _wait_for_overlay(page, timeout=1500)
    except Exception:
        pass
    return _visible_marker_present()


class JobTimeoutError(RuntimeError):
    """Raised if a worker spends too long on a single product."""


class NoOffersError(RuntimeError):
    """Raised when Kaspi explicitly shows that the product has no offers."""


@contextlib.contextmanager
def _job_timeout(seconds: int, label: str | None = None):
    if seconds <= 0:
        yield
        return

    def _handle_timeout(signum, frame):
        raise JobTimeoutError(f"Task timeout ({seconds}s) for {label or 'product'}")

    previous = signal.signal(signal.SIGALRM, _handle_timeout)
    signal.setitimer(signal.ITIMER_REAL, seconds)
    try:
        yield
    finally:
        signal.setitimer(signal.ITIMER_REAL, 0)
        signal.signal(signal.SIGALRM, previous)


try:
    from office365.runtime.auth.user_credential import UserCredential
    from office365.runtime.auth.client_credential import ClientCredential
    from office365.sharepoint.client_context import ClientContext
    from office365.sharepoint.files.file import File
    import msal
except ImportError:
    ClientContext = None
    UserCredential = None
    ClientCredential = None
    File = None
    msal = None


def _split_telegram_text(text: str, max_len: int = TELEGRAM_SAFE_CHUNK) -> list[str]:
    """Разбиваем длинный текст на части ниже лимита Telegram."""
    if not text:
        return []
    max_len = max(1, min(max_len, 4096))
    chunks: list[str] = []
    current = ""
    for line in text.splitlines():
        line = line.rstrip()
        if not current:
            if len(line) <= max_len:
                current = line
            else:
                for i in range(0, len(line), max_len):
                    chunks.append(line[i:i + max_len])
                current = ""
            continue
        projected = len(current) + 1 + len(line)
        if projected <= max_len:
            current = f"{current}\n{line}"
        else:
            chunks.append(current)
            if len(line) <= max_len:
                current = line
            else:
                for i in range(0, len(line), max_len):
                    chunks.append(line[i:i + max_len])
                current = ""
    if current:
        chunks.append(current)
    return [chunk for chunk in chunks if chunk]


def extract_price(text: str):
    digits = re.sub(r"[^\d]", "", text)
    return int(digits) if digits else None


def extract_product_code(url: Any) -> str:
    """Получаем код товара из URL каспи (/p/<name>-123456/)."""
    if url is None:
        return ""
    try:
        if pd.isna(url):
            return ""
    except Exception:
        pass
    url = str(url)
    match = re.search(r"/p/[^/]*?-(\d+)(?:/|$)", url)
    return match.group(1) if match else ""


def normalize_seller_name(name: str) -> str:
    """Приведение имени продавца к единому виду для сравнения."""
    text = (name or "").strip()
    lookalikes = str.maketrans({
        "a": "а", "A": "а",
        "e": "е", "E": "е",
        "o": "о", "O": "о",
        "p": "р", "P": "р",
        "c": "с", "C": "с",
        "x": "х", "X": "х",
        "y": "у", "Y": "у",
        "k": "к", "K": "к",
        "h": "н", "H": "н",
        "b": "в", "B": "в",
        "m": "м", "M": "м",
        "t": "т", "T": "т",
    })
    text = text.translate(lookalikes)
    return re.sub(r"\s+", " ", text).lower()


def seller_matches(actual_norm: str, expected_norm: str) -> bool:
    """Гибкое совпадение продавца: точное или подстрочное (для переименований)."""
    if not expected_norm:
        return bool(actual_norm)
    if actual_norm == expected_norm:
        return True
    return expected_norm in actual_norm or actual_norm in expected_norm


MISSION_PARTNER_SELLERS_DEFAULT = (
    "Аптека от А до Я;"
    "Аптека MSP;"
    "Аптека MSP Алматы;"
    "Аптека MSP Астана;"
    "Аптека MSP Шымкент;"
    "ФАРМАКОМ"
)


def parse_seller_list(raw: str | None, default_raw: str = "") -> list[str]:
    text = raw if raw is not None else default_raw
    return [s.strip() for s in re.split(r"[;,/|]", text or "") if s.strip()]


def seller_in_list(seller_name: str, sellers: list[str] | set[str] | tuple[str, ...]) -> bool:
    actual_norm = normalize_seller_name(seller_name)
    for candidate in sellers:
        candidate_norm = normalize_seller_name(candidate)
        if seller_matches(actual_norm, candidate_norm):
            return True
    return False


MISSION_SCREENSHOT_SPECS = [
    {"seller": "Аптека MSP Алматы", "region": "Алматы", "label": "Аптека MSP — Алматы", "slug": "msp_almaty"},
    {"seller": "Аптека MSP Астана", "region": "Астана", "label": "Аптека MSP — Астана", "slug": "msp_astana"},
    {"seller": "Аптека MSP Шымкент", "region": "Шымкент", "label": "Аптека MSP — Шымкент", "slug": "msp_shymkent"},
    {"seller": "ФАРМАКОМ", "region": "Алматы", "label": "ФАРМАКОМ — Алматы", "slug": "pharmakom_almaty"},
    {"seller": "ФАРМАКОМ", "region": "Астана", "label": "ФАРМАКОМ — Астана", "slug": "pharmakom_astana"},
    {"seller": "ФАРМАКОМ", "region": "Шымкент", "label": "ФАРМАКОМ — Шымкент", "slug": "pharmakom_shymkent"},
    {"seller": "Аптека от А до Я", "region": None, "label": "Туровская", "slug": "turovskaya"},
]


def _normalize_text(value: Any) -> str:
    text = str(value or "").replace("\u00a0", " ").strip()
    return " ".join(text.split())


def _normalize_mission_status_text(value: Any) -> str:
    text = str(value or "").replace("\u00a0", " ").strip().lower()
    return " ".join(text.split())


def _mission_status_state(value: Any) -> str:
    text = _normalize_mission_status_text(value)
    if not text:
        return "neutral"
    if text.startswith("ok") or text.startswith("ок"):
        return "ok"
    if text.startswith("внимание") or "не самое выгодное" in text or "не лучшее" in text:
        return "attention"
    if text.startswith("проблема") or text.startswith("дороже") or text.startswith("дешевле"):
        return "alert"
    if "нет" in text or "отсутств" in text or "ошибка" in text:
        return "alert"
    return "neutral"


def _filter_problematic_mission_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "status" not in df.columns:
        return df.copy()
    states = df["status"].fillna("").astype(str).map(_mission_status_state)
    problematic = df.loc[states.eq("alert")].copy()
    return problematic


def _filter_alert_mission_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "status" not in df.columns:
        return df.copy()
    states = df["status"].fillna("").astype(str).map(_mission_status_state)
    alerts = df.loc[states.eq("alert")].copy()
    return alerts


def _filter_mission_preview_slice(df: pd.DataFrame, seller: str, region: str | None = None) -> pd.DataFrame:
    if df.empty or "seller" not in df.columns:
        return df.iloc[0:0].copy()
    expected_seller = normalize_seller_name(seller)
    seller_mask = df["seller"].fillna("").astype(str).map(
        lambda raw: seller_matches(normalize_seller_name(raw), expected_seller)
    )
    mask = seller_mask
    if region:
        if "region" not in df.columns:
            return df.iloc[0:0].copy()
        region_norm = _normalize_text(region).lower()
        region_mask = df["region"].fillna("").astype(str).map(lambda raw: _normalize_text(raw).lower() == region_norm)
        mask = seller_mask & region_mask
    return df.loc[mask].copy()


def render_mission_slice_images(
    df: pd.DataFrame,
    output_dir: Path,
    timestamp: str,
    specs: list[dict[str, str | None]] | None = None,
) -> list[dict[str, Any]]:
    """Рендерит отдельные mission screenshots по аптекам/городам."""
    if df.empty:
        return []
    output_dir.mkdir(parents=True, exist_ok=True)
    rendered: list[dict[str, Any]] = []
    for spec in specs or MISSION_SCREENSHOT_SPECS:
        seller = str(spec.get("seller") or "").strip()
        region = spec.get("region")
        label = str(spec.get("label") or seller or "Mission")
        slug = str(spec.get("slug") or re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_") or "mission")
        slice_df = _filter_mission_preview_slice(df, seller=seller, region=region)
        slice_df = _filter_alert_mission_rows(slice_df)
        if slice_df.empty:
            continue
        image_path = output_dir / f"mission_preview_{timestamp}__{slug}.png"
        render_mission_image(slice_df, image_path)
        caption = f"{label} — {len(slice_df)} SKU"
        rendered.append({"path": image_path, "caption": caption, "rows": len(slice_df), "label": label})
    return rendered


def resolve_kaspi_link(url: str, timeout: int = 8) -> str:
    """Разворачиваем короткие l.kaspi.kz ссылки в полноценный product URL."""
    if not url:
        return url
    try:
        parsed = urlparse(url if "://" in url else f"https://{url}")
        if parsed.netloc.lower() == "l.kaspi.kz" or parsed.path.startswith("/shp/"):
            target = url if "://" in url else f"https://{url}"
            resp = _requests_get(target, allow_redirects=True, timeout=timeout)
            if resp.url:
                return resp.url
        elif not parsed.scheme:
            # добавляем https для прямых kaspi.kz ссылок без схемы
            return f"https://{url}"
    except Exception:
        pass
    return url


def _url_with_city(url: str, city: str | None) -> str:
    """Гарантируем, что ссылка содержит нужный параметр города ?c=..."""
    if not city:
        return url
    code = CITY_CODES.get(city.strip())
    if not code:
        return url
    parsed = urlparse(url if "://" in url else f"https://{url}")
    query = parse_qs(parsed.query)
    query["c"] = [code]
    new_query = urlencode(query, doseq=True)
    return parsed._replace(query=new_query).geturl()


def parse_expected_sellers(raw_value) -> list[str]:
    """Парсим колонку expected_sellers из CSV (разделители ; или ,)."""
    if raw_value is None:
        return []
    try:
        if pd.isna(raw_value):
            return []
    except Exception:
        pass
    parts = re.split(r"[;,]", str(raw_value))
    return [part.strip() for part in parts if part and part.strip()]


def parse_price_value(raw_value) -> int | None:
    """Приводит цену вида "3 650" или "3\u00a0650" к целому."""
    if raw_value is None:
        return None
    try:
        if pd.isna(raw_value):
            return None
    except Exception:
        pass
    if isinstance(raw_value, (int, float)):
        try:
            return int(raw_value)
        except Exception:
            return None
    digits = re.sub(r"[^0-9]", "", str(raw_value))
    if not digits:
        return None
    try:
        return int(digits)
    except Exception:
        return None


def format_kzt(value: int | None) -> str:
    if value is None:
        return "—"
    return f"{value:,} ₸".replace(",", " ")


def _matched_records_for_meta(records: list[dict], meta: dict) -> list[dict]:
    """Фильтруем записи по городу и карточке."""
    product_code = meta.get("product_code")
    product_url = meta.get("product_url")
    city = meta.get("city")
    matched = []
    for rec in records:
        if city and rec.get("region") != city:
            continue
        if product_code and rec.get("product_code") == product_code:
            matched.append(rec)
            continue
        if product_url and rec.get("product_url") == product_url:
            matched.append(rec)
    return matched


def find_missing_expected_sellers(records: list[dict], job_meta_list: list[dict]) -> list[dict]:
    """Сравниваем найденных продавцов с ожидаемыми и собираем алерты."""
    alerts = []
    for meta in job_meta_list:
        expected = meta.get("expected_sellers") or []
        if not expected:
            continue
        expected_norm = {normalize_seller_name(name) for name in expected if name}
        if not expected_norm:
            continue
        matched_records = _matched_records_for_meta(records, meta)
        found_norm = {
            normalize_seller_name(rec.get("seller", ""))
            for rec in matched_records
            if rec.get("seller")
        }
        missing = [name for name in expected if normalize_seller_name(name) not in found_norm]
        if missing:
            found_raw = sorted({rec.get("seller", "") for rec in matched_records if rec.get("seller")})
            product_label = ""
            if matched_records:
                product_label = (
                    matched_records[0].get("product")
                    or matched_records[0].get("input_product")
                    or ""
                )
            if not product_label:
                product_label = meta.get("product_name") or meta.get("product_url")
            alerts.append({
                "city": meta.get("city"),
                "product": product_label,
                "product_code": meta.get("product_code") or "",
                "product_url": meta.get("product_url"),
                "missing": missing,
                "found": found_raw,
            })
    return alerts


def build_price_alerts(records: list[dict], job_meta_list: list[dict]) -> tuple[list[str], list[str]]:
    """Возвращает (нарушение контрактной цены, лучшие цены конкурентов)."""
    contract_alerts: list[str] = []
    better_price_alerts: list[str] = []
    for meta in job_meta_list:
        target_price = meta.get("expected_price_kzt")
        if not target_price:
            continue
        matched_records = _matched_records_for_meta(records, meta)
        if not matched_records:
            continue
        expected = meta.get("expected_sellers") or []
        expected_norm = {normalize_seller_name(name) for name in expected if name}
        product_label = meta.get("product_name") or matched_records[0].get("product") or matched_records[0].get("input_product")
        if not product_label:
            product_label = meta.get("product_url")

        contract_records = [
            rec for rec in matched_records
            if normalize_seller_name(rec.get("seller", "")) in expected_norm
        ] if expected_norm else []

        for rec in contract_records:
            price = rec.get("price_kzt")
            if price is None:
                continue
            if abs(price - target_price) > PRICE_TOLERANCE_KZT:
                contract_alerts.append(
                    f"💰 {product_label} — {meta.get('city')} — {rec.get('seller')}: ожидал {format_kzt(target_price)}, сейчас {format_kzt(price)}"
                )

        priced_records = [rec for rec in matched_records if rec.get("price_kzt") is not None]
        if not priced_records:
            continue
        best_offer = min(priced_records, key=lambda rec: rec.get("price_kzt") or 10**12)
        best_price = best_offer.get("price_kzt")
        if best_price is None:
            continue
        if best_price < target_price:
            diff = target_price - best_price
            better_price_alerts.append(
                f"🏷️ {product_label} — {meta.get('city')} — {best_offer.get('seller')} продаёт за {format_kzt(best_price)}"
                f" (цель {format_kzt(target_price)}, выгоднее на {format_kzt(diff)})"
            )

    return contract_alerts, better_price_alerts


def render_mission_image(df: pd.DataFrame, path: Path):
    if df.empty:
        return
    status_series = df.get("status", pd.Series(dtype=str)).fillna("").astype(str)
    unique_regions = sorted({_normalize_text(v) for v in df.get("region", pd.Series(dtype=str)).dropna().tolist() if _normalize_text(v)})
    unique_sellers = sorted({_normalize_text(v) for v in df.get("seller", pd.Series(dtype=str)).dropna().tolist() if _normalize_text(v)})

    base_specs = [
        ("region", "Город", 12, False, "left"),
        ("seller", "Аптека", 26, True, "left"),
        ("product", "Товар", 56, True, "left"),
        ("price_kzt", "Цель", 14, False, "right"),
        ("actual_price_kzt", "Наша цена", 14, False, "right"),
        ("status", "Статус", 28, True, "left"),
        ("best_price_kzt", "Рынок", 14, False, "right"),
        ("better_price_analysis", "Лучший оффер", 42, True, "left"),
        ("second_price_analysis", "След. оффер", 42, True, "left"),
        ("scraped_at", "Время", 18, False, "right"),
    ]
    specs = [
        spec
        for spec in base_specs
        if not (spec[0] == "region" and len(unique_regions) == 1)
        and not (spec[0] == "seller" and len(unique_sellers) == 1)
    ]

    base_dir = Path(__file__).resolve().parent
    regular_candidates = [
        base_dir / "fonts" / "DejaVuSans.ttf",
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    bold_candidates = [
        base_dir / "fonts" / "DejaVuSans-Bold.ttf",
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
    ]

    base_total_width = sum((width * 8) + 24 for _, _, width, _, _ in specs) + 92
    preview_scale = max(2.35, min(2.9, 3840 / max(1, base_total_width)))

    def S(value: float) -> int:
        return max(1, int(round(value * preview_scale)))

    def load_font(size: int, bold: bool = False):
        candidates = bold_candidates if bold else regular_candidates
        scaled_size = S(size)
        for font_path in candidates:
            if not font_path.exists():
                continue
            try:
                return ImageFont.truetype(str(font_path), scaled_size)
            except Exception as exc:
                print(f"⚠️ Could not load font {font_path}: {exc}")
        return ImageFont.load_default()

    def measure_line_height(font, extra: int = 0) -> int:
        bbox = font.getbbox("Ag")
        return max(S(14), bbox[3] - bbox[1] + S(extra))

    def measure_text_width(font, text: str) -> int:
        text = str(text or "")
        if not text:
            return 0
        try:
            return int(math.ceil(font.getlength(text)))
        except Exception:
            bbox = font.getbbox(text)
            return max(0, bbox[2] - bbox[0])

    def draw_shadowed_panel(draw_ctx, rect, radius: int, fill, shadow=(17, 24, 39, 22), outline=None):
        x1, y1, x2, y2 = rect
        shadow_rect = [x1 + S(6), y1 + S(12), x2 + S(6), y2 + S(12)]
        draw_ctx.rounded_rectangle(shadow_rect, radius=radius, fill=shadow)
        draw_ctx.rounded_rectangle(rect, radius=radius, fill=fill, outline=outline)

    def fmt_value(col: str, val):
        if pd.isna(val):
            return "—"
        if col in {"actual_price_kzt", "best_price_kzt", "price_kzt"}:
            return format_kzt(int(val))
        return str(val)

    title_font = load_font(30, bold=True)
    subtitle_font = load_font(15)
    eyebrow_font = load_font(12, bold=True)
    metric_value_font = load_font(22, bold=True)
    metric_label_font = load_font(12, bold=True)
    metric_note_font = load_font(11)
    table_header_font = load_font(13, bold=True)
    table_font = load_font(14)
    table_font_bold = load_font(14, bold=True)

    table_line_height = measure_line_height(table_font, extra=5)
    header_line_height = measure_line_height(table_header_font, extra=2)
    cell_padding_x = S(12)
    cell_padding_y = S(10)
    col_widths_px = [S((width * 8) + 24) for _, _, width, _, _ in specs]
    body_fonts = [table_font_bold if col == "product" else table_font for col, *_rest in specs]
    body_line_heights = [measure_line_height(font, extra=5) for font in body_fonts]
    wrap_cache: dict[tuple[str, int, str], list[str]] = {}

    def fit_text_ellipsis(text: str, font, max_width: int) -> str:
        text = str(text or "")
        if measure_text_width(font, text) <= max_width:
            return text
        ellipsis = "…"
        clipped = text
        while clipped and measure_text_width(font, clipped + ellipsis) > max_width:
            clipped = clipped[:-1]
        return (clipped + ellipsis) if clipped else ellipsis

    def wrap_text_by_width(text: str, font, max_width: int) -> list[str]:
        raw = str(text or "")
        if not raw:
            return [""]
        wrapped_lines: list[str] = []
        for paragraph in raw.splitlines() or [""]:
            paragraph = " ".join(paragraph.split())
            if not paragraph:
                wrapped_lines.append("")
                continue
            words = paragraph.split(" ")
            current = ""
            for word in words:
                candidate = word if not current else f"{current} {word}"
                if measure_text_width(font, candidate) <= max_width:
                    current = candidate
                    continue
                if current:
                    wrapped_lines.append(current)
                    current = ""
                if measure_text_width(font, word) <= max_width:
                    current = word
                    continue
                chunk = ""
                for char in word:
                    probe = chunk + char
                    if not chunk or measure_text_width(font, probe) <= max_width:
                        chunk = probe
                    else:
                        wrapped_lines.append(chunk)
                        chunk = char
                current = chunk
            if current:
                wrapped_lines.append(current)
        return wrapped_lines or [""]

    def prep_lines(value, col_width_px: int, wrap_enabled: bool, font, cache_key_hint: str):
        max_width = max(S(48), col_width_px - cell_padding_x * 2)
        text = str(value) if value not in (None, "") else ""
        if not wrap_enabled:
            return [fit_text_ellipsis(text, font, max_width)]
        cache_key = (text, max_width, cache_key_hint)
        if cache_key in wrap_cache:
            return wrap_cache[cache_key]
        wrapped = wrap_text_by_width(text, font, max_width)
        wrap_cache[cache_key] = wrapped
        return wrapped

    table_rows: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        cell_lines: list[list[str]] = []
        cell_heights: list[int] = []
        for col_idx, (col, _label, _width_chars, wrap_enabled, _align) in enumerate(specs):
            value = fmt_value(col, row.get(col))
            font_for_col = body_fonts[col_idx]
            lines = prep_lines(value, col_widths_px[col_idx], wrap_enabled, font_for_col, col)
            cell_lines.append(lines)
            cell_heights.append(len(lines) * body_line_heights[col_idx])
        table_rows.append(
            {
                "cells": cell_lines,
                "state": _mission_status_state(row.get("status")),
                "height": max(cell_heights) + cell_padding_y * 2,
            }
        )

    row_heights: list[int] = [header_line_height + cell_padding_y * 2]
    row_heights.extend(int(item["height"]) for item in table_rows)

    states = status_series.map(_mission_status_state)
    total_rows = int(len(df))
    attention_count = int((states == "attention").sum())
    alert_count = int((states == "alert").sum())
    problem_count = alert_count
    snapshot_time = ""
    if "scraped_at" in df.columns:
        scraped_values = [str(v).strip() for v in df["scraped_at"].dropna().tolist() if str(v).strip()]
        if scraped_values:
            snapshot_time = scraped_values[-1]
    if not snapshot_time:
        snapshot_time = datetime.now(ALMATY_TZ).strftime("%d.%m.%Y %H:%M")
    actual_prices = pd.to_numeric(df.get("actual_price_kzt"), errors="coerce")
    best_prices = pd.to_numeric(df.get("best_price_kzt"), errors="coerce")
    positive_gap = (actual_prices - best_prices).dropna()
    positive_gap = positive_gap[positive_gap > 0]
    avg_gap_text = format_kzt(int(round(float(positive_gap.mean())))) if not positive_gap.empty else "—"
    max_gap_text = format_kzt(int(round(float(positive_gap.max())))) if not positive_gap.empty else "—"
    missing_seller_count = int(status_series.map(_normalize_mission_status_text).eq("продавец отсутствует").sum())
    no_offer_count = int(status_series.map(_normalize_mission_status_text).eq("нет продавцов на карточке").sum())
    context_seller = unique_sellers[0] if len(unique_sellers) == 1 else f"{len(unique_sellers)} аптек"
    context_region = unique_regions[0] if len(unique_regions) == 1 else f"{len(unique_regions)} городов"
    hero_title = " • ".join(part for part in [context_seller, context_region] if part) or "Mission Issues"
    hero_subtitle = f"Mission snapshot • {snapshot_time}"

    outer_padding = S(34)
    hero_padding = S(28)
    hero_gap = S(18)
    metric_gap = S(14)
    metric_height = S(88)
    metric_specs = [
        {"label": "Проблемы", "value": str(problem_count), "note": "красные SKU", "tone": "alert"},
        {"label": "Внимание", "value": str(attention_count), "note": "не лучший оффер", "tone": "attention"},
        {"label": "Нет продавца", "value": str(missing_seller_count), "note": "seller missing", "tone": "alert"},
        {"label": "Нет офферов", "value": str(no_offer_count), "note": "no offers", "tone": "alert"},
        {"label": "Ср. gap", "value": avg_gap_text, "note": "к рынку", "tone": "neutral"},
        {"label": "Макс. gap", "value": max_gap_text, "note": "к рынку", "tone": "neutral"},
    ]
    table_width = sum(col_widths_px) + S(34)
    total_width = max(S(base_total_width), table_width + outer_padding * 2)
    metrics_total_width = total_width - outer_padding * 2 - hero_padding * 2
    metric_width = int((metrics_total_width - metric_gap * (len(metric_specs) - 1)) / len(metric_specs))
    hero_height = S(178) + metric_height
    table_header_height = row_heights[0]
    table_body_height = sum(row_heights[1:])
    table_container_padding = S(18)
    table_height = table_container_padding * 2 + table_header_height + table_body_height + max(0, len(table_rows) - 1) * S(2)
    total_height = outer_padding * 2 + hero_height + hero_gap + table_height

    img = Image.new("RGBA", (total_width, total_height), (245, 247, 251, 255))
    draw = ImageDraw.Draw(img)

    draw.rectangle([0, 0, total_width, total_height], fill=(245, 247, 251, 255))
    draw.ellipse([-S(220), -S(280), S(980), S(760)], fill=(214, 227, 245, 155))
    draw.ellipse([int(total_width * 0.7), -S(160), total_width + S(260), S(620)], fill=(227, 235, 247, 120))

    hero_rect = [outer_padding, outer_padding, total_width - outer_padding, outer_padding + hero_height]
    draw_shadowed_panel(draw, hero_rect, radius=S(30), fill=(252, 253, 255, 255), outline=(220, 228, 240, 255))
    hero_x1, hero_y1, hero_x2, hero_y2 = hero_rect
    draw.rounded_rectangle([hero_x1, hero_y1, hero_x2, hero_y1 + S(14)], radius=S(14), fill=(29, 78, 216, 255))

    eyebrow_y = hero_y1 + hero_padding
    draw.text((hero_x1 + hero_padding, eyebrow_y), "MISSION ISSUE SNAPSHOT", font=eyebrow_font, fill=(54, 93, 161, 255))
    title_y = eyebrow_y + measure_line_height(eyebrow_font, extra=10)
    draw.text((hero_x1 + hero_padding, title_y), hero_title, font=title_font, fill=(15, 23, 42, 255))
    subtitle_y = title_y + measure_line_height(title_font, extra=10)
    draw.text((hero_x1 + hero_padding, subtitle_y), hero_subtitle, font=subtitle_font, fill=(71, 85, 105, 255))

    only_alert_rows = total_rows > 0 and bool((states == "alert").all())
    pill_text = "Only red-problem SKU" if only_alert_rows else "Problems + attention"
    pill_width = measure_text_width(eyebrow_font, pill_text) + S(28)
    pill_rect = [hero_x2 - hero_padding - pill_width, eyebrow_y, hero_x2 - hero_padding, eyebrow_y + S(34)]
    draw.rounded_rectangle(pill_rect, radius=S(17), fill=(255, 239, 236, 255), outline=(240, 170, 160, 255))
    draw.text((pill_rect[0] + S(14), pill_rect[1] + S(8)), pill_text, font=eyebrow_font, fill=(127, 29, 29, 255))

    metric_palette = {
        "neutral": {"fill": (255, 255, 255, 255), "border": (206, 217, 232, 255), "accent": (71, 85, 105, 255)},
        "attention": {"fill": (255, 248, 235, 255), "border": (234, 179, 8, 255), "accent": (133, 77, 14, 255)},
        "alert": {"fill": (255, 239, 236, 255), "border": (239, 68, 68, 255), "accent": (153, 27, 27, 255)},
    }
    metrics_y = hero_y2 - hero_padding - metric_height
    metric_x = hero_x1 + hero_padding
    for metric in metric_specs:
        palette = metric_palette.get(metric["tone"], metric_palette["neutral"])
        card_rect = [metric_x, metrics_y, metric_x + metric_width, metrics_y + metric_height]
        draw.rounded_rectangle(card_rect, radius=S(20), fill=palette["fill"], outline=palette["border"], width=S(2))
        draw.rounded_rectangle([card_rect[0], card_rect[1], card_rect[2], card_rect[1] + S(8)], radius=S(8), fill=palette["border"])
        draw.text((card_rect[0] + S(16), card_rect[1] + S(18)), metric["label"], font=metric_label_font, fill=(15, 23, 42, 255))
        draw.text((card_rect[0] + S(16), card_rect[1] + S(40)), metric["value"], font=metric_value_font, fill=(0, 0, 0, 255))
        draw.text((card_rect[0] + S(16), card_rect[1] + S(68)), metric["note"], font=metric_note_font, fill=(71, 85, 105, 255))
        metric_x += metric_width + metric_gap

    table_rect = [outer_padding, hero_y2 + hero_gap, total_width - outer_padding, total_height - outer_padding]
    draw_shadowed_panel(draw, table_rect, radius=S(28), fill=(255, 255, 255, 255), outline=(222, 230, 242, 255))
    table_x1, table_y1, table_x2, table_y2 = table_rect
    draw.rounded_rectangle([table_x1, table_y1, table_x2, table_y1 + S(10)], radius=S(10), fill=(15, 23, 42, 255))

    header_rect = [
        table_x1 + table_container_padding,
        table_y1 + table_container_padding + S(8),
        table_x2 - table_container_padding,
        table_y1 + table_container_padding + S(8) + table_header_height,
    ]
    draw.rounded_rectangle(header_rect, radius=S(18), fill=(30, 41, 59, 255))

    header_x = header_rect[0] + S(2)
    for col_idx, (_col, label, _width_chars, _wrap_enabled, align) in enumerate(specs):
        col_width = col_widths_px[col_idx]
        label = fit_text_ellipsis(label, table_header_font, col_width - cell_padding_x * 2)
        text_width = measure_text_width(table_header_font, label)
        text_y = header_rect[1] + (table_header_height - header_line_height) / 2
        if align == "right":
            text_x = header_x + col_width - cell_padding_x - text_width
        else:
            text_x = header_x + cell_padding_x
        draw.text((text_x, text_y), label, font=table_header_font, fill=(248, 250, 252, 255))
        header_x += col_width

    row_palette = {
        "ok": {"fill": (236, 253, 245, 255), "accent": (34, 197, 94, 255), "text": (17, 24, 39, 255)},
        "attention": {"fill": (255, 248, 235, 255), "accent": (245, 158, 11, 255), "text": (17, 24, 39, 255)},
        "alert": {"fill": (255, 239, 236, 255), "accent": (239, 68, 68, 255), "text": (0, 0, 0, 255)},
        "neutral": {"fill": (255, 255, 255, 255), "accent": (203, 213, 225, 255), "text": (17, 24, 39, 255)},
    }
    row_y = header_rect[3] + S(10)
    for row_index, row in enumerate(table_rows):
        row_height = row_heights[row_index + 1]
        palette = row_palette.get(row["state"], row_palette["neutral"])
        row_rect = [header_rect[0], row_y, header_rect[2], row_y + row_height]
        draw.rounded_rectangle(row_rect, radius=S(14), fill=palette["fill"], outline=(232, 238, 245, 255))
        draw.rounded_rectangle([row_rect[0], row_rect[1], row_rect[0] + S(8), row_rect[3]], radius=S(8), fill=palette["accent"])

        cell_x = row_rect[0] + S(2)
        for col_idx, (col, _label, _width_chars, _wrap_enabled, align) in enumerate(specs):
            col_width = col_widths_px[col_idx]
            lines = row["cells"][col_idx]
            font_for_col = body_fonts[col_idx]
            effective_line_height = body_line_heights[col_idx]
            text_y = row_y + cell_padding_y
            for line in lines:
                text_width = measure_text_width(font_for_col, line)
                if align == "right":
                    text_x = cell_x + col_width - cell_padding_x - text_width
                else:
                    text_x = cell_x + cell_padding_x
                draw.text((text_x, text_y), line, font=font_for_col, fill=palette["text"])
                text_y += effective_line_height
            cell_x += col_width
        row_y += row_height + S(3)

    path.parent.mkdir(parents=True, exist_ok=True)
    img.save(path)


def _cleanup_old_previews(results_dir: Path, keep: int = 10):
    """Удаляем старые mission_preview_*.png, оставляя только последние N."""
    try:
        previews = sorted(
            results_dir.glob("mission_preview_*.png"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        for old in previews[keep:]:
            try:
                old.unlink()
            except Exception as exc:
                print(f"⚠️ Could not remove old preview {old}: {exc}")
    except Exception as exc:
        print(f"⚠️ Cleanup previews failed: {exc}")


def _combine_excel(existing_df: pd.DataFrame, new_df: pd.DataFrame) -> pd.DataFrame:
    has_existing_structure = len(existing_df.columns) > 0
    frames = [existing_df, new_df] if has_existing_structure else [new_df]
    combined_df = pd.concat(frames, ignore_index=True, sort=False)

    column_order = existing_df.columns.tolist() if has_existing_structure else new_df.columns.tolist()
    for col in new_df.columns:
        if col not in column_order:
            column_order.append(col)
    return combined_df.reindex(columns=column_order)


def _parse_positive_int_env(raw: str | None) -> int | None:
    """Возвращает положительное целое из строки/ENV или None."""
    if not raw:
        return None
    raw = str(raw).strip()
    if not raw:
        return None
    try:
        value = int(raw)
        return value if value > 0 else None
    except Exception:
        return None


def _bundle_safe_int(val: Any) -> int | None:
    try:
        if pd.isna(val):
            return None
        s = str(val).replace(" ", "")
        return int(float(s))
    except Exception:
        return None


def _is_missing_cell(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _bundle_cache_key(region: str, seller: str, bundle_url: str) -> str:
    return "||".join(
        [
            str(region or "").strip(),
            normalize_seller_name(str(seller or "").strip()),
            str(bundle_url or "").strip(),
        ]
    )


def _load_bundle_cache(path: Path = BUNDLE_CACHE_PATH) -> dict[str, dict[str, Any]]:
    try:
        if not path.exists():
            return {}
        payload = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(payload, dict):
            return {}
        rows = payload.get("rows")
        return rows if isinstance(rows, dict) else {}
    except Exception as exc:
        print(f"⚠️ Could not read bundle cache {path}: {exc}")
        return {}


def _save_bundle_cache(rows: dict[str, dict[str, Any]], path: Path = BUNDLE_CACHE_PATH) -> None:
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "version": BUNDLE_CACHE_VERSION,
            "updated_at": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
            "rows": rows,
        }
        tmp_path = path.with_suffix(path.suffix + ".tmp")
        tmp_path.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        tmp_path.replace(path)
    except Exception as exc:
        print(f"⚠️ Could not save bundle cache {path}: {exc}")


def _join_bundle_notes(*parts: Any) -> str:
    clean: list[str] = []
    for part in parts:
        text = str(part or "").strip()
        if text and text not in clean:
            clean.append(text)
    return " | ".join(clean)


def _bundle_cache_note(cache_entry: dict[str, Any] | None) -> str:
    if not cache_entry:
        return ""
    stamp = str(cache_entry.get("cached_at") or "").strip()
    if not stamp:
        return "Показаны последние успешные данные по 1 SKU"
    try:
        dt = datetime.fromisoformat(stamp)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ALMATY_TZ)
        else:
            dt = dt.astimezone(ALMATY_TZ)
        return f"Показаны последние успешные данные по 1 SKU от {dt.strftime('%d.%m %H:%M')}"
    except Exception:
        return "Показаны последние успешные данные по 1 SKU"


def _bundle_build_result_row(
    row: dict[str, Any],
    bundle_url: str,
    bundle_qty: int,
    bundle_price: int | None,
    base_price: int | None,
    best_price: int | None,
    best_seller: str | None,
    note: str = "",
) -> dict[str, Any]:
    bundle_unit = math.ceil(bundle_price / bundle_qty) if bundle_price is not None and bundle_qty else None
    same_sku_total = (base_price * bundle_qty) if base_price is not None and bundle_qty else None
    best_total = (best_price * bundle_qty) if best_price is not None and bundle_qty else None
    delta_single = bundle_unit - base_price if bundle_unit is not None and base_price is not None else None
    delta_best = bundle_price - best_total if bundle_price is not None and best_total is not None else None
    savings_same_total = same_sku_total - bundle_price if same_sku_total is not None and bundle_price is not None else None
    savings_best_total = best_total - bundle_price if best_total is not None and bundle_price is not None else None
    discount_pct = None
    try:
        if best_total and bundle_price is not None:
            discount_pct = round((best_total - bundle_price) / best_total * 100, 2)
    except Exception:
        discount_pct = None

    out_row = dict(row)
    out_row.update({
        "bundle_of": bundle_url or None,
        "bundle_qty": bundle_qty,
        "bundle_1_sku_price_kzt": base_price,
        "bundle_price_kzt": bundle_price,
        "bundle_unit_price_kzt": bundle_unit,
        "bundle_same_sku_total_kzt": same_sku_total,
        "bundle_savings_vs_same_sku_total_kzt": savings_same_total,
        "bundle_vs_single_kzt": delta_single,
        "bundle_best_seller": best_seller,
        "1_sku_best_price_kzt": best_price,
        "bundle_best_price_kzt": best_total,
        "bundle_savings_vs_best_total_kzt": savings_best_total,
        "bundle_vs_best_kzt": delta_best,
        "bundle_discount_pct": (f"{float(discount_pct):.2f}%") if discount_pct is not None else None,
        "bundle_note": note or "",
    })
    return out_row


def _bundle_fetch_base_price(
    page,
    url: str,
    city: str,
    seller_hint: str,
    exclude_sellers: list[str] | None = None,
) -> tuple[int | None, tuple[int | None, str | None], tuple[int | None, str | None]]:
    """Возвращает (цена продавца, лучший по всем, лучший по исключениям)."""
    recs = scrape_single_product(page, url, input_name="base", city=city) or []
    target_norm = normalize_seller_name(seller_hint)
    prices = []
    best_overall = []
    best_filtered = []
    for rec in recs:
        seller = normalize_seller_name(rec.get("seller", ""))
        price = _bundle_safe_int(rec.get("price_kzt"))
        if price:
            best_overall.append((price, rec.get("seller")))
            if not seller_in_list(rec.get("seller", ""), exclude_sellers or []):
                best_filtered.append((price, rec.get("seller")))
        if seller_matches(seller, target_norm) and price:
            prices.append(price)
    best_overall_pair = min(best_overall, key=lambda x: x[0]) if best_overall else (None, None)
    best_filtered_pair = min(best_filtered, key=lambda x: x[0]) if best_filtered else (None, None)
    return (min(prices) if prices else None), best_overall_pair, best_filtered_pair


def _build_bundle_dataframe_from_rows(
    rows: list[dict[str, Any]],
    limit_rows: int | None = None,
) -> pd.DataFrame | None:
    """Собирает данные для бандлов (1 SKU) из списка строк."""
    if not rows:
        return None
    if limit_rows and limit_rows > 0:
        rows = rows[:limit_rows]
    results: list[dict[str, Any]] = []
    bundle_cache = _load_bundle_cache()
    cache_updated = False

    exclude_raw = os.environ.get("MISSION_PARTNER_SELLERS")
    if exclude_raw is None:
        exclude_raw = os.environ.get("BUNDLE_EXCLUDE_SELLERS")
    exclude_sellers = parse_seller_list(exclude_raw, MISSION_PARTNER_SELLERS_DEFAULT)

    def _row_bundle_context(row: dict[str, Any]) -> tuple[str, int, int | None, str, str, str]:
        bundle_url_raw = row.get("bundle_of")
        bundle_url = "" if _is_missing_cell(bundle_url_raw) else str(bundle_url_raw).strip()
        bundle_qty = _bundle_safe_int(row.get("bundle_qty")) or 1
        if not bundle_url:
            fallback_url = str(row.get("product_url") or row.get("product_link") or "").strip()
            if fallback_url:
                bundle_url = fallback_url
                if _is_missing_cell(row.get("bundle_qty")):
                    bundle_qty = 1
        bundle_price = _bundle_safe_int(row.get("price_kzt"))
        seller_hint = str(row.get("seller") or "")
        city = str(row.get("region") or "Алматы")
        cache_key = _bundle_cache_key(city, seller_hint, bundle_url)
        return bundle_url, bundle_qty, bundle_price, seller_hint, city, cache_key

    def _append_row_from_cache(
        row: dict[str, Any],
        bundle_url: str,
        bundle_qty: int,
        bundle_price: int | None,
        cache_key: str,
        reason: str = "",
    ) -> None:
        cache_entry = bundle_cache.get(cache_key)
        base_price = _bundle_safe_int((cache_entry or {}).get("bundle_1_sku_price_kzt"))
        best_price = _bundle_safe_int((cache_entry or {}).get("1_sku_best_price_kzt"))
        best_seller = str((cache_entry or {}).get("bundle_best_seller") or "").strip() or None
        note = _join_bundle_notes(reason, _bundle_cache_note(cache_entry))
        results.append(
            _bundle_build_result_row(
                row=row,
                bundle_url=bundle_url,
                bundle_qty=bundle_qty,
                bundle_price=bundle_price,
                base_price=base_price,
                best_price=best_price,
                best_seller=best_seller,
                note=note,
            )
        )

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True, args=["--disable-dev-shm-usage"])
            context = browser.new_context(
                viewport={"width": 1600, "height": 900},
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
            )
            for row in rows:
                bundle_url, bundle_qty, bundle_price, seller_hint, city, cache_key = _row_bundle_context(row)
                if not bundle_url:
                    continue
                live_base_price = None
                live_best_price = None
                live_best_seller = None
                note = ""
                page = context.new_page()
                try:
                    live_base_price, best_overall, best_filtered = _bundle_fetch_base_price(
                        page,
                        bundle_url,
                        city,
                        seller_hint,
                        exclude_sellers=exclude_sellers,
                    )
                    best_pair = best_filtered if best_filtered[0] is not None else best_overall
                    live_best_price, live_best_seller = best_pair
                    if live_base_price is None and live_best_price is None:
                        note = "Базовый SKU не найден"
                except Exception as exc:
                    exc_text = str(exc)
                    if "Kaspi returned error page" in exc_text or "ERR_NAME_NOT_RESOLVED" in exc_text:
                        note = "Каспи не дает парсить 1 SKU"
                    else:
                        note = "Нет данных по 1 SKU"
                    print(f"⚠️ Bundle base error for {bundle_url} ({city}): {exc}")
                finally:
                    page.close()

                base_price = live_base_price
                best_price = live_best_price
                best_seller = live_best_seller
                cache_entry = bundle_cache.get(cache_key)
                cache_used = False
                if cache_entry:
                    cached_base = _bundle_safe_int(cache_entry.get("bundle_1_sku_price_kzt"))
                    cached_best = _bundle_safe_int(cache_entry.get("1_sku_best_price_kzt"))
                    cached_seller = str(cache_entry.get("bundle_best_seller") or "").strip() or None
                    if base_price is None and cached_base is not None:
                        base_price = cached_base
                        cache_used = True
                    if best_price is None and cached_best is not None:
                        best_price = cached_best
                        cache_used = True
                    if not best_seller and cached_seller:
                        best_seller = cached_seller
                        cache_used = True
                    if cache_used:
                        note = _join_bundle_notes(note, _bundle_cache_note(cache_entry))

                results.append(
                    _bundle_build_result_row(
                        row=row,
                        bundle_url=bundle_url,
                        bundle_qty=bundle_qty,
                        bundle_price=bundle_price,
                        base_price=base_price,
                        best_price=best_price,
                        best_seller=best_seller,
                        note=note,
                    )
                )

                if live_base_price is not None or live_best_price is not None or live_best_seller:
                    bundle_cache[cache_key] = {
                        "region": city,
                        "seller": seller_hint,
                        "bundle_of": bundle_url,
                        "bundle_1_sku_price_kzt": live_base_price,
                        "bundle_best_seller": live_best_seller,
                        "1_sku_best_price_kzt": live_best_price,
                        "cached_at": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
                    }
                    cache_updated = True
            context.close()
            browser.close()
    except Exception as exc:
        print(f"⚠️ Bundle session failed, falling back to cache: {exc}")
        for row in rows:
            bundle_url, bundle_qty, bundle_price, _seller_hint, _city, cache_key = _row_bundle_context(row)
            if not bundle_url:
                continue
            _append_row_from_cache(
                row=row,
                bundle_url=bundle_url,
                bundle_qty=bundle_qty,
                bundle_price=bundle_price,
                cache_key=cache_key,
                reason="Каспи временно не дал пересчитать 1 SKU",
            )

    if cache_updated:
        _save_bundle_cache(bundle_cache)

    if not results:
        return None

    out_df = pd.DataFrame(results)
    desired_order = [
        "region",
        "seller",
        "input_product",
        "product_url",
        "price_kzt",
        "bundle_of",
        "bundle_qty",
        "bundle_1_sku_price_kzt",
        "bundle_price_kzt",
        "bundle_unit_price_kzt",
        "bundle_same_sku_total_kzt",
        "bundle_savings_vs_same_sku_total_kzt",
        "bundle_vs_single_kzt",
        "bundle_best_seller",
        "1_sku_best_price_kzt",
        "bundle_best_price_kzt",
        "bundle_savings_vs_best_total_kzt",
        "bundle_vs_best_kzt",
        "bundle_discount_pct",
        "bundle_note",
    ]
    for col in desired_order:
        if col not in out_df.columns:
            out_df[col] = None
    out_df = out_df.reindex(columns=desired_order)
    return out_df


def _build_bundle_dataframe(csv_path: Path, limit_rows: int | None = None) -> pd.DataFrame | None:
    """Собирает данные для бандлов (1 SKU) и возвращает DataFrame для отдельного листа."""
    if not csv_path.exists():
        print(f"ℹ️ Bundle CSV не найден: {csv_path}")
        return None
    df = pd.read_csv(csv_path)
    if "bundle_of" not in df.columns and "bundle_qty" not in df.columns:
        print(f"ℹ️ Bundle CSV без колонок bundle_of / bundle_qty: {csv_path}")
        return None
    if limit_rows and limit_rows > 0:
        df = df.head(limit_rows)
    rows = df.to_dict(orient="records")
    return _build_bundle_dataframe_from_rows(rows, limit_rows=None)


def _attach_bundle_columns(df: pd.DataFrame, base_dir: Path) -> pd.DataFrame:
    """Добавляет bundle-колонки к миссии по ключу region+seller+product_url."""
    enabled_flag = (
        os.environ.get("MISSION_FEBRUARY_BUNDLE_ENABLED")
        or "1"
    ).strip().lower()
    if enabled_flag in {"0", "false", "no"}:
        return df

    limit_rows = _parse_positive_int_env(
        os.environ.get("MISSION_FEBRUARY_BUNDLE_LIMIT")
    )
    bundle_df = _build_bundle_dataframe_from_rows(
        df.to_dict(orient="records"),
        limit_rows=limit_rows,
    )
    if bundle_df is None or bundle_df.empty:
        return df

    join_cols = ["region", "seller", "product_url"]
    if not all(col in df.columns for col in join_cols):
        return df
    if not all(col in bundle_df.columns for col in join_cols):
        return df

    extra_cols = BUNDLE_COLUMNS

    base_columns = df.columns.tolist()
    base_df = df.copy()
    df = df.copy()
    bundle_df = bundle_df.copy()

    def _norm(val: Any) -> str:
        if val is None or pd.isna(val):
            return ""
        return str(val).strip()

    join_keys = [f"__join_{col}" for col in join_cols]
    for col, key in zip(join_cols, join_keys):
        df[key] = df[col].apply(_norm)
        bundle_df[key] = bundle_df[col].apply(_norm)

    available_extra_cols = [col for col in extra_cols if col in bundle_df.columns]
    merge_cols = available_extra_cols
    bundle_subset = bundle_df[join_keys + merge_cols].drop_duplicates(subset=join_keys)
    if merge_cols:
        rename_map = {col: f"__bundle_{col}" for col in merge_cols}
        bundle_subset = bundle_subset.rename(columns=rename_map)
    merged = df.merge(bundle_subset, how="left", on=join_keys)
    merged.drop(columns=join_keys, inplace=True, errors="ignore")

    for col in merge_cols:
        source_col = f"__bundle_{col}"
        if source_col in merged.columns:
            if col in merged.columns:
                merged[col] = merged[col].combine_first(merged[source_col])
            else:
                merged[col] = merged[source_col]
            merged.drop(columns=[source_col], inplace=True, errors="ignore")

    for col in extra_cols:
        if col in base_df.columns and col not in merged.columns:
            merged[col] = base_df[col]
        elif col not in merged.columns:
            merged[col] = None
    merged = merged.reindex(columns=base_columns + [col for col in extra_cols if col not in base_columns])
    return merged


def _ensure_bundle_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for col in BUNDLE_COLUMNS:
        if col not in df.columns:
            df[col] = None
    return df


def _reorder_mission_columns(df: pd.DataFrame) -> pd.DataFrame:
    desired = [
        "region",
        "seller",
        "input_product",
        "product",
        "product_url",
        "review_count",
        "review_progress",
        "price_kzt",
        "actual_price_kzt",
        "status",
        "best_price_kzt",
        "better_price_analysis",
        "second_price_analysis",
        "scraped_at",
    ] + BUNDLE_COLUMNS
    existing = df.columns.tolist()
    ordered = [col for col in desired if col in existing]
    tail = [col for col in existing if col not in ordered]
    return df.reindex(columns=ordered + tail)


def _write_dataframe_table(
    df: pd.DataFrame,
    path: Path,
    table_name: str = "KaspiTable",
    sheet_name: str = "Sheet1",
    highlight_rows: list[int] | None = None,
    highlight_columns: list[str] | None = None,
    extra_sheets: dict[str, pd.DataFrame] | None = None,
):
    safe_name = re.sub(r"[^A-Za-z0-9_]", "_", table_name) or "KaspiTable"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        if df.shape[1] == 0:
            return
        max_row = df.shape[0] + 1
        max_col = df.shape[1]
        table_ref = f"A1:{get_column_letter(max_col)}{max_row}"
        table = Table(displayName=safe_name[:31], ref=table_ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        if extra_sheets:
            for sheet, extra_df in extra_sheets.items():
                sheet_title = (sheet or "Sheet2")[:31]
                extra_df.to_excel(writer, index=False, sheet_name=sheet_title)

    wb = load_workbook(path)
    ws = wb[sheet_name]

    red_fill = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")

    cols_to_highlight: list[int] = []
    if highlight_columns:
        for col_name in highlight_columns:
            if col_name in df.columns:
                cols_to_highlight.append(df.columns.get_loc(col_name) + 1)
    else:
        for col_name in ("price_kzt", "price_match_status", "price_delta_kzt", "status"):
            if col_name in df.columns:
                cols_to_highlight.append(df.columns.get_loc(col_name) + 1)

    attention_rows: list[int] = []
    if "status" in df.columns:
        attention_mask = df["status"].fillna("").str.contains(
            r"внимание|не самое выгодное|не лучшее",
            case=False,
            regex=True,
        )
        attention_rows = df.index[attention_mask].tolist()

    attention_row_set = set(attention_rows)

    if highlight_rows:
        for row_idx in highlight_rows:
            if row_idx in attention_row_set:
                continue
            excel_row = row_idx + 2
            for col_idx in cols_to_highlight:
                ws.cell(row=excel_row, column=col_idx).fill = red_fill

    if attention_rows:
        for row_idx in attention_rows:
            excel_row = row_idx + 2
            for col_idx in cols_to_highlight:
                ws.cell(row=excel_row, column=col_idx).fill = warn_fill

    if "status" in df.columns:
        status_col_idx = df.columns.get_loc("status") + 1
        ok_rows = df.index[df["status"].fillna("").str.lower().str.startswith("ok")].tolist()
        for row_idx in ok_rows:
            excel_row = row_idx + 2
            ws.cell(row=excel_row, column=status_col_idx).fill = green_fill
            for col_idx in cols_to_highlight:
                if col_idx == status_col_idx:
                    continue
                ws.cell(row=excel_row, column=col_idx).fill = green_fill

    wb.save(path)


def _snapshot_key(record: dict) -> str:
    product_code = record.get("product_code") or extract_product_code(record.get("product_url", ""))
    return f"{record.get('region','')}|{product_code or record.get('product_url','')}|{normalize_seller_name(record.get('seller',''))}"


def load_snapshot(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        mtime = path.stat().st_mtime
        age_hours = (datetime.now(ALMATY_TZ).timestamp() - mtime) / 3600
        if age_hours > SNAPSHOT_TTL_HOURS:
            return {}
        with open(path, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return {}


class KaspiAnalysisSession:
    """Shared Playwright session for product analysis/search."""

    def __init__(self, headless: bool = True):
        self._headless = headless
        self._viewport = {"width": 1600, "height": 900}
        self._playwright = None
        self._browser = None
        self._context = None

    def __enter__(self):
        self._ensure_context()
        return self

    def __exit__(self, exc_type, exc, tb):
        self.close()

    def _ensure_context(self):
        if self._context:
            return
        pw = sync_playwright().start()
        browser = pw.chromium.launch(headless=self._headless)
        context = browser.new_context(viewport=self._viewport)
        self._playwright = pw
        self._browser = browser
        self._context = context

    def close(self):
        if self._context:
            try:
                self._context.close()
            except Exception:
                pass
            self._context = None
        if self._browser:
            try:
                self._browser.close()
            except Exception:
                pass
            self._browser = None
        if self._playwright:
            try:
                self._playwright.stop()
            except Exception:
                pass
            self._playwright = None

    def analyze_product(self, url: str, city: str | None = None) -> dict:
        self._ensure_context()
        page = self._context.new_page()
        try:
            return _collect_product_analysis(page, url, city=city)
        finally:
            try:
                page.close()
            except Exception:
                pass

    def search_product(self, query: str, required_terms: list[str] | None = None, max_results: int = 12) -> dict | None:
        self._ensure_context()
        page = self._context.new_page()
        try:
            return _search_product_link(page, query, required_terms=required_terms, max_results=max_results)
        finally:
            try:
                page.close()
            except Exception:
                pass

    def search_keyword_results(self, search_url: str, max_results: int = 20) -> list[dict]:
        self._ensure_context()
        page = self._context.new_page()
        try:
            return _collect_keyword_search_results(page, search_url, max_results=max_results)
        finally:
            try:
                page.close()
            except Exception:
                pass


def _collect_product_analysis(page, url: str, city: str | None = None) -> dict:
    resolved_url = _url_with_city(resolve_kaspi_link(url), city)
    try:
        page.set_default_timeout(SLOW_NAV_TIMEOUT)
        page.goto(resolved_url, wait_until="domcontentloaded", timeout=SLOW_NAV_TIMEOUT)
        page.wait_for_timeout(POST_LOAD_DELAY_FAST)
    except TimeoutError:
        page.set_default_timeout(SLOW_NAV_TIMEOUT)
        page.goto(resolved_url, timeout=SLOW_NAV_TIMEOUT)
        page.wait_for_timeout(POST_LOAD_DELAY_SLOW)

    if city:
        try:
            page.wait_for_selector("text=Выберите ваш город", timeout=5000)
            page.click(f"a:has-text('{city}')")
            page.wait_for_timeout(1000)
        except Exception:
            pass

    try:
        product_meta = page.evaluate("() => window.digitalData?.product || {}") or {}
    except Exception:
        product_meta = {}
    try:
        breadcrumbs = page.evaluate("() => (window.BACKEND?.components?.item?.breadcrumbs) || []") or []
    except Exception:
        breadcrumbs = []

    product_id = product_meta.get("id") or extract_product_code(resolved_url)
    review_payload = None
    if product_id:
        try:
            review_payload = page.evaluate(
                """async (pid) => {
                    try {
                        const resp = await fetch(`https://kaspi.kz/yml/review-view/api/v1/reviews/product/${pid}?withAgg=true`, {credentials: 'include'});
                        if (!resp.ok) {
                            return null;
                        }
                        return await resp.json();
                    } catch (err) {
                        return null;
                    }
                }""",
                product_id,
            )
        except Exception:
            review_payload = None

    summary = (review_payload or {}).get("summary") or {}
    statistic = summary.get("statistic") or []
    review_count = sum(int(item.get("count", 0) or 0) for item in statistic) or int(product_meta.get("reviewCount") or 0)
    rating_value = summary.get("global")
    reviews = []
    for entry in (review_payload or {}).get("data", [])[:5]:
        comment = entry.get("comment") or {}
        reviews.append({
            "author": entry.get("author"),
            "rating": entry.get("rating"),
            "text": (comment.get("text") or "").strip(),
        })

    category_titles = [crumb.get("title") for crumb in breadcrumbs if crumb.get("title")]
    if not category_titles and product_meta.get("category"):
        category_titles = product_meta["category"]
    category_path = " → ".join(category_titles) if category_titles else ""

    price = product_meta.get("unitSalePrice") or product_meta.get("unitPrice")
    try:
        price = int(price)
    except Exception:
        price = None

    return {
        "id": product_id,
        "name": product_meta.get("name") or "",
        "product_url": resolved_url,
        "price_kzt": price,
        "currency": product_meta.get("currency") or "KZT",
        "category_path": category_path,
        "rating": rating_value,
        "review_count": review_count,
        "reviews": reviews,
        "review_summary": summary,
        "breadcrumbs": breadcrumbs,
        "thumbnail": product_meta.get("thumbnailUrl") or (product_meta.get("primaryImage") or {}).get("large"),
    }


def _search_product_link(page, query: str, required_terms: list[str] | None = None, max_results: int = 12) -> dict | None:
    normalized_query = (query or "").strip()
    if not normalized_query:
        return None
    search_url = f"https://kaspi.kz/shop/search/?text={quote_plus(normalized_query)}"
    try:
        page.set_default_timeout(SLOW_NAV_TIMEOUT)
        page.goto(search_url, timeout=SLOW_NAV_TIMEOUT)
        page.wait_for_timeout(POST_LOAD_DELAY_FAST)
    except TimeoutError:
        try:
            page.set_default_timeout(SLOW_NAV_TIMEOUT)
            page.goto(search_url, timeout=SLOW_NAV_TIMEOUT)
            page.wait_for_timeout(POST_LOAD_DELAY_SLOW)
        except TimeoutError:
            return None

    try:
        page.wait_for_selector("[data-product-id]", timeout=SELLER_WAIT_TIMEOUT)
    except TimeoutError:
        return None

    cards = page.locator("[data-product-id]")
    total = min(cards.count(), max_results)
    required = [term.lower() for term in (required_terms or []) if term]
    fallback = None
    for idx in range(total):
        card = cards.nth(idx)
        try:
            name = card.locator(".item-card__name").inner_text().strip()
        except Exception:
            continue
        href = None
        try:
            for link in card.locator("a").all():
                link_href = link.get_attribute("href")
                if link_href and "/shop/p/" in link_href:
                    href = link_href
                    break
        except Exception:
            href = None
        if not href:
            continue
        record = {
            "name": name,
            "url": urljoin("https://kaspi.kz", href),
        }
        if not required or all(term in name.lower() for term in required):
            return record
        if not fallback:
            fallback = record
    return fallback


def _collect_keyword_search_results(page, search_url: str, max_results: int = 20) -> list[dict]:
    target_url = (search_url or "").strip()
    if not target_url:
        return []
    try:
        page.set_default_timeout(SLOW_NAV_TIMEOUT)
        page.goto(target_url, timeout=SLOW_NAV_TIMEOUT)
        page.wait_for_timeout(POST_LOAD_DELAY_FAST)
    except TimeoutError:
        try:
            page.set_default_timeout(SLOW_NAV_TIMEOUT)
            page.goto(target_url, timeout=SLOW_NAV_TIMEOUT)
            page.wait_for_timeout(POST_LOAD_DELAY_SLOW)
        except TimeoutError:
            return []

    try:
        page.wait_for_selector("[data-product-id]", timeout=SELLER_WAIT_TIMEOUT)
    except TimeoutError:
        return []

    cards = page.locator("[data-product-id]")
    total = min(cards.count(), max_results)
    results: list[dict] = []
    for idx in range(total):
        card = cards.nth(idx)
        try:
            name = card.locator(".item-card__name").inner_text().strip()
        except Exception:
            continue
        href = None
        try:
            for link in card.locator("a").all():
                link_href = link.get_attribute("href")
                if link_href and "/shop/p/" in link_href:
                    href = link_href
                    break
        except Exception:
            href = None
        if not href:
            continue
        results.append({
            "position": idx + 1,
            "product_name": name,
            "product_url": urljoin("https://kaspi.kz", href),
        })
    return results


def _clean_cell_text(value) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def _load_keyword_queries(base_dir: Path) -> list[dict]:
    raw_path = os.environ.get("SEARCH_KEYWORDS_CSV", "search_keywords.csv")
    path = Path(raw_path)
    if not path.is_absolute():
        path = base_dir / path
    if not path.exists():
        return []
    try:
        df = pd.read_csv(path)
    except Exception as exc:
        print(f"Keyword CSV read failed: {exc}")
        return []
    if df.empty:
        return []
    columns = {str(col).strip().lower(): col for col in df.columns}
    keyword_col = columns.get("keyword") or columns.get("query")
    url_col = columns.get("search_url") or columns.get("url")
    if not keyword_col and df.columns.size:
        keyword_col = df.columns[0]
    rows: list[dict] = []
    for _, row in df.iterrows():
        keyword = _clean_cell_text(row.get(keyword_col) if keyword_col else "")
        search_url = _clean_cell_text(row.get(url_col) if url_col else "")
        if not keyword and not search_url:
            continue
        rows.append({"keyword": keyword, "search_url": search_url})
    return rows


def _build_keyword_search_url(keyword: str, search_url: str | None) -> str | None:
    if search_url:
        url = search_url.strip()
        if "{query}" in url:
            if not keyword:
                return None
            url = url.replace("{query}", quote_plus(keyword))
        if "{keyword}" in url:
            if not keyword:
                return None
            url = url.replace("{keyword}", quote_plus(keyword))
        return url
    if not keyword:
        return None
    extra_query = (os.environ.get("SEARCH_EXTRA_QUERY") or "").strip().lstrip("?&")
    base = f"https://kaspi.kz/shop/search/?text={quote_plus(keyword)}"
    if extra_query:
        base = f"{base}&{extra_query}"
    return base


def _extract_product_codes_from_df(products_df: pd.DataFrame) -> set[str]:
    codes: set[str] = set()
    for col in ("product_link", "product_url"):
        if col not in products_df.columns:
            continue
        for raw in products_df[col].dropna():
            code = extract_product_code(str(raw))
            if code:
                codes.add(code)
    return codes


def _build_keyword_search_report(
    base_dir: Path,
    results_dir: Path,
    timestamp: str,
    own_codes: set[str],
    chat_id: str | None = None,
) -> Path | None:
    keyword_rows = _load_keyword_queries(base_dir)
    if not keyword_rows:
        return None
    max_results = int(os.environ.get("SEARCH_MAX_RESULTS", "20"))
    rows: list[dict] = []
    with KaspiAnalysisSession(headless=True) as session:
        for entry in keyword_rows:
            keyword = entry.get("keyword", "")
            search_url = _build_keyword_search_url(keyword, entry.get("search_url"))
            if not search_url:
                continue
            results = session.search_keyword_results(search_url, max_results=max_results)
            for item in results:
                product_url = item.get("product_url") or ""
                product_code = extract_product_code(product_url)
                is_own = bool(product_code and product_code in own_codes)
                rows.append({
                    "keyword": keyword,
                    "search_url": search_url,
                    "position": item.get("position"),
                    "product_name": item.get("product_name"),
                    "product_url": product_url,
                    "product_code": product_code or "",
                    "status": "OK" if is_own else "WRONG",
                })
    if not rows:
        return None
    df = pd.DataFrame(rows, columns=[
        "keyword",
        "search_url",
        "position",
        "product_name",
        "product_url",
        "product_code",
        "status",
    ])
    wrong_rows = df.index[df["status"] != "OK"].tolist()
    output_path = results_dir / f"kaspi_search_{timestamp}.xlsx"
    _write_dataframe_table(
        df,
        output_path,
        table_name="kaspi_search",
        highlight_rows=wrong_rows,
        highlight_columns=["status", "product_name", "product_code"],
    )
    print(f"Keyword search file saved: {output_path.name} in {results_dir}")
    try:
        send_telegram_file(output_path, caption="Kaspi keyword search", chat_id=chat_id)
    except Exception as exc:
        print(f"Could not send keyword search file: {exc}")
    return output_path


def save_snapshot(path: Path, snapshot: dict):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(snapshot, fh, ensure_ascii=False, indent=2)


def _graph_upload(df: pd.DataFrame) -> bool:
    """Upload via Microsoft Graph if credentials are provided."""
    if not msal:
        return False

    client_id = os.environ.get("GRAPH_CLIENT_ID") or os.environ.get("SHAREPOINT_CLIENT_ID")
    client_secret = os.environ.get("GRAPH_CLIENT_SECRET") or os.environ.get("SHAREPOINT_CLIENT_SECRET")
    tenant_id = os.environ.get("GRAPH_TENANT_ID") or os.environ.get("SHAREPOINT_TENANT_ID") or DEFAULT_TENANT_ID
    upload_path = os.environ.get("GRAPH_UPLOAD_PATH") or DEFAULT_GRAPH_UPLOAD_PATH
    target_user = (
        os.environ.get("GRAPH_USER_ID")
        or os.environ.get("GRAPH_USER_UPN")
        or os.environ.get("SHAREPOINT_USERNAME")
        or DEFAULT_SHAREPOINT_USERNAME
    )

    # Если нет client credentials, пробуем resource owner password (устарело, может блокироваться при MFA)
    username = os.environ.get("GRAPH_USERNAME") or os.environ.get("SHAREPOINT_USERNAME")
    password = os.environ.get("GRAPH_PASSWORD") or os.environ.get("SHAREPOINT_PASSWORD")

    scopes = ["https://graph.microsoft.com/.default"]
    app = None
    token = None

    if client_id and client_secret:
        try:
            authority = f"https://login.microsoftonline.com/{tenant_id}"
            app = msal.ConfidentialClientApplication(
                client_id=client_id,
                authority=authority,
                client_credential=client_secret,
            )
            token = app.acquire_token_silent(scopes, account=None)
            if not token:
                token = app.acquire_token_for_client(scopes=scopes)
            if "access_token" not in token:
                raise RuntimeError(token.get("error_description") or "No access token")
            print("🔐 Using Graph client_credentials flow")
        except Exception as exc:
            print(f"⚠️ Graph client_credentials failed: {exc}")

    if token is None and client_id and username and password:
        try:
            authority = f"https://login.microsoftonline.com/{tenant_id}"
            app = msal.PublicClientApplication(client_id=client_id, authority=authority)
            token = app.acquire_token_by_username_password(
                username=username,
                password=password,
                scopes=["Files.ReadWrite.All", "offline_access"],
            )
            if "access_token" not in token:
                raise RuntimeError(token.get("error_description") or "No access token")
            print("🔐 Using Graph username/password flow")
        except Exception as exc:
            print(f"⚠️ Graph username/password failed: {exc}")

    if not token or "access_token" not in token:
        return False

    headers = {"Authorization": f"Bearer {token['access_token']}"}
    download_url = f"https://graph.microsoft.com/v1.0/users/{target_user}/drive/root:/{upload_path}:/content"

    existing_df = pd.DataFrame()
    try:
        dl_resp = requests.get(download_url, headers=headers)
        if dl_resp.status_code == 200:
            with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp_dl:
                tmp_dl.write(dl_resp.content)
                tmp_dl.flush()
                existing_df = pd.read_excel(tmp_dl.name)
        elif dl_resp.status_code != 404:
            print(f"⚠️ Graph download failed: HTTP {dl_resp.status_code} {dl_resp.text}")
    except Exception as exc:
        print(f"⚠️ Graph download error: {exc}")

    combined_df = _combine_excel(existing_df, df)

    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp_out:
            combined_df.to_excel(tmp_out.name, index=False)
            tmp_out.flush()
            with open(tmp_out.name, "rb") as fh:
                put_resp = requests.put(download_url, headers=headers, data=fh.read())
        if 200 <= put_resp.status_code < 300:
            print("✅ Uploaded rows to SharePoint via Graph")
            return True
        print(f"⚠️ Graph upload failed: HTTP {put_resp.status_code} {put_resp.text}")
    except Exception as exc:
        print(f"⚠️ Graph upload error: {exc}")

    return False


def send_telegram_message(text: str, chat_id: str | None = None):
    """Send a plain-text message to Telegram. Fails hard if not configured."""
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = chat_id or os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        raise RuntimeError("TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID must be set for Telegram alerts")
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    chunks = _split_telegram_text(text, TELEGRAM_SAFE_CHUNK)
    if not chunks:
        return

    def _send_chunk(chunk: str, allow_resplit: bool = True):
        resp = requests.post(url, data={"chat_id": chat_id, "text": chunk})
        if resp.status_code == 200:
            return
        if resp.status_code == 413 and allow_resplit and len(chunk) > 1200:
            smaller_chunks = _split_telegram_text(chunk, max_len=1200)
            for sub in smaller_chunks:
                _send_chunk(sub, allow_resplit=False)
            return
        raise RuntimeError(f"Telegram send failed: {resp.status_code} {resp.text}")

    try:
        for idx, chunk in enumerate(chunks, start=1):
            _send_chunk(chunk, allow_resplit=True)
    except Exception as exc:
        raise RuntimeError(f"Telegram send error: {exc}") from exc


def send_results_email(
    file_path: Path,
    subject: str,
    body: str,
    recipients: Sequence[str] | None = None,
):
    host = os.environ.get("EMAIL_SMTP_HOST")
    username = os.environ.get("EMAIL_SMTP_USER")
    password = os.environ.get("EMAIL_SMTP_PASSWORD")
    from_addr = os.environ.get("EMAIL_FROM") or username
    port = int(os.environ.get("EMAIL_SMTP_PORT", "587"))
    if not recipients:
        base = os.environ.get("EMAIL_DEFAULT_RECIPIENTS", "")
        recipients = [addr.strip() for addr in base.split(",") if addr.strip()]
    if not recipients:
        recipients = ["data.receiver.eurasia@stada.com"]
    if not (host and username and password and from_addr):
        print("ℹ️ Email credentials not configured, skipping email send")
        return False
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = ", ".join(recipients)
    msg.set_content(body or "")
    try:
        with open(file_path, "rb") as fh:
            data = fh.read()
        msg.add_attachment(
            data,
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=file_path.name,
        )
    except Exception as exc:
        print(f"⚠️ Failed to attach file for email: {exc}")
        return False
    try:
        with smtplib.SMTP(host, port, timeout=20) as server:
            server.starttls()
            server.login(username, password)
            server.send_message(msg)
        print("📧 Sent results email")
        return True
    except Exception as exc:
        print(f"⚠️ Email send failed: {exc}")
        return False


def send_telegram_file(file_path: Path, caption: str | None = None, chat_id: str | None = None):
    """Send a file to Telegram chat. Fails hard if not configured."""
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = chat_id or os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        raise RuntimeError("TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID must be set for Telegram alerts")
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    try:
        with open(file_path, "rb") as fh:
            files = {"document": fh}
            data = {"chat_id": chat_id}
            if caption:
                data["caption"] = caption
            resp = requests.post(url, data=data, files=files, timeout=60)
        if resp.status_code != 200:
            raise RuntimeError(f"Telegram file send failed: {resp.status_code} {resp.text}")
    except Exception as exc:
        raise RuntimeError(f"Telegram file send error: {exc}") from exc


def send_telegram_photo(photo_path: Path, caption: str | None = None, chat_id: str | None = None):
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = chat_id or os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        raise RuntimeError("TELEGRAM_BOT_TOKEN and TELEGRAM_CHAT_ID must be set for Telegram alerts")
    url = f"https://api.telegram.org/bot{token}/sendPhoto"
    try:
        with open(photo_path, "rb") as fh:
            files = {"photo": fh}
            data = {"chat_id": chat_id}
            if caption:
                data["caption"] = caption
            resp = requests.post(url, data=data, files=files, timeout=60)
        if resp.status_code != 200:
            raise RuntimeError(f"Telegram photo send failed: {resp.status_code} {resp.text}")
    except Exception as exc:
        raise RuntimeError(f"Telegram photo send error: {exc}") from exc


def append_results_to_sharepoint(df: pd.DataFrame):
    """Append scraped rows to the configured SharePoint workbook."""
    site_url = os.environ.get("SHAREPOINT_SITE_URL") or DEFAULT_SHAREPOINT_SITE_URL
    file_url = os.environ.get("SHAREPOINT_FILE_RELATIVE_URL") or DEFAULT_SHAREPOINT_FILE_URL
    username = os.environ.get("SHAREPOINT_USERNAME") or DEFAULT_SHAREPOINT_USERNAME
    password = os.environ.get("SHAREPOINT_PASSWORD") or DEFAULT_SHAREPOINT_PASSWORD
    client_id = os.environ.get("SHAREPOINT_CLIENT_ID")
    client_secret = os.environ.get("SHAREPOINT_CLIENT_SECRET")
    tenant_id = os.environ.get("SHAREPOINT_TENANT_ID") or DEFAULT_TENANT_ID

    # 0) Graph API (client credentials или username/password) — самая надёжная автоматизация при наличии app registration
    if _graph_upload(df):
        return

    if not ClientContext or not File:
        print("ℹ️ Install 'office365-rest-python-client' to enable SharePoint uploads")
        return

    ctx = None

    # 1) client credentials (если заданы id/secret)
    if client_id and client_secret and ClientCredential:
        try:
            creds = ClientCredential(client_id, client_secret)
            ctx = ClientContext(site_url).with_credentials(creds)
            print("🔐 Using SharePoint client credentials flow")
        except Exception as exc:
            print(f"⚠️ Client credentials auth failed: {exc}")

    # 2) user/password (fallback)
    if ctx is None and username and password and UserCredential:
        try:
            creds = UserCredential(username, password)
            ctx = ClientContext(site_url).with_credentials(creds)
            print("🔐 Using SharePoint user/password auth")
        except Exception as exc:
            print(f"⚠️ User/password auth failed: {exc}")

    if ctx is None:
        print("ℹ️ No valid SharePoint auth configured; skipping cloud upload")
        return

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir) / "kaspi_upload.xlsx"
        existing_df = pd.DataFrame()
        try:
            response = File.open_binary(ctx, file_url)
            with open(tmp_path, "wb") as download_file:
                download_file.write(response.content)
            existing_df = pd.read_excel(tmp_path)
        except Exception as exc:
            print(f"ℹ️ Could not download existing workbook ({exc}); creating a new file")

        has_existing_structure = len(existing_df.columns) > 0
        frames = [existing_df, df] if has_existing_structure else [df]
        combined_df = pd.concat(frames, ignore_index=True, sort=False)

        column_order = existing_df.columns.tolist() if has_existing_structure else df.columns.tolist()
        for col in df.columns:
            if col not in column_order:
                column_order.append(col)
        combined_df = combined_df.reindex(columns=column_order)

        combined_df.to_excel(tmp_path, index=False)
        try:
            with open(tmp_path, "rb") as upload_file:
                File.save_binary(ctx, file_url, upload_file.read())
            print("✅ Uploaded rows to SharePoint workbook")
        except Exception as exc:
            print(f"⚠️ Failed to upload to SharePoint: {exc}")
            return


def scrape_single_product(page, url: str, input_name: str | None = None, city: str = "Алматы", scraped_date: str | None = None):
    resolved_url = _url_with_city(resolve_kaspi_link(url), city)
    requested_name = (input_name or "").strip()
    product_name = None
    # первая попытка: быстрый заход
    try:
        page.set_default_timeout(FAST_NAV_TIMEOUT)  # короче таймауты для скорости
        _jitter_sleep(0.1, 0.4)
        page.goto(resolved_url, wait_until="domcontentloaded", timeout=FAST_NAV_TIMEOUT)
        _wait_after_navigation(page, POST_LOAD_DELAY_FAST)
    except TimeoutError:
        print(f"⏱️ First attempt timed out for {input_name or url}, retrying with relaxed settings...")
        try:
            page.set_default_timeout(SLOW_NAV_TIMEOUT)
            page.goto(resolved_url, timeout=SLOW_NAV_TIMEOUT)  # без wait_until
            _wait_after_navigation(page, POST_LOAD_DELAY_SLOW)
        except TimeoutError:
            raise RuntimeError(f"Page.goto failed twice for {input_name or url}")

    page_product_name_selectors = [
        "h1[data-product-name]",
        "h1[itemprop='name']",
        "[data-test='productName']",
        "h1",
    ]
    for selector in page_product_name_selectors:
        locator = page.locator(selector).first
        if locator.count():
            try:
                text = locator.inner_text().strip()
                if text:
                    product_name = text
            except Exception:
                pass
            if product_name:
                break
    if not product_name:
        try:
            product_name = page.title().strip()
        except Exception:
            product_name = ""
    # если с карточки не забрали название, используем исходное имя
    if not product_name:
        product_name = requested_name
    if product_name and "что-то пошло не так" in product_name.lower():
        raise RuntimeError(f"Kaspi returned error page for {input_name or url}")

    product_code = extract_product_code(resolved_url or url)
    if not product_code:
        meta_sku = page.locator("meta[itemprop='sku'], meta[name='sku']").first
        if meta_sku.count():
            product_code = meta_sku.get_attribute("content") or ""

    try:
        page.wait_for_selector("text=Выберите ваш город", timeout=5000)
        page.click(f"a:has-text('{city}')")
        page.wait_for_timeout(800)
        _wait_for_overlay(page)
    except Exception:
        pass  # city modal did not appear

    review_count_value = _extract_review_display_count(page)

    tab_selectors = [
        "li[data-tab='offers']",
        "[data-test='tabOffers']",
        "text=Продавцы",
    ]

    def _find_sellers_tab(wait_timeout: int) -> Any:
        for tab_selector in tab_selectors:
            try:
                page.wait_for_selector(tab_selector, timeout=wait_timeout)
            except TimeoutError:
                continue
            locator = page.locator(tab_selector).first
            if locator.count():
                return locator
        return None

    def _click_sellers_tab(wait_timeout: int) -> bool:
        tab = _find_sellers_tab(wait_timeout)
        if not tab:
            return False
        return _safe_click(tab, page, f"sellers tab for {input_name or url}")

    seller_table_rows = "table.sellers-table__self tbody tr"
    seller_card = "div[data-test='sellerItem']"
    seller_wait_selector = f"{seller_table_rows}, {seller_card}"
    seller_discovery_attempts = max(1, SELLER_DISCOVERY_ATTEMPTS)
    seller_discovery_reload_every = max(0, SELLER_DISCOVERY_RELOAD_EVERY)
    seller_discovery_pause_ms = max(200, SELLER_DISCOVERY_PAUSE_MS)
    sellers_ready = False
    for discover_attempt in range(1, seller_discovery_attempts + 1):
        if _has_no_offers_text(page):
            print(f"ℹ️ No offers marker found for {input_name or url} ({city}), skipping retries.")
            raise NoOffersError(f"No offers marker found for {input_name or url} ({city})")

        clicked = _click_sellers_tab(wait_timeout=SELLER_WAIT_TIMEOUT if discover_attempt == 1 else 2500)
        if not clicked:
            print(
                f"⚠️ Sellers tab not found for {input_name or url}, "
                f"retry {discover_attempt}/{seller_discovery_attempts}"
            )
        else:
            try:
                page.wait_for_selector(seller_wait_selector, timeout=SELLER_WAIT_TIMEOUT)
                _wait_for_overlay(page)
            except TimeoutError:
                print(
                    f"⚠️ No sellers block found for {input_name or url}, "
                    f"retry {discover_attempt}/{seller_discovery_attempts}"
                )
            if page.locator(seller_table_rows).count() or page.locator(seller_card).count():
                sellers_ready = True
                break

        if discover_attempt >= seller_discovery_attempts:
            break
        if _has_no_offers_text(page):
            print(f"ℹ️ No offers marker found for {input_name or url} ({city}), stopping.")
            raise NoOffersError(f"No offers marker found for {input_name or url} ({city})")

        if seller_discovery_reload_every and discover_attempt % seller_discovery_reload_every == 0:
            try:
                page.goto(resolved_url, wait_until="domcontentloaded", timeout=SLOW_NAV_TIMEOUT)
                _wait_after_navigation(page, POST_LOAD_DELAY_FAST)
            except Exception as exc:
                print(f"⚠️ Could not reload product page for {input_name or url}: {exc}")
        else:
            page.wait_for_timeout(seller_discovery_pause_ms + random.randint(120, 520))

    if not sellers_ready:
        if _has_no_offers_text(page):
            print(f"ℹ️ No offers marker found for {input_name or url} ({city}), stopping.")
            raise NoOffersError(f"No offers marker found for {input_name or url} ({city})")
        raise RuntimeError(
            f"Sellers were not loaded after {seller_discovery_attempts} attempts for {input_name or url}"
        )

    results = []
    pagination_attempts = 0
    while True:
        if page.locator(seller_table_rows).count():
            sellers = page.locator(seller_table_rows)
            count = sellers.count()
            for i in range(count):
                row = sellers.nth(i)
                seller_link = row.locator("td").first.locator("a")
                if not seller_link.count():
                    continue
                name = seller_link.first.inner_text().strip()

                price_cell = row.locator("div.sellers-table__price-cell-text:not(._installments-price)").first
                if not price_cell.count():
                    price_cell = row.locator("div.sellers-table__price-cell-text").first
                if not price_cell.count():
                    continue
                price_text = price_cell.inner_text().replace("\xa0", " ").strip()

                installment_cell = row.locator("div.sellers-table__price-cell-text._installments-price").first
                installment_text = (
                    installment_cell.inner_text().replace("\xa0", " ").strip()
                    if installment_cell.count()
                    else ""
                )

                delivery_options = row.locator("div.sellers-table__delivery-cell-option")
                delivery_types = []
                delivery_details = []
                option_count = delivery_options.count()
                for j in range(option_count):
                    option_text = delivery_options.nth(j).inner_text().replace("\xa0", " ").strip()
                    if not option_text:
                        continue
                    parts = [part.strip() for part in option_text.split(",")]
                    if parts:
                        delivery_types.append(parts[0])
                        if len(parts) > 1:
                            detail = ", ".join(parts[1:]).strip()
                            if detail:
                                delivery_details.append(f"{parts[0]}: {detail}")

                numeric_price = extract_price(price_text)
                numeric_installment = extract_price(installment_text) if installment_text else None

                results.append({
                    "input_product": requested_name,
                    "product": product_name,
                    "product_url": url,
                    "seller": name,
                    "product_code": product_code,
                    "price_text": price_text,
                    "price_kzt": numeric_price,
                    "installment_text": installment_text,
                    "installment_kzt": numeric_installment,
                    "delivery_types": "; ".join(dict.fromkeys(delivery_types)),
                    "delivery_details": " | ".join(delivery_details),
                    "scraped_date": scraped_date or datetime.now(ALMATY_TZ).strftime("%d.%m.%Y"),
                    "region": city,
                    "review_count": review_count_value,
                })
        elif page.locator(seller_card).count():
            sellers = page.locator(seller_card)
            count = sellers.count()
            for i in range(count):
                seller = sellers.nth(i)
                name = ""
                price_text = ""
                installment_text = ""
                try:
                    name = seller.locator("a").first.inner_text().strip()
                except Exception:
                    pass
                price_locators = [
                    "[data-test='item-price']",
                    "span:has-text('₸')",
                    "div:has-text('₸')",
                ]
                for selector in price_locators:
                    loc = seller.locator(selector).first
                    if loc.count():
                        price_text = loc.inner_text().replace("\xa0", " ").strip()
                        break
                installment_loc = seller.locator("[data-test='item-installment-price']").first
                if installment_loc.count():
                    installment_text = installment_loc.inner_text().replace("\xa0", " ").strip()

                delivery_blocks = seller.locator("[data-test='deliveryOption'], .delivery-option")
                delivery_types = []
                delivery_details = []
                option_count = delivery_blocks.count()
                for j in range(option_count):
                    option_text = delivery_blocks.nth(j).inner_text().replace("\xa0", " ").strip()
                    if not option_text:
                        continue
                    parts = [part.strip() for part in option_text.split(",")]
                    if parts:
                        delivery_types.append(parts[0])
                        if len(parts) > 1:
                            detail = ", ".join(parts[1:]).strip()
                            if detail:
                                delivery_details.append(f"{parts[0]}: {detail}")

                numeric_price = extract_price(price_text)
                numeric_installment = extract_price(installment_text) if installment_text else None

                results.append({
                    "input_product": requested_name,
                    "product": product_name,
                    "product_url": url,
                    "seller": name,
                    "product_code": product_code,
                    "price_text": price_text,
                    "price_kzt": numeric_price,
                    "installment_text": installment_text,
                    "installment_kzt": numeric_installment,
                    "delivery_types": "; ".join(dict.fromkeys(delivery_types)),
                    "delivery_details": " | ".join(delivery_details),
                    "scraped_date": scraped_date or datetime.now(ALMATY_TZ).strftime("%d.%m.%Y"),
                    "region": city,
                    "review_count": review_count_value,
                })
        else:
            print(f"⚠️ Seller list disappeared for {input_name or url}")
            if not results:
                if _has_no_offers_text(page):
                    print(f"ℹ️ No offers marker found for {input_name or url} ({city}), stopping.")
                    raise NoOffersError(f"No offers marker found for {input_name or url} ({city})")
                raise RuntimeError(f"Seller list disappeared before any data was collected for {input_name or url}")
            break

        next_btn = page.locator(".pagination__el", has_text="Следующая")
        if not next_btn.count():
            next_btn = page.locator("button:has-text('Следующая'), a:has-text('Следующая')")
        if not next_btn.count() or "_disabled" in (next_btn.first.get_attribute("class") or ""):
            break
        if pagination_attempts >= MAX_PAGINATION_RETRIES:
            print(f"⚠️ Pagination limit reached for {input_name or url}")
            break
        if not _safe_click(next_btn.first, page, "pagination button"):
            break
        pagination_attempts += 1
        wait_delay = POST_PAGINATION_DELAY + random.randint(150, 500)
        page.wait_for_timeout(wait_delay)
        _wait_for_overlay(page)
        try:
            page.wait_for_selector(seller_wait_selector, timeout=SELLER_WAIT_TIMEOUT)
        except TimeoutError:
            break

    return results


def _photo_quality_note(photo_count: int) -> str:
    if photo_count >= 6:
        return f"📸 {photo_count} фото — выглядит богато."
    if photo_count >= 3:
        return f"📸 {photo_count} фото — достаточно, но можно лучше."
    if photo_count == 0:
        return "📸 Фото не найдены."
    return f"📸 Только {photo_count} фото — карточка выглядит бедно."


PHARMA_CATEGORY_KEYS = {"аптека", "лекарств"}


def _is_pharma_category(meta: dict) -> bool:
    path = " ".join(meta.get("category_path") or [])
    path_lower = path.lower()
    return any(key in path_lower for key in PHARMA_CATEGORY_KEYS)


def _format_kzt(value) -> str:
    if value is None:
        return ""
    try:
        return f"{int(value):,} ₸".replace(",", " ")
    except Exception:
        return str(value)


def _extract_review_display_count(page) -> int | None:
    selectors = [
        "[data-test='tabReviews']",
        "[data-tab='reviews']",
        "a:has-text('Отзывы')",
        "button:has-text('Отзывы')",
        "span:has-text('Отзывы')",
    ]
    for selector in selectors:
        try:
            locator = page.locator(selector).first
        except Exception:
            continue
        if not locator or not locator.count():
            continue
        try:
            text = locator.inner_text().strip()
        except Exception:
            continue
        match = re.search(r"\((\d+)\)", text)
        if not match:
            match = re.search(r"(\d+)", text)
        if match:
            try:
                return int(match.group(1))
            except ValueError:
                continue
    return None


def _fetch_review_count_via_api(page, product_id: str | None) -> int | None:
    if not product_id:
        return None
    try:
        payload = page.evaluate(
            """async (pid) => {
                try {
                    const resp = await fetch(`https://kaspi.kz/yml/review-view/api/v1/reviews/product/${pid}?withAgg=true`, {credentials: 'include'});
                    if (!resp.ok) return null;
                    return await resp.json();
                } catch (err) {
                    return null;
                }
            }""",
            product_id,
        )
    except Exception:
        return None
    summary = (payload or {}).get("summary") or {}
    statistic = summary.get("statistic") or []
    if statistic:
        try:
            return sum(int(item.get("count", 0) or 0) for item in statistic)
        except Exception:
            return None
    try:
        return int((payload or {}).get("count") or 0) or None
    except Exception:
        return None


def _review_api_url(product_id: str, params: dict | None = None) -> str:
    base = f"https://kaspi.kz/yml/review-view/api/v1/reviews/product/{product_id}"
    merged = {"withAgg": "true"}
    if params:
        merged.update({k: v for k, v in params.items() if v is not None})
    query = urlencode(merged, doseq=True)
    return f"{base}?{query}" if query else base


def _fetch_review_payload(page, url: str) -> dict | None:
    try:
        return page.evaluate(
            """async (targetUrl) => {
                try {
                    const resp = await fetch(targetUrl, {credentials: 'include'});
                    if (!resp.ok) {
                        return null;
                    }
                    return await resp.json();
                } catch (err) {
                    return null;
                }
            }""",
            url,
        )
    except Exception:
        return None


def _review_entry_id(entry: dict) -> str:
    for key in ("id", "reviewId", "uuid"):
        val = entry.get(key)
        if val:
            return str(val)
    comment = entry.get("comment") or {}
    text = ""
    if isinstance(comment, dict):
        text = (comment.get("text") or comment.get("comment") or "").strip()
    return "|".join(
        str(item or "")
        for item in (
            entry.get("author"),
            entry.get("rating"),
            entry.get("createdTime") or entry.get("createdDate") or entry.get("date"),
            text[:40],
        )
    )


def _infer_review_total(payload: dict | None) -> int | None:
    if not payload:
        return None
    for key in ("total", "totalCount", "totalReviews", "count"):
        val = payload.get(key)
        if isinstance(val, (int, float)) and val > 0:
            return int(val)
    pagination = payload.get("pagination") or {}
    for key in ("total", "totalCount", "totalElements", "totalItems"):
        val = pagination.get(key)
        if isinstance(val, (int, float)) and val > 0:
            return int(val)
    summary = payload.get("summary") or {}
    statistic = summary.get("statistic") or []
    if statistic:
        try:
            return int(sum(int(item.get("count", 0) or 0) for item in statistic))
        except Exception:
            return None
    return None


def _choose_review_pagination_strategy(page, product_id: str, first_payload: dict, page_size: int) -> dict | None:
    first_entries = first_payload.get("data") or []
    if not first_entries:
        return None
    first_ids = [_review_entry_id(entry) for entry in first_entries]
    if len(first_entries) < page_size:
        return None

    strategies = [
        ("page_size", lambda page_num, size: {"page": page_num, "size": size}),
        ("page_pageSize", lambda page_num, size: {"page": page_num, "pageSize": size}),
        ("page_limit", lambda page_num, size: {"page": page_num, "limit": size}),
        ("offset_limit", lambda page_num, size: {"offset": page_num * size, "limit": size}),
        ("start_size", lambda page_num, size: {"start": page_num * size, "size": size}),
        ("currentPage_pageSize", lambda page_num, size: {"currentPage": page_num, "pageSize": size}),
    ]
    for label, builder in strategies:
        params = builder(1, page_size)
        url = _review_api_url(product_id, params)
        payload = _fetch_review_payload(page, url)
        entries = (payload or {}).get("data") or []
        if not entries:
            continue
        entry_ids = [_review_entry_id(entry) for entry in entries]
        if entry_ids != first_ids:
            return {"label": label, "builder": builder}
    return None


def _parse_review_date(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        ts = int(value)
        if ts > 10**12:
            ts = int(ts / 1000)
        try:
            return datetime.fromtimestamp(ts, tz=ALMATY_TZ).date().isoformat()
        except Exception:
            return str(value)
    text = str(value).strip()
    if not text:
        return ""
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except Exception:
        pass
    match = re.search(r"(\\d{4}-\\d{2}-\\d{2})", text)
    if match:
        return match.group(1)
    match = re.search(r"(\\d{2})[./-](\\d{2})[./-](\\d{4})", text)
    if match:
        day, month, year = match.group(1), match.group(2), match.group(3)
        return f"{year}-{month}-{day}"
    return text


def _pick_first(entry: dict, keys: list[str]) -> Any:
    for key in keys:
        if key in entry and entry.get(key) not in (None, ""):
            return entry.get(key)
    return None


def _normalize_review_text(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, (list, tuple)):
        return " ".join(str(item).strip() for item in value if str(item).strip())
    if isinstance(value, dict):
        return " ".join(str(item).strip() for item in value.values() if str(item).strip())
    return str(value).strip()


def _extract_review_row(entry: dict) -> dict:
    comment = entry.get("comment") or {}
    if not isinstance(comment, dict):
        comment = {}
    pros = _normalize_review_text(
        _pick_first(comment, ["pros", "advantages", "plus", "positive"]) or _pick_first(entry, ["pros", "advantages"])
    )
    cons = _normalize_review_text(
        _pick_first(comment, ["cons", "disadvantages", "minus", "negative"]) or _pick_first(entry, ["cons", "disadvantages"])
    )
    text = _normalize_review_text(
        _pick_first(comment, ["text", "comment", "description"]) or _pick_first(entry, ["text", "comment"])
    )
    helpful = _pick_first(entry, ["helpfulCount", "usefulCount", "likeCount", "likes", "useful", "usefulReviewsCount"])
    try:
        helpful = int(helpful) if helpful not in (None, "") else None
    except Exception:
        helpful = None
    review_date = _parse_review_date(
        _pick_first(entry, ["createdTime", "createdDate", "date", "createdAt"])
        or _pick_first(comment, ["createdTime", "createdDate", "date"])
    )
    return {
        "review_id": _pick_first(entry, ["id", "reviewId", "uuid"]) or "",
        "author": _normalize_review_text(_pick_first(entry, ["author", "user", "authorName"]) or ""),
        "rating": _pick_first(entry, ["rating", "score"]) or "",
        "pros": pros,
        "cons": cons,
        "comment": text,
        "helpful_count": helpful if helpful is not None else "",
        "review_date": review_date,
    }


def _open_product_page_for_reviews(page, url: str, city: str | None):
    resolved_url = _url_with_city(resolve_kaspi_link(url), city)
    try:
        page.set_default_timeout(FAST_NAV_TIMEOUT)
        _jitter_sleep(0.1, 0.4)
        page.goto(resolved_url, wait_until="domcontentloaded", timeout=FAST_NAV_TIMEOUT)
        _wait_after_navigation(page, POST_LOAD_DELAY_FAST)
    except TimeoutError:
        page.set_default_timeout(SLOW_NAV_TIMEOUT)
        page.goto(resolved_url, timeout=SLOW_NAV_TIMEOUT)
        _wait_after_navigation(page, POST_LOAD_DELAY_SLOW)
    try:
        page.wait_for_function(
            "() => window.digitalData?.product || (window.BACKEND && window.BACKEND.components && window.BACKEND.components.item)",
            timeout=6000,
        )
    except Exception:
        pass
    if city:
        try:
            page.wait_for_selector("text=Выберите ваш город", timeout=5000)
            page.click(f"a:has-text('{city}')")
            page.wait_for_timeout(800)
            _wait_for_overlay(page)
        except Exception:
            pass
    return resolved_url


def _resolve_product_id_from_page(page, fallback_url: str) -> str:
    product_id = extract_product_code(fallback_url)
    if product_id:
        return product_id
    try:
        product_id = page.evaluate(
            "() => window.digitalData?.product?.id || window.BACKEND?.components?.item?.card?.id || ''"
        )
    except Exception:
        product_id = ""
    return str(product_id or "")


def _collect_all_reviews(page, product_id: str, page_size: int, max_pages: int, max_reviews: int | None):
    base_url = _review_api_url(product_id, {"withAgg": "true"})
    first_payload = _fetch_review_payload(page, base_url) or {}
    reviews = list(first_payload.get("data") or [])
    total_count = _infer_review_total(first_payload)

    if not reviews:
        warmup_strategies = [
            ("page_size", lambda page_num, size: {"page": page_num, "size": size}),
            ("page_pageSize", lambda page_num, size: {"page": page_num, "pageSize": size}),
            ("page_limit", lambda page_num, size: {"page": page_num, "limit": size}),
            ("offset_limit", lambda page_num, size: {"offset": page_num * size, "limit": size}),
        ]
        for _, builder in warmup_strategies:
            url = _review_api_url(product_id, builder(0, page_size))
            payload = _fetch_review_payload(page, url) or {}
            data = payload.get("data") or []
            if data:
                first_payload = payload
                reviews = list(data)
                total_count = _infer_review_total(first_payload)
                break

    strategy = _choose_review_pagination_strategy(page, product_id, first_payload, page_size)

    if not strategy and total_count and len(reviews) < total_count:
        strategy = {"label": "fallback_page_size", "builder": lambda page_num, size: {"page": page_num, "size": size}}

    seen_ids = {_review_entry_id(entry) for entry in reviews}
    page_num = 1
    while strategy and page_num < max_pages:
        params = strategy["builder"](page_num, page_size)
        url = _review_api_url(product_id, params)
        payload = _fetch_review_payload(page, url) or {}
        entries = payload.get("data") or []
        if not entries:
            break
        new_entries = []
        for entry in entries:
            entry_id = _review_entry_id(entry)
            if entry_id in seen_ids:
                continue
            seen_ids.add(entry_id)
            new_entries.append(entry)
        if not new_entries:
            break
        reviews.extend(new_entries)
        if total_count and len(reviews) >= total_count:
            break
        if max_reviews and len(reviews) >= max_reviews:
            break
        page_num += 1
        _jitter_sleep(0.2, 0.6)
    if max_reviews:
        return reviews[:max_reviews]
    return reviews


def _init_reviews_db(conn: sqlite3.Connection):
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS reviews (
            review_uid TEXT PRIMARY KEY,
            product_id TEXT,
            sku TEXT,
            review_id TEXT,
            author TEXT,
            rating REAL,
            pros TEXT,
            cons TEXT,
            comment TEXT,
            helpful_count INTEGER,
            review_date TEXT,
            date_source TEXT,
            scraped_at TEXT,
            product_url TEXT
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_reviews_date ON reviews(review_date)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_reviews_sku ON reviews(sku)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_reviews_product ON reviews(product_id)")
    conn.commit()


def _store_review_rows(
    conn: sqlite3.Connection,
    rows: list[dict],
):
    if not rows:
        return
    conn.executemany(
        """
        INSERT OR REPLACE INTO reviews (
            review_uid,
            product_id,
            sku,
            review_id,
            author,
            rating,
            pros,
            cons,
            comment,
            helpful_count,
            review_date,
            date_source,
            scraped_at,
            product_url
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row.get("review_uid"),
                row.get("product_id"),
                row.get("sku"),
                row.get("review_id"),
                row.get("author"),
                row.get("rating"),
                row.get("pros"),
                row.get("cons"),
                row.get("comment"),
                row.get("helpful_count"),
                row.get("review_date"),
                row.get("date_source"),
                row.get("scraped_at"),
                row.get("product_url"),
            )
            for row in rows
        ],
    )
    conn.commit()


def _month_label_from_iso(ym: str) -> str:
    """Convert YYYY-MM to Jan-25 label."""
    try:
        year, month = ym.split("-", 1)
        month_idx = int(month)
        if 1 <= month_idx <= 12 and len(year) == 4:
            month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
            return f"{month_names[month_idx - 1]}-{year[-2:]}"
    except Exception:
        pass
    return ym


def _parse_iso_month(ym: str) -> tuple[int, int] | None:
    text = str(ym or "").strip()
    m = re.match(r"^(\d{4})-(\d{2})$", text)
    if not m:
        return None
    year = int(m.group(1))
    month = int(m.group(2))
    if not (1 <= month <= 12):
        return None
    return (year, month)


def _iter_months(start_ym: str, end_ym: str) -> list[str]:
    start = _parse_iso_month(start_ym)
    end = _parse_iso_month(end_ym)
    if not start or not end:
        return []
    y, m = start
    end_y, end_m = end
    out: list[str] = []
    while (y, m) <= (end_y, end_m):
        out.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            y += 1
            m = 1
    return out


def _append_comment_review_month_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Append `Reviews>` and monthly counts for reviews with non-empty comment text."""
    if df.empty or "product_code" not in df.columns:
        return df
    if not REVIEWS_DB_PATH.exists():
        return df

    product_codes = sorted(
        {
            str(code).strip()
            for code in df["product_code"].dropna().tolist()
            if str(code).strip() and str(code).strip().lower() != "nan"
        }
    )
    if not product_codes:
        return df

    start_month = (os.environ.get("REVIEWS_MONTHS_FROM") or "").strip()
    if not start_month:
        # By default include the full available history from reviews.db.
        start_month = "0000-00"

    placeholders = ",".join("?" for _ in product_codes)
    sql = f"""
        WITH normalized AS (
            SELECT
                product_id,
                comment,
                CASE
                    WHEN review_date LIKE '__.__.____' THEN
                        substr(review_date, 7, 4) || '-' || substr(review_date, 4, 2) || '-' || substr(review_date, 1, 2)
                    WHEN review_date LIKE '__/__/____' THEN
                        substr(review_date, 7, 4) || '-' || substr(review_date, 4, 2) || '-' || substr(review_date, 1, 2)
                    ELSE review_date
                END AS review_date_norm
            FROM reviews
            WHERE product_id IN ({placeholders})
        )
        SELECT
            product_id,
            substr(review_date_norm, 1, 7) AS ym,
            COUNT(*) AS cnt
        FROM normalized
        WHERE length(trim(coalesce(comment, ''))) > 0
          AND review_date_norm GLOB '????-??-??'
          AND substr(review_date_norm, 1, 7) >= ?
        GROUP BY product_id, ym
        ORDER BY ym
    """

    rows: list[tuple[str, str, int]] = []
    try:
        with sqlite3.connect(REVIEWS_DB_PATH) as conn:
            rows = conn.execute(sql, product_codes + [start_month]).fetchall()
    except Exception as exc:
        print(f"⚠️ Could not build monthly review-comment columns: {exc}")
        return df

    month_keys: list[str] = []
    count_map: dict[str, dict[str, int]] = {}
    for product_id, ym, cnt in rows:
        product_key = str(product_id or "").strip()
        ym_key = str(ym or "").strip()
        if not product_key or not ym_key or not _parse_iso_month(ym_key):
            continue
        if ym_key not in month_keys:
            month_keys.append(ym_key)
        product_bucket = count_map.setdefault(product_key, {})
        product_bucket[ym_key] = int(cnt or 0)

    include_current_month = (os.environ.get("REVIEWS_INCLUDE_CURRENT_MONTH", "1") or "1").strip().lower()
    if include_current_month not in {"0", "false", "no"}:
        current_ym = datetime.now(ALMATY_TZ).strftime("%Y-%m")
        if month_keys:
            for ym in _iter_months(month_keys[0], current_ym):
                if ym not in month_keys:
                    month_keys.append(ym)
        elif _parse_iso_month(current_ym):
            month_keys.append(current_ym)

    if not month_keys:
        return df

    insert_after = "price_match_status" if "price_match_status" in df.columns else "product_code"
    insert_at = df.columns.get_loc(insert_after) + 1
    if "Reviews>" not in df.columns:
        df.insert(insert_at, "Reviews>", "")
        insert_at += 1

    product_keys = df["product_code"].astype(str).str.strip()
    for ym_key in month_keys:
        month_label = _month_label_from_iso(ym_key)
        values = [
            int(count_map.get(code, {}).get(ym_key, 0))
            if code and code.lower() != "nan" and code != "<NA>"
            else 0
            for code in product_keys
        ]
        if month_label in df.columns:
            df[month_label] = pd.Series(values, index=df.index, dtype="Int64")
        else:
            df.insert(insert_at, month_label, pd.Series(values, index=df.index, dtype="Int64"))
            insert_at += 1
    return df


def _backfill_reviews_db_for_products(df: pd.DataFrame):
    """Sync reviews.db for tracked products before exporting monthly review columns."""
    if df.empty or "product_url" not in df.columns:
        return
    auto_flag = (os.environ.get("MAIN_REPORT_SYNC_REVIEWS", "1") or "1").strip().lower()
    if auto_flag in {"0", "false", "no"}:
        return
    sync_mode = (os.environ.get("MAIN_REPORT_SYNC_REVIEWS_MODE", "stale") or "stale").strip().lower()
    if sync_mode not in {"missing", "stale", "all"}:
        sync_mode = "stale"
    stale_days_raw = (os.environ.get("MAIN_REPORT_SYNC_REVIEWS_STALE_DAYS", "1") or "1").strip()
    try:
        stale_days = max(0, int(stale_days_raw))
    except Exception:
        stale_days = 1

    work = df.copy()
    if "product_code" in work.columns:
        work["product_code"] = work["product_code"].astype(str).str.strip()
    else:
        work["product_code"] = ""
    work["product_url"] = work["product_url"].astype(str).str.strip()
    work["sku"] = ""
    if "input_product" in work.columns:
        work["sku"] = work["input_product"].astype(str).fillna("").str.strip()
    if "product" in work.columns:
        fallback = work["product"].astype(str).fillna("").str.strip()
        work.loc[work["sku"] == "", "sku"] = fallback[work["sku"] == ""]

    rows = []
    seen_urls: set[str] = set()
    for _, row in work.iterrows():
        url = (row.get("product_url") or "").strip()
        if not url or url.lower() in {"nan", "<na>"}:
            continue
        if url in seen_urls:
            continue
        seen_urls.add(url)
        code = (row.get("product_code") or "").strip()
        if not code or code.lower() in {"nan", "<na>"}:
            code = extract_product_code(url)
        sku = (row.get("sku") or "").strip()
        rows.append({"product_code": code, "product_url": url, "sku": sku})
    if not rows:
        return

    codes = sorted({r["product_code"] for r in rows if r["product_code"]})
    existing: set[str] = set()
    latest_scraped: dict[str, str] = {}
    if codes and REVIEWS_DB_PATH.exists():
        placeholders = ",".join("?" for _ in codes)
        comments_sql = f"""
            SELECT DISTINCT product_id
            FROM reviews
            WHERE product_id IN ({placeholders})
              AND length(trim(coalesce(comment, ''))) > 0
        """
        latest_sql = f"""
            SELECT product_id, max(substr(scraped_at, 1, 10)) AS last_scraped_date
            FROM reviews
            WHERE product_id IN ({placeholders})
            GROUP BY product_id
        """
        try:
            with sqlite3.connect(REVIEWS_DB_PATH) as conn:
                db_rows = conn.execute(comments_sql, codes).fetchall()
                latest_rows = conn.execute(latest_sql, codes).fetchall()
            existing = {str(item[0]).strip() for item in db_rows if item and item[0]}
            latest_scraped = {
                str(item[0]).strip(): str(item[1]).strip()
                for item in latest_rows
                if item and item[0] and item[1]
            }
        except Exception as exc:
            print(f"⚠️ Could not read reviews db before backfill: {exc}")

    stale_cutoff = (datetime.now(ALMATY_TZ).date() - timedelta(days=stale_days)).isoformat()
    targets = []
    missing_cnt = 0
    stale_cnt = 0
    for row in rows:
        code = row.get("product_code") or ""
        if sync_mode == "all":
            targets.append(row)
            continue
        if not code or code not in existing:
            targets.append(row)
            missing_cnt += 1
            continue
        if sync_mode == "stale":
            last_scraped = latest_scraped.get(code, "")
            if not last_scraped or last_scraped < stale_cutoff:
                targets.append(row)
                stale_cnt += 1

    if not targets:
        return

    limit_raw = (os.environ.get("MAIN_REPORT_SYNC_REVIEWS_LIMIT", "0") or "0").strip()
    try:
        limit = int(limit_raw)
    except Exception:
        limit = 0
    if limit > 0:
        targets = targets[:limit]
    if not targets:
        return

    with tempfile.NamedTemporaryFile("w", suffix=".csv", delete=False, encoding="utf-8", newline="") as tmp:
        writer = csv.DictWriter(tmp, fieldnames=["sku", "product_url"])
        writer.writeheader()
        for item in targets:
            writer.writerow({"sku": item["sku"], "product_url": item["product_url"]})
        tmp_path = Path(tmp.name)

    try:
        meta = [f"mode={sync_mode}"]
        if sync_mode in {"missing", "stale"}:
            meta.append(f"missing={missing_cnt}")
        if sync_mode == "stale":
            meta.append(f"stale={stale_cnt}")
            meta.append(f"stale_days={stale_days}")
        print(f"🔄 Reviews sync for {len(targets)} SKU before monthly export ({', '.join(meta)})...")
        run_reviews_scrape(reviews_csv=tmp_path, city=os.environ.get("REVIEWS_CITY", "Алматы"))
    except Exception as exc:
        print(f"⚠️ Reviews sync failed: {exc}")
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass


def _month_col_sort_key(label: str) -> tuple[int, int]:
    """Sort Jan-25 style columns chronologically."""
    month_map = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
    }
    m = re.match(r"^([A-Za-z]{3})-(\d{2})$", str(label).strip())
    if not m:
        return (9999, 99)
    mon = month_map.get(m.group(1).title(), 99)
    year = 2000 + int(m.group(2))
    return (year, mon)


def _build_reviews_summary_sheet(df: pd.DataFrame) -> pd.DataFrame | None:
    """Build one-row-per-SKU review sheet with best seller/price and monthly comment counts."""
    required = {"product_code", "price_kzt", "seller"}
    if df.empty or not required.issubset(set(df.columns)):
        return None

    work = df.copy()
    work["product_code"] = work["product_code"].astype(str).str.strip()
    work = work[work["product_code"].notna() & (work["product_code"] != "") & (work["product_code"].str.lower() != "nan")]
    if work.empty:
        return None

    work["price_kzt_num"] = pd.to_numeric(work["price_kzt"], errors="coerce")
    best = work.sort_values(
        by=["product_code", "price_kzt_num", "seller"],
        na_position="last",
        kind="stable",
    ).drop_duplicates(subset=["product_code"], keep="first")

    base_cols: list[str] = []
    for col in ("product_code", "input_product", "product", "product_url"):
        if col in best.columns:
            base_cols.append(col)
    reviews_df = best[base_cols].copy()
    reviews_df["best_price_kzt"] = best["price_kzt_num"].astype("Int64")
    reviews_df["best_price_seller"] = best["seller"].astype(str).str.strip()

    # Build monthly comment columns directly for the review tab
    # (main sheet no longer carries these columns).
    month_seed = reviews_df[["product_code"]].copy()
    month_seed = _append_comment_review_month_columns(month_seed)
    month_cols = [c for c in month_seed.columns if re.match(r"^[A-Za-z]{3}-\d{2}$", str(c))]
    month_cols = sorted(month_cols, key=_month_col_sort_key)
    if month_cols:
        month_values = month_seed[["product_code"] + month_cols].copy()
        reviews_df = reviews_df.merge(month_values, on="product_code", how="left")
        for col in month_cols:
            reviews_df[col] = pd.to_numeric(reviews_df[col], errors="coerce").fillna(0).astype("Int64")

    order = [c for c in ("product_code", "input_product", "product", "product_url") if c in reviews_df.columns]
    order += ["best_price_kzt", "best_price_seller"]
    order += month_cols
    reviews_df = reviews_df.reindex(columns=order)
    return reviews_df


def _simplify_seller_prices(records: list[dict]) -> list[dict]:
    simplified: list[dict] = []
    for rec in records or []:
        price = rec.get("price_kzt")
        price_text = rec.get("price_text") or ""
        seller = rec.get("seller") or "Без названия"
        if price is None and not price_text:
            continue
        simplified.append({
            "seller": seller.strip() or "Без названия",
            "price_kzt": price,
            "price_text": price_text.strip(),
        })
    simplified.sort(key=lambda item: ((item.get("price_kzt") if item.get("price_kzt") is not None else 10 ** 12), item.get("seller", "")))
    return simplified


def _calculate_price_stats(prices: list[dict]) -> dict:
    numeric_values = [item["price_kzt"] for item in prices if isinstance(item.get("price_kzt"), (int, float))]
    if not numeric_values:
        return {}
    min_price = min(numeric_values)
    max_price = max(numeric_values)
    unique_levels = sorted({value for value in numeric_values})
    min_step = 0
    if len(unique_levels) > 1:
        deltas = [b - a for a, b in zip(unique_levels, unique_levels[1:]) if b - a > 0]
        if deltas:
            min_step = min(deltas)
    spread = max_price - min_price
    min_sellers = sorted({item["seller"] for item in prices if item.get("price_kzt") == min_price and item.get("seller")})
    max_sellers = sorted({item["seller"] for item in prices if item.get("price_kzt") == max_price and item.get("seller")})
    seller_count = len({item.get("seller") for item in prices if item.get("seller")})
    return {
        "min_price": min_price,
        "max_price": max_price,
        "spread": spread,
        "min_step": min_step,
        "seller_count": seller_count,
        "min_sellers": min_sellers,
        "max_sellers": max_sellers,
    }


def _summarize_review_sentiment(histogram: list[dict]) -> str | None:
    if not histogram:
        return None
    buckets: list[tuple[int, int]] = []
    total = 0
    positive = 0
    for entry in histogram:
        stars = entry.get("rating") or entry.get("value") or entry.get("score") or entry.get("rate")
        count = entry.get("count") or entry.get("quantity") or 0
        try:
            stars = int(stars)
            count = int(count)
        except Exception:
            continue
        if count <= 0:
            continue
        total += count
        if stars >= 4:
            positive += count
        buckets.append((stars, count))
    if not buckets or total == 0:
        return None
    buckets.sort(reverse=True)
    parts = [f"{stars}★×{count}" for stars, count in buckets]
    ratio = positive / total if total else 0
    if ratio >= 0.75:
        tone = "очень позитивно"
    elif ratio >= 0.55:
        tone = "в целом позитивно"
    elif ratio >= 0.35:
        tone = "смешанные оценки"
    else:
        tone = "скорее негативно"
    return f"{', '.join(parts)} — {tone}"


def analyze_product_card(product_url: str, city: str = "Алматы") -> dict:
    """Собираем метаданные по карточке Kaspi: рейтинг, отзывы, контент."""
    resolved_url = resolve_kaspi_link(product_url.strip())
    meta: dict = {
        "input_url": product_url,
        "resolved_url": resolved_url,
        "city": city,
        "timestamp": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
    }
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(viewport={"width": 1600, "height": 900})
        page = context.new_page()
        try:
            page.set_default_timeout(FAST_NAV_TIMEOUT)
            page.goto(resolved_url, wait_until="domcontentloaded", timeout=FAST_NAV_TIMEOUT)
            page.wait_for_timeout(POST_LOAD_DELAY_FAST)
        except TimeoutError:
            page.set_default_timeout(SLOW_NAV_TIMEOUT)
            page.goto(resolved_url, timeout=SLOW_NAV_TIMEOUT)
            page.wait_for_timeout(POST_LOAD_DELAY_SLOW)

        try:
            page.wait_for_selector("text=Выберите ваш город", timeout=5000)
            page.click(f"a:has-text('{city}')")
            page.wait_for_timeout(1000)
        except Exception:
            pass

        try:
            page.wait_for_function("() => window.digitalData && window.digitalData.product", timeout=6000)
        except Exception:
            pass
        try:
            page.wait_for_function("() => window.BACKEND && window.BACKEND.components && window.BACKEND.components.item", timeout=6000)
        except Exception:
            pass
        meta["resolved_url"] = page.url
        digital_product = page.evaluate("() => window.digitalData?.product || {}") or {}
        backend_data = page.evaluate("() => window.BACKEND?.components?.item || {}") or {}
        gallery = backend_data.get("galleryImages") or []
        breadcrumbs = backend_data.get("breadcrumbs") or []
        category_path = [crumb.get("title") for crumb in breadcrumbs if crumb.get("title")]
        if not category_path:
            category_path = digital_product.get("category") or []

        product_id = digital_product.get("id") or backend_data.get("card", {}).get("id") or extract_product_code(meta["resolved_url"])
        review_count_display = _extract_review_display_count(page)
        has_description = bool((backend_data.get("description") or "").strip())
        if not has_description:
            for desc in backend_data.get("descriptions") or []:
                if (desc.get("text") or "").strip():
                    has_description = True
                    break
        has_characteristics = bool(backend_data.get("specifications"))
        review_summary = None
        review_entries = []
        if product_id:
            review_summary = page.evaluate(
                """async (productId) => {
                    try {
                        const resp = await fetch(`https://kaspi.kz/yml/review-view/api/v1/reviews/product/${productId}?withAgg=true`, {credentials: 'include'});
                        if (!resp.ok) {
                            return null;
                        }
                        return await resp.json();
                    } catch (err) {
                        return null;
                    }
                }""",
                product_id,
            )
            if review_summary:
                review_entries = review_summary.get("data") or []
        seller_prices: list[dict] = []
        sellers_page = None
        try:
            sellers_page = context.new_page()
            sellers_raw = scrape_single_product(sellers_page, meta["resolved_url"], city=city)
            seller_prices = _simplify_seller_prices(sellers_raw or [])
        except NoOffersError:
            seller_prices = []
        except Exception:
            seller_prices = []
        finally:
            if sellers_page:
                try:
                    sellers_page.close()
                except Exception:
                    pass
        context.close()
        browser.close()

    photo_count = len(gallery)
    photo_note = _photo_quality_note(photo_count)
    rating_value = None
    rating_histogram = []
    review_count = digital_product.get("reviewCount")
    if review_summary and review_summary.get("summary"):
        rating_value = review_summary["summary"].get("global")
        stats = review_summary["summary"].get("statistic") or []
        rating_histogram = stats
        if review_count in (None, 0):
            review_count = sum(item.get("count", 0) for item in stats)
    if (review_count_display in (None, 0)) and review_entries:
        review_count_display = len(review_entries)

    supplier_offers = digital_product.get("supplierOffers") or []
    best_offer = supplier_offers[0] if supplier_offers else {}
    has_rich_content = bool(backend_data.get("richContent")) or bool(backend_data.get("videos")) or bool(backend_data.get("teasers"))
    video_count = len(backend_data.get("videos") or [])

    price_stats = _calculate_price_stats(seller_prices)
    review_sentiment = _summarize_review_sentiment(rating_histogram)

    meta.update(
        {
            "product_id": product_id or "",
            "product_name": digital_product.get("name") or backend_data.get("card", {}).get("title") or "",
            "price": digital_product.get("unitSalePrice") or (backend_data.get("card", {}).get("price")),
            "rating_value": rating_value,
            "review_count": review_count or 0,
            "review_count_display": review_count_display or review_count or 0,
            "rating_histogram": rating_histogram,
            "category_path": category_path,
            "photo_count": photo_count,
            "photo_note": photo_note,
            "has_rich_content": has_rich_content,
            "video_count": video_count,
            "supplier_offer": best_offer,
            "brand": digital_product.get("brand") or backend_data.get("card", {}).get("promoConditions", {}).get("brand"),
            "gallery_images": gallery,
            "seller_prices": seller_prices,
            "price_stats": price_stats,
            "has_description": has_description,
            "has_characteristics": has_characteristics,
            "review_sentiment": review_sentiment,
        }
    )
    return meta


def format_product_analysis(meta: dict) -> str:
    pharma = _is_pharma_category(meta)
    lines = [
        f"🛍️ Товар: {meta.get('product_name') or 'без названия'}",
        f"🔗 {meta.get('resolved_url')}",
    ]
    if meta.get("category_path"):
        lines.append("📂 Категория: " + " › ".join(meta["category_path"]))
    rating = meta.get("rating_value")
    review_count = meta.get("review_count_display") or meta.get("review_count")
    if rating:
        lines.append(f"⭐️ Рейтинг: {rating}/5 из {review_count} отзывов")
    elif review_count:
        lines.append(f"⭐️ Отзывов: {review_count}, но рейтинг не найден")
    stats = meta.get("price_stats") or {}
    if stats.get("min_price") is not None:
        min_price = _format_kzt(stats.get("min_price"))
        max_price = _format_kzt(stats.get("max_price"))
        if stats.get("min_price") == stats.get("max_price"):
            lines.append(f"💸 Цена: {min_price} · продавцов: {stats.get('seller_count') or 0}")
        else:
            spread_val = stats.get("spread")
            step_val = stats.get("min_step")
            spread = _format_kzt(spread_val) if spread_val is not None else "—"
            step = _format_kzt(step_val) if step_val is not None else "—"
            lines.append(f"💸 Цены: {min_price} – {max_price} (разброс {spread}, шаг {step})")
        if stats.get("min_sellers"):
            lines.append("   Минимум у: " + ", ".join(stats["min_sellers"][:3]))
        if stats.get("max_sellers"):
            lines.append("   Максимум у: " + ", ".join(stats["max_sellers"][:3]))
    elif meta.get("price"):
        lines.append(f"💰 Цена: {_format_kzt(meta.get('price'))}")
    seller_prices = meta.get("seller_prices") or []
    if seller_prices:
        preview = []
        for entry in seller_prices[:5]:
            price_text = entry.get("price_text") or _format_kzt(entry.get("price_kzt"))
            preview.append(f"{entry.get('seller')}: {price_text}")
        lines.append("🏷️ Продавцы (топ): " + "; ".join(preview))
    lines.append(meta.get("photo_note", ""))
    if meta.get("video_count"):
        lines.append("🎬 Видео: есть")
    else:
        if pharma:
            lines.append("🎬 Видео не найдено (для лекарств это допустимо).")
        else:
            lines.append("🎬 Видео не найдено")
    if pharma:
        lines.append("🧩 Rich content/A+ не используются в аптечной категории — закон запрещает такие блоки.")
    else:
        lines.append("🧩 Rich content / A+: есть" if meta.get("has_rich_content") else "🧩 Rich content не обнаружен")
    lines.append("📝 Описание: есть" if meta.get("has_description") else "📝 Описание не заполнено")
    lines.append("📑 Характеристики: есть" if meta.get("has_characteristics") else "📑 Характеристик нет")
    if meta.get("review_sentiment"):
        lines.append(f"🗣️ Среднее мнение: {meta['review_sentiment']}")
    offer = meta.get("supplier_offer") or {}
    if offer.get("supplier"):
        sup = offer["supplier"]
        sup_name = sup.get("name") or "продавец"
        sup_rating = sup.get("rating")
        sup_reviews = sup.get("reviewCount")
        seller_line = f"🏪 Продавец: {sup_name}"
        if sup_rating:
            seller_line += f" (рейтинг {sup_rating}"
            if sup_reviews:
                seller_line += f", {sup_reviews} отзывов"
            seller_line += ")"
        lines.append(seller_line)
    return "\n".join(line for line in lines if line)


def gpt_product_opinion(meta: dict) -> str | None:
    if not OPENAI_API_KEY:
        return None
    pharma = _is_pharma_category(meta)
    price_stats = meta.get("price_stats") or {}
    if price_stats.get("min_price") is not None:
        price_line = (
            f"Цены: {_format_kzt(price_stats.get('min_price'))} – {_format_kzt(price_stats.get('max_price'))}, "
            f"разброс {_format_kzt(price_stats.get('spread'))}, шаг {_format_kzt(price_stats.get('min_step'))}, "
            f"продавцов {price_stats.get('seller_count')}"
        )
    else:
        price_line = f"Цена: {_format_kzt(meta.get('price')) or 'нет данных'}"
    min_price = price_stats.get("min_price") or meta.get("price")
    if min_price:
        leader_price = max(int(min_price * 0.95), 0) if isinstance(min_price, (int, float)) else min_price
    else:
        leader_price = None
    leader_desc = ""
    if leader_price:
        leader_desc = (
            f"SKU №1 в категории держит цену около {_format_kzt(leader_price)}"
            " , рейтинг ~4.9 и минимум 6 подробных фото."
        )
    payload_lines = [
        f"Название: {meta.get('product_name')}",
        f"Категория: {' > '.join(meta.get('category_path') or [])}",
        f"Рейтинг: {meta.get('rating_value') or 'нет данных'}",
        f"Отзывов: {meta.get('review_count_display') or meta.get('review_count')}",
        f"Настроение отзывов: {meta.get('review_sentiment') or 'нет информации'}",
        price_line,
        f"Фото: {meta.get('photo_count')} ({meta.get('photo_note')})",
        f"Rich content: {'да' if meta.get('has_rich_content') else 'нет'}",
        f"Видео: {meta.get('video_count')}",
        f"Описание: {'есть' if meta.get('has_description') else 'нет'}",
        f"Характеристики: {'есть' if meta.get('has_characteristics') else 'нет'}",
    ]
    if leader_desc:
        payload_lines.append(leader_desc)
    if pharma:
        payload_lines.append("Аптечная категория: rich content/A+ запрещены законом, отсутствие таких блоков — норма.")
    payload_lines = [
        line for line in payload_lines if line
    ]
    if meta.get("supplier_offer"):
        payload_lines.append(f"Продавец: {meta['supplier_offer'].get('supplier', {}).get('name')}")
    prompt = (
        "Ты старший специалист Kaspi Магазина по карточкам товаров. Проанализируй данные и ответь от своего лица.\n"
        "- Оцени ценовой диапазон и конкурентность, обязательно сравни с SKU №1 из категории (используй данные выше).\n"
        "- Оцени визуал карточки (фото, описание, rich/A+, видео) и если это аптека, не предлагай запрещённые законом блоки.\n"
        "- Сформулируй мнение по отзывам/настроению покупателей.\n"
        "- Заверши общим впечатлением и конкретным советом, пиши живым языком, без Markdown и звёздочек.\n"
        "Ответ должен быть 2–3 коротких абзаца.\n\n"
        + "\n".join(payload_lines)
    )
    body = {
        "model": OPENAI_MODEL,
        "messages": [
            {"role": "system", "content": "Ты специалист Kaspi.kz по качеству карточек товаров и говоришь по-русски."},
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.2,
        "max_tokens": 300,
    }
    try:
        resp = requests.post(
            "https://api.openai.com/v1/chat/completions",
            headers={"Authorization": f"Bearer {OPENAI_API_KEY}"},
            json=body,
            timeout=20,
        )
        data = resp.json()
        if resp.status_code == 200 and data.get("choices"):
            return data["choices"][0]["message"]["content"].strip()
        return f"GPT ответил ошибкой: {resp.status_code} {data}"
    except Exception as exc:
        return f"GPT запрос не удался: {exc}"


def run_product_analysis(product_url: str, city: str = "Алматы"):
    meta = analyze_product_card(product_url, city=city)
    summary = format_product_analysis(meta)
    ai_opinion = gpt_product_opinion(meta)
    if ai_opinion:
        report = f"{summary}\n\n🤖 Отчет специалиста:\n{ai_opinion}"
    else:
        report = f"{summary}\n\n🤖 Отчет специалиста недоступен (нет OPENAI_API_KEY)."
    print(report)


def _scrape_product_job(args):
    """Запуск одного товара в отдельном процессе/окне для параллельной работы."""
    url, provided_name, city, scrape_date, _expected_sellers = args
    label = provided_name or url
    job_started = time.perf_counter()
    for attempt in range(MAX_BROWSER_RETRIES):
        attempt_started = time.perf_counter()
        try:
            print(f"⏱️ Start job: {label} ({city}) attempt {attempt + 1}/{MAX_BROWSER_RETRIES}")
            with _job_timeout(JOB_TIMEOUT_SECONDS, label):
                _jitter_sleep()
                with sync_playwright() as p:
                    browser = p.chromium.launch(headless=True)
                    context = browser.new_context(
                        viewport={"width": 1600, "height": 900},
                        user_agent=_random_user_agent(),
                    )
                    page = context.new_page()
                    res = scrape_single_product(page, url, provided_name, city=city, scraped_date=scrape_date)
                    context.close()
                    browser.close()
                    attempt_elapsed = time.perf_counter() - attempt_started
                    total_elapsed = time.perf_counter() - job_started
                    result_rows = res or []
                    status = "ok" if result_rows else "empty"
                    print(f"⏱️ Done job: {label} ({city}) attempt {attempt + 1} in {attempt_elapsed:.1f}s (total {total_elapsed:.1f}s)")
                    _append_scrape_metric(
                        {
                            "ts": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
                            "city": city,
                            "product": provided_name or "",
                            "url": url,
                            "attempt": attempt + 1,
                            "status": status,
                            "elapsed_s": f"{attempt_elapsed:.1f}",
                            "total_s": f"{total_elapsed:.1f}",
                            "error": "",
                        }
                    )
                    return {
                        "records": result_rows,
                        "state": status,
                        "error": "",
                    }
        except NoOffersError as exc:
            attempt_elapsed = time.perf_counter() - attempt_started
            print(f"ℹ️ No offers for {label} ({city}) attempt {attempt + 1} after {attempt_elapsed:.1f}s")
            _append_scrape_metric(
                {
                    "ts": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
                    "city": city,
                    "product": provided_name or "",
                    "url": url,
                    "attempt": attempt + 1,
                    "status": "no_offers",
                    "elapsed_s": f"{attempt_elapsed:.1f}",
                    "total_s": f"{time.perf_counter() - job_started:.1f}",
                    "error": str(exc)[:200],
                }
            )
            return {
                "records": [],
                "state": "no_offers",
                "error": str(exc)[:200],
            }
        except JobTimeoutError as exc:
            attempt_elapsed = time.perf_counter() - attempt_started
            print(f"⏱️ Timeout job: {label} ({city}) attempt {attempt + 1} after {attempt_elapsed:.1f}s")
            _append_scrape_metric(
                {
                    "ts": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
                    "city": city,
                    "product": provided_name or "",
                    "url": url,
                    "attempt": attempt + 1,
                    "status": "timeout",
                    "elapsed_s": f"{attempt_elapsed:.1f}",
                    "total_s": f"{time.perf_counter() - job_started:.1f}",
                    "error": str(exc)[:200],
                }
            )
            if attempt + 1 == MAX_BROWSER_RETRIES:
                print(f"⏰ Job timeout for {label} ({city}): {exc}")
                raise
            print(f"⏰ Job timeout for {label} ({city}), retrying ({attempt + 1}/{MAX_BROWSER_RETRIES})")
            time.sleep(1 + attempt)
        except Exception as exc:
            attempt_elapsed = time.perf_counter() - attempt_started
            print(f"⏱️ Error job: {label} ({city}) attempt {attempt + 1} after {attempt_elapsed:.1f}s: {exc}")
            _append_scrape_metric(
                {
                    "ts": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
                    "city": city,
                    "product": provided_name or "",
                    "url": url,
                    "attempt": attempt + 1,
                    "status": "error",
                    "elapsed_s": f"{attempt_elapsed:.1f}",
                    "total_s": f"{time.perf_counter() - job_started:.1f}",
                    "error": str(exc)[:200],
                }
            )
            if attempt + 1 == MAX_BROWSER_RETRIES:
                print(f"❌ Failed in worker for {provided_name or url} ({city}): {exc}")
                raise
            backoff = 1 + attempt
            print(f"⚠️ Worker retry {attempt + 1} for {provided_name or url} ({city}) after error: {exc}")
            time.sleep(backoff)
    raise RuntimeError(f"Worker exhausted retries for {label} ({city})")


def scrape_products_from_csv(
    csv_paths: Path | Sequence[Path],
    cities: Sequence[str] | None = None,
    stop_event: threading.Event | None = None,
    progress_callback=None,
    chat_id: str | None = None,
    alert_only: bool = False,
    include_base: bool = True,
    change_alerts: bool = True,
    output_prefix: str | None = None,
):
    if isinstance(csv_paths, (str, Path)):
        csv_paths = [Path(csv_paths)]
    csv_paths = [Path(p) for p in csv_paths]
    base_dir = Path(__file__).resolve().parent

    mission_mode = not change_alerts

    frames = []
    for path in csv_paths:
        if not path.exists():
            print(f"ℹ️ CSV not found, skipping: {path}")
            continue
        df = pd.read_csv(path)
        if "product_link" not in df.columns:
            if "product_url" in df.columns:
                df["product_link"] = df["product_url"]
            else:
                print(f"ℹ️ CSV {path} skipped: no 'product_link' column")
                continue
        frames.append(df)

    if not frames:
        raise FileNotFoundError(f"No valid CSV files found in: {', '.join(str(p) for p in csv_paths)}")

    products_df = pd.concat(frames, ignore_index=True, sort=False)
    if not mission_mode:
        products_df = products_df.drop_duplicates(subset=["product_link"]).reset_index(drop=True)
    else:
        products_df = products_df.reset_index(drop=True)

    all_records = []
    job_meta_map = {}
    job_meta_list = []
    scrape_state_map: dict[str, dict[str, str]] = {}
    cities = list(cities) if cities else ["Алматы", "Астана", "Шымкент"]

    def _apply_meta_to_records(product_records: list[dict], meta: dict):
        for rec in product_records:
            if not rec.get("input_product"):
                rec["input_product"] = meta.get("product_name") or ""
            if not rec.get("product_code"):
                rec["product_code"] = meta.get("product_code") or ""
            if not rec.get("product_url"):
                rec["product_url"] = meta.get("product_url") or ""
            rec["expected_sellers"] = "; ".join(meta.get("expected_sellers", []))
            if meta.get("expected_price_kzt") is not None:
                rec["expected_price_kzt"] = meta.get("expected_price_kzt")

    def _product_key_from_meta(meta: dict) -> str:
        prod_key = meta.get("product_code") or extract_product_code(meta.get("product_url", "")) or meta.get("product_url")
        return f"{meta.get('city','')}|{prod_key or ''}"

    def _product_key_from_record(rec: dict) -> str:
        prod_key = rec.get("product_code") or extract_product_code(rec.get("product_url", "")) or rec.get("product_url")
        return f"{rec.get('region','')}|{prod_key or ''}"

    scrape_date = datetime.now(ALMATY_TZ).strftime("%d.%m.%Y")
    requested_workers = int(os.environ.get("SCRAPER_WORKERS", "6"))
    cpu_cap = max(2, (os.cpu_count() or 2))
    max_workers = max(2, min(requested_workers, cpu_cap, 6))
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    progress_state = _load_progress_state(PROGRESS_PATH)
    progress_state.update({
        "started_at": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
        "stage": "init",
        "stage_detail": "",
        # обнуляем прогресс для текущего запуска, чтобы не оставалось стейта от прошлых
        "cities": {
            city: {
                "total": 0,
                "done": 0,
                "completed": False,
                "workers": 0,
                "empty": 0,
                "error": 0,
                "crash": 0,
                "last_error": "",
            }
            for city in cities
        },
    })
    _save_progress_state(PROGRESS_PATH, progress_state)

    def _set_stage(stage: str, detail: str = ""):
        progress_state["stage"] = stage
        progress_state["stage_detail"] = detail
        history = progress_state.get("stage_history")
        if not isinstance(history, list):
            history = []
        history.append(
            {
                "ts": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
                "stage": stage,
                "detail": detail,
            }
        )
        if len(history) > 200:
            history = history[-200:]
        progress_state["stage_history"] = history
        _save_progress_state(PROGRESS_PATH, progress_state)

    for city in cities:
        if stop_event and stop_event.is_set():
            break
        _set_stage("scrape_city", city)
        jobs = []
        for _, row in products_df.iterrows():
            product_url = resolve_kaspi_link(str(row["product_link"]).strip())
            if not product_url:
                continue
            provided_name = str(row.get("product_name") or row.get("input_product") or "").strip()
            expected_raw = row.get("expected_sellers") if "expected_sellers" in row else row.get("seller", "")
            expected_sellers = tuple(parse_expected_sellers(expected_raw))
            row_regions_raw = str(row.get("region") or "").strip()
            if row_regions_raw:
                allowed_regions = [part.strip() for part in re.split(r"[;,/|]", row_regions_raw) if part and part.strip()]
                if allowed_regions and city not in allowed_regions:
                    continue
            expected_price = parse_price_value(row.get("target_price_kzt"))
            if expected_price is None:
                expected_price = parse_price_value(row.get("price_kzt"))
            product_code = extract_product_code(product_url)
            job = (product_url, provided_name, city, scrape_date, expected_sellers)
            jobs.append(job)
            input_product_value = row.get("input_product") or row.get("product_name") or provided_name
            original_region = row.get("region")
            source_order = int(row.name)
            meta = {
                "product_url": product_url,
                "product_name": provided_name,
                "city": city,
                "expected_sellers": list(expected_sellers),
                "product_code": product_code,
                "expected_price_kzt": expected_price,
                "mission_row_index": source_order,
                "input_row": {
                    "region": original_region,
                    "seller": row.get("seller"),
                    "input_product": input_product_value,
                    "product_url": row.get("product_url") or row.get("product_link") or product_url,
                    "price_kzt": parse_price_value(row.get("price_kzt")),
                    "bundle_of": row.get("bundle_of"),
                    "bundle_qty": row.get("bundle_qty"),
                },
            }
            job_meta_map[job] = meta
            job_meta_list.append(meta)
        if not jobs:
            continue

        city_state = progress_state.setdefault("cities", {}).get(city, {})
        city_state.update(
            {
                "total": len(jobs),
                "done": min(city_state.get("done", 0), len(jobs)),
                "completed": False,
                "empty": 0,
                "error": 0,
                "crash": 0,
                "last_error": "",
            }
        )
        progress_state["cities"][city] = city_state
        _save_progress_state(PROGRESS_PATH, progress_state)

        if progress_callback:
            progress_callback("city_start", city=city, total=len(jobs))

        city_attempt = 0
        while True:
            workers_for_city = min(max_workers, len(jobs))
            print(f"🚀 Starting parallel scrape for {len(jobs)} items in {city} with {workers_for_city} workers")
            progress_state["cities"][city]["workers"] = workers_for_city
            _save_progress_state(PROGRESS_PATH, progress_state)
            city_records: list[dict] = []
            transient_city_error = None
            try:
                with ProcessPoolExecutor(max_workers=workers_for_city) as executor:
                    future_map = {executor.submit(_scrape_product_job, job): job for job in jobs}
                    for future in as_completed(future_map):
                        if stop_event and stop_event.is_set():
                            executor.shutdown(cancel_futures=True)
                            break
                        job = future_map[future]
                        product_url, provided_name, city_label, _, _ = job
                        meta = job_meta_map.get(job, {})
                        label = provided_name or product_url
                        meta_key = _product_key_from_meta(meta)
                        try:
                            job_result = future.result()
                        except Exception as exc:
                            print(f"❌ Worker crashed for {label} ({city_label}): {exc}")
                            scrape_state_map[meta_key] = {
                                "state": "error",
                                "error": str(exc)[:200],
                            }
                            if _is_transient_worker_error(exc):
                                transient_city_error = exc
                                executor.shutdown(cancel_futures=True)
                                break
                            _update_last_processed({
                                "city": city_label,
                                "product": label,
                                "status": "error",
                                "error": str(exc),
                            })
                            progress_entry = progress_state["cities"].setdefault(city_label, {"total": len(jobs), "done": 0})
                            progress_entry["error"] = progress_entry.get("error", 0) + 1
                            progress_entry["last_error"] = str(exc)[:300]
                            progress_entry["done"] = progress_entry.get("done", 0) + 1
                            progress_state["last_job"] = {
                                "city": city_label,
                                "product": label,
                                "status": "error",
                            }
                            _save_progress_state(PROGRESS_PATH, progress_state)
                            continue
                        if isinstance(job_result, dict):
                            product_records = job_result.get("records") or []
                            scrape_state_map[meta_key] = {
                                "state": str(job_result.get("state") or ("ok" if product_records else "empty")),
                                "error": str(job_result.get("error") or "")[:200],
                            }
                        else:
                            product_records = job_result or []
                            scrape_state_map[meta_key] = {
                                "state": "ok" if product_records else "empty",
                                "error": "",
                            }
                        if not product_records:
                            scrape_state_name = (scrape_state_map.get(meta_key, {}).get("state") or "").strip().lower()
                            if scrape_state_name == "no_offers":
                                print(f"ℹ️ No offers confirmed for {label} ({city_label})")
                                _update_last_processed({"city": city_label, "product": label, "status": "no_offers"})
                            else:
                                print(f"⚠️ No seller data for {label} ({city_label})")
                                _update_last_processed({"city": city_label, "product": label, "status": "empty"})
                            progress_state["cities"][city_label]["empty"] = progress_state["cities"][city_label].get("empty", 0) + 1
                        else:
                            _apply_meta_to_records(product_records, meta)
                            city_records.extend(product_records)
                            _update_last_processed({
                                "city": city_label,
                                "product": label,
                                "status": "ok",
                                "sellers": len(product_records),
                            })

                        progress_state["cities"][city_label]["done"] = progress_state["cities"][city_label].get("done", 0) + 1
                        progress_state["last_job"] = {
                            "city": city_label,
                            "product": label,
                            "status": "ok" if product_records else "empty",
                        }
                        _save_progress_state(PROGRESS_PATH, progress_state)
                        if progress_callback:
                            progress_callback("progress", city=city_label, done=len(all_records) + len(city_records))
            except Exception as exc:
                if _is_transient_worker_error(exc):
                    transient_city_error = exc
                else:
                    raise

            if transient_city_error and city_attempt < SCRAPER_CITY_RETRIES:
                city_attempt += 1
                print(f"⚠️ City scrape crashed ({city}) due to {transient_city_error}. Retrying {city_attempt}/{SCRAPER_CITY_RETRIES}...")
                progress_state["cities"][city]["done"] = 0
                progress_state["cities"][city]["completed"] = False
                progress_state["cities"][city]["crash"] = progress_state["cities"][city].get("crash", 0) + 1
                progress_state["cities"][city]["last_error"] = str(transient_city_error)[:300]
                _save_progress_state(PROGRESS_PATH, progress_state)
                time.sleep(SCRAPER_CITY_RETRY_BACKOFF_SEC + city_attempt * 5)
                continue

            if transient_city_error and city_attempt >= SCRAPER_CITY_RETRIES:
                print(f"❌ City scrape failed after retries for {city}: {transient_city_error}")
                progress_state["cities"][city]["crash"] = progress_state["cities"][city].get("crash", 0) + 1
                progress_state["cities"][city]["last_error"] = str(transient_city_error)[:300]
            all_records.extend(city_records)
            break

        if progress_callback:
            progress_callback("city_done", city=city, total=len(jobs))
        progress_state["cities"][city]["completed"] = True
        _save_progress_state(PROGRESS_PATH, progress_state)
        _set_stage("city_done", city)

        if stop_event and stop_event.is_set():
            break

    if not stop_event or not stop_event.is_set():
        retry_flag = (os.environ.get("MISSION_SHYMKENT_RETRY", "1") or "1").strip().lower()
        if mission_mode and retry_flag not in {"0", "false", "no"} and "Шымкент" in cities:
            _set_stage("retry_shymkent", "Шымкент")
            retry_jobs = [job for job in job_meta_map.keys() if job[2] == "Шымкент"]
            if retry_jobs:
                print("🔁 Повторный запуск для Шымкента после основного прогона...")
                retry_map: dict[str, list[dict]] = {}
                workers_for_city = min(max_workers, len(retry_jobs))
                with ProcessPoolExecutor(max_workers=workers_for_city) as executor:
                    future_map = {executor.submit(_scrape_product_job, job): job for job in retry_jobs}
                    for future in as_completed(future_map):
                        job = future_map[future]
                        meta = job_meta_map.get(job, {})
                        key = _product_key_from_meta(meta)
                        try:
                            product_records = future.result()
                        except Exception as exc:
                            print(f"❌ Retry failed for {meta.get('product_name') or job[0]} (Шымкент): {exc}")
                            retry_map.setdefault(key, [])
                            continue
                        if product_records:
                            _apply_meta_to_records(product_records, meta)
                        retry_map[key] = product_records or []

                if retry_map:
                    other_records: list[dict] = []
                    existing_map: dict[str, list[dict]] = {}
                    for rec in all_records:
                        if rec.get("region") == "Шымкент":
                            existing_map.setdefault(_product_key_from_record(rec), []).append(rec)
                        else:
                            other_records.append(rec)
                    merged_shymkent: list[dict] = []
                    for key in set(existing_map) | set(retry_map):
                        new_records = retry_map.get(key)
                        if new_records:
                            merged_shymkent.extend(new_records)
                        else:
                            merged_shymkent.extend(existing_map.get(key, []))
                    all_records = other_records + merged_shymkent

    if mission_mode and all_records:
        retry_empty_flag = (os.environ.get("MISSION_RETRY_EMPTY", "1") or "1").strip().lower()
        if retry_empty_flag not in {"0", "false", "no"}:
            _set_stage("retry_empty", "")
            retry_limit = _parse_positive_int_env(os.environ.get("MISSION_RETRY_EMPTY_LIMIT")) or 6
            missing_meta = [meta for meta in job_meta_list if not _matched_records_for_meta(all_records, meta)]
            if missing_meta:
                retry_meta = missing_meta[:retry_limit]
                print(f"🔁 Повторный сбор карточек без продавцов: {len(retry_meta)}/{len(missing_meta)}")
                try:
                    with sync_playwright() as p:
                        browser = p.chromium.launch(headless=True)
                        context = browser.new_context(
                            viewport={"width": 1600, "height": 900},
                            user_agent=_random_user_agent(),
                        )
                        page = context.new_page()
                        for meta in retry_meta:
                            label = meta.get("product_name") or meta.get("product_url")
                            meta_key = _product_key_from_meta(meta)
                            try:
                                records = scrape_single_product(
                                    page,
                                    meta.get("product_url"),
                                    input_name=meta.get("product_name"),
                                    city=meta.get("city") or "Алматы",
                                    scraped_date=scrape_date,
                                ) or []
                            except NoOffersError as exc:
                                scrape_state_map[meta_key] = {
                                    "state": "no_offers",
                                    "error": str(exc)[:200],
                                }
                                continue
                            except Exception as exc:
                                print(f"⚠️ Retry failed for {label}: {exc}")
                                scrape_state_map[meta_key] = {
                                    "state": "error",
                                    "error": str(exc)[:200],
                                }
                                continue
                            if records:
                                _apply_meta_to_records(records, meta)
                                all_records.extend(records)
                                scrape_state_map[meta_key] = {"state": "ok", "error": ""}
                            else:
                                scrape_state_map[meta_key] = {
                                    "state": "empty",
                                    "error": "",
                                }
                        context.close()
                        browser.close()
                except Exception as exc:
                    print(f"⚠️ Retry session failed: {exc}")

    if stop_event and stop_event.is_set():
        _set_stage("stopped", "")
        if progress_callback:
            progress_callback("stopped", city=None, done=len(all_records))
        return
    if not all_records and not mission_mode:
        _set_stage("no_data", "")
        print("⚠️ No data collected; skipping Excel export")
        if progress_callback:
            progress_callback("no_data", city=None, done=0)
        return

    missing_alerts = find_missing_expected_sellers(all_records, job_meta_list)
    contract_price_alerts, better_price_alerts = build_price_alerts(all_records, job_meta_list)

    mismatch_rows: list[int] = []
    highlight_cols: list[str] | None = None

    if mission_mode:
        _set_stage("build_report", "mission")
        mission_rows: list[dict] = []
        bundle_lookup: dict[tuple[str, str, str], dict[str, Any]] = {}
        review_progress_state = _load_review_progress_state(REVIEW_PROGRESS_PATH)
        build_timeout_sec = int(os.environ.get("MISSION_BUILD_TIMEOUT_SEC", "180"))
        build_deadline = time.time() + max(15, build_timeout_sec)
        build_timed_out = False
        build_error: str | None = None

        try:
            if os.environ.get("DEBUG_MISSION_INPUT", "0").strip() == "1" and job_meta_list:
                sample = job_meta_list[0].get("input_row") or {}
                print(f"🧪 Mission input sample: {sample}")

            def _sorted_offers(records: list[dict]) -> list[dict]:
                """Список цен с одним значением на продавца (берём минимальную)."""
                best_by_seller: dict[str, dict] = {}
                for rec in records:
                    price = rec.get("price_kzt")
                    if price is None:
                        continue
                    norm = normalize_seller_name(rec.get("seller", ""))
                    current = best_by_seller.get(norm)
                    if current is None or (current.get("price_kzt") or 10**12) > price:
                        best_by_seller[norm] = rec
                return sorted(best_by_seller.values(), key=lambda rec: rec.get("price_kzt"))

            exclude_raw = os.environ.get("MISSION_PARTNER_SELLERS")
            if exclude_raw is None:
                exclude_raw = os.environ.get("MISSION_EXCLUDE_SELLERS")
            exclude_sellers = parse_seller_list(exclude_raw, MISSION_PARTNER_SELLERS_DEFAULT)
            mission_timestamp = datetime.now(ALMATY_TZ).strftime("%d.%m.%Y %H:%M")
            if "bundle_of" in products_df.columns or "bundle_qty" in products_df.columns:
                for _, row in products_df.iterrows():
                    product_url_value = row.get("product_url") or row.get("product_link")
                    if _is_missing_cell(product_url_value):
                        continue
                    key = (
                        str(row.get("region") or "").strip(),
                        str(row.get("seller") or "").strip(),
                        str(product_url_value).strip(),
                    )
                    bundle_lookup[key] = {
                        "bundle_of": row.get("bundle_of"),
                        "bundle_qty": row.get("bundle_qty"),
                    }
            for idx, meta in enumerate(job_meta_list, start=1):
                if time.time() > build_deadline:
                    build_timed_out = True
                    print("⚠️ Mission report build timed out, saving partial results.")
                    _set_stage("build_report_timeout", f"{idx}/{len(job_meta_list)}")
                    break
                if idx % 10 == 0:
                    _set_stage("build_report", f"{idx}/{len(job_meta_list)}")
                base = meta.get("input_row") or {}
                seller_hint = (base.get("seller") or (meta.get("expected_sellers") or [""])[0] or "").strip()
                seller_norm = normalize_seller_name(seller_hint)
                scrape_state = scrape_state_map.get(_product_key_from_meta(meta), {})
                scrape_state_name = (scrape_state.get("state") or "").strip().lower()
                scrape_error = (scrape_state.get("error") or "").strip()
                matched = _matched_records_for_meta(all_records, meta)
                price_levels = _sorted_offers(matched)
                filtered_levels = [
                    rec for rec in price_levels
                    if not seller_in_list(rec.get("seller", ""), exclude_sellers)
                ]
                excluded_levels = [
                    rec for rec in price_levels
                    if seller_in_list(rec.get("seller", ""), exclude_sellers)
                ]
                actual_rec = next(
                    (
                        rec
                        for rec in price_levels
                        if seller_matches(normalize_seller_name(rec.get("seller", "")), seller_norm)
                    ),
                    None,
                )
                actual_price = actual_rec.get("price_kzt") if actual_rec else None
                input_expected = base.get("price_kzt")
                expected_price = meta.get("expected_price_kzt")
                if expected_price is None:
                    expected_price = input_expected

                review_counts = []
                for rec in matched:
                    val = rec.get("review_count")
                    if val is None:
                        continue
                    try:
                        val_int = int(val)
                    except Exception:
                        continue
                    if val_int >= 0:
                        review_counts.append(val_int)
                review_count_value = max(review_counts) if review_counts else None
                review_progress_text = None
                review_key = meta.get("product_code") or extract_product_code(
                    base.get("product_url") or meta.get("product_url") or ""
                ) or (base.get("product_url") or meta.get("product_url") or "")
                if review_count_value is not None and review_key:
                    prev = review_progress_state.get(str(review_key), {}).get("count")
                    if prev is None:
                        review_progress_text = str(review_count_value)
                    else:
                        delta = review_count_value - int(prev)
                        sign = "+" if delta >= 0 else ""
                        review_progress_text = f"{review_count_value} ({sign}{delta})"
                    review_progress_state[str(review_key)] = {
                        "count": int(review_count_value),
                        "updated_at": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
                    }
                elif review_key:
                    prev = review_progress_state.get(str(review_key), {}).get("count")
                    if prev is not None:
                        review_count_value = int(prev)
                        review_progress_text = str(review_count_value)
                best_offer = filtered_levels[0] if filtered_levels else None
                best_price = best_offer.get("price_kzt") if best_offer else None
                best_seller = best_offer.get("seller") if best_offer else None
                better_analysis = ""
                if not price_levels:
                    if scrape_state_name == "no_offers":
                        better_analysis = "Нет продавцов"
                    elif scrape_state_name == "error":
                        better_analysis = "Ошибка скрейпа"
                    else:
                        better_analysis = "Не удалось получить продавцов"
                elif best_price is None:
                    better_analysis = "Нет внешних офферов"
                else:
                    diff_txt = ""
                    if actual_price is not None:
                        gap = actual_price - best_price
                        if gap > 0:
                            diff_txt = f" (+{format_kzt(gap)})"
                        elif gap < 0:
                            diff_txt = f" (−{format_kzt(-gap)})"
                    better_analysis = f"{best_seller or 'нет'} — {format_kzt(best_price)}{diff_txt}"

                second_analysis = ""
                second_offer = filtered_levels[1] if len(filtered_levels) > 1 else None
                if second_offer:
                    ref_price = best_price if best_price is not None else second_offer.get("price_kzt")
                    diff = second_offer.get("price_kzt") - (ref_price or 0)
                    if diff > 0:
                        diff_txt = f" (+{format_kzt(diff)})"
                    elif diff < 0:
                        diff_txt = f" (−{format_kzt(-diff)})"
                    else:
                        diff_txt = ""
                    is_best_ours = actual_rec and best_seller and normalize_seller_name(best_seller) == seller_norm
                    prefix = "После нас:" if is_best_ours else "Следующий:"
                    second_analysis = f"{prefix} {second_offer.get('seller')} — {format_kzt(second_offer.get('price_kzt'))}{diff_txt}"
                else:
                    second_analysis = "Нет второй цены"

                if excluded_levels:
                    shown = excluded_levels[:2]
                    items = "; ".join(
                        f"{rec.get('seller')} — {format_kzt(rec.get('price_kzt'))}" for rec in shown
                    )
                    excluded_note = f"Исключены: {items}"
                    second_analysis = f"{second_analysis}\n{excluded_note}" if second_analysis else excluded_note

                if not matched and scrape_state_name == "no_offers":
                    status_text = "Нет продавцов на карточке"
                elif not matched and scrape_state_name in {"error", "empty"}:
                    status_text = "Ошибка скрейпа"
                elif not matched:
                    status_text = "Ошибка скрейпа"
                elif actual_rec is None:
                    status_text = "Продавец отсутствует"
                elif actual_price is None:
                    status_text = "Нет цены"
                elif expected_price is None:
                    status_text = "Нет целевой цены"
                else:
                    diff = actual_price - expected_price
                    if diff == 0:
                        status_text = "OK"
                    elif abs(diff) <= PRICE_TOLERANCE_KZT:
                        status_text = f"OK (в допуске ±{format_kzt(PRICE_TOLERANCE_KZT)}, отклонение {format_kzt(diff)})"
                    elif diff > 0:
                        status_text = f"ПРОБЛЕМА: наша цена выше цели на {format_kzt(diff)}"
                    else:
                        status_text = f"ПРОБЛЕМА: наша цена ниже цели на {format_kzt(-diff)}"

                if status_text == "Ошибка скрейпа":
                    second_analysis = scrape_error[:180] if scrape_error else "Нужно повторить карточку"

                if status_text.strip().lower().startswith("ok") and actual_price is not None and best_price is not None:
                    best_gap = actual_price - best_price
                    if best_gap > PRICE_TOLERANCE_KZT:
                        status_text = f"Внимание: не самое выгодное предложение (+{format_kzt(best_gap)})"

                input_product_value = base.get("input_product")
                if not isinstance(input_product_value, str) or not input_product_value.strip():
                    input_product_value = meta.get("product_name") or ""

                bundle_of = base.get("bundle_of")
                bundle_qty = base.get("bundle_qty")
                if bundle_lookup:
                    region_key = base.get("region")
                    if _is_missing_cell(region_key):
                        region_key = meta.get("city")
                    seller_key = (base.get("seller") or seller_hint or "").strip()
                    product_url_key = base.get("product_url") or meta.get("product_url") or ""
                    lookup = bundle_lookup.get(
                        (str(region_key or "").strip(), seller_key, str(product_url_key).strip())
                    )
                    if lookup:
                        if _is_missing_cell(bundle_of):
                            bundle_of = lookup.get("bundle_of")
                        if _is_missing_cell(bundle_qty):
                            bundle_qty = lookup.get("bundle_qty")

                mission_rows.append({
                    "_order": meta.get("mission_row_index", 0),
                    "region": base.get("region") or meta.get("city"),
                    "seller": seller_hint,
                    "input_product": input_product_value,
                    "product": (actual_rec and actual_rec.get("product"))
                    or (price_levels[0].get("product") if price_levels else meta.get("product_name") or ""),
                    "product_url": base.get("product_url") or meta.get("product_url"),
                    "review_count": review_count_value,
                    "review_progress": review_progress_text,
                    "price_kzt": expected_price,
                    "actual_price_kzt": actual_price,
                    "status": status_text,
                    "best_price_kzt": best_price,
                    "better_price_analysis": better_analysis,
                    "second_price_analysis": second_analysis,
                    "scraped_at": mission_timestamp,
                    "bundle_of": bundle_of,
                    "bundle_qty": bundle_qty,
                })

            mission_rows = sorted(mission_rows, key=lambda row: row.get("_order", 0))
            for row in mission_rows:
                row.pop("_order", None)
            df = pd.DataFrame(mission_rows, columns=[
                "region",
                "seller",
                "input_product",
                "product",
                "product_url",
                "review_count",
                "review_progress",
                "price_kzt",
                "actual_price_kzt",
                "status",
                "best_price_kzt",
                "better_price_analysis",
                "second_price_analysis",
                "scraped_at",
            ] + BUNDLE_COLUMNS)
            df["price_kzt"] = df["price_kzt"].astype("Int64")
            df["actual_price_kzt"] = df["actual_price_kzt"].astype("Int64")
            df["best_price_kzt"] = df["best_price_kzt"].astype("Int64")
            if os.environ.get("DEBUG_MISSION_INPUT", "0").strip() == "1":
                print("🧪 Mission bundle columns preview:", df[["bundle_of", "bundle_qty"]].head(3).to_dict(orient="records"))
            ok_mask = df["status"].fillna("").str.strip().str.lower().str.startswith("ok")
            mismatch_rows = df.index[~ok_mask].tolist()
            highlight_cols = ["actual_price_kzt", "status"]
            _save_review_progress_state(REVIEW_PROGRESS_PATH, review_progress_state)
            if build_timed_out and df.empty:
                _set_stage("build_report_timeout", "empty")
        except Exception as exc:
            build_error = str(exc)
            _set_stage("build_report_error", build_error[:120])
            print(f"⚠️ Mission report build failed: {exc}")
            df = pd.DataFrame(
                [{
                    "region": "",
                    "seller": "",
                    "input_product": "",
                    "product": "",
                    "product_url": "",
                    "review_count": None,
                    "review_progress": "",
                    "price_kzt": None,
                    "actual_price_kzt": None,
                    "status": f"ERROR: {build_error}",
                    "best_price_kzt": None,
                    "better_price_analysis": "",
                    "second_price_analysis": "",
                    "scraped_at": datetime.now(ALMATY_TZ).strftime("%d.%m.%Y %H:%M"),
                }],
                columns=[
                    "region",
                    "seller",
                    "input_product",
                    "product",
                    "product_url",
                    "review_count",
                    "review_progress",
                    "price_kzt",
                    "actual_price_kzt",
                    "status",
                    "best_price_kzt",
                    "better_price_analysis",
                    "second_price_analysis",
                    "scraped_at",
                ] + BUNDLE_COLUMNS,
            )
            mismatch_rows = [0]
            highlight_cols = ["status"]
    else:
        _set_stage("build_report", "main")
        df = pd.DataFrame(all_records, columns=[
            "input_product",
            "product",
            "product_url",
            "seller",
            "expected_sellers",
            "price_kzt",
            "expected_price_kzt",
            "delivery_details",
            "scraped_date",
            "region",
            "product_code",
            "review_count",
        ])

        # Preserve integers while allowing missing values
        df["price_kzt"] = df["price_kzt"].astype("Int64")
        if "expected_price_kzt" not in df.columns:
            df["expected_price_kzt"] = pd.Series(dtype="Int64")
        df["expected_price_kzt"] = df["expected_price_kzt"].astype("Int64")
        if "review_count" not in df.columns:
            df["review_count"] = pd.Series(dtype="Int64")
        df["review_count"] = df["review_count"].astype("Int64")

        df["price_delta_kzt"] = df["price_kzt"] - df["expected_price_kzt"]
        mismatch_mask = df["expected_price_kzt"].notna() & df["price_kzt"].notna()
        df.loc[~mismatch_mask, "price_delta_kzt"] = pd.NA
        df["price_delta_kzt"] = df["price_delta_kzt"].astype("Int64")

        def _delta_status(row):
            if pd.isna(row["expected_price_kzt"]) or pd.isna(row["price_kzt"]):
                return "—"
            if row["price_delta_kzt"] == 0:
                return "OK"
            return "Дороже" if row["price_delta_kzt"] > 0 else "Дешевле"

        df["price_match_status"] = df.apply(_delta_status, axis=1)
        _backfill_reviews_db_for_products(df)
        mismatch_rows = df.index[mismatch_mask & (df["price_delta_kzt"] != 0)].tolist()
        highlight_cols = None

    # Сравнение со снимком
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    snapshot_path = STATE_DIR / "last_snapshot.json"
    if output_prefix:
        norm_prefix = re.sub(r"[^a-zA-Z0-9_.-]+", "_", str(output_prefix)).strip("._")
        default_mission_prefix = (os.environ.get("MISSION_FILE_PREFIX") or "mission_april").strip()
        if norm_prefix and norm_prefix not in {"kaspi_prices", default_mission_prefix}:
            snapshot_path = STATE_DIR / f"last_snapshot_{norm_prefix}.json"
    prev_snapshot = load_snapshot(snapshot_path)
    first_snapshot = len(prev_snapshot) == 0
    curr_snapshot = {}
    changes = []

    product_keys = set()
    for rec in all_records:
        key = _snapshot_key(rec)
        prod_key = rec.get("product_code") or extract_product_code(rec.get("product_url", "")) or rec.get("product_url")
        if prod_key:
            product_keys.add(f"{rec.get('region','')}|{prod_key}")
        curr_snapshot[key] = {
            "price_kzt": rec.get("price_kzt"),
            "installment_kzt": rec.get("installment_kzt"),
            "seller": rec.get("seller"),
            "product": rec.get("product"),
            "product_code": rec.get("product_code"),
            "region": rec.get("region"),
            "product_url": rec.get("product_url"),
        }

    if not first_snapshot:
        for key, cur in curr_snapshot.items():
            prev = prev_snapshot.get(key)
            if prev:
                if prev.get("price_kzt") != cur.get("price_kzt") or prev.get("installment_kzt") != cur.get("installment_kzt"):
                    changes.append(f"💸 {cur.get('product')} — {cur.get('region')} — {cur.get('seller')}: цена {prev.get('price_kzt')} → {cur.get('price_kzt')}")
            else:
                changes.append(f"🆕 {cur.get('product')} — {cur.get('region')} — новый продавец {cur.get('seller')} цена {cur.get('price_kzt')}")

        for key, prev in prev_snapshot.items():
            if key not in curr_snapshot:
                prev_prod_key = f"{prev.get('region','')}|{prev.get('product_code') or extract_product_code(prev.get('product_url','')) or prev.get('product_url')}"
                # если этого продукта нет в текущем списке скрейпа, пропускаем
                if prev_prod_key not in product_keys:
                    continue
                changes.append(f"❌ {prev.get('product')} — {prev.get('region')} — продавец пропал: {prev.get('seller')}")
    else:
        print("ℹ️ Нет предыдущего снимка, сохраняю базу и алерты по изменениям будут со следующего запуска.")

    save_snapshot(snapshot_path, curr_snapshot)

    results_dir = Path(__file__).resolve().parent / "RESULTS"
    results_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(ALMATY_TZ).strftime("%Y-%m-%d_%H-%M-%S")
    default_mission_prefix = (os.environ.get("MISSION_FILE_PREFIX") or "mission_april").strip() or "mission_april"
    prefix = output_prefix or (default_mission_prefix if mission_mode else "kaspi_prices")
    output_path = results_dir / f"{prefix}_{timestamp}.xlsx"

    alerts_sent = False
    alert_lines: list[str] = []
    if change_alerts:
        if changes:
            alert_lines.append("⚠️ Изменения:")
            alert_lines.extend(changes)
            alert_lines.append("")
        if missing_alerts:
            alert_lines.append("⚠️ Отсутствуют ожидаемые продавцы:")
            for alert in missing_alerts:
                sku_info = f" | SKU: {alert['product_code']}" if alert.get("product_code") else ""
                alert_lines.append(f"{alert['product']} — {alert['city']}{sku_info}")
                alert_lines.append(f"Ожидал: {', '.join(alert['missing'])}")
                current_line = ", ".join(alert.get("found") or ["нет продавцов на карточке"])
                alert_lines.append(f"Сейчас: {current_line}")
                alert_lines.append("")
        if contract_price_alerts:
            alert_lines.append("⚠️ Несоответствие заданной цене:")
            alert_lines.extend(contract_price_alerts)
            alert_lines.append("")
        if better_price_alerts:
            alert_lines.append("⚠️ Есть более выгодные предложения:")
            alert_lines.extend(better_price_alerts)
            alert_lines.append("")
        if first_snapshot and not alert_lines and alert_only:
            alert_lines.append("📌 Сохранил базовый снимок. Изменения по ценам и продавцам пришлю со следующего запуска.")
        if alert_lines:
            if TELEGRAM_FILE_ONLY:
                print("ℹ️ Telegram alerts suppressed (TELEGRAM_FILE_ONLY=1).")
            else:
                try:
                    send_telegram_message("\n".join(alert_lines).strip(), chat_id=chat_id)
                    alerts_sent = True
                except Exception as exc:
                    print(f"⚠️ Could not send Telegram alerts: {exc}")

    if mission_mode:
        try:
            df = _attach_bundle_columns(df, base_dir)
        except Exception as exc:
            print(f"⚠️ Ошибка добавления колонок Bundle 1 SKU: {exc}")
        df = _ensure_bundle_columns(df)
        df = _reorder_mission_columns(df)

    should_write_file = (not alert_only) or mission_mode
    image_path: Path | None = None
    slice_preview_images: list[dict[str, Any]] = []
    if should_write_file:
        extra_sheets: dict[str, pd.DataFrame] | None = None
        if not mission_mode:
            reviews_sheet = _build_reviews_summary_sheet(df)
            if reviews_sheet is not None and not reviews_sheet.empty:
                extra_sheets = {"Reviews": reviews_sheet}
        _set_stage("export_start", output_path.name)
        try:
            _write_dataframe_table(
                df,
                output_path,
                table_name="kaspi_data",
                highlight_rows=mismatch_rows,
                highlight_columns=highlight_cols,
                extra_sheets=extra_sheets,
            )
        except Exception as exc:
            print(f"⚠️ Export failed, falling back to plain Excel: {exc}")
            _set_stage("export_fallback", str(exc)[:120])
            try:
                df.to_excel(output_path, index=False)
            except Exception as exc2:
                _set_stage("export_error", str(exc2)[:120])
                raise
        _set_stage("export_done", output_path.name)
        if mission_mode:
            image_path = results_dir / f"mission_preview_{timestamp}.png"
            try:
                _set_stage("preview_start", image_path.name)
                render_mission_image(df, image_path)
            except Exception as exc:
                print(f"⚠️ Could not render mission preview: {exc}")
                image_path = None
            try:
                slice_preview_images = render_mission_slice_images(df, results_dir, timestamp=timestamp)
            except Exception as exc:
                print(f"⚠️ Could not render mission slice previews: {exc}")
                slice_preview_images = []
            _set_stage("preview_done", image_path.name if image_path else output_path.name)
        print(f"✅ Done! Saved file {output_path.name} in {results_dir}")
        summary = (
            f"Kaspi scrape finished\n"
            f"Rows: {len(df)}\n"
            f"Cities: {', '.join(cities)}\n"
            f"File: {output_path.name}"
        )
        telegram_caption = None if mission_mode else summary
        if not (mission_mode and MISSION_SUPPRESS_TELEGRAM):
            _set_stage("send_start", output_path.name)
            try:
                send_telegram_file(output_path, caption=telegram_caption, chat_id=chat_id)
            except Exception as exc:
                print(f"⚠️ Could not send Telegram file: {exc}")
            if mission_mode:
                previews_to_send = slice_preview_images
                if not previews_to_send and image_path and image_path.exists():
                    previews_to_send = [{"path": image_path, "caption": "Сводка по миссии"}]
                for preview in previews_to_send:
                    preview_path = preview.get("path")
                    preview_caption = preview.get("caption") or "Сводка по миссии"
                    if not preview_path or not Path(preview_path).exists():
                        continue
                    try:
                        send_telegram_photo(Path(preview_path), caption=preview_caption, chat_id=chat_id)
                    except Exception as exc:
                        print(f"⚠️ Could not send mission preview {preview_path}: {exc}")
            _set_stage("send_done", output_path.name)
        else:
            print("ℹ️ Mission Telegram sends suppressed (MISSION_SUPPRESS_TELEGRAM=1).")
            _set_stage("send_skipped", "mission_suppressed")
        if mission_mode and image_path and image_path.exists():
            keep_n = int(os.environ.get("MISSION_PREVIEW_KEEP", "10") or "10")
            _cleanup_old_previews(results_dir, keep=max(1, keep_n))
        if not alert_only:
            try:
                send_results_email(output_path, subject="Kaspi results", body=summary)
            except Exception as exc:
                print(f"⚠️ Could not send email: {exc}")
            if include_base:
                STATE_DIR.mkdir(parents=True, exist_ok=True)
                marker = STATE_DIR / "latest_full.txt"
                try:
                    marker.write_text(str(output_path), encoding="utf-8")
                except Exception:
                    pass
    else:
        if not alerts_sent:
            print("ℹ️ alert_only enabled and no changes; nothing sent.")

    try:
        _set_stage("finished", "")
        if PROGRESS_PATH.exists():
            PROGRESS_PATH.unlink()
    except Exception:
        pass

def _resolve_csv_paths(
    base_dir: Path,
    extra_csv_paths: Path | Sequence[Path] | str | None = None,
    include_base: bool = True,
) -> list[Path]:
    base_csv = base_dir / "my_products.csv"
    paths: list[Path] = [base_csv] if include_base else []

    def add_path(p: Path):
        if not p.is_absolute():
            p = base_dir / p
        if p.exists():
            paths.append(p)
        else:
            print(f"ℹ️ Extra CSV not found, skipping: {p}")

    if extra_csv_paths:
        if isinstance(extra_csv_paths, (str, Path)):
            extra_csv_paths = [extra_csv_paths]
        for part in extra_csv_paths:
            if isinstance(part, str):
                for split_part in re.split(r"[;,]", part):
                    split_part = split_part.strip()
                    if split_part:
                        add_path(Path(split_part))
            else:
                add_path(Path(part))
    else:
        extra_env = os.environ.get("EXTRA_PRODUCTS_CSV")
        if extra_env:
            for part in re.split(r"[;,]", extra_env):
                part = part.strip()
                if part:
                    add_path(Path(part))
        else:
            default_extra = base_dir / "my_products_extra.csv"
            if default_extra.exists():
                add_path(default_extra)

    return paths


def _load_reviews_jobs(csv_path: Path) -> list[dict]:
    jobs: list[dict] = []
    if not csv_path.exists():
        print(f"Reviews: CSV not found: {csv_path}")
        return jobs
    with open(csv_path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        if not reader.fieldnames:
            print(f"Reviews: CSV has no header: {csv_path}")
            return jobs
        for row in reader:
            url = (row.get("link") or row.get("url") or row.get("product_url") or row.get("product_link") or "").strip()
            sku = (row.get("sku") or row.get("name") or row.get("product") or row.get("title") or "").strip()
            if not url:
                continue
            jobs.append({"sku": sku, "url": url})
    return jobs


def run_reviews_scrape(
    reviews_csv: Path | str | None = None,
    city: str | None = None,
    output_dir: Path | None = None,
):
    base_dir = Path(__file__).resolve().parent
    reviews_csv = Path(reviews_csv) if reviews_csv else base_dir / "reviews.csv"
    output_dir = output_dir or (base_dir / "RESULTS")
    output_dir.mkdir(parents=True, exist_ok=True)

    jobs = _load_reviews_jobs(Path(reviews_csv))
    if not jobs:
        print("Reviews: no jobs found, exiting.")
        return None

    city = city or os.environ.get("REVIEWS_CITY", "Алматы")
    page_size = int(os.environ.get("REVIEWS_PAGE_SIZE", "20") or "20")
    max_pages = int(os.environ.get("REVIEWS_MAX_PAGES", "200") or "200")
    max_reviews = os.environ.get("REVIEWS_MAX_PER_PRODUCT")
    max_reviews = int(max_reviews) if max_reviews else None

    timestamp = datetime.now(ALMATY_TZ).strftime("%Y-%m-%d_%H-%M-%S")
    scraped_at = datetime.now(ALMATY_TZ).isoformat(timespec="seconds")
    output_path = output_dir / f"reviews_{timestamp}.csv"

    fieldnames = [
        "sku",
        "product_url",
        "product_id",
        "review_id",
        "author",
        "rating",
        "pros",
        "cons",
        "comment",
        "helpful_count",
        "review_date",
    ]

    REVIEWS_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(REVIEWS_DB_PATH)
    try:
        _init_reviews_db(conn)
        with open(output_path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=fieldnames)
            writer.writeheader()

            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True, args=["--disable-dev-shm-usage"])
                context = browser.new_context(
                    viewport={"width": 1600, "height": 900},
                    user_agent=_random_user_agent(),
                )
                for idx, job in enumerate(jobs, start=1):
                    sku = job.get("sku") or ""
                    url = job.get("url") or ""
                    page = context.new_page()
                    try:
                        resolved_url = _open_product_page_for_reviews(page, url, city)
                        product_id = _resolve_product_id_from_page(page, resolved_url or url)
                        if not product_id:
                            print(f"⚠️ Reviews: product id not found for {sku or url}")
                            continue
                        entries = _collect_all_reviews(page, product_id, page_size, max_pages, max_reviews)
                        print(f"✅ Reviews: {sku or product_id}: {len(entries)} отзывов")
                        db_rows = []
                        for entry in entries:
                            row = _extract_review_row(entry)
                            review_uid = _review_entry_id(entry)
                            review_date = row.get("review_date") or ""
                            if review_date:
                                date_source = "review"
                            else:
                                review_date = scraped_at[:10]
                                date_source = "scrape"
                            writer.writerow(
                                {
                                    "sku": sku,
                                    "product_url": resolved_url or url,
                                    "product_id": product_id,
                                    **row,
                                }
                            )
                            db_rows.append(
                                {
                                    "review_uid": review_uid,
                                    "product_id": product_id,
                                    "sku": sku,
                                    "review_id": row.get("review_id") or review_uid,
                                    "author": row.get("author"),
                                    "rating": row.get("rating"),
                                    "pros": row.get("pros"),
                                    "cons": row.get("cons"),
                                    "comment": row.get("comment"),
                                    "helpful_count": row.get("helpful_count") or None,
                                    "review_date": review_date,
                                    "date_source": date_source,
                                    "scraped_at": scraped_at,
                                    "product_url": resolved_url or url,
                                }
                            )
                        _store_review_rows(conn, db_rows)
                    except Exception as exc:
                        print(f"⚠️ Reviews: error for {sku or url}: {exc}")
                    finally:
                        page.close()
                context.close()
                browser.close()
    finally:
        conn.close()

    print(f"✅ Reviews: saved {output_path}")
    return output_path


def run_keyword_search_report(
    extra_csv_paths: Path | Sequence[Path] | str | None = None,
    include_base: bool = True,
    chat_id: str | None = None,
):
    """Запуск отдельного поиска по ключевым словам."""
    base_dir = Path(__file__).resolve().parent
    csvs = _resolve_csv_paths(base_dir, extra_csv_paths, include_base=include_base)
    frames = []
    for path in csvs:
        if not path.exists():
            print(f"Keyword search: CSV not found, skipping: {path}")
            continue
        try:
            df = pd.read_csv(path)
        except Exception as exc:
            print(f"Keyword search: failed to read {path}: {exc}")
            continue
        if "product_link" not in df.columns:
            if "product_url" in df.columns:
                df["product_link"] = df["product_url"]
            else:
                print(f"Keyword search: CSV {path} skipped: no 'product_link' column")
                continue
        frames.append(df)

    if not frames:
        print("Keyword search: no CSV data, skipping")
        return None

    products_df = pd.concat(frames, ignore_index=True, sort=False)
    own_product_codes = _extract_product_codes_from_df(products_df)
    if not own_product_codes:
        print("Keyword search: no product codes found, skipping")
        return None

    results_dir = base_dir / "RESULTS"
    results_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(ALMATY_TZ).strftime("%Y-%m-%d_%H-%M-%S")
    return _build_keyword_search_report(base_dir, results_dir, timestamp, own_product_codes, chat_id=chat_id)


def run_kaspi_scrape(
    cities: Sequence[str] | None = None,
    extra_csv_paths: Path | Sequence[Path] | str | None = None,
    include_base: bool = True,
    chat_id: str | None = None,
    stop_event: threading.Event | None = None,
    progress_callback=None,
    alert_only: bool = False,
    change_alerts: bool = True,
    output_prefix: str | None = None,
    priority: str = "primary",
):
    """Удобный вызов из чат-бота: старт скрейпа и алертов."""
    with _exclusive_scrape_lock(priority=priority):
        base_dir = Path(__file__).resolve().parent
        csvs = _resolve_csv_paths(base_dir, extra_csv_paths, include_base=include_base)
        if progress_callback:
            progress_callback("start", city=None, done=0)
        scrape_products_from_csv(
            csvs,
            cities=cities or ["Алматы", "Астана", "Шымкент"],
            stop_event=stop_event,
            progress_callback=progress_callback,
            chat_id=chat_id,
            alert_only=alert_only,
            include_base=include_base,
            change_alerts=change_alerts,
            output_prefix=output_prefix,
        )
        if stop_event and stop_event.is_set():
            return
        if progress_callback:
            progress_callback("finished", city=None, done=None)


def bot_help_text() -> str:
    """Подсказка для /help в боте."""
    base_dir = Path(__file__).resolve().parent
    default_extra = base_dir / "my_products_extra.csv"
    return (
        "Что могу:\n"
        "• /add <ссылка> | <имя товара> | <продавцы через ;> — добавить товар в временный список\n"
        "• /run — запустить сбор по основному списку + временным\n"
        "• /run_extra — запустить сбор только по временным\n"
        "• /run_keywords — поиск по ключевым словам (search_keywords.csv)\n"
        "• /run_natasha_prices — скрейп по prices_natasha.csv (для остальных отдаёт последний файл)\n"
        "• /research <запрос> — быстрый интернет-ресёрч (Tavily + GPT)\n"
        "• /research_deep <запрос> — глубокий ресёрч с расширенным списком источников\n"
        "• /research_more <вопрос> — уточнить последний ресёрч\n"
        "• /help — показать эту подсказку\n"
        f"По умолчанию города: Алматы, Астана, Шымкент. Доп. список: {default_extra.name} или переменная EXTRA_PRODUCTS_CSV. "
        "Список конкурентов: переменная COMPETITOR_PRODUCTS_CSV или файл competitor_products.csv."
    )


if __name__ == "__main__":
    if len(sys.argv) > 1:
        cmd = sys.argv[1].lower()
        if cmd == "analyze":
            if len(sys.argv) < 3:
                print("Usage: python Scraper_Kaspi.py analyze <product_url> [city]", file=sys.stderr)
                sys.exit(1)
            url_arg = sys.argv[2]
            city_arg = sys.argv[3] if len(sys.argv) > 3 else "Алматы"
            run_product_analysis(url_arg, city=city_arg)
        elif cmd == "reviews":
            reviews_csv = sys.argv[2] if len(sys.argv) > 2 else None
            city_arg = sys.argv[3] if len(sys.argv) > 3 else None
            run_reviews_scrape(reviews_csv=reviews_csv, city=city_arg)
        elif cmd == "competitors":
            csv_arg = sys.argv[2] if len(sys.argv) > 2 else (os.environ.get("COMPETITOR_PRODUCTS_CSV") or "competitor_products.csv")
            city_arg = sys.argv[3] if len(sys.argv) > 3 else (os.environ.get("COMPETITOR_CITY") or "Алматы")
            prefix_arg = sys.argv[4] if len(sys.argv) > 4 else (os.environ.get("COMPETITOR_FILE_PREFIX") or "kaspi_competitors_almaty")
            run_kaspi_scrape(
                cities=[city_arg],
                extra_csv_paths=csv_arg,
                include_base=False,
                output_prefix=prefix_arg,
                priority="secondary",
            )
        else:
            run_kaspi_scrape(cities=["Алматы", "Астана", "Шымкент"])
    else:
        run_kaspi_scrape(cities=["Алматы", "Астана", "Шымкент"])
