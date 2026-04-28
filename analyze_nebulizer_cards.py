#!/usr/bin/env python3
from __future__ import annotations

import json
import math
import os
import re
import time
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from html import unescape
from pathlib import Path
from typing import Any
from urllib.parse import urlencode
from zoneinfo import ZoneInfo

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
RESULTS_DIR = BASE_DIR / "RESULTS"
ALMATY_TZ = ZoneInfo("Asia/Almaty")
CITY_CODE = "750000000"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://kaspi.kz/",
}

SESSION = requests.Session()

IMPORTANT_SPEC_KEYS = [
    "Тип",
    "Для кого",
    "Питание",
    "Скорость распыления",
    "Объем емкости для лекарств",
    "Остаточный объем лекарства",
    "Время непрерывной работы",
    "Уровень шума",
    "Размеры",
    "Вес",
    "Комплектация",
    "Дополнительно",
    "Гарантийный срок",
]

THEME_KEYWORDS = {
    "Шум/тишина": ["шум", "шумит", "громк", "тихий", "тихо", "шумноват"],
    "Компактность/вес": ["компакт", "легк", "маленьк", "портатив", "удобно брать", "не тяжел"],
    "Эффективность лечения": ["помога", "эффект", "леч", "астм", "кашл", "дыш", "облегч"],
    "Качество/надежность": ["качеств", "работает", "не работает", "слом", "брак", "исправ", "надеж"],
    "Цена/ценность": ["цена", "дорог", "дешев", "недорог", "стоим", "по цене"],
    "Доставка/упаковка": ["доставка", "упаков", "пришел", "пришёл", "коробк", "упакова"],
    "Комплектация": ["комплект", "маск", "мундштук", "насадк", "трубк", "фильтр"],
    "Питание/зарядка": ["аккум", "заряд", "батаре", "type-c", "сеть", "провод"],
}


COMPETITOR_PRODUCTS = [
    {
        "group": "Конкурент",
        "name": "Omron CompAir C101 Essential NE-C101-E компрессорный ингалятор для всех возрастов",
        "url": "https://kaspi.kz/shop/p/omron-compair-c101-essential-ne-c101-e-kompressornyi-ingaljator-dlja-vseh-vozrastov-100159096/?c=750000000",
    },
    {
        "group": "Конкурент",
        "name": "MAXRON YK-N2AA меш-ингалятор для всех возрастов",
        "url": "https://kaspi.kz/shop/p/maxron-yk-n2aa-mesh-ingaljator-dlja-vseh-vozrastov-138268537/?c=750000000",
    },
    {
        "group": "Конкурент",
        "name": "Omron Comp Air NE-C28 Plus компрессорный ингалятор для всех возрастов",
        "url": "https://kaspi.kz/shop/p/omron-comp-air-ne-c28-plus-kompressornyi-ingaljator-dlja-vseh-vozrastov-16500007/?c=750000000",
    },
    {
        "group": "Конкурент",
        "name": "MAXRON YK-N3AA (черный) меш-ингалятор для всех возрастов",
        "url": "https://kaspi.kz/shop/p/maxron-yk-n3aa-chernyi-mesh-ingaljator-dlja-vseh-vozrastov-138268542/?c=750000000",
    },
    {
        "group": "Конкурент",
        "name": "MAXRON YK-N3AA (белый) меш-ингалятор для всех возрастов",
        "url": "https://kaspi.kz/shop/p/maxron-yk-n3aa-belyi-mesh-ingaljator-dlja-vseh-vozrastov-138268544/?c=750000000",
    },
    {
        "group": "Конкурент",
        "name": "Bai-Bala UN100 с таймером (перезаряжаемый) меш-ингалятор для всех возрастов",
        "url": "https://kaspi.kz/shop/p/bai-bala-un100-s-taimerom-perezarjazhaemyi-mesh-ingaljator-dlja-vseh-vozrastov-142955172/?c=750000000",
    },
    {
        "group": "Конкурент",
        "name": "Bai-Bala UN300 меш-ингалятор для детей",
        "url": "https://kaspi.kz/shop/p/bai-bala-un300-mesh-ingaljator-dlja-detei-131113334/?c=750000000",
    },
    {
        "group": "Конкурент",
        "name": "Bai-Bala CNB 69021 ингалятор-небулайзер",
        "url": "https://kaspi.kz/shop/p/bai-bala-cnb-69021-ingaljator-nebulaizer-121453676/?c=750000000",
    },
    {
        "group": "Конкурент",
        "name": "Bai-Bala CNB69009 компрессорный ингалятор для всех возрастов",
        "url": "https://kaspi.kz/shop/p/bai-bala-cnb69009-kompressornyi-ingaljator-dlja-vseh-vozrastov-130252710/?c=750000000",
    },
    {
        "group": "Конкурент",
        "name": "Gamma Effect New компрессорный ингалятор для всех возрастов",
        "url": "https://kaspi.kz/shop/p/gamma-effect-new-kompressornyi-ingaljator-dlja-vseh-vozrastov-16500012/?c=750000000",
    },
    {
        "group": "Конкурент",
        "name": "Gamma Nemo компрессорный ингалятор для всех возрастов",
        "url": "https://kaspi.kz/shop/p/gamma-nemo-kompressornyi-ingaljator-dlja-vseh-vozrastov-16500026/?c=750000000",
    },
]

OWN_PRODUCTS = [
    {
        "group": "Мой товар",
        "name": "CNB69016 небулайзер для всех возрастов",
        "url": "https://kaspi.kz/shop/p/cnb69016-nebulaizer-dlja-vseh-vozrastov-161688981/?maSource=dynamicLink&ref=shared_link",
    },
    {
        "group": "Мой товар",
        "name": "CNB69020 небулайзер для всех возрастов",
        "url": "https://kaspi.kz/shop/p/cnb69020-nebulaizer-dlja-vseh-vozrastov-161689013/?maSource=dynamicLink&ref=shared_link",
    },
    {
        "group": "Мой товар",
        "name": "CNB69011 белый M небулайзер для всех возрастов",
        "url": "https://kaspi.kz/shop/p/cnb69011-belyi-m-nebulaizer-dlja-vseh-vozrastov-161689040/?maSource=dynamicLink&ref=shared_link",
    },
]


@dataclass
class ProductBundle:
    row: dict[str, Any]
    specs_rows: list[dict[str, Any]]
    reviews_rows: list[dict[str, Any]]


def normalize_url(url: str) -> str:
    url = url.strip()
    if not url:
        return url
    if "?c=" not in url and "?" not in url:
        return f"{url}?c={CITY_CODE}"
    return url


def extract_product_code(url: str) -> str:
    match = re.search(r"-(\d+)(?:/|\?|$)", url)
    if match:
        return match.group(1)
    match = re.search(r"/p/[^/]+-(\d+)", url)
    if match:
        return match.group(1)
    return ""


def http_get(url: str, *, params: dict[str, Any] | None = None, expect_json: bool = False) -> Any:
    response = SESSION.get(url, headers=HEADERS, params=params, timeout=30)
    response.raise_for_status()
    if expect_json:
        return response.json()
    return response.text


def clean_html_text(raw_html: str) -> str:
    if not raw_html:
        return ""
    text = raw_html
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"</p>\s*<p>", "\n\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = unescape(text)
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def extract_backend_item(html_text: str) -> dict[str, Any]:
    marker = "BACKEND.components.item = "
    start = html_text.find(marker)
    if start == -1:
        raise ValueError("BACKEND.components.item not found")
    start += len(marker)
    tail = html_text.find("window.digitalData", start)
    if tail == -1:
        raise ValueError("window.digitalData not found after BACKEND.components.item")
    blob = html_text[start:tail]
    blob = blob.split("</script><script>")[0].strip()
    return json.loads(blob)


def extract_digital_product(html_text: str) -> dict[str, Any]:
    match = re.search(r"window\.digitalData\.product\s*=\s*(\{.*?\});", html_text, re.S)
    if not match:
        return {}
    try:
        return json.loads(match.group(1))
    except json.JSONDecodeError:
        return {}


def parse_specs(item_data: dict[str, Any], product_name: str, product_id: str) -> tuple[list[dict[str, Any]], dict[str, str]]:
    rows: list[dict[str, Any]] = []
    spec_map: dict[str, str] = {}
    for group in item_data.get("specifications") or []:
        group_name = (group.get("name") or "").strip()
        for feature in group.get("features") or []:
            key = (feature.get("name") or "").strip()
            values = [
                str(value.get("value")).strip()
                for value in feature.get("featureValues") or []
                if str(value.get("value") or "").strip()
            ]
            value_text = "; ".join(values)
            if not key:
                continue
            spec_map[key] = value_text
            rows.append(
                {
                    "product_id": product_id,
                    "product_name": product_name,
                    "group_name": group_name,
                    "spec_name": key,
                    "spec_value": value_text,
                }
            )
    return rows, spec_map


def fetch_card(product: dict[str, str]) -> tuple[dict[str, Any], list[dict[str, Any]], dict[str, str]]:
    url = normalize_url(product["url"])
    html_text = http_get(url)
    item = extract_backend_item(html_text)
    digital = extract_digital_product(html_text)
    card = item.get("card") or {}
    product_id = str(card.get("id") or digital.get("id") or extract_product_code(url))
    product_name = card.get("title") or digital.get("name") or product["name"]
    specs_rows, spec_map = parse_specs(item, product_name, product_id)
    description_html = item.get("description") or ""
    if not description_html:
        for entry in item.get("descriptions") or []:
            text_html = entry.get("text") or ""
            if text_html.strip():
                description_html = text_html
                break
    description_text = clean_html_text(description_html)

    breadcrumbs = [entry.get("title") for entry in item.get("breadcrumbs") or [] if entry.get("title")]
    category_path = [entry for entry in breadcrumbs if entry != product_name]
    if category_path and category_path[0] == "Kaspi Магазин":
        category_path = category_path[1:]
    category_leaf = category_path[-1] if category_path else ""
    card_created = card.get("createdTime") or ""

    row = {
        "group": product["group"],
        "input_name": product["name"],
        "product_name": product_name,
        "product_id": product_id,
        "product_url": url,
        "price_kzt": card.get("price"),
        "brand": card.get("promoConditions", {}).get("brand") or digital.get("brand") or "",
        "created_time": card_created,
        "card_age_days": age_days(card_created),
        "category_path": " > ".join(category_path),
        "category_leaf": category_leaf,
        "photo_count": len(item.get("galleryImages") or []),
        "has_video": bool(item.get("videoId") or item.get("videos")),
        "video_id": item.get("videoId") or "",
        "has_description": bool(description_text),
        "description_chars": len(description_text),
        "description_preview": description_text[:500],
        "spec_count": len(specs_rows),
        "spec_groups_count": len(item.get("specifications") or []),
        "rich_content": bool(item.get("richContent") or item.get("videos") or item.get("teasers")),
        "type": spec_map.get("Тип") or infer_type(product_name),
        "for_whom": spec_map.get("Для кого") or "",
        "power": spec_map.get("Питание") or "",
        "spray_rate": spec_map.get("Скорость распыления") or "",
        "medicine_capacity": spec_map.get("Объем емкости для лекарств") or "",
        "residual_volume": spec_map.get("Остаточный объем лекарства") or "",
        "continuous_work": spec_map.get("Время непрерывной работы") or "",
        "noise_db": spec_map.get("Уровень шума") or "",
        "size": spec_map.get("Размеры") or "",
        "weight": spec_map.get("Вес") or "",
        "warranty": spec_map.get("Гарантийный срок") or "",
        "package": spec_map.get("Комплектация") or "",
        "extra": spec_map.get("Дополнительно") or "",
    }
    return row, specs_rows, spec_map


def age_days(value: str) -> int | None:
    if not value:
        return None
    try:
        created = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    now = datetime.now(tz=created.tzinfo)
    return max(0, (now - created).days)


def infer_type(name: str) -> str:
    lowered = name.lower()
    if "меш" in lowered or "mesh" in lowered:
        return "меш-ингалятор"
    if "компресс" in lowered:
        return "компрессорный ингалятор"
    return "небулайзер"


def parse_review_date(value: str) -> datetime | None:
    if not value:
        return None
    value = value.strip()
    try:
        return datetime.strptime(value, "%d.%m.%Y")
    except ValueError:
        pass
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def fetch_review_summary(product_id: str) -> tuple[dict[str, Any], dict[str, int]]:
    summary_url = f"https://kaspi.kz/yml/creview/rest/misc/product/{product_id}/summary"
    groups_url = f"https://kaspi.kz/yml/creview/rest/misc/product/{product_id}/groupsummary"
    summary_payload = http_get(summary_url, expect_json=True).get("data") or {}
    groups_payload = http_get(groups_url, expect_json=True).get("data") or []
    groups = {entry.get("id"): int(entry.get("total") or 0) for entry in groups_payload if entry.get("id")}
    return summary_payload, groups


def normalize_comment_block(raw: Any) -> str:
    if not raw:
        return ""
    if isinstance(raw, str):
        return raw.strip()
    if isinstance(raw, list):
        return " ".join(str(part).strip() for part in raw if str(part).strip())
    if isinstance(raw, dict):
        return " ".join(str(part).strip() for part in raw.values() if str(part).strip())
    return str(raw).strip()


def parse_review(entry: dict[str, Any], product_row: dict[str, Any]) -> dict[str, Any]:
    comment = entry.get("comment") or {}
    feedback = entry.get("feedback") or {}
    plus_text = normalize_comment_block(comment.get("plus") or comment.get("pros") or comment.get("advantages"))
    minus_text = normalize_comment_block(comment.get("minus") or comment.get("cons") or comment.get("disadvantages"))
    body_text = normalize_comment_block(comment.get("text") or comment.get("comment") or comment.get("description"))
    helpful_count = (
        entry.get("helpfulCount")
        or entry.get("usefulCount")
        or entry.get("likes")
        or feedback.get("positive")
        or 0
    )
    try:
        helpful_count = int(helpful_count)
    except (TypeError, ValueError):
        helpful_count = 0
    review_date = str(entry.get("date") or entry.get("createdDate") or entry.get("createdTime") or "")
    gallery = entry.get("galleryImages") or []
    return {
        "group": product_row["group"],
        "product_id": product_row["product_id"],
        "product_name": product_row["product_name"],
        "product_url": product_row["product_url"],
        "review_id": entry.get("id") or "",
        "author": entry.get("author") or "",
        "review_date": review_date,
        "review_date_iso": parsed_date_to_iso(review_date),
        "rating": entry.get("rating"),
        "helpful_count": helpful_count,
        "plus": plus_text,
        "minus": minus_text,
        "comment": body_text,
        "has_text": bool(plus_text or minus_text or body_text),
        "has_photo": bool(gallery),
        "photo_count": len(gallery),
        "merchant_name": (entry.get("merchant") or {}).get("name") or "",
        "merchant_code": (entry.get("merchant") or {}).get("code") or "",
        "order_number": entry.get("orderNumber") or "",
        "comment_kind": comment_kind(plus_text, minus_text, body_text),
    }


def parsed_date_to_iso(value: str) -> str:
    parsed = parse_review_date(value)
    return parsed.date().isoformat() if parsed else ""


def comment_kind(plus_text: str, minus_text: str, body_text: str) -> str:
    if plus_text and minus_text:
        return "Плюсы и минусы"
    if plus_text:
        return "Плюсы"
    if minus_text:
        return "Минусы"
    if body_text:
        return "Текстовый комментарий"
    return "Без текста"


def fetch_all_reviews(product_row: dict[str, Any], expected_total: int) -> list[dict[str, Any]]:
    if expected_total <= 0:
        return []
    base_url = f"https://kaspi.kz/yml/review-view/api/v1/reviews/product/{product_row['product_id']}"
    limit = 200
    rows: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    max_pages = max(1, math.ceil(expected_total / limit) + 1)
    for page in range(max_pages):
        params = {"withAgg": "true", "page": page, "limit": limit}
        payload = http_get(base_url, params=params, expect_json=True)
        entries = payload.get("data") or []
        if not entries:
            break
        fresh = 0
        for entry in entries:
            review_id = str(entry.get("id") or "")
            if review_id and review_id in seen_ids:
                continue
            if review_id:
                seen_ids.add(review_id)
            rows.append(parse_review(entry, product_row))
            fresh += 1
        if len(entries) < limit or fresh == 0 or len(rows) >= expected_total:
            break
        time.sleep(0.15)
    return rows


def theme_summary(reviews: list[dict[str, Any]], predicate) -> str:
    counter: Counter[str] = Counter()
    for review in reviews:
        if not predicate(review):
            continue
        text = " ".join(
            part.strip().lower()
            for part in [review.get("plus", ""), review.get("minus", ""), review.get("comment", "")]
            if part and str(part).strip()
        )
        if not text:
            continue
        for theme, keywords in THEME_KEYWORDS.items():
            if any(keyword in text for keyword in keywords):
                counter[theme] += 1
    if not counter:
        return ""
    return ", ".join(f"{theme} ({count})" for theme, count in counter.most_common(3))


def build_card_score(row: dict[str, Any]) -> tuple[int, str, str]:
    score = 0
    strengths: list[str] = []
    gaps: list[str] = []

    if row["photo_count"] >= 5:
        score += 15
        strengths.append("много фото")
    elif row["photo_count"] >= 3:
        score += 10
        strengths.append("достаточно фото")
    elif row["photo_count"] >= 1:
        score += 5
        gaps.append("фото мало")
    else:
        gaps.append("нет фото")

    if row["has_description"]:
        score += 15 if row["description_chars"] >= 250 else 10
        strengths.append("есть описание")
    else:
        gaps.append("нет описания")

    if row["spec_count"] >= 12:
        score += 20
        strengths.append("характеристики заполнены подробно")
    elif row["spec_count"] >= 8:
        score += 14
        strengths.append("характеристики заполнены")
    elif row["spec_count"] >= 1:
        score += 6
        gaps.append("характеристик мало")
    else:
        gaps.append("нет характеристик")

    if row["has_video"]:
        score += 8
        strengths.append("есть видео")
    else:
        gaps.append("нет видео")

    reviews_count = int(row.get("reviews_count") or 0)
    if reviews_count >= 1000:
        score += 25
        strengths.append("очень сильная соцдоказательность")
    elif reviews_count >= 300:
        score += 20
        strengths.append("сильная база отзывов")
    elif reviews_count >= 50:
        score += 12
        strengths.append("есть заметная база отзывов")
    elif reviews_count >= 1:
        score += 5
        gaps.append("отзывов мало")
    else:
        gaps.append("нет отзывов")

    photo_reviews = int(row.get("review_images_count") or 0)
    if photo_reviews >= 200:
        score += 10
        strengths.append("много отзывов с фото")
    elif photo_reviews >= 20:
        score += 6
        strengths.append("есть отзывы с фото")
    elif photo_reviews > 0:
        score += 3
        strengths.append("есть немного отзывов с фото")
    else:
        gaps.append("нет отзывов с фото")

    if row.get("brand"):
        score += 4
    else:
        gaps.append("бренд в карточке неочевиден")

    if row.get("price_kzt") is not None:
        score += 3

    return min(score, 100), "; ".join(strengths), "; ".join(gaps)


def rating_distribution(summary: dict[str, Any]) -> dict[int, int]:
    distribution = {5: 0, 4: 0, 3: 0, 2: 0, 1: 0}
    for entry in summary.get("statistic") or []:
        rate = int(entry.get("rate") or 0)
        if rate in distribution:
            distribution[rate] = int(entry.get("count") or 0)
    return distribution


def build_product_row(card_row: dict[str, Any], reviews: list[dict[str, Any]], summary: dict[str, Any], groups: dict[str, int]) -> dict[str, Any]:
    distribution = rating_distribution(summary)
    reviews_count = int(summary.get("reviewsCount") or 0)
    text_reviews = sum(1 for review in reviews if review["has_text"])
    helpful_reviews = [review for review in reviews if review["helpful_count"] > 0]
    helpful_votes_total = sum(review["helpful_count"] for review in reviews)
    latest_review_date = max((review["review_date_iso"] for review in reviews if review["review_date_iso"]), default="")
    earliest_review_date = min((review["review_date_iso"] for review in reviews if review["review_date_iso"]), default="")
    star_positive = distribution[5] + distribution[4]
    star_neutral = distribution[3]
    star_negative = distribution[2] + distribution[1]
    groups_total = int(groups.get("all") or 0)
    group_positive = int(groups.get("positive") or 0)
    group_negative = int(groups.get("negative") or 0)
    group_neutral = max(0, groups_total - group_positive - group_negative)

    top_helpful = sorted(helpful_reviews, key=lambda row: (row["helpful_count"], row["review_date_iso"]), reverse=True)[:3]
    top_helpful_text = " | ".join(
        f"{review['author']} {review['review_date']} ({review['helpful_count']}): "
        f"{truncate_review(review)}"
        for review in top_helpful
    )

    row = dict(card_row)
    row.update(
        {
            "rating_global": summary.get("global"),
            "rating_count": int(summary.get("ratingCount") or 0),
            "reviews_count": reviews_count,
            "review_images_count": int(summary.get("reviewImagesCount") or 0),
            "adjusted_rating": summary.get("adjustedRating"),
            "reviews_with_text": text_reviews,
            "reviews_with_text_share": safe_share(text_reviews, reviews_count),
            "reviews_with_photo": sum(1 for review in reviews if review["has_photo"]),
            "reviews_with_helpful_votes": len(helpful_reviews),
            "helpful_votes_total": helpful_votes_total,
            "latest_review_date": latest_review_date,
            "earliest_review_date": earliest_review_date,
            "group_positive_count": group_positive,
            "group_negative_count": group_negative,
            "group_neutral_count": group_neutral,
            "stars_positive_count": star_positive,
            "stars_neutral_count": star_neutral,
            "stars_negative_count": star_negative,
            "star_5_count": distribution[5],
            "star_4_count": distribution[4],
            "star_3_count": distribution[3],
            "star_2_count": distribution[2],
            "star_1_count": distribution[1],
            "top_positive_themes": theme_summary(reviews, lambda review: (review.get("rating") or 0) >= 4),
            "top_negative_themes": theme_summary(reviews, lambda review: (review.get("rating") or 0) <= 2),
            "top_helpful_reviews": top_helpful_text,
        }
    )
    score, strengths, gaps = build_card_score(row)
    row["card_score_100"] = score
    row["card_strengths"] = strengths
    row["card_gaps"] = gaps
    row["price_segment"] = price_segment(row.get("price_kzt"))
    return row


def truncate_review(review: dict[str, Any]) -> str:
    text = " ".join(part for part in [review.get("plus", ""), review.get("minus", ""), review.get("comment", "")] if part)
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= 120:
        return text
    return text[:117] + "..."


def safe_share(numerator: int, denominator: int) -> float:
    if not denominator:
        return 0.0
    return round(numerator / denominator, 4)


def price_segment(price_kzt: Any) -> str:
    try:
        value = int(price_kzt)
    except (TypeError, ValueError):
        return ""
    if value < 12000:
        return "До 12 тыс."
    if value < 18000:
        return "12-18 тыс."
    if value < 25000:
        return "18-25 тыс."
    return "25 тыс. и выше"


def diff_specs(own_row: dict[str, Any], competitor_row: dict[str, Any]) -> str:
    parts: list[str] = []
    key_to_column = {
        "Тип": "type",
        "Для кого": "for_whom",
        "Питание": "power",
        "Скорость распыления": "spray_rate",
        "Объем емкости для лекарств": "medicine_capacity",
        "Остаточный объем лекарства": "residual_volume",
        "Время непрерывной работы": "continuous_work",
        "Уровень шума": "noise_db",
        "Вес": "weight",
        "Гарантийный срок": "warranty",
    }
    for label, column in key_to_column.items():
        left = str(own_row.get(column) or "").strip()
        right = str(competitor_row.get(column) or "").strip()
        if left and right and left != right:
            parts.append(f"{label}: {left} vs {right}")
        if len(parts) >= 5:
            break
    if not parts:
        own_price = own_row.get("price_kzt")
        competitor_price = competitor_row.get("price_kzt")
        if own_price is not None and competitor_price is not None:
            parts.append(f"Цена: {own_price} vs {competitor_price} KZT")
    return " | ".join(parts)


def build_comparison_sheet(products_df: pd.DataFrame) -> pd.DataFrame:
    own_df = products_df.loc[products_df["group"] == "Мой товар"].copy()
    competitors_df = products_df.loc[products_df["group"] == "Конкурент"].copy()
    rows: list[dict[str, Any]] = []
    for _, own_row in own_df.iterrows():
        subset = competitors_df.copy()
        same_type = subset["type"].fillna("").eq(own_row["type"])
        if same_type.any():
            subset = subset.loc[same_type].copy()
        subset["price_gap_abs"] = (subset["price_kzt"].fillna(0) - int(own_row["price_kzt"] or 0)).abs()
        subset = subset.sort_values(["price_gap_abs", "reviews_count", "card_score_100"], ascending=[True, False, False])
        for _, competitor_row in subset.head(4).iterrows():
            rows.append(
                {
                    "my_product": own_row["product_name"],
                    "my_price_kzt": own_row["price_kzt"],
                    "my_type": own_row["type"],
                    "my_reviews": own_row["reviews_count"],
                    "competitor_product": competitor_row["product_name"],
                    "competitor_price_kzt": competitor_row["price_kzt"],
                    "competitor_type": competitor_row["type"],
                    "competitor_reviews": competitor_row["reviews_count"],
                    "price_delta_kzt": int(competitor_row["price_kzt"] or 0) - int(own_row["price_kzt"] or 0),
                    "card_score_delta": int(competitor_row["card_score_100"] or 0) - int(own_row["card_score_100"] or 0),
                    "review_delta": int(competitor_row["reviews_count"] or 0) - int(own_row["reviews_count"] or 0),
                    "key_differences": diff_specs(own_row.to_dict(), competitor_row.to_dict()),
                    "competitor_card_advantage": competitor_row["card_strengths"],
                    "my_card_gap": own_row["card_gaps"],
                }
            )
    return pd.DataFrame(rows)


def build_recommendations_sheet(products_df: pd.DataFrame, comparison_df: pd.DataFrame) -> pd.DataFrame:
    own_df = products_df.loc[products_df["group"] == "Мой товар"].copy()
    recommendations: list[dict[str, Any]] = []
    for _, row in own_df.iterrows():
        recommendations.append(
            {
                "product_name": row["product_name"],
                "priority": "Высокий",
                "issue": "Соцдоказательность",
                "recommendation": (
                    "Запустить сбор первых отзывов и фотоотзывов: у карточки 0 отзывов, "
                    "а у ключевых конкурентов уже накоплены сотни и тысячи."
                    if not row["reviews_count"]
                    else "Усиливать базу отзывов и фотоотзывов, чтобы сократить разрыв с лидерами."
                ),
            }
        )
        if not row["has_description"]:
            recommendations.append(
                {
                    "product_name": row["product_name"],
                    "priority": "Высокий",
                    "issue": "Описание",
                    "recommendation": (
                        "Добавить продающее описание: кому подходит, тип ингалятора, "
                        "питание, шум, скорость распыления, комплектацию, противопоказания и сценарии использования."
                    ),
                }
            )
        if not row["has_video"]:
            recommendations.append(
                {
                    "product_name": row["product_name"],
                    "priority": "Средний",
                    "issue": "Видео",
                    "recommendation": "Добавить короткое видео с распаковкой и работой устройства.",
                }
            )
        if row["photo_count"] < 5:
            recommendations.append(
                {
                    "product_name": row["product_name"],
                    "priority": "Средний",
                    "issue": "Галерея",
                    "recommendation": "Расширить галерею до 5-7 фото: общий вид, комплектация, крупные планы, габариты, питание.",
                }
            )
        if row["spec_count"] < 10:
            recommendations.append(
                {
                    "product_name": row["product_name"],
                    "priority": "Средний",
                    "issue": "Характеристики",
                    "recommendation": "Расширить список характеристик и вынести ключевые параметры в видимую часть карточки.",
                }
            )
        nearest = comparison_df.loc[comparison_df["my_product"] == row["product_name"]].sort_values("price_delta_kzt")
        if not nearest.empty:
            competitor = nearest.iloc[0]
            recommendations.append(
                {
                    "product_name": row["product_name"],
                    "priority": "Средний",
                    "issue": "Позиционирование",
                    "recommendation": (
                        f"Ближайший конкурент по цене: {competitor['competitor_product']} "
                        f"({competitor['competitor_price_kzt']} KZT, отзывов: {competitor['competitor_reviews']}). "
                        "Нужно явно показать в карточке, чем ваш товар лучше или за счет чего оправдана цена."
                    ),
                }
            )
    return pd.DataFrame(recommendations)


def build_markdown(products_df: pd.DataFrame, comparison_df: pd.DataFrame) -> str:
    own_df = products_df.loc[products_df["group"] == "Мой товар"].copy().sort_values("price_kzt")
    competitors_df = products_df.loc[products_df["group"] == "Конкурент"].copy().sort_values("price_kzt")
    strongest = competitors_df.sort_values(["reviews_count", "card_score_100"], ascending=[False, False]).iloc[0]
    cheapest_competitor = competitors_df.iloc[0]
    highest_score = competitors_df.sort_values("card_score_100", ascending=False).iloc[0]

    lines = [
        "# Обзор карточек небулайзеров на Kaspi",
        "",
        f"Дата анализа: {datetime.now(ALMATY_TZ).strftime('%d.%m.%Y %H:%M')} Алматы",
        "",
        "## Ключевые выводы",
        (
            f"- Все ваши карточки созданы недавно: "
            + ", ".join(
                f"{row['product_name']} ({format_date(row['created_time'])})"
                for _, row in own_df.iterrows()
            )
        ),
        (
            f"- По соцдоказательности рынок ведут {strongest['product_name']} "
            f"({strongest['reviews_count']} отзывов, рейтинг {strongest['rating_global']})."
        ),
        (
            f"- Самый дешевый конкурент в выборке: {cheapest_competitor['product_name']} "
            f"за {cheapest_competitor['price_kzt']} KZT."
        ),
        (
            f"- Самая проработанная карточка по контенту и отзывам: {highest_score['product_name']} "
            f"(score {highest_score['card_score_100']}/100)."
        ),
        "",
        "## Ваши товары",
    ]
    for _, row in own_df.iterrows():
        lines.extend(
            [
                f"### {row['product_name']}",
                f"- Цена: {row['price_kzt']} KZT",
                f"- Категория: {row['category_path']}",
                f"- Тип: {row['type'] or 'не указан'}",
                f"- Отзывы: {row['reviews_count']} | Рейтинг: {row['rating_global'] or 'нет'}",
                f"- Контент карточки: фото {row['photo_count']}, описание {'есть' if row['has_description'] else 'нет'}, "
                f"видео {'есть' if row['has_video'] else 'нет'}, характеристик {row['spec_count']}",
                f"- Слабые места: {row['card_gaps'] or 'критичных пробелов не найдено'}",
                "",
            ]
        )

    lines.append("## Конкуренты")
    for _, row in competitors_df.iterrows():
        lines.extend(
            [
                f"### {row['product_name']}",
                f"- Цена: {row['price_kzt']} KZT",
                f"- Тип: {row['type'] or 'не указан'} | Для кого: {row['for_whom'] or 'не указано'}",
                f"- Категория: {row['category_path']}",
                f"- Отзывы: {row['reviews_count']} | Рейтинг: {row['rating_global'] or 'нет'}",
                (
                    f"- Позитив/негатив: {row['group_positive_count']} / {row['group_negative_count']} "
                    f"(по group summary), 5★ {row['star_5_count']}, 1★ {row['star_1_count']}"
                ),
                f"- Карточка: score {row['card_score_100']}/100 | сильные стороны: {row['card_strengths'] or '—'}",
                f"- Темы отзывов: + {row['top_positive_themes'] or 'нет выраженного паттерна'} | "
                f"- {row['top_negative_themes'] or 'негатива мало или он однотипно не выражен'}",
                "",
            ]
        )

    lines.append("## Сравнение с вашими товарами")
    for own_name in own_df["product_name"]:
        lines.append(f"### {own_name}")
        subset = comparison_df.loc[comparison_df["my_product"] == own_name].head(3)
        for _, row in subset.iterrows():
            lines.append(
                f"- {row['competitor_product']}: цена {row['competitor_price_kzt']} KZT "
                f"(дельта {row['price_delta_kzt']}), отзывов {row['competitor_reviews']}, "
                f"ключевые отличия: {row['key_differences']}"
            )
        lines.append("")
    return "\n".join(lines).strip() + "\n"


def format_date(value: str) -> str:
    if not value:
        return ""
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).strftime("%d.%m.%Y")
    except ValueError:
        return value


def autosize_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    own_fill = PatternFill("solid", fgColor="E2F0D9")
    competitor_fill = PatternFill("solid", fgColor="FCE4D6")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in sheet.columns:
            width = 0
            for idx, cell in enumerate(column, start=1):
                if idx > 200:
                    break
                value = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(value), 120))
            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(width + 2, 12), 60)
        header_map = {cell.value: cell.column for cell in sheet[1]}
        group_col = header_map.get("group")
        if group_col:
            for row in range(2, sheet.max_row + 1):
                group_val = str(sheet.cell(row=row, column=group_col).value or "")
                fill = own_fill if group_val == "Мой товар" else competitor_fill
                for col in range(1, sheet.max_column + 1):
                    sheet.cell(row=row, column=col).fill = fill
    workbook.save(path)


def send_message(text: str) -> None:
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        raise RuntimeError("TELEGRAM_BOT_TOKEN or TELEGRAM_CHAT_ID is not set")
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    response = SESSION.post(url, data={"chat_id": chat_id, "text": text}, timeout=60)
    response.raise_for_status()


def send_file(file_path: Path, caption: str | None = None) -> None:
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        raise RuntimeError("TELEGRAM_BOT_TOKEN or TELEGRAM_CHAT_ID is not set")
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    with open(file_path, "rb") as handle:
        response = SESSION.post(
            url,
            data={"chat_id": chat_id, "caption": caption or ""},
            files={"document": handle},
            timeout=120,
        )
    response.raise_for_status()


def main() -> None:
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)
    products = OWN_PRODUCTS + COMPETITOR_PRODUCTS

    bundles: list[ProductBundle] = []
    all_reviews: list[dict[str, Any]] = []
    all_specs: list[dict[str, Any]] = []

    for index, product in enumerate(products, start=1):
        print(f"[{index}/{len(products)}] {product['name']}", flush=True)
        card_row, specs_rows, _ = fetch_card(product)
        summary, groups = fetch_review_summary(card_row["product_id"])
        reviews_rows = fetch_all_reviews(card_row, int(summary.get("reviewsCount") or 0))
        product_row = build_product_row(card_row, reviews_rows, summary, groups)
        bundles.append(ProductBundle(row=product_row, specs_rows=specs_rows, reviews_rows=reviews_rows))
        all_specs.extend(specs_rows)
        all_reviews.extend(reviews_rows)
        time.sleep(0.15)

    products_df = pd.DataFrame([bundle.row for bundle in bundles]).sort_values(
        ["group", "price_kzt", "reviews_count"], ascending=[True, True, False]
    )
    specs_df = pd.DataFrame(all_specs).sort_values(["product_name", "group_name", "spec_name"])
    reviews_df = pd.DataFrame(all_reviews).sort_values(
        ["product_name", "review_date_iso", "helpful_count"], ascending=[True, False, False]
    )
    comparison_df = build_comparison_sheet(products_df)
    recommendations_df = build_recommendations_sheet(products_df, comparison_df)

    markdown = build_markdown(products_df, comparison_df)

    timestamp = datetime.now(ALMATY_TZ).strftime("%Y-%m-%d_%H-%M-%S")
    excel_path = RESULTS_DIR / f"nebulizer_card_audit_{timestamp}.xlsx"
    reviews_csv_path = RESULTS_DIR / f"nebulizer_all_reviews_{timestamp}.csv"
    report_md_path = RESULTS_DIR / f"nebulizer_card_audit_{timestamp}.md"

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        products_df.to_excel(writer, sheet_name="Products", index=False)
        comparison_df.to_excel(writer, sheet_name="Comparison", index=False)
        recommendations_df.to_excel(writer, sheet_name="Recommendations", index=False)
        specs_df.to_excel(writer, sheet_name="Specs", index=False)
        reviews_df.to_excel(writer, sheet_name="AllReviews", index=False)

    reviews_df.to_csv(reviews_csv_path, index=False)
    report_md_path.write_text(markdown, encoding="utf-8")
    autosize_workbook(excel_path)

    send_message(
        "Kaspi отчет по небулайзерам готов.\n"
        f"Дата: {datetime.now(ALMATY_TZ).strftime('%d.%m.%Y %H:%M')} Алматы\n"
        f"Карточек: {len(products_df)}\n"
        f"Отзывов в выгрузке: {len(reviews_df)}"
    )
    send_file(excel_path, caption="Kaspi: аудит карточек небулайзеров")
    send_file(reviews_csv_path, caption="Kaspi: все отзывы по небулайзерам")
    send_file(report_md_path, caption="Kaspi: текстовый summary")

    print(excel_path)
    print(reviews_csv_path)
    print(report_md_path)


if __name__ == "__main__":
    main()
