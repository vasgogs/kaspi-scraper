#!/usr/bin/env python3
from __future__ import annotations

import json
import math
import os
import re
import time
from collections import Counter
from datetime import datetime
from html import unescape
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
RESULTS_DIR = BASE_DIR / "RESULTS"
ALMATY_TZ = ZoneInfo("Asia/Almaty")
CITY_CODE = "750000000"
ZONE_ID = "Magnum_ZONE1"
CATEGORY_CODES = ["Supports and walkers", "Rehabilitation", "Pharmacy", "Categories"]
CATEGORY_QUERY = f":category:supports and walkers:availableInZones:{ZONE_ID}"
CATEGORY_URL = f"https://kaspi.kz/shop/c/supports%20and%20walkers/?c={CITY_CODE}"
CATEGORY_URL_RATING = f"https://kaspi.kz/shop/c/supports%20and%20walkers/?sort=rating&c={CITY_CODE}"
LISTING_API_URL = "https://kaspi.kz/yml/product-view/pl/filters"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/136.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://kaspi.kz/",
}

SESSION = requests.Session()

TOP_RATING_PRODUCTS = [
    {
        "group": "Top rating",
        "name": "Ходунки FS 913 L до 120 кг",
        "url": "https://kaspi.kz/shop/p/hodunki-fs-913-l-do-120-kg-102956788/?c=750000000",
    },
    {
        "group": "Top rating",
        "name": "Kazimport ходунки xod-020 до 130",
        "url": "https://kaspi.kz/shop/p/kazimport-hodunki-xod-020-do-130-120666120/?c=750000000",
    },
    {
        "group": "Top rating",
        "name": "Med Tech ходунки Regular до 150",
        "url": "https://kaspi.kz/shop/p/med-tech-hodunki-regular-do-150-138922753/?c=750000000",
    },
    {
        "group": "Top rating",
        "name": "Опорные ходунки Med Tech с колесами 150 кг",
        "url": "https://kaspi.kz/shop/p/opornye-hodunki-med-tech-s-kolesami-150-kg-139241209/?c=750000000",
    },
    {
        "group": "Top rating",
        "name": "Зенетар ходунки 4882 до 100 кг",
        "url": "https://kaspi.kz/shop/p/zenetar-hodunki-4882-do-100-kg-112823596/?c=750000000",
    },
]

OWN_PRODUCTS = [
    {
        "group": "Мой SKU",
        "name": "Osako ходунки S shape с образные 01 до 120",
        "url": "https://kaspi.kz/shop/p/osako-hodunki-s-shape-s-obraznye-01-do-120-161855849/?c=750000000",
    },
    {
        "group": "Мой SKU",
        "name": "Osako ходунки Regular обычные 01 до 120",
        "url": "https://kaspi.kz/shop/p/osako-hodunki-regular-obychnye-01-do-120-161855791/?c=750000000",
    },
    {
        "group": "Мой SKU",
        "name": "Osako ходунки W-wheels с колесами 01 до 120",
        "url": "https://kaspi.kz/shop/p/osako-hodunki-w-wheels-s-kolesami-01-do-120-161855875/?c=750000000",
    },
]

THEME_KEYWORDS = {
    "Устойчивость/надежность": ["устойчив", "крепк", "надеж", "надеж", "прочный", "не шата", "не люфт"],
    "Комфорт/удобство": ["удоб", "комфорт", "рукоят", "ручк", "эргоном", "подходит"],
    "Легкость/вес": ["легк", "тяжел", "вес", "не тяжел"],
    "Складной механизм/регулировка": ["складыв", "собира", "разбира", "регулиров", "высот", "настро"],
    "Колеса/маневренность": ["колес", "едет", "кат", "маневр", "поворач", "крут"],
    "Размеры/рост": ["рост", "широк", "узк", "высок", "низк", "габарит"],
    "Качество материалов": ["материал", "металл", "алюмин", "скрип", "резин", "люфт"],
    "Цена/ценность": ["цена", "дорог", "дешев", "стоим", "цен"],
    "Доставка/упаковка": ["доставка", "упаков", "короб", "пришел", "пришёл"],
    "Брак/дефекты": ["брак", "слом", "дефект", "поврежд", "не хватает", "не комплект"],
}


def load_env_from_file() -> None:
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("'\"")
        if key and key not in os.environ:
            os.environ[key] = value


def normalize_url(url: str) -> str:
    url = url.strip()
    if not url:
        return url
    if "?c=" not in url and "?" not in url:
        return f"{url}?c={CITY_CODE}"
    return url


def extract_product_code(url: str) -> str:
    match = re.search(r"-(\d+)(?:/|\?|$)", url)
    if match:
        return match.group(1)
    return ""


def http_get(url: str, *, params: dict[str, Any] | None = None, expect_json: bool = False) -> Any:
    response = SESSION.get(url, headers=HEADERS, params=params, timeout=30)
    response.raise_for_status()
    if expect_json:
        return response.json()
    return response.text


def http_post_json(url: str, payload: dict[str, Any]) -> dict[str, Any]:
    headers = dict(HEADERS)
    headers["Content-Type"] = "application/json"
    response = SESSION.post(url, headers=headers, json=payload, timeout=30)
    response.raise_for_status()
    return response.json()


def clean_html_text(raw_html: str) -> str:
    if not raw_html:
        return ""
    text = raw_html
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.I)
    text = re.sub(r"</p>\s*<p>", "\n\n", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = unescape(text)
    text = text.replace("\xa0", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def extract_backend_item(html_text: str) -> dict[str, Any]:
    marker = "BACKEND.components.item = "
    start = html_text.find(marker)
    if start == -1:
        raise ValueError("BACKEND.components.item not found")
    start += len(marker)
    tail = html_text.find("window.digitalData", start)
    if tail == -1:
        raise ValueError("window.digitalData not found after BACKEND.components.item")
    blob = html_text[start:tail]
    blob = blob.split("</script><script>")[0].strip()
    return json.loads(blob)


def extract_digital_product(html_text: str) -> dict[str, Any]:
    match = re.search(r"window\.digitalData\.product\s*=\s*(\{.*?\});", html_text, re.S)
    if not match:
        return {}
    try:
        return json.loads(match.group(1))
    except json.JSONDecodeError:
        return {}


def parse_specs(item_data: dict[str, Any], product_name: str, product_id: str) -> tuple[list[dict[str, Any]], dict[str, str]]:
    rows: list[dict[str, Any]] = []
    spec_map: dict[str, str] = {}
    for group in item_data.get("specifications") or []:
        group_name = (group.get("name") or "").strip()
        for feature in group.get("features") or []:
            key = (feature.get("name") or "").strip()
            values = [
                str(value.get("value")).strip()
                for value in feature.get("featureValues") or []
                if str(value.get("value") or "").strip()
            ]
            value_text = "; ".join(values)
            if not key:
                continue
            spec_map[key] = value_text
            rows.append(
                {
                    "product_id": product_id,
                    "product_name": product_name,
                    "group_name": group_name,
                    "spec_name": key,
                    "spec_value": value_text,
                }
            )
    return rows, spec_map


def parse_number(value: Any) -> int | None:
    if value is None:
        return None
    digits = re.findall(r"\d+(?:[.,]\d+)?", str(value))
    if not digits:
        return None
    try:
        return int(float(digits[0].replace(",", ".")))
    except ValueError:
        return None


def infer_main_type(name: str, spec_map: dict[str, str]) -> str:
    explicit = str(spec_map.get("Тип") or "").strip()
    if explicit:
        return explicit
    lowered = name.lower()
    if "ходунк" in lowered:
        return "ходунки"
    if "поруч" in lowered:
        return "поручень"
    if "ступень" in lowered:
        return "ступенька"
    if "опор" in lowered:
        return "опора"
    return ""


def infer_subtype(name: str, spec_map: dict[str, str]) -> str:
    lowered = " ".join(
        part.lower()
        for part in [
            name,
            spec_map.get("Тип опоры", ""),
            spec_map.get("Особенности", ""),
            spec_map.get("Тип", ""),
        ]
        if part
    )
    if any(token in lowered for token in ["колес", "wheel", "роллатор"]):
        return "с колесами"
    if any(token in lowered for token in ["s shape", "shape", "с образ", "с-образ", "s-wall", "swall"]):
        return "s-shape"
    if "regular" in lowered or "обыч" in lowered:
        return "regular"
    if "шага" in lowered:
        return "шагающие"
    return "базовые"


def age_days(value: str) -> int | None:
    if not value:
        return None
    try:
        created = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    now = datetime.now(tz=created.tzinfo)
    return max(0, (now - created).days)


def fetch_card(product: dict[str, str]) -> tuple[dict[str, Any], list[dict[str, Any]], dict[str, str]]:
    url = normalize_url(product["url"])
    html_text = http_get(url)
    item = extract_backend_item(html_text)
    digital = extract_digital_product(html_text)
    card = item.get("card") or {}
    product_id = str(card.get("id") or digital.get("id") or extract_product_code(url))
    product_name = card.get("title") or digital.get("name") or product["name"]
    specs_rows, spec_map = parse_specs(item, product_name, product_id)

    description_html = item.get("description") or ""
    if not description_html:
        for entry in item.get("descriptions") or []:
            text_html = entry.get("text") or ""
            if text_html.strip():
                description_html = text_html
                break
    description_text = clean_html_text(description_html)

    breadcrumbs = [entry.get("title") for entry in item.get("breadcrumbs") or [] if entry.get("title")]
    category_path = [entry for entry in breadcrumbs if entry != product_name]
    if category_path and category_path[0] == "Kaspi Магазин":
        category_path = category_path[1:]
    category_leaf = category_path[-1] if category_path else ""

    row = {
        "group": product["group"],
        "input_name": product["name"],
        "product_name": product_name,
        "product_id": product_id,
        "product_url": url,
        "price_kzt": card.get("price"),
        "brand": card.get("promoConditions", {}).get("brand") or digital.get("brand") or "",
        "created_time": card.get("createdTime") or "",
        "card_age_days": age_days(card.get("createdTime") or ""),
        "category_path": " > ".join(category_path),
        "category_leaf": category_leaf,
        "photo_count": len(item.get("galleryImages") or []),
        "has_video": bool(item.get("videoId") or item.get("videos")),
        "video_id": item.get("videoId") or "",
        "has_description": bool(description_text),
        "description_chars": len(description_text),
        "description_preview": description_text[:500],
        "spec_count": len(specs_rows),
        "spec_groups_count": len(item.get("specifications") or []),
        "rich_content": bool(item.get("richContent") or item.get("videos") or item.get("teasers")),
        "main_type": infer_main_type(product_name, spec_map),
        "support_type": spec_map.get("Тип опоры") or "",
        "purpose": spec_map.get("Назначение") or "",
        "country": spec_map.get("Страна производитель") or "",
        "max_load_kg": parse_number(spec_map.get("Максимальная нагрузка") or product_name),
        "max_height_cm": parse_number(spec_map.get("Максимальный рост, см") or ""),
        "features": spec_map.get("Особенности") or "",
        "wheel_count": parse_number(spec_map.get("Количество колес") or ""),
        "subtype": infer_subtype(product_name, spec_map),
    }
    return row, specs_rows, spec_map


def build_title_score(product_name: str, brand: str, subtype: str, max_load_kg: Any) -> tuple[int, str, str]:
    score = 0
    strengths: list[str] = []
    gaps: list[str] = []
    lowered = product_name.lower()

    if brand and brand.lower() in lowered:
        score += 20
        strengths.append("бренд в заголовке")
    else:
        gaps.append("бренд не читается в заголовке")

    has_model_token = bool(
        re.search(r"\b[A-Za-zА-Яа-я-]*\d+[A-Za-zА-Яа-я0-9-]*\b", product_name)
        or re.search(r"\b(?:regular|shape|wheel|fs|xod|swl|mlk)\b", product_name, re.I)
    )
    if has_model_token:
        score += 20
        strengths.append("есть модель/артикул")
    else:
        gaps.append("нет явной модели")

    if re.search(r"до\s*\d+\s*кг", lowered):
        score += 20
        strengths.append("нагрузка в заголовке")
    elif max_load_kg:
        gaps.append("нагрузка не вынесена в заголовок")

    if subtype and subtype in lowered:
        score += 15
        strengths.append("подтип читается")
    elif subtype != "базовые":
        gaps.append("подтип неочевиден")

    length = len(product_name)
    if 30 <= length <= 90:
        score += 15
        strengths.append("нормальная длина")
    else:
        gaps.append("длина заголовка неоптимальна")

    if "ходунки ходунки" in lowered:
        score -= 10
        gaps.append("дублирование слов")
    if re.search(r"regular.+обыч", lowered) or re.search(r"shape.+образ", lowered) or re.search(r"wheel.+колес", lowered):
        score -= 10
        gaps.append("смешение EN/RU и дубль смысла")
    else:
        score += 10
        strengths.append("без лишних дублей")

    return max(0, min(100, score)), ", ".join(strengths), ", ".join(gaps)


def parse_review_date(value: str) -> datetime | None:
    if not value:
        return None
    value = value.strip()
    try:
        return datetime.strptime(value, "%d.%m.%Y")
    except ValueError:
        pass
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None


def parsed_date_to_iso(value: str) -> str:
    parsed = parse_review_date(value)
    return parsed.date().isoformat() if parsed else ""


def fetch_review_summary(product_id: str) -> tuple[dict[str, Any], dict[str, int]]:
    summary_url = f"https://kaspi.kz/yml/creview/rest/misc/product/{product_id}/summary"
    groups_url = f"https://kaspi.kz/yml/creview/rest/misc/product/{product_id}/groupsummary"
    summary_payload = http_get(summary_url, expect_json=True).get("data") or {}
    groups_payload = http_get(groups_url, expect_json=True).get("data") or []
    groups = {entry.get("id"): int(entry.get("total") or 0) for entry in groups_payload if entry.get("id")}
    return summary_payload, groups


def normalize_comment_block(raw: Any) -> str:
    if not raw:
        return ""
    if isinstance(raw, str):
        return raw.strip()
    if isinstance(raw, list):
        return " ".join(str(part).strip() for part in raw if str(part).strip())
    if isinstance(raw, dict):
        return " ".join(str(part).strip() for part in raw.values() if str(part).strip())
    return str(raw).strip()


def comment_kind(plus_text: str, minus_text: str, body_text: str) -> str:
    if plus_text and minus_text:
        return "Плюсы и минусы"
    if plus_text:
        return "Плюсы"
    if minus_text:
        return "Минусы"
    if body_text:
        return "Текст"
    return "Без текста"


def parse_review(entry: dict[str, Any], product_row: dict[str, Any]) -> dict[str, Any]:
    comment = entry.get("comment") or {}
    feedback = entry.get("feedback") or {}
    plus_text = normalize_comment_block(comment.get("plus") or comment.get("pros") or comment.get("advantages"))
    minus_text = normalize_comment_block(comment.get("minus") or comment.get("cons") or comment.get("disadvantages"))
    body_text = normalize_comment_block(comment.get("text") or comment.get("comment") or comment.get("description"))
    helpful_count = (
        entry.get("helpfulCount")
        or entry.get("usefulCount")
        or entry.get("likes")
        or feedback.get("positive")
        or 0
    )
    try:
        helpful_count = int(helpful_count)
    except (TypeError, ValueError):
        helpful_count = 0
    review_date = str(entry.get("date") or entry.get("createdDate") or entry.get("createdTime") or "")
    gallery = entry.get("galleryImages") or []
    return {
        "group": product_row["group"],
        "product_id": product_row["product_id"],
        "product_name": product_row["product_name"],
        "product_url": product_row["product_url"],
        "review_id": entry.get("id") or "",
        "author": entry.get("author") or "",
        "review_date": review_date,
        "review_date_iso": parsed_date_to_iso(review_date),
        "rating": entry.get("rating"),
        "helpful_count": helpful_count,
        "plus": plus_text,
        "minus": minus_text,
        "comment": body_text,
        "has_text": bool(plus_text or minus_text or body_text),
        "has_photo": bool(gallery),
        "photo_count": len(gallery),
        "merchant_name": (entry.get("merchant") or {}).get("name") or "",
        "merchant_code": (entry.get("merchant") or {}).get("code") or "",
        "comment_kind": comment_kind(plus_text, minus_text, body_text),
    }


def fetch_all_reviews(product_row: dict[str, Any], expected_total: int) -> list[dict[str, Any]]:
    if expected_total <= 0:
        return []
    base_url = f"https://kaspi.kz/yml/review-view/api/v1/reviews/product/{product_row['product_id']}"
    limit = 200
    rows: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    max_pages = max(1, math.ceil(expected_total / limit) + 1)
    for page in range(max_pages):
        params = {"withAgg": "true", "page": page, "limit": limit}
        payload = http_get(base_url, params=params, expect_json=True)
        entries = payload.get("data") or []
        if not entries:
            break
        fresh = 0
        for entry in entries:
            review_id = str(entry.get("id") or "")
            if review_id and review_id in seen_ids:
                continue
            if review_id:
                seen_ids.add(review_id)
            rows.append(parse_review(entry, product_row))
            fresh += 1
        if len(entries) < limit or fresh == 0 or len(rows) >= expected_total:
            break
        time.sleep(0.15)
    return rows


def truncate_review(review: dict[str, Any]) -> str:
    text = " ".join(part for part in [review.get("plus", ""), review.get("minus", ""), review.get("comment", "")] if part)
    text = re.sub(r"\s+", " ", text).strip()
    if len(text) <= 120:
        return text
    return text[:117] + "..."


def rating_distribution(summary: dict[str, Any]) -> dict[int, int]:
    distribution = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
    for item in summary.get("distribution") or []:
        try:
            star = int(item.get("rating"))
        except (TypeError, ValueError):
            continue
        distribution[star] = int(item.get("count") or 0)
    return distribution


def theme_summary(reviews: list[dict[str, Any]], predicate) -> str:
    counts: Counter[str] = Counter()
    for review in reviews:
        if not predicate(review):
            continue
        text = " ".join(
            part.lower()
            for part in [review.get("plus", ""), review.get("minus", ""), review.get("comment", "")]
            if part
        )
        if not text:
            continue
        for theme, keywords in THEME_KEYWORDS.items():
            if any(keyword in text for keyword in keywords):
                counts[theme] += 1
    if not counts:
        return ""
    return "; ".join(f"{theme} ({count})" for theme, count in counts.most_common(3))


def safe_share(numerator: int, denominator: int) -> float:
    if not denominator:
        return 0.0
    return round(numerator / denominator, 4)


def round_price(value: float | int | None) -> int | None:
    if value is None:
        return None
    return int(round(float(value) / 10.0) * 10)


def price_segment(price_kzt: Any) -> str:
    try:
        value = int(price_kzt)
    except (TypeError, ValueError):
        return ""
    if value < 10000:
        return "до 10k"
    if value < 15000:
        return "10k-15k"
    if value < 20000:
        return "15k-20k"
    if value < 50000:
        return "20k-50k"
    return "50k+"


def build_content_score(row: dict[str, Any], title_score: int) -> tuple[int, str, str]:
    score = 0
    strengths: list[str] = []
    gaps: list[str] = []

    photos = int(row.get("photo_count") or 0)
    if photos >= 5:
        score += 25
        strengths.append("5+ фото")
    elif photos == 4:
        score += 20
    elif photos == 3:
        score += 15
    elif photos == 2:
        score += 10
    elif photos == 1:
        score += 5
    else:
        gaps.append("нет фото")

    desc_chars = int(row.get("description_chars") or 0)
    if desc_chars >= 300:
        score += 25
        strengths.append("подробное описание")
    elif desc_chars >= 120:
        score += 18
    elif desc_chars > 0:
        score += 10
    else:
        gaps.append("нет продающего описания")

    specs = int(row.get("spec_count") or 0)
    if specs >= 8:
        score += 20
        strengths.append("много характеристик")
    elif specs >= 5:
        score += 15
    elif specs >= 3:
        score += 10
    else:
        gaps.append("мало характеристик")

    if row.get("has_video"):
        score += 10
        strengths.append("есть видео")
    else:
        gaps.append("нет видео")

    score += round(title_score * 0.2)
    if title_score >= 75:
        strengths.append("понятный заголовок")
    elif title_score < 60:
        gaps.append("заголовок слабее рынка")

    return min(100, score), ", ".join(dict.fromkeys(strengths)), ", ".join(dict.fromkeys(gaps))


def build_social_score(summary: dict[str, Any], reviews: list[dict[str, Any]]) -> int:
    reviews_count = int(summary.get("reviewsCount") or 0)
    rating = float(summary.get("global") or 0)
    review_images = int(summary.get("reviewImagesCount") or 0)
    latest_review_date = max((review["review_date_iso"] for review in reviews if review["review_date_iso"]), default="")

    score = 0
    if reviews_count >= 200:
        score += 40
    elif reviews_count >= 100:
        score += 32
    elif reviews_count >= 50:
        score += 24
    elif reviews_count >= 10:
        score += 14
    elif reviews_count >= 1:
        score += 6

    if rating >= 4.95:
        score += 20
    elif rating >= 4.9:
        score += 18
    elif rating >= 4.8:
        score += 16
    elif rating >= 4.6:
        score += 12
    elif rating > 0:
        score += 8

    if review_images >= 20:
        score += 20
    elif review_images >= 10:
        score += 15
    elif review_images >= 3:
        score += 10
    elif review_images >= 1:
        score += 5

    if latest_review_date:
        try:
            latest = datetime.fromisoformat(latest_review_date)
            days_ago = (datetime.now() - latest).days
            if days_ago <= 30:
                score += 20
            elif days_ago <= 90:
                score += 15
            elif days_ago <= 180:
                score += 10
            else:
                score += 5
        except ValueError:
            pass

    return min(100, score)


def fetch_offer_stats(product_row: dict[str, Any]) -> dict[str, Any]:
    url = f"https://kaspi.kz/yml/offer-view/offers/{product_row['product_id']}"
    payload = {
        "cityId": CITY_CODE,
        "id": product_row["product_id"],
        "merchantUID": [],
        "limit": 50,
        "page": 0,
        "product": {
            "brand": product_row.get("brand") or "No brand",
            "categoryCodes": CATEGORY_CODES,
            "baseProductCodes": [],
            "groups": None,
            "productSeries": [],
        },
        "sortOption": "PRICE",
        "highRating": None,
        "searchText": None,
        "isExcellentMerchant": False,
        "zoneId": [ZONE_ID],
        "installationId": "-1",
    }
    try:
        response = http_post_json(url, payload)
    except Exception:
        return {
            "merchant_count": None,
            "min_offer_price": None,
            "max_offer_price": None,
            "median_offer_price": None,
            "price_spread_kzt": None,
            "price_spread_pct": None,
            "best_offer_merchant": "",
            "best_offer_merchant_rating": None,
            "best_offer_merchant_reviews": None,
            "today_offer_count": None,
            "tomorrow_offer_count": None,
            "high_rating_offer_present": None,
            "excellent_merchant_present": None,
        }

    offers = response.get("offers") or []
    prices = [int(float(offer.get("price") or 0)) for offer in offers if offer.get("price") is not None]
    merchant_count = int(response.get("offersCount") or len(offers) or 0)
    best_offer = offers[0] if offers else {}
    facets = response.get("deliveryDurationFacetValues") or {}
    min_offer_price = min(prices) if prices else None
    max_offer_price = max(prices) if prices else None
    median_offer_price = int(pd.Series(prices).median()) if prices else None
    price_spread_kzt = (max_offer_price - min_offer_price) if prices else None
    price_spread_pct = round((price_spread_kzt / min_offer_price) * 100, 2) if prices and min_offer_price else None

    return {
        "merchant_count": merchant_count,
        "min_offer_price": min_offer_price,
        "max_offer_price": max_offer_price,
        "median_offer_price": median_offer_price,
        "price_spread_kzt": price_spread_kzt,
        "price_spread_pct": price_spread_pct,
        "best_offer_merchant": best_offer.get("merchantName") or "",
        "best_offer_merchant_rating": best_offer.get("merchantRating"),
        "best_offer_merchant_reviews": best_offer.get("merchantReviewsQuantity"),
        "today_offer_count": facets.get("TODAY"),
        "tomorrow_offer_count": facets.get("TOMORROW"),
        "high_rating_offer_present": response.get("highRatingPresent"),
        "excellent_merchant_present": response.get("excellentMerchantPresent"),
    }


def build_commercial_score(offer_stats: dict[str, Any], row: dict[str, Any]) -> int:
    merchant_count = int(offer_stats.get("merchant_count") or 0)
    score = 0
    if merchant_count >= 10:
        score += 35
    elif merchant_count >= 5:
        score += 28
    elif merchant_count >= 3:
        score += 22
    elif merchant_count >= 2:
        score += 16
    elif merchant_count == 1:
        score += 10

    spread = offer_stats.get("price_spread_pct")
    if spread is not None:
        if spread <= 2:
            score += 20
        elif spread <= 5:
            score += 16
        elif spread <= 10:
            score += 10
        else:
            score += 4

    today_offers = int(offer_stats.get("today_offer_count") or 0)
    tomorrow_offers = int(offer_stats.get("tomorrow_offer_count") or 0)
    if today_offers >= 1:
        score += 20
    elif tomorrow_offers >= 1:
        score += 14

    merchant_rating = offer_stats.get("best_offer_merchant_rating")
    if merchant_rating is not None:
        try:
            rating = float(merchant_rating)
        except (TypeError, ValueError):
            rating = 0.0
        if rating >= 4.9:
            score += 15
        elif rating >= 4.8:
            score += 12
        elif rating >= 4.6:
            score += 8

    if row.get("price_kzt"):
        score += 10

    return min(100, score)


def build_product_row(
    card_row: dict[str, Any],
    reviews: list[dict[str, Any]],
    summary: dict[str, Any],
    groups: dict[str, int],
    offer_stats: dict[str, Any],
) -> dict[str, Any]:
    distribution = rating_distribution(summary)
    reviews_count = int(summary.get("reviewsCount") or 0)
    text_reviews = sum(1 for review in reviews if review["has_text"])
    helpful_reviews = [review for review in reviews if review["helpful_count"] > 0]
    helpful_votes_total = sum(review["helpful_count"] for review in reviews)
    latest_review_date = max((review["review_date_iso"] for review in reviews if review["review_date_iso"]), default="")
    earliest_review_date = min((review["review_date_iso"] for review in reviews if review["review_date_iso"]), default="")
    top_helpful = sorted(helpful_reviews, key=lambda row: (row["helpful_count"], row["review_date_iso"]), reverse=True)[:3]
    top_helpful_text = " | ".join(
        f"{review['author']} {review['review_date']} ({review['helpful_count']}): {truncate_review(review)}"
        for review in top_helpful
    )

    title_score, title_strengths, title_gaps = build_title_score(
        card_row["product_name"],
        card_row.get("brand") or "",
        card_row.get("subtype") or "",
        card_row.get("max_load_kg"),
    )
    content_score, content_strengths, content_gaps = build_content_score(card_row, title_score)
    social_score = build_social_score(summary, reviews)
    commercial_score = build_commercial_score(offer_stats, card_row)
    card_score = round(content_score * 0.45 + social_score * 0.35 + commercial_score * 0.20)

    row = dict(card_row)
    row.update(
        {
            "rating_global": summary.get("global"),
            "rating_count": int(summary.get("ratingCount") or 0),
            "reviews_count": reviews_count,
            "review_images_count": int(summary.get("reviewImagesCount") or 0),
            "adjusted_rating": summary.get("adjustedRating"),
            "reviews_with_text": text_reviews,
            "reviews_with_text_share": safe_share(text_reviews, reviews_count),
            "reviews_with_photo": sum(1 for review in reviews if review["has_photo"]),
            "reviews_with_helpful_votes": len(helpful_reviews),
            "helpful_votes_total": helpful_votes_total,
            "latest_review_date": latest_review_date,
            "earliest_review_date": earliest_review_date,
            "group_positive_count": groups.get("POSITIVE", 0),
            "group_negative_count": groups.get("NEGATIVE", 0),
            "group_neutral_count": groups.get("NEUTRAL", 0),
            "star_5_count": distribution[5],
            "star_4_count": distribution[4],
            "star_3_count": distribution[3],
            "star_2_count": distribution[2],
            "star_1_count": distribution[1],
            "top_positive_themes": theme_summary(reviews, lambda review: int(review.get("rating") or 0) >= 4),
            "top_negative_themes": theme_summary(reviews, lambda review: int(review.get("rating") or 0) <= 2),
            "top_helpful_reviews": top_helpful_text,
            "title_score_100": title_score,
            "title_strengths": title_strengths,
            "title_gaps": title_gaps,
            "content_score_100": content_score,
            "content_strengths": content_strengths,
            "content_gaps": content_gaps,
            "social_score_100": social_score,
            "commercial_score_100": commercial_score,
            "card_score_100": card_score,
            "price_segment": price_segment(card_row.get("price_kzt")),
        }
    )
    row.update(offer_stats)
    return row


def listing_query(brand: str | None = None) -> str:
    if not brand:
        return CATEGORY_QUERY
    return f":category:supports and walkers:manufacturerName:{brand}:availableInZones:{ZONE_ID}"


def fetch_listing_page(sort: str = "", page: int = 0, brand: str | None = None) -> dict[str, Any]:
    params = {
        "q": listing_query(brand),
        "page": page,
        "all": "false",
        "fl": "true",
        "ui": "d",
        "i": "-1",
        "c": CITY_CODE,
    }
    if sort:
        params["sort"] = sort
    return http_get(LISTING_API_URL, params=params, expect_json=True).get("data") or {}


def collect_visible_listing(sort: str = "", brand: str | None = None, max_pages: int = 2) -> list[dict[str, Any]]:
    cards: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for page in range(max_pages):
        data = fetch_listing_page(sort=sort, page=page, brand=brand)
        page_cards = data.get("cards") or []
        if not page_cards:
            break
        start_rank = 1 + page * int(data.get("limit") or len(page_cards) or 12)
        for offset, card in enumerate(page_cards):
            card_id = str(card.get("id") or "")
            if not card_id or card_id in seen_ids:
                continue
            seen_ids.add(card_id)
            enriched = dict(card)
            enriched["visible_rank"] = start_rank + offset
            enriched["visible_sort"] = sort or "popular"
            enriched["product_url"] = f"https://kaspi.kz/shop{card.get('shopLink')}"
            enriched["subtype"] = infer_subtype(str(card.get("title") or ""), {})
            cards.append(enriched)
    return cards


def build_filters_sheet(filters: list[dict[str, Any]], total: int) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for filter_item in filters:
        for entry in filter_item.get("rows") or []:
            count = int(entry.get("count") or 0)
            rows.append(
                {
                    "filter_name": filter_item.get("name"),
                    "filter_id": filter_item.get("id"),
                    "value": entry.get("name"),
                    "count": count,
                    "share_of_total": round(count / total, 4) if total else 0.0,
                    "hidden": entry.get("hide"),
                }
            )
    df = pd.DataFrame(rows)
    if df.empty:
        return df
    return df.sort_values(["filter_name", "count", "value"], ascending=[True, False, True]).reset_index(drop=True)


def build_listing_df(cards: list[dict[str, Any]], label: str) -> pd.DataFrame:
    rows = []
    for card in cards:
        rows.append(
            {
                "listing": label,
                "visible_rank": card.get("visible_rank"),
                "product_id": card.get("id"),
                "product_name": card.get("title"),
                "brand": card.get("brand"),
                "subtype": card.get("subtype"),
                "price_kzt": card.get("unitPrice"),
                "rating": card.get("rating"),
                "reviews_count": card.get("reviewsQuantity"),
                "merchant_id": card.get("bestMerchant"),
                "delivery_duration": card.get("deliveryDuration"),
                "product_url": card.get("product_url"),
            }
        )
    return pd.DataFrame(rows)


def visible_rank_map(cards: list[dict[str, Any]]) -> dict[str, int]:
    return {str(card.get("id")): int(card.get("visible_rank") or 0) for card in cards if card.get("id")}


def collect_osako_pool() -> list[dict[str, Any]]:
    cards = []
    for sort in ["", "rating"]:
        cards.extend(collect_visible_listing(sort=sort, brand="Osako", max_pages=2))
    deduped: dict[str, dict[str, Any]] = {}
    for card in cards:
        card_id = str(card.get("id") or "")
        current = deduped.get(card_id)
        if not current:
            deduped[card_id] = card
            continue
        if (card.get("reviewsQuantity") or 0) > (current.get("reviewsQuantity") or 0):
            deduped[card_id] = card
    return list(deduped.values())


def select_peer_cards(
    own_cards: list[dict[str, Any]],
    top_market_cards: list[dict[str, Any]],
    osako_cards: list[dict[str, Any]],
) -> list[dict[str, str]]:
    own_ids = {row["product_id"] for row in own_cards}
    top_rating_ids = {extract_product_code(product["url"]) for product in TOP_RATING_PRODUCTS}
    results: dict[str, dict[str, str]] = {}

    def add_card(card: dict[str, Any], group: str) -> None:
        card_id = str(card.get("id") or "")
        if not card_id or card_id in own_ids or card_id in top_rating_ids or card_id in results:
            return
        results[card_id] = {
            "group": group,
            "name": str(card.get("title") or ""),
            "url": str(card.get("product_url") or ""),
        }

    for own in own_cards:
        subtype = own.get("subtype") or ""
        own_price = int(own.get("price_kzt") or 0)

        same_brand = [
            card
            for card in osako_cards
            if str(card.get("id")) not in own_ids
            and int(card.get("reviewsQuantity") or 0) > 0
            and infer_subtype(str(card.get("title") or ""), {}) == subtype
        ]
        same_brand.sort(
            key=lambda card: (
                -int(card.get("reviewsQuantity") or 0),
                abs(int(card.get("unitPrice") or 0) - own_price),
            )
        )
        for card in same_brand[:2]:
            add_card(card, "Osako benchmark")

        external = [
            card
            for card in top_market_cards
            if str(card.get("id")) not in own_ids
            and str(card.get("brand") or "").lower() != "osako"
            and infer_subtype(str(card.get("title") or ""), {}) == subtype
        ]
        if not external:
            external = [
                card
                for card in top_market_cards
                if str(card.get("id")) not in own_ids and str(card.get("brand") or "").lower() != "osako"
            ]
        external.sort(
            key=lambda card: (
                -float(card.get("rating") or 0),
                -int(card.get("reviewsQuantity") or 0),
                abs(int(card.get("unitPrice") or 0) - own_price),
            )
        )
        for card in external[:1]:
            add_card(card, "Market peer")

    return list(results.values())


def build_comparison_sheet(products_df: pd.DataFrame) -> pd.DataFrame:
    own_df = products_df.loc[products_df["group"] == "Мой SKU"].copy()
    peer_df = products_df.loc[products_df["group"] != "Мой SKU"].copy()
    rows: list[dict[str, Any]] = []

    for _, own_row in own_df.iterrows():
        subtype = own_row["subtype"]
        same_subtype = peer_df.loc[peer_df["subtype"] == subtype].copy()
        if same_subtype.empty:
            same_subtype = peer_df.copy()
        same_subtype["price_gap_abs"] = (same_subtype["price_kzt"].fillna(0) - int(own_row["price_kzt"] or 0)).abs()
        same_subtype = same_subtype.sort_values(
            ["group", "price_gap_abs", "reviews_count", "card_score_100"],
            ascending=[True, True, False, False],
        )
        for _, peer_row in same_subtype.head(4).iterrows():
            rows.append(
                {
                    "my_product": own_row["product_name"],
                    "my_url": own_row["product_url"],
                    "my_subtype": own_row["subtype"],
                    "my_price_kzt": own_row["price_kzt"],
                    "my_reviews": own_row["reviews_count"],
                    "my_content_score": own_row["content_score_100"],
                    "peer_group": peer_row["group"],
                    "peer_product": peer_row["product_name"],
                    "peer_url": peer_row["product_url"],
                    "peer_price_kzt": peer_row["price_kzt"],
                    "peer_reviews": peer_row["reviews_count"],
                    "peer_content_score": peer_row["content_score_100"],
                    "price_delta_kzt": int(peer_row["price_kzt"] or 0) - int(own_row["price_kzt"] or 0),
                    "review_delta": int(peer_row["reviews_count"] or 0) - int(own_row["reviews_count"] or 0),
                    "content_score_delta": int(peer_row["content_score_100"] or 0) - int(own_row["content_score_100"] or 0),
                    "card_score_delta": int(peer_row["card_score_100"] or 0) - int(own_row["card_score_100"] or 0),
                    "peer_advantage": peer_row["content_strengths"] or peer_row["title_strengths"],
                    "my_gap": own_row["content_gaps"] or own_row["title_gaps"],
                }
            )
    return pd.DataFrame(rows)


def recommended_price_range(own_row: pd.Series, comparison_df: pd.DataFrame) -> str:
    subset = comparison_df.loc[comparison_df["my_product"] == own_row["product_name"]].copy()
    if subset.empty:
        return ""
    peer_prices = [int(value) for value in subset["peer_price_kzt"].dropna().tolist() if int(value) > 0]
    if not peer_prices:
        return ""
    target_floor = round_price(min(peer_prices) * 0.95)
    target_ceiling = round_price(min(peer_prices))
    own_price = int(own_row.get("price_kzt") or 0)
    if int(own_row.get("reviews_count") or 0) > 0:
        target_floor = round_price(min(peer_prices) * 0.98)
        target_ceiling = round_price(min(peer_prices) * 1.02)
    if own_price <= target_ceiling:
        return f"{target_floor}-{target_ceiling}"
    return f"{target_floor}-{target_ceiling}"


def build_recommendations_sheet(products_df: pd.DataFrame, comparison_df: pd.DataFrame) -> pd.DataFrame:
    own_df = products_df.loc[products_df["group"] == "Мой SKU"].copy()
    recommendations: list[dict[str, Any]] = []

    for _, row in own_df.iterrows():
        price_range = recommended_price_range(row, comparison_df)

        recommendations.append(
            {
                "scope": row["product_name"],
                "priority": "Высокий",
                "stream": "Отзывы",
                "issue": "0 отзывов = нулевая соцдоказательность",
                "action": "Собрать первые 10-15 честных отзывов и минимум 3-5 фотоотзывов от реальных покупателей.",
                "kpi": "10+ отзывов, 3+ фотоотзыва, рейтинг 4.8+",
            }
        )

        if price_range:
            recommendations.append(
                {
                    "scope": row["product_name"],
                    "priority": "Высокий",
                    "stream": "Цена",
                    "issue": "Новый SKU дороже или не дешевле карточек с отзывами.",
                    "action": f"Запустить товар в ценовом коридоре {price_range} ₸ до накопления базовой репутации.",
                    "kpi": "Конверсия в заказ выше текущей; 10+ заказов на SKU",
                }
            )

        if int(row.get("title_score_100") or 0) < 70:
            recommendations.append(
                {
                    "scope": row["product_name"],
                    "priority": "Высокий",
                    "stream": "Контент",
                    "issue": "Заголовок хуже читается, чем у лидеров.",
                    "action": "Переписать title в формате: Бренд + тип + модель + ключевая особенность + нагрузка.",
                    "kpi": "Title score 80+",
                }
            )

        if not row.get("has_description") or int(row.get("description_chars") or 0) < 150:
            recommendations.append(
                {
                    "scope": row["product_name"],
                    "priority": "Высокий",
                    "stream": "Контент",
                    "issue": "Недостаточно продающего описания.",
                    "action": "Добавить описание со сценариями использования, ростом/нагрузкой, материалом, складным механизмом и инструкцией по подбору.",
                    "kpi": "Описание 250+ символов",
                }
            )

        if int(row.get("spec_count") or 0) < 8:
            recommendations.append(
                {
                    "scope": row["product_name"],
                    "priority": "Средний",
                    "stream": "Характеристики",
                    "issue": "Мало характеристик в карточке.",
                    "action": "Заполнить максимальную нагрузку, рост, ширину, вес изделия, материал, наличие колес, складной механизм, страну производства.",
                    "kpi": "8+ характеристик",
                }
            )

        recommendations.append(
            {
                "scope": row["product_name"],
                "priority": "Средний",
                "stream": "CRM / post-purchase",
                "issue": "Нет механики сбора отзывов.",
                "action": "После доставки на 3-5 день делать нейтральный follow-up: помощь со сборкой, затем просьба оценить товар, если все подошло.",
                "kpi": "Ответ на follow-up 20%+, review rate 8%+",
            }
        )

    recommendations.extend(
        [
            {
                "scope": "Портфель Osako",
                "priority": "Высокий",
                "stream": "Ассортимент",
                "issue": "В категории много близких Osako SKU, они каннибализируют друг друга.",
                "action": "Сократить дубляж моделей и оставить 1 hero-SKU на подтип: regular, s-shape, wheels.",
                "kpi": "Меньше внутренних дублей, больше отзывов на hero-SKU",
            },
            {
                "scope": "Портфель Osako",
                "priority": "Высокий",
                "stream": "Отзывы",
                "issue": "Отзывы распыляются между похожими карточками.",
                "action": "Лить трафик и CRM-коммуникации в 3 приоритетные карточки, а не во все новые SKU одновременно.",
                "kpi": "20+ отзывов на hero-SKU за 30-45 дней",
            },
            {
                "scope": "Портфель Osako",
                "priority": "Средний",
                "stream": "Визуал",
                "issue": "Покупателю сложно быстро понять отличия между моделями.",
                "action": "Сделать первое фото с крупным differentiator: Regular / S-shape / Wheels, плюс нагрузка и рост.",
                "kpi": "Рост CTR из листинга",
            },
        ]
    )

    return pd.DataFrame(recommendations)


def build_summary_sheet(
    total_skus: int,
    filters_df: pd.DataFrame,
    top_popular_df: pd.DataFrame,
    top_rating_df: pd.DataFrame,
    products_df: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []

    def get_filter_value(filter_name: str, value: str) -> int:
        subset = filters_df.loc[(filters_df["filter_name"] == filter_name) & (filters_df["value"] == value)]
        if subset.empty:
            return 0
        return int(subset.iloc[0]["count"])

    rows.extend(
        [
            {"metric": "Дата анализа", "value": datetime.now(ALMATY_TZ).strftime("%d.%m.%Y %H:%M"), "note": "Алматы"},
            {"metric": "Категория", "value": "Опоры и ходунки", "note": CATEGORY_URL},
            {"metric": "SKU в категории", "value": total_skus, "note": "по filter endpoint Kaspi"},
            {"metric": "Доля типа 'ходунки'", "value": get_filter_value("Тип", "ходунки"), "note": f"{round(get_filter_value('Тип', 'ходунки') / total_skus * 100, 1) if total_skus else 0}%"},
            {"metric": "SKU в ценовом сегменте 10k-50k", "value": get_filter_value("Цена", "10 000 - 49 999 т"), "note": f"{round(get_filter_value('Цена', '10 000 - 49 999 т') / total_skus * 100, 1) if total_skus else 0}%"},
            {"metric": "Osako SKU в категории", "value": get_filter_value("Бренд", "Osako"), "note": "широкая полка бренда"},
            {"metric": "Топ продавец по числу SKU", "value": "OSAKO BRAND STORE", "note": str(get_filter_value("Продавцы", "OSAKO BRAND STORE"))},
        ]
    )

    if not top_popular_df.empty:
        top_pop = top_popular_df.iloc[0]
        rows.append(
            {
                "metric": "Лидер по популярности",
                "value": top_pop["product_name"],
                "note": f"{top_pop['price_kzt']} ₸ | {top_pop['reviews_count']} отзывов",
            }
        )
    if not top_rating_df.empty:
        top_rate = top_rating_df.iloc[0]
        rows.append(
            {
                "metric": "Лидер по рейтингу",
                "value": top_rate["product_name"],
                "note": f"{top_rate['price_kzt']} ₸ | рейтинг {top_rate['rating']}",
            }
        )

    own_df = products_df.loc[products_df["group"] == "Мой SKU"].copy()
    if not own_df.empty:
        rows.append(
            {
                "metric": "Мои SKU с отзывами",
                "value": int((own_df["reviews_count"].fillna(0) > 0).sum()),
                "note": f"из {len(own_df)} SKU",
            }
        )
    return pd.DataFrame(rows)


def format_date(value: str) -> str:
    if not value:
        return ""
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).strftime("%d.%m.%Y")
    except ValueError:
        return value


def build_markdown(
    total_skus: int,
    filters_df: pd.DataFrame,
    top_popular_df: pd.DataFrame,
    top_rating_df: pd.DataFrame,
    products_df: pd.DataFrame,
    comparison_df: pd.DataFrame,
) -> str:
    own_df = products_df.loc[products_df["group"] == "Мой SKU"].copy()
    peer_df = products_df.loc[products_df["group"] != "Мой SKU"].copy()

    lines = [
        "# Kaspi: аудит категории «Опоры и ходунки»",
        "",
        f"Дата анализа: {datetime.now(ALMATY_TZ).strftime('%d.%m.%Y %H:%M')} Алматы",
        f"Категория: {CATEGORY_URL}",
        f"Высокий рейтинг: {CATEGORY_URL_RATING}",
        "",
        "## Ключевые выводы",
        f"- В категории {total_skus} SKU. Основной спрос визуально и по фильтрам сидит в сегменте ходунков и цене 10-50 тыс. ₸.",
        f"- Бренд Osako уже занимает широкую полку: {int(filters_df.loc[(filters_df['filter_name'] == 'Бренд') & (filters_df['value'] == 'Osako'), 'count'].iloc[0]) if not filters_df.empty and not filters_df.loc[(filters_df['filter_name'] == 'Бренд') & (filters_df['value'] == 'Osako')].empty else 0} SKU.",
        "- Проблема новых SKU не только во внешней конкуренции, но и во внутренней каннибализации более старых Osako-карточек с отзывами и сопоставимой ценой.",
        "",
        "## Мои SKU",
    ]

    for _, row in own_df.iterrows():
        lines.extend(
            [
                f"### {row['product_name']}",
                f"- Цена: {row['price_kzt']} ₸",
                f"- Подтип: {row['subtype']}",
                f"- Отзывы: {row['reviews_count']} | рейтинг: {row['rating_global'] or 'нет'}",
                f"- Контент: фото {row['photo_count']}, описание {'есть' if row['has_description'] else 'нет'}, видео {'есть' if row['has_video'] else 'нет'}, характеристик {row['spec_count']}",
                f"- Заголовок: score {row['title_score_100']}/100 | пробелы: {row['title_gaps'] or 'критичных нет'}",
                f"- Ключевые пробелы: {row['content_gaps'] or 'нет'}",
                "",
            ]
        )

    lines.append("## С кем реально конкурируют новые SKU")
    for own_name in own_df["product_name"]:
        lines.append(f"### {own_name}")
        subset = comparison_df.loc[comparison_df["my_product"] == own_name].head(3)
        for _, row in subset.iterrows():
            lines.append(
                f"- {row['peer_group']}: {row['peer_product']} | цена {row['peer_price_kzt']} ₸ | "
                f"отзывов {row['peer_reviews']} | дельта цены {row['price_delta_kzt']} ₸"
            )
        lines.append("")

    if not top_rating_df.empty:
        top = top_rating_df.iloc[0]
        lines.extend(
            [
                "## Лидер рейтинга",
                f"- {top['product_name']} | {top['price_kzt']} ₸ | {top['reviews_count']} отзывов | рейтинг {top['rating']}",
                "",
            ]
        )

    if not peer_df.empty:
        strongest = peer_df.sort_values(["reviews_count", "card_score_100"], ascending=[False, False]).iloc[0]
        lines.extend(
            [
                "## Главный сигнал рынка",
                f"- Самые сильные карточки выигрывают не только ценой. Они одновременно держат понятный title, базу отзывов и уже накопленное доверие.",
                f"- Наиболее сильный peer в выборке: {strongest['product_name']} "
                f"({strongest['reviews_count']} отзывов, score {strongest['card_score_100']}/100).",
            ]
        )

    return "\n".join(lines).strip() + "\n"


def autosize_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    own_fill = PatternFill("solid", fgColor="E2F0D9")
    benchmark_fill = PatternFill("solid", fgColor="FFF2CC")
    top_fill = PatternFill("solid", fgColor="FCE4D6")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column in sheet.columns:
            width = 0
            for idx, cell in enumerate(column, start=1):
                if idx > 250:
                    break
                value = "" if cell.value is None else str(cell.value)
                width = max(width, min(len(value), 120))
            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(max(width + 2, 12), 70)

        headers = {cell.value: cell.column for cell in sheet[1]}
        group_col = headers.get("group")
        if not group_col:
            continue
        for row_idx in range(2, sheet.max_row + 1):
            group_val = str(sheet.cell(row=row_idx, column=group_col).value or "")
            if group_val == "Мой SKU":
                fill = own_fill
            elif group_val == "Osako benchmark":
                fill = benchmark_fill
            elif group_val == "Top rating":
                fill = top_fill
            else:
                continue
            for col_idx in range(1, sheet.max_column + 1):
                sheet.cell(row=row_idx, column=col_idx).fill = fill

    workbook.save(path)


def send_message(text: str) -> None:
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        raise RuntimeError("TELEGRAM_BOT_TOKEN or TELEGRAM_CHAT_ID is not set")
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    response = SESSION.post(url, data={"chat_id": chat_id, "text": text}, timeout=60)
    response.raise_for_status()


def send_file(file_path: Path, caption: str | None = None) -> None:
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    chat_id = os.environ.get("TELEGRAM_CHAT_ID")
    if not token or not chat_id:
        raise RuntimeError("TELEGRAM_BOT_TOKEN or TELEGRAM_CHAT_ID is not set")
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    with open(file_path, "rb") as handle:
        response = SESSION.post(
            url,
            data={"chat_id": chat_id, "caption": caption or ""},
            files={"document": handle},
            timeout=180,
        )
    response.raise_for_status()


def main() -> None:
    load_env_from_file()
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    listing_data = fetch_listing_page(sort="", page=0)
    total_skus = int(listing_data.get("total") or 0)
    filters = listing_data.get("filters") or []
    filters_df = build_filters_sheet(filters, total_skus)

    top_popular_cards = collect_visible_listing(sort="", max_pages=2)
    top_rating_cards = collect_visible_listing(sort="rating", max_pages=2)
    top_market_pool = list({str(card.get("id")): card for card in (top_popular_cards + top_rating_cards)}.values())

    own_card_rows: list[dict[str, Any]] = []
    own_specs: list[dict[str, Any]] = []
    own_spec_maps: dict[str, dict[str, str]] = {}
    for product in OWN_PRODUCTS:
        card_row, specs_rows, spec_map = fetch_card(product)
        own_card_rows.append(card_row)
        own_specs.extend(specs_rows)
        own_spec_maps[card_row["product_id"]] = spec_map

    osako_pool = collect_osako_pool()
    peer_products = select_peer_cards(own_card_rows, top_market_pool, osako_pool)
    analyzed_products = OWN_PRODUCTS + TOP_RATING_PRODUCTS + peer_products

    bundles: list[dict[str, Any]] = []
    all_specs: list[dict[str, Any]] = []
    all_reviews: list[dict[str, Any]] = []

    popular_rank = visible_rank_map(top_popular_cards)
    rating_rank = visible_rank_map(top_rating_cards)
    osako_visible_rank = visible_rank_map(osako_pool)

    for index, product in enumerate(analyzed_products, start=1):
        print(f"[{index}/{len(analyzed_products)}] {product['name']}", flush=True)
        card_row, specs_rows, spec_map = fetch_card(product)
        summary, groups = fetch_review_summary(card_row["product_id"])
        reviews_rows = fetch_all_reviews(card_row, int(summary.get("reviewsCount") or 0))
        offer_stats = fetch_offer_stats(card_row)
        product_row = build_product_row(card_row, reviews_rows, summary, groups, offer_stats)
        product_row["visible_rank_popular_top24"] = popular_rank.get(card_row["product_id"])
        product_row["visible_rank_rating_top24"] = rating_rank.get(card_row["product_id"])
        product_row["visible_rank_osako_pool"] = osako_visible_rank.get(card_row["product_id"])

        bundles.append(product_row)
        all_specs.extend(specs_rows)
        all_reviews.extend(reviews_rows)
        time.sleep(0.2)

    products_df = pd.DataFrame(bundles)
    if not products_df.empty:
        group_priority = {"Мой SKU": 0, "Top rating": 1, "Osako benchmark": 2, "Market peer": 3}
        products_df["group_priority"] = products_df["group"].map(group_priority).fillna(99)
        products_df = (
            products_df.sort_values(
                ["product_id", "group_priority", "reviews_count", "card_score_100"],
                ascending=[True, True, False, False],
            )
            .drop_duplicates(subset=["product_id"], keep="first")
            .sort_values(["group_priority", "reviews_count", "card_score_100"], ascending=[True, False, False])
            .drop(columns=["group_priority"])
            .reset_index(drop=True)
        )

    specs_df = pd.DataFrame(all_specs)
    if not specs_df.empty:
        specs_df = specs_df.drop_duplicates().sort_values(["product_name", "group_name", "spec_name"])

    reviews_df = pd.DataFrame(all_reviews)
    if not reviews_df.empty:
        review_dedupe_cols = ["product_id", "review_id"] if "review_id" in reviews_df.columns else ["product_id", "review_date_iso", "author"]
        reviews_df = (
            reviews_df.drop_duplicates(subset=review_dedupe_cols, keep="first")
            .sort_values(["product_name", "review_date_iso", "helpful_count"], ascending=[True, False, False])
            .reset_index(drop=True)
        )

    top_popular_df = build_listing_df(top_popular_cards, "popular")
    top_rating_df = build_listing_df(top_rating_cards, "rating")
    comparison_df = build_comparison_sheet(products_df)
    recommendations_df = build_recommendations_sheet(products_df, comparison_df)
    summary_df = build_summary_sheet(total_skus, filters_df, top_popular_df, top_rating_df, products_df)
    markdown = build_markdown(total_skus, filters_df, top_popular_df, top_rating_df, products_df, comparison_df)

    timestamp = datetime.now(ALMATY_TZ).strftime("%Y-%m-%d_%H-%M-%S")
    excel_path = RESULTS_DIR / f"walkers_card_audit_{timestamp}.xlsx"
    report_md_path = RESULTS_DIR / f"walkers_card_audit_{timestamp}.md"

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        products_df.to_excel(writer, sheet_name="Products", index=False)
        comparison_df.to_excel(writer, sheet_name="Comparison", index=False)
        recommendations_df.to_excel(writer, sheet_name="Recommendations", index=False)
        top_popular_df.to_excel(writer, sheet_name="TopPopular", index=False)
        top_rating_df.to_excel(writer, sheet_name="TopRating", index=False)
        filters_df.to_excel(writer, sheet_name="CategoryFilters", index=False)
        specs_df.to_excel(writer, sheet_name="Specs", index=False)
        if not reviews_df.empty:
            reviews_df.to_excel(writer, sheet_name="AllReviews", index=False)

    report_md_path.write_text(markdown, encoding="utf-8")
    autosize_workbook(excel_path)

    telegram_sent = False
    telegram_error = ""
    try:
        send_message(
            "Kaspi: аудит категории «Опоры и ходунки» готов.\n"
            f"Дата: {datetime.now(ALMATY_TZ).strftime('%d.%m.%Y %H:%M')} Алматы\n"
            f"SKU в категории: {total_skus}\n"
            f"Карточек в детальном разборе: {len(products_df)}"
        )
        send_file(excel_path, caption="Kaspi: аудит категории Опоры и ходунки")
        send_file(report_md_path, caption="Kaspi: текстовое summary по категории")
        telegram_sent = True
    except Exception as exc:
        telegram_error = str(exc)

    print(excel_path)
    print(report_md_path)
    print(f"telegram_sent={telegram_sent}")
    if telegram_error:
        print(f"telegram_error={telegram_error}")


if __name__ == "__main__":
    main()
