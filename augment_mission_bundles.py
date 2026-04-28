#!/usr/bin/env python3
"""Дополняем отчёт миссии колонками по бандлам (bundle_of/bundle_qty)."""

from __future__ import annotations

import argparse
import math
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
from playwright.sync_api import sync_playwright

from Scraper_Kaspi import scrape_single_product, normalize_seller_name, seller_matches


def _safe_int(val: Any) -> int | None:
    try:
        if pd.isna(val):
            return None
        s = str(val).replace(" ", "")
        return int(float(s))
    except Exception:
        return None


def fetch_prices(page, url: str, city: str, seller_hint: str) -> tuple[int | None, int | None, str | None]:
    """Цена нужного продавца и лучшая цена по рынку (и продавец лучшей цены)."""
    recs = scrape_single_product(page, url, input_name="base", city=city) or []
    target_norm = normalize_seller_name(seller_hint)
    prices = []
    best_overall = []
    for rec in recs:
        seller = normalize_seller_name(rec.get("seller", ""))
        price = _safe_int(rec.get("price_kzt"))
        if price:
            best_overall.append((price, rec.get("seller")))
        if seller_matches(seller, target_norm) and price:
            prices.append(price)
    base_price = min(prices) if prices else None
    best_price, best_seller = (min(best_overall, key=lambda x: x[0]) if best_overall else (None, None))
    return base_price, best_price, best_seller


def augment_mission(path: Path, limit: int | None = None) -> Path:
    df = pd.read_excel(path)
    if "bundle_of" not in df.columns or "bundle_qty" not in df.columns:
        raise SystemExit("Файл не содержит колонок bundle_of / bundle_qty")

    rows = df.to_dict(orient="records")
    if limit:
        rows = rows[:limit]

    results: list[dict[str, Any]] = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--disable-dev-shm-usage"])
        context = browser.new_context(
            viewport={"width": 1600, "height": 900},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36",
        )
        for row in rows:
            bundle_url = str(row.get("bundle_of") or "").strip()
            bundle_qty = _safe_int(row.get("bundle_qty")) or 1
            seller_hint = str(row.get("seller") or "")
            city = str(row.get("region") or "Алматы")
            bundle_price = _safe_int(row.get("price_kzt"))
            base_price = None
            best_price = None
            best_seller = None
            if bundle_url:
                page = context.new_page()
                try:
                    base_price, best_price, best_seller = fetch_prices(page, bundle_url, city, seller_hint)
                except Exception as exc:
                    # отмечаем ошибку в заметке
                    row["bundle_note"] = f"Ошибка base: {exc}"
                finally:
                    page.close()
            bundle_unit = math.ceil(bundle_price / bundle_qty) if bundle_price is not None and bundle_qty else None
            delta_single = bundle_unit - base_price if bundle_unit is not None and base_price is not None else None
            bundle_best_total = best_price * bundle_qty if best_price is not None else None
            delta_best = bundle_price - bundle_best_total if bundle_price is not None and bundle_best_total is not None else None
            row.update({
                "bundle_1_sku_price_kzt": base_price,
                "bundle_price_kzt": bundle_price,
                "bundle_unit_price_kzt": bundle_unit,
                "bundle_vs_single_kzt": delta_single,
                "bundle_best_seller": best_seller,
                "1_sku_best_price_kzt": best_price,
                "bundle_best_price_kzt": bundle_best_total,
                "bundle_vs_best_kzt": delta_best,
            })
            results.append(row)
        context.close()
        browser.close()

    out_df = pd.DataFrame(results)
    # скидка в процентах (положительная — бандл дешевле)
    def _discount(row):
        best_total = row.get("bundle_best_price_kzt")
        bundle_price = row.get("bundle_price_kzt")
        if best_total and bundle_price:
            try:
                return round((best_total - bundle_price) / best_total * 100, 2)
            except Exception:
                return None
        return None

    out_df["bundle_discount_pct"] = out_df.apply(_discount, axis=1)

    desired_order = [
        "region", "seller", "input_product", "product_url", "price_kzt",
        "bundle_of", "bundle_qty",
        "bundle_price_kzt", "bundle_unit_price_kzt", "bundle_1_sku_price_kzt",
        "bundle_vs_single_kzt",
        "bundle_best_seller", "1_sku_best_price_kzt", "bundle_best_price_kzt", "bundle_vs_best_kzt",
        "bundle_discount_pct",
    ]
    for col in desired_order:
        if col not in out_df.columns:
            out_df[col] = None
    out_df = out_df.reindex(columns=desired_order)

    out_path = path.with_name(path.stem + "_bundles.xlsx")
    out_df.to_excel(out_path, index=False)

    # применим формат процента для колонки bundle_discount_pct
    wb = load_workbook(out_path)
    ws = wb.active
    for cell in ws["Q"][1:]:  # колонка Q — bundle_discount_pct
        cell.number_format = numbers.FORMAT_PERCENTAGE_00
    wb.save(out_path)

    return out_path


def main():
    parser = argparse.ArgumentParser(description="Augment mission report with bundle comparisons")
    parser.add_argument("--file", type=Path, required=True, help="Путь к файлу mission_january_*.xlsx")
    parser.add_argument("--limit", type=int, help="Ограничение строк (для теста)")
    args = parser.parse_args()

    out_path = augment_mission(args.file, limit=args.limit)
    print(f"✅ Saved {out_path}")


if __name__ == "__main__":
    main()
