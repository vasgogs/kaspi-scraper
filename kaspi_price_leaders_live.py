#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import fmean
from typing import Any
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook

ALMATY_TZ = ZoneInfo("Asia/Almaty")
FILE_RE = re.compile(r"^kaspi_prices_(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2}-\d{2})\.xlsx$")
LIVE_JSON_NAME = "kaspi_price_leaders_live_latest.json"
LIVE_CSV_NAME = "kaspi_price_leaders_live_latest.csv"
CSV_FIELDS = [
    "month",
    "region",
    "product_code",
    "product_name",
    "seller",
    "wins",
    "product_snapshots",
    "win_share_pct",
    "seller_presence_snapshots",
    "seller_presence_share_pct",
    "avg_sellers_per_snapshot",
    "seller_min_price_kzt",
    "seller_max_price_kzt",
    "is_qualified",
]


def _safe_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def _slugify_region(region: str) -> str:
    known = {
        "алматы": "almaty",
        "астана": "astana",
        "шымкент": "shymkent",
    }
    key = region.strip().lower()
    if key in known:
        return known[key]
    cleaned = re.sub(r"[^a-z0-9]+", "_", key)
    cleaned = cleaned.strip("_")
    return cleaned or "region"


def _iter_snapshot_files(results_dir: Path) -> list[tuple[datetime, Path]]:
    rows: list[tuple[datetime, Path]] = []
    for path in results_dir.glob("kaspi_prices_*.xlsx"):
        match = FILE_RE.match(path.name)
        if not match:
            continue
        dt = datetime.strptime(
            f"{match.group(1)} {match.group(2)}",
            "%Y-%m-%d %H-%M-%S",
        )
        rows.append((dt, path))
    rows.sort(key=lambda x: x[0])
    return rows


def _read_snapshot_offers(path: Path, region: str) -> tuple[dict[str, dict[str, int]], dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        idx = {name: i for i, name in enumerate(header)}
        for col in ("region", "product_code", "seller", "price_kzt"):
            if col not in idx:
                raise RuntimeError(f"{path.name}: missing required column {col}")

        offers: dict[str, dict[str, int]] = defaultdict(dict)
        names: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_region = str(row[idx["region"]] or "").strip()
            if row_region != region:
                continue
            product_code = str(row[idx["product_code"]] or "").strip()
            seller = str(row[idx["seller"]] or "").strip()
            price = _safe_int(row[idx["price_kzt"]])
            if not product_code or not seller or price is None:
                continue

            title = ""
            if "product" in idx:
                title = str(row[idx["product"]] or "").strip()
            if not title and "input_product" in idx:
                title = str(row[idx["input_product"]] or "").strip()
            if title and product_code not in names:
                names[product_code] = title

            prev = offers[product_code].get(seller)
            if prev is None or price < prev:
                offers[product_code][seller] = price
        return offers, names
    finally:
        wb.close()


def _sort_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        rows,
        key=lambda r: (
            -float(r.get("win_share_pct") or 0),
            -int(r.get("wins") or 0),
            -int(r.get("product_snapshots") or 0),
            -float(r.get("avg_sellers_per_snapshot") or 0),
            str(r.get("seller") or ""),
            str(r.get("product_name") or ""),
        ),
    )


def _build_summary(
    results_dir: Path,
    region: str,
    months_limit: int,
    top_n: int,
    min_product_snapshots: int,
    min_avg_sellers: float,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    files = _iter_snapshot_files(results_dir)
    if not files:
        raise RuntimeError("No kaspi_prices_YYYY-MM-DD_HH-MM-SS.xlsx files found")

    month_files_count: Counter[str] = Counter()
    month_latest_dt: dict[str, datetime] = {}
    month_product_snapshots: dict[str, dict[str, list[dict[str, int]]]] = defaultdict(lambda: defaultdict(list))
    product_name_by_code: dict[str, str] = {}

    for snap_dt, path in files:
        offers, names = _read_snapshot_offers(path, region)
        month = snap_dt.strftime("%Y-%m")
        month_files_count[month] += 1
        month_latest_dt[month] = snap_dt
        for code, title in names.items():
            if code not in product_name_by_code and title:
                product_name_by_code[code] = title
        for code, seller_prices in offers.items():
            month_product_snapshots[month][code].append(seller_prices)

    months_desc = sorted(month_product_snapshots.keys(), reverse=True)
    if months_limit > 0:
        months_desc = months_desc[:months_limit]

    all_rows: list[dict[str, Any]] = []
    month_rows: dict[str, list[dict[str, Any]]] = {}
    months_meta: list[dict[str, Any]] = []

    for month in months_desc:
        per_product = month_product_snapshots[month]
        rows: list[dict[str, Any]] = []
        for product_code, snapshots in per_product.items():
            if not snapshots:
                continue
            product_snapshots = len(snapshots)
            avg_sellers = fmean(len(prices) for prices in snapshots)
            wins: Counter[str] = Counter()
            presence: Counter[str] = Counter()
            seller_prices: dict[str, list[int]] = defaultdict(list)

            for seller_prices_map in snapshots:
                if not seller_prices_map:
                    continue
                best_price = min(seller_prices_map.values())
                for seller, price in seller_prices_map.items():
                    presence[seller] += 1
                    seller_prices[seller].append(price)
                    if price == best_price:
                        wins[seller] += 1

            for seller, win_count in wins.items():
                pres_count = int(presence.get(seller) or 0)
                prices = seller_prices.get(seller) or []
                win_share = round((win_count / product_snapshots) * 100, 2)
                presence_share = round((pres_count / product_snapshots) * 100, 2)
                row = {
                    "month": month,
                    "region": region,
                    "product_code": product_code,
                    "product_name": product_name_by_code.get(product_code, ""),
                    "seller": seller,
                    "wins": int(win_count),
                    "product_snapshots": int(product_snapshots),
                    "win_share_pct": win_share,
                    "seller_presence_snapshots": pres_count,
                    "seller_presence_share_pct": presence_share,
                    "avg_sellers_per_snapshot": round(float(avg_sellers), 2),
                    "seller_min_price_kzt": min(prices) if prices else None,
                    "seller_max_price_kzt": max(prices) if prices else None,
                }
                row["is_qualified"] = bool(
                    row["product_snapshots"] >= min_product_snapshots
                    and row["avg_sellers_per_snapshot"] >= min_avg_sellers
                )
                rows.append(row)
                all_rows.append(row)

        rows = _sort_rows(rows)
        qualified = [r for r in rows if r["is_qualified"]]
        top_rows = qualified[:top_n] if qualified else rows[:top_n]
        month_rows[month] = top_rows
        months_meta.append(
            {
                "month": month,
                "snapshot_files": int(month_files_count.get(month, 0)),
                "products_count": int(len(per_product)),
                "pairs_count": int(len(rows)),
                "qualified_pairs_count": int(len(qualified)),
                "latest_snapshot_dt": month_latest_dt.get(month).isoformat(timespec="seconds")
                if month_latest_dt.get(month)
                else "",
            }
        )

    now = datetime.now(ALMATY_TZ)
    latest_dt, latest_path = files[-1]
    payload = {
        "generated_at": now.isoformat(timespec="seconds"),
        "generated_at_human": now.strftime("%d.%m.%Y %H:%M"),
        "region": region,
        "filters": {
            "min_product_snapshots": int(min_product_snapshots),
            "min_avg_sellers": float(min_avg_sellers),
            "top_n": int(top_n),
            "months_limit": int(months_limit),
        },
        "files_total": int(len(files)),
        "latest_snapshot_file": latest_path.name,
        "latest_snapshot_dt": latest_dt.isoformat(timespec="seconds"),
        "months": months_meta,
        "selected_month": months_desc[0] if months_desc else "",
        "month_rows": month_rows,
    }
    return payload, _sort_rows(all_rows)


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=CSV_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k) for k in CSV_FIELDS})


def _split_message(text: str, max_len: int = 3200) -> list[str]:
    text = (text or "").strip()
    if not text:
        return []
    if len(text) <= max_len:
        return [text]
    parts: list[str] = []
    buff = ""
    for line in text.splitlines():
        line = line.rstrip()
        cand = line if not buff else f"{buff}\n{line}"
        if len(cand) <= max_len:
            buff = cand
            continue
        if buff:
            parts.append(buff)
        buff = line
        while len(buff) > max_len:
            parts.append(buff[:max_len])
            buff = buff[max_len:]
    if buff:
        parts.append(buff)
    return parts


def _send_telegram_message(token: str, chat_id: str, text: str) -> None:
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    for chunk in _split_message(text):
        resp = requests.post(
            url,
            data={"chat_id": chat_id, "text": chunk, "disable_web_page_preview": True},
            timeout=25,
        )
        resp.raise_for_status()


def _send_telegram_document(token: str, chat_id: str, file_path: Path, caption: str | None = None) -> None:
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    with open(file_path, "rb") as fh:
        data = {"chat_id": chat_id}
        if caption:
            data["caption"] = caption
        resp = requests.post(url, data=data, files={"document": (file_path.name, fh)}, timeout=40)
        resp.raise_for_status()


def _build_telegram_text(payload: dict[str, Any], telegram_top_n: int) -> str:
    month = str(payload.get("selected_month") or "")
    month_meta = next((m for m in (payload.get("months") or []) if m.get("month") == month), {})
    rows = (payload.get("month_rows") or {}).get(month, [])[: max(1, telegram_top_n)]

    lines = [
        "Kaspi Price Leader Live",
        f"Region: {payload.get('region') or 'n/a'}",
        f"Generated: {payload.get('generated_at_human') or payload.get('generated_at') or 'n/a'}",
        f"Month: {month or 'n/a'}",
        (
            "Snapshots / Products / Qualified: "
            f"{month_meta.get('snapshot_files', 0)} / {month_meta.get('products_count', 0)} / {month_meta.get('qualified_pairs_count', 0)}"
        ),
        "",
        f"Top {max(1, telegram_top_n)} sellers by best-price share:",
    ]
    if not rows:
        lines.append("No rows for selected month.")
    else:
        for idx, row in enumerate(rows, start=1):
            lines.append(
                (
                    f"{idx}. {row.get('seller') or '—'} | {row.get('win_share_pct', 0)}% "
                    f"({row.get('wins', 0)}/{row.get('product_snapshots', 0)}) | "
                    f"SKU {row.get('product_code') or '—'} | {row.get('product_name') or '—'}"
                )
            )
    return "\n".join(lines).strip()


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build live monthly best-price leaderboard from kaspi_prices reports")
    parser.add_argument(
        "--results-dir",
        default="RESULTS",
        help="Directory with kaspi_prices_*.xlsx files (default: %(default)s)",
    )
    parser.add_argument("--region", default=os.environ.get("PRICE_LEADERS_REGION", "Алматы"), help="Region filter")
    parser.add_argument("--months-limit", type=int, default=int(os.environ.get("PRICE_LEADERS_MONTHS_LIMIT", "3")))
    parser.add_argument("--top-n", type=int, default=int(os.environ.get("PRICE_LEADERS_TOP_N", "60")))
    parser.add_argument(
        "--min-product-snapshots",
        type=int,
        default=int(os.environ.get("PRICE_LEADERS_MIN_PRODUCT_SNAPSHOTS", "8")),
    )
    parser.add_argument(
        "--min-avg-sellers",
        type=float,
        default=float(os.environ.get("PRICE_LEADERS_MIN_AVG_SELLERS", "2.0")),
    )
    parser.add_argument("--send-telegram", action="store_true", help="Send summary and CSV to Telegram")
    parser.add_argument("--telegram-top-n", type=int, default=int(os.environ.get("PRICE_LEADERS_TELEGRAM_TOP_N", "15")))
    parser.add_argument("--telegram-bot-token", default="", help="Override TELEGRAM_BOT_TOKEN")
    parser.add_argument("--telegram-chat-id", default="", help="Override TELEGRAM_CHAT_ID")
    return parser.parse_args()


def main() -> int:
    args = _parse_args()

    base_dir = Path(__file__).resolve().parent
    results_dir = Path(args.results_dir)
    if not results_dir.is_absolute():
        results_dir = (base_dir / results_dir).resolve()
    results_dir.mkdir(parents=True, exist_ok=True)

    payload, all_rows = _build_summary(
        results_dir=results_dir,
        region=str(args.region).strip(),
        months_limit=max(1, int(args.months_limit)),
        top_n=max(1, int(args.top_n)),
        min_product_snapshots=max(1, int(args.min_product_snapshots)),
        min_avg_sellers=max(0.0, float(args.min_avg_sellers)),
    )

    timestamp = datetime.now(ALMATY_TZ).strftime("%Y-%m-%d_%H-%M-%S")
    region_slug = _slugify_region(str(args.region))
    details_name = f"kaspi_price_leaders_{region_slug}_{timestamp}_details.csv"
    summary_name = f"kaspi_price_leaders_{region_slug}_{timestamp}_summary.json"
    details_path = results_dir / details_name
    summary_path = results_dir / summary_name

    _write_csv(details_path, all_rows)
    payload["details_file"] = details_name
    payload["summary_file"] = summary_name

    with open(summary_path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)

    live_json = results_dir / LIVE_JSON_NAME
    live_csv = results_dir / LIVE_CSV_NAME
    shutil.copyfile(summary_path, live_json)
    shutil.copyfile(details_path, live_csv)
    payload["details_file_latest"] = live_csv.name
    payload["summary_file_latest"] = live_json.name

    with open(live_json, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)

    print(f"summary: {summary_path}")
    print(f"details: {details_path}")
    print(f"live_json: {live_json}")
    print(f"live_csv: {live_csv}")

    if args.send_telegram:
        token = str(args.telegram_bot_token or os.environ.get("TELEGRAM_BOT_TOKEN", "")).strip()
        chat_id = str(args.telegram_chat_id or os.environ.get("TELEGRAM_CHAT_ID", "")).strip()
        if not token or not chat_id:
            raise RuntimeError("Telegram send requested, but TELEGRAM_BOT_TOKEN/TELEGRAM_CHAT_ID is missing")
        text = _build_telegram_text(payload, telegram_top_n=max(1, int(args.telegram_top_n)))
        _send_telegram_message(token, chat_id, text)
        _send_telegram_document(token, chat_id, live_csv, caption="Kaspi price leaders live (CSV)")
        print("telegram: sent")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
