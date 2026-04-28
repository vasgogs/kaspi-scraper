#!/usr/bin/env python3
"""Рассылка отчёта миссии по аптекам с выборкой их строк."""

import argparse
import asyncio
import html
import json
import os
import re
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Sequence

import pandas as pd
import requests
from telethon import TelegramClient
from telethon.tl.functions.contacts import ImportContactsRequest
from telethon.tl.types import InputPhoneContact
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font
from PIL import Image
from Scraper_Kaspi import render_mission_image, _filter_alert_mission_rows
from zoneinfo import ZoneInfo

BASE_DIR = Path(__file__).resolve().parent
RESULTS_DIR = BASE_DIR / "RESULTS"
STATE_DIR = BASE_DIR / "state"
CAMPAIGN_SUMMARY_STATE_PATH = STATE_DIR / "campaign_daily_summary_state.json"
ALMATY_TZ = ZoneInfo("Asia/Almaty")
BOT_PREVIEW_MAX_DIM = 2400


def parse_list(raw: str | None) -> list[str]:
    if not raw:
        return []
    parts: list[str] = []
    for chunk in raw.replace(";", ",").split(","):
        value = chunk.strip()
        if value and value != "...":
            parts.append(value)
    return parts


async def ensure_phone_contacts(client: TelegramClient, targets: Sequence[str]) -> dict[str, str]:
    """Добавляем телефоны в контакты и возвращаем карту phone -> peer id."""
    phones = [t for t in targets if t.startswith("+")]
    if not phones:
        return {}
    contacts = [
        InputPhoneContact(client_id=i, phone=phone, first_name=f"Broadcast{i}", last_name="")
        for i, phone in enumerate(phones)
    ]
    result = await client(ImportContactsRequest(contacts))
    mapping: dict[str, str] = {}
    for imported, phone in zip(result.imported, phones):
        if imported.user_id:
            mapping[phone] = imported.user_id
    return mapping


async def resolve_target_peer(client: TelegramClient, phone_map: dict[str, str], target: str):
    peer = phone_map.get(target) if target.startswith("+") else target
    if isinstance(peer, str) and re.fullmatch(r"-?\d+", peer.strip()):
        numeric_peer = int(peer.strip())
        try:
            peer = await client.get_entity(numeric_peer)
        except Exception:
            peer = numeric_peer
    return peer


def latest_mission_file(prefix: str) -> Path | None:
    results = BASE_DIR / "RESULTS"
    if not results.exists():
        return None
    files = sorted(results.glob(f"{prefix}_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def normalize(text: str) -> str:
    return " ".join((text or "").strip().lower().split())


def load_env_file():
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("'\"")
        if key and key not in os.environ:
            os.environ[key] = value


def slugify_region(value: str) -> str:
    region = normalize(value)
    mapping = {
        "алматы": "almaty",
        "астана": "astana",
        "шымкент": "shymkent",
    }
    if region in mapping:
        return mapping[region]
    return "".join(ch if ch.isalnum() else "_" for ch in region).strip("_") or "region"


def first_non_empty_targets(*raw_values: str | None) -> list[str]:
    for raw in raw_values:
        parsed = parse_list(raw)
        if parsed:
            return parsed
    return []


def latest_campaign_summary_file() -> Path | None:
    if not RESULTS_DIR.exists():
        return None
    files = sorted(
        RESULTS_DIR.glob("mission_campaign_quality_*_summary.json"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return files[0] if files else None


def campaign_related_path(summary_path: Path, suffix: str) -> Path:
    stem = summary_path.name[: -len("_summary.json")]
    return summary_path.with_name(f"{stem}_{suffix}")


def format_pct(value) -> str:
    try:
        return f"{float(value):.1f}%"
    except Exception:
        return "—"


def format_hours(value) -> str:
    try:
        num = float(value)
    except Exception:
        return "—"
    digits = 0 if num >= 100 else 1
    return f"{num:.{digits}f}".replace(".", ",") + " ч"


def format_snapshot_label(value: str) -> str:
    text = str(value or "").strip()
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2}) (\d{2}):(\d{2})", text)
    if not match:
        return text or "—"
    year, month, day, hour, minute = match.groups()
    return f"{day}.{month}.{year} {hour}:{minute}"


def is_numeric_chat_target(target: str) -> bool:
    return bool(re.fullmatch(r"-?\d+", str(target or "").strip()))


def bot_token() -> str:
    return str(os.environ.get("TELEGRAM_BOT_TOKEN") or "").strip()


def telegram_bot_post(method: str, data: dict, files=None, timeout: int = 60):
    token = bot_token()
    if not token:
        raise RuntimeError("TELEGRAM_BOT_TOKEN not set")
    url = f"https://api.telegram.org/bot{token}/{method}"
    return requests.post(url, data=data, files=files, timeout=timeout)


def resize_photo_for_bot(source_path: Path, output_path: Path) -> Path:
    with Image.open(source_path) as img:
        width, height = img.size
        max_dim = max(width, height)
        if max_dim <= BOT_PREVIEW_MAX_DIM:
            if source_path == output_path:
                return source_path
            img.save(output_path, format="PNG", optimize=True)
            return output_path
        scale = BOT_PREVIEW_MAX_DIM / float(max_dim)
        resized = img.resize((max(1, int(width * scale)), max(1, int(height * scale))), Image.LANCZOS)
        resized.save(output_path, format="PNG", optimize=True)
    return output_path


def send_bot_document(chat_id: str, file_path: Path, caption: str | None = None):
    with open(file_path, "rb") as fh:
        files = {"document": fh}
        data = {"chat_id": str(chat_id)}
        if caption:
            data["caption"] = caption
        resp = telegram_bot_post("sendDocument", data=data, files=files, timeout=90)
    if resp.status_code != 200:
        raise RuntimeError(f"Bot sendDocument failed: {resp.status_code} {resp.text}")


def send_bot_photo(chat_id: str, file_path: Path, caption: str | None = None):
    with open(file_path, "rb") as fh:
        files = {"photo": fh}
        data = {"chat_id": str(chat_id)}
        if caption:
            data["caption"] = caption
        resp = telegram_bot_post("sendPhoto", data=data, files=files, timeout=90)
    if resp.status_code != 200:
        raise RuntimeError(f"Bot sendPhoto failed: {resp.status_code} {resp.text}")


def _load_json(path: Path) -> dict | None:
    try:
        with open(path, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
        return payload if isinstance(payload, dict) else None
    except Exception:
        return None


def _campaign_summary_state() -> dict:
    try:
        return json.loads(CAMPAIGN_SUMMARY_STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _write_campaign_summary_state(payload: dict):
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    CAMPAIGN_SUMMARY_STATE_PATH.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def build_campaign_daily_summary_text(summary_path: Path) -> str | None:
    summary = _load_json(summary_path)
    if not summary:
        return None
    per_seller_path = campaign_related_path(summary_path, "per_seller.csv")
    per_region_path = campaign_related_path(summary_path, "per_region.csv")
    per_row_path = campaign_related_path(summary_path, "per_row.csv")
    if not per_seller_path.exists() or not per_region_path.exists() or not per_row_path.exists():
        return None

    sellers = pd.read_csv(per_seller_path, encoding="utf-8-sig")
    regions = pd.read_csv(per_region_path, encoding="utf-8-sig")
    rows = pd.read_csv(per_row_path, encoding="utf-8-sig")
    if sellers.empty or regions.empty or rows.empty:
        return None

    def best_price_metric(frame: pd.DataFrame, row) -> float | None:
        if "best_price_pct_market" in frame.columns:
            value = row.get("best_price_pct_market")
            if pd.notna(value):
                return float(value)
        value = row.get("best_price_pct_observed")
        if pd.notna(value):
            return float(value)
        return None

    def best_price_text(frame: pd.DataFrame, row) -> str:
        value = best_price_metric(frame, row)
        return format_pct(value)

    sellers["__best_price_metric"] = 0.0
    if "best_price_pct_observed" in sellers.columns:
        sellers["__best_price_metric"] = pd.to_numeric(sellers["best_price_pct_observed"], errors="coerce").fillna(0.0)
    if "best_price_pct_market" in sellers.columns:
        sellers["__best_price_metric"] = pd.to_numeric(sellers["best_price_pct_market"], errors="coerce").fillna(
            sellers["__best_price_metric"]
        )
    sellers = sellers.sort_values(
        by=["correct_price_pct_observed", "__best_price_metric", "availability_pct_observed"],
        ascending=[False, False, False],
    )
    regions["__rank"] = regions["region"].map({"Алматы": 0, "Астана": 1, "Шымкент": 2}).fillna(9)
    regions = regions.sort_values(by=["__rank", "region"])
    rows = rows.sort_values(
        by=["problem_pct_observed", "correct_price_pct_observed", "availability_pct_observed"],
        ascending=[False, True, True],
    )

    top_sellers = []
    for _, row in sellers.head(5).iterrows():
        top_sellers.append(
            f"• {row['seller']}: SKU availability {format_pct(row['availability_pct_observed'])} • pricing discipline {format_pct(row['correct_price_pct_observed'])} • best price {best_price_text(sellers, row)}"
        )

    top_regions = []
    for _, row in regions.iterrows():
        top_regions.append(
            f"• {row['region']}: SKU availability {format_pct(row['availability_pct_observed'])} • pricing discipline {format_pct(row['correct_price_pct_observed'])} • best price {best_price_text(regions, row)}"
        )

    risk_rows = []
    for _, row in rows.head(5).iterrows():
        title = html.unescape(str(row.get("input_product") or row.get("product") or "").strip() or "SKU")
        risk_rows.append(
            f"• {row['seller']} / {title}: проблема {format_pct(row['problem_pct_observed'])} • pricing discipline {format_pct(row['correct_price_pct_observed'])} • best price {best_price_text(rows, row)}"
        )

    return "\n".join(
        [
            f"Campaign summary ({datetime.now(ALMATY_TZ):%d.%m})",
            f"Окно данных: {format_hours(summary.get('observed_window_hours'))} • обновлено {format_snapshot_label(summary.get('data_coverage_end', ''))}",
            f"SKU availability {format_pct(summary.get('availability_pct_observed'))} • pricing discipline {format_pct(summary.get('correct_price_pct_observed'))} • best price {format_pct(summary.get('best_price_pct_market', summary.get('best_price_pct_observed')))} • проблема {format_pct(summary.get('problem_pct_observed'))}",
            "",
            "Аптеки:",
            *(top_sellers or ["• Нет данных"]),
            "",
            "Города:",
            *(top_regions or ["• Нет данных"]),
            "",
            "Риски:",
            *(risk_rows or ["• Нет данных"]),
        ]
    )


def should_send_campaign_daily_summary(summary_path: Path) -> bool:
    today_key = datetime.now(ALMATY_TZ).date().isoformat()
    state = _campaign_summary_state()
    return not (
        state.get("date") == today_key and state.get("source_file") == summary_path.name
    )


def filter_rows(df: pd.DataFrame, seller_keywords: list[str], regions: list[str] | None = None) -> pd.DataFrame:
    seller_mask = df["seller"].fillna("").apply(lambda x: any(k in normalize(x) for k in seller_keywords))
    region_mask = df["region"].fillna("").apply(lambda x: normalize(x) in {normalize(r) for r in regions}) if regions else True
    if isinstance(region_mask, bool):
        mask = seller_mask
    else:
        mask = seller_mask & region_mask
    return df[mask].copy()


def write_styled_slice(df: pd.DataFrame, path: Path):
    """Записываем с простым цветовым кодом статуса."""
    wb = Workbook()
    ws = wb.active
    headers = df.columns.tolist()
    ws.append(headers)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    ok_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    bad_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    for _, row in df.iterrows():
        values = [row.get(col, "") for col in headers]
        ws.append(values)
        status_text = " ".join(str(row.get("status") or "").replace("\u00a0", " ").lower().split())
        if status_text.startswith("ok") or status_text.startswith("ок"):
            fill = ok_fill
        elif status_text.startswith("внимание") or "не самое выгодное" in status_text or "не лучшее" in status_text:
            fill = warn_fill
        elif status_text.startswith("проблема") or status_text.startswith("дороже") or status_text.startswith("дешевле") or "нет" in status_text or "отсутств" in status_text or "ошибка" in status_text:
            fill = bad_fill
        else:
            fill = warn_fill
        for cell in ws[ws.max_row]:
            cell.fill = fill

    # немного ширины
    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(40, len(str(header)) + 4))

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


async def send_file(
    client: TelegramClient,
    targets: list[str],
    df: pd.DataFrame,
    caption: str,
    filename: str,
    *,
    send_document: bool = True,
    send_preview: bool = True,
):
    phone_map = await ensure_phone_contacts(client, targets)
    with tempfile.TemporaryDirectory() as tmpdir:
        preview_df = pd.DataFrame()
        try:
            preview_df = _filter_alert_mission_rows(df)
        except Exception as exc:
            print(f"⚠️ Could not filter alert rows for {filename}: {exc}")
            preview_df = pd.DataFrame()

        if preview_df.empty:
            print(f"ℹ️ {filename}: нет проблемных строк, пропускаем отправку файла и скрина.")
            return

        file_path: Path | None = None
        if send_document:
            file_path = Path(tmpdir) / filename
            write_styled_slice(df, file_path)

        preview_path: Path | None = None
        if send_preview:
            try:
                preview_path = Path(tmpdir) / (filename.replace(".xlsx", "_preview.png"))
                render_mission_image(preview_df, preview_path)
            except Exception as exc:
                print(f"⚠️ Could not render preview for {filename}: {exc}")
                preview_path = None

        preview_photo_path: Path | None = None
        if preview_path and preview_path.exists():
            try:
                preview_photo_path = resize_photo_for_bot(
                    preview_path,
                    Path(tmpdir) / (preview_path.stem + "_bot.png"),
                )
            except Exception as exc:
                print(f"⚠️ Could not resize preview for bot delivery {filename}: {exc}")
                preview_photo_path = preview_path

        for target in targets:
            if is_numeric_chat_target(target) and bot_token():
                if send_document and file_path is not None:
                    try:
                        send_bot_document(target, file_path, caption=caption)
                        print(f"✅ Sent to {target} via bot ({filename})")
                    except Exception as exc:
                        print(f"❌ Failed to send to {target} via bot: {exc}")
                if preview_photo_path and preview_photo_path.exists():
                    try:
                        send_bot_photo(target, preview_photo_path, caption=f"{caption} — red alerts only")
                        print(f"✅ Sent preview to {target} via bot ({preview_photo_path.name})")
                    except Exception as exc:
                        print(f"⚠️ Failed to send preview to {target} via bot: {exc}")
                continue

            peer = await resolve_target_peer(client, phone_map, target)
            if send_document and file_path is not None:
                try:
                    await client.send_file(peer, file_path, caption=caption)
                    print(f"✅ Sent to {target} ({filename})")
                except Exception as exc:
                    print(f"❌ Failed to send to {target}: {exc}")
            if preview_path and preview_path.exists():
                try:
                    await client.send_file(
                        peer,
                        preview_path,
                        caption=f"{caption} — red alerts only",
                        force_document=bool(send_document),
                    )
                    print(f"✅ Sent preview to {target} ({preview_path.name})")
                except Exception as exc:
                    print(f"⚠️ Failed to send preview to {target}: {exc}")


async def send_text_message(client: TelegramClient, targets: list[str], text: str) -> bool:
    if not targets or not text.strip():
        return False
    phone_map = await ensure_phone_contacts(client, targets)
    sent = False
    for target in targets:
        if is_numeric_chat_target(target) and bot_token():
            try:
                resp = telegram_bot_post("sendMessage", data={"chat_id": str(target), "text": text}, timeout=60)
                if resp.status_code != 200:
                    raise RuntimeError(f"{resp.status_code} {resp.text}")
                print(f"✅ Sent message to {target} via bot")
                sent = True
            except Exception as exc:
                print(f"❌ Failed to send message to {target} via bot: {exc}")
            continue
        peer = await resolve_target_peer(client, phone_map, target)
        try:
            await client.send_message(peer, text)
            print(f"✅ Sent message to {target}")
            sent = True
        except Exception as exc:
            print(f"❌ Failed to send message to {target}: {exc}")
    return sent


async def main():
    load_env_file()
    parser = argparse.ArgumentParser(description="Broadcast mission report slices by pharmacy")
    parser.add_argument("--file", type=Path, help="Путь к файлу миссии (.xlsx). Если не задан, возьмём последний.")
    parser.add_argument(
        "--prefix",
        default=os.environ.get("MISSION_FILE_PREFIX", "mission_april"),
        help="Префикс файла миссии (по умолчанию mission_april)",
    )
    parser.add_argument("--adoy-name", default="Аптека от А до Я", help="Название для фильтра Аптека от А до Я")
    parser.add_argument("--msp-name", default="Аптека MSP", help="Название для фильтра Аптека MSP")
    parser.add_argument("--msp-regions", default="Алматы,Астана,Шымкент", help="Регионы для MSP через запятую")
    parser.add_argument("--pharmakom-name", default="ФАРМАКОМ", help="Название для фильтра ФАРМАКОМ")
    parser.add_argument("--pharmakom-regions", default="Алматы,Астана,Шымкент", help="Регионы для ФАРМАКОМ через запятую")
    parser.add_argument("--tur-name", default="Аптека от А до Я", help="Ключ для фильтра Туровской аптеки (Аптека от А до Я)")
    parser.add_argument("--adoy-targets", help="Цели для Аптека от А до Я (чат id / @ / телефон через запятую)")
    parser.add_argument("--msp-targets", help="Цели для Аптека MSP (чат id / @ / телефон через запятую)")
    parser.add_argument("--pharmakom-targets", help="Цели для ФАРМАКОМ (чат id / @ / телефон через запятую)")
    parser.add_argument("--tur-targets", help="Цели для Туровской (чат id / @ / телефон через запятую)")
    parser.add_argument("--campaign-summary-targets", help="Цели для daily summary по кампании (чат id / @ / телефон через запятую)")
    parser.add_argument("--api-id", type=int, default=int(os.environ.get("TELETHON_API_ID", "0")), help="api_id из my.telegram.org")
    parser.add_argument("--api-hash", default=os.environ.get("TELETHON_API_HASH", ""), help="api_hash из my.telegram.org")
    parser.add_argument("--phone", default=os.environ.get("TELETHON_PHONE", ""), help="Номер аккаунта для логина (+7...)")
    args = parser.parse_args()

    targets_adoy = first_non_empty_targets(args.adoy_targets, os.environ.get("ADOY_TARGETS"))
    targets_msp = first_non_empty_targets(args.msp_targets, os.environ.get("MSP_TARGETS"))
    targets_pharmakom = first_non_empty_targets(
        args.pharmakom_targets,
        os.environ.get("PHARMAKOM_TARGETS"),
        os.environ.get("BROADCAST_CHAT_IDS"),
        os.environ.get("MISSION_APRIL_CHAT_ID"),
        os.environ.get("MISSION_FEBRUARY_CHAT_ID"),
        os.environ.get("TELEGRAM_CHAT_ID"),
    )
    targets_tur = first_non_empty_targets(args.tur_targets, os.environ.get("TUR_TARGETS"))
    targets_campaign_summary = first_non_empty_targets(
        args.campaign_summary_targets,
        os.environ.get("CAMPAIGN_SUMMARY_TARGETS"),
        os.environ.get("BROADCAST_CHAT_IDS"),
        os.environ.get("MISSION_APRIL_CHAT_ID"),
        os.environ.get("MISSION_FEBRUARY_CHAT_ID"),
        os.environ.get("TELEGRAM_CHAT_ID"),
    )
    if not targets_adoy and not targets_msp and not targets_pharmakom and not targets_tur:
        raise SystemExit("Нужно указать получателей (--adoy-targets/--msp-targets/--pharmakom-targets/--tur-targets или env ADOY_TARGETS/MSP_TARGETS/PHARMAKOM_TARGETS/TUR_TARGETS)")
    if not args.api_id or not args.api_hash or not args.phone:
        raise SystemExit("Нужны api_id, api_hash, phone")

    mission_path = args.file
    if not mission_path:
        mission_path = latest_mission_file(args.prefix)
    if not mission_path or not mission_path.exists():
        raise SystemExit("Не найден файл миссии. Укажи --file или убедись, что есть результаты.")

    df = pd.read_excel(mission_path)
    required_cols = {"seller", "region"}
    if not required_cols.issubset(df.columns):
        raise SystemExit(f"В файле нет нужных колонок: {required_cols}")

    # Фильтры
    msp_regions = [r.strip() for r in args.msp_regions.split(",") if r.strip()]
    pharmakom_regions = [r.strip() for r in args.pharmakom_regions.split(",") if r.strip()]
    adoy_df = filter_rows(df, [normalize(args.adoy_name)], regions=None)
    msp_region_frames = {
        region: filter_rows(df, [normalize(args.msp_name)], regions=[region])
        for region in msp_regions
    }
    msp_df = pd.concat(
        [frame for frame in msp_region_frames.values() if not frame.empty],
        ignore_index=True,
    ) if any(not frame.empty for frame in msp_region_frames.values()) else df.iloc[0:0].copy()
    pharmakom_region_frames = {
        region: filter_rows(df, [normalize(args.pharmakom_name)], regions=[region])
        for region in pharmakom_regions
    }
    pharmakom_df = pd.concat(
        [frame for frame in pharmakom_region_frames.values() if not frame.empty],
        ignore_index=True,
    ) if any(not frame.empty for frame in pharmakom_region_frames.values()) else df.iloc[0:0].copy()
    tur_df = filter_rows(df, [normalize(args.tur_name)], regions=None)

    tasks = []
    async with TelegramClient(Path(os.environ.get("TELETHON_SESSION", "userbot_session")), args.api_id, args.api_hash) as client:
        await client.start(phone=args.phone)
        if targets_adoy and not adoy_df.empty:
            caption = f"Отчёт миссии для {args.adoy_name} — {len(adoy_df)} строк(и)"
            tasks.append(send_file(client, targets_adoy, adoy_df, caption, "mission_adoy.xlsx"))
        if targets_msp and not msp_df.empty:
            for region in msp_regions:
                region_df = msp_region_frames.get(region)
                if region_df is None or region_df.empty:
                    print(f"⚠️ Нет данных для {args.msp_name} {region}, отдельную отбивку не отправляем.")
                    continue
                caption = f"Отчёт миссии для {args.msp_name} {region} — {len(region_df)} строк(и)"
                filename = f"mission_msp_{slugify_region(region)}.xlsx"
                tasks.append(send_file(client, targets_msp, region_df, caption, filename))
        if targets_pharmakom and not pharmakom_df.empty:
            caption = f"Общий отчёт миссии для {args.pharmakom_name} — {len(pharmakom_df)} строк(и)"
            tasks.append(
                send_file(
                    client,
                    targets_pharmakom,
                    pharmakom_df,
                    caption,
                    "mission_pharmakom.xlsx",
                    send_document=True,
                    send_preview=False,
                )
            )
            for region in pharmakom_regions:
                region_df = pharmakom_region_frames.get(region)
                if region_df is None or region_df.empty:
                    print(f"⚠️ Нет данных для {args.pharmakom_name} {region}, отдельную отбивку не отправляем.")
                    continue
                caption = f"Отчёт миссии для {args.pharmakom_name} {region} — {len(region_df)} строк(и)"
                filename = f"mission_pharmakom_{slugify_region(region)}.xlsx"
                tasks.append(
                    send_file(
                        client,
                        targets_pharmakom,
                        region_df,
                        caption,
                        filename,
                        send_document=False,
                    )
                )
        if targets_tur and not tur_df.empty:
            caption = f"Отчёт миссии (Туровская) — {len(tur_df)} строк(и)"
            tasks.append(send_file(client, targets_tur, tur_df, caption, "mission_tur.xlsx"))
        if tasks:
            await asyncio.gather(*tasks)
        else:
            print("⚠️ Нет данных для отправки (пустые выборки).")

        campaign_summary_file = latest_campaign_summary_file()
        if targets_campaign_summary and campaign_summary_file and should_send_campaign_daily_summary(campaign_summary_file):
            try:
                summary_text = build_campaign_daily_summary_text(campaign_summary_file)
                if summary_text:
                    sent = await send_text_message(client, targets_campaign_summary, summary_text)
                    if sent:
                        _write_campaign_summary_state(
                            {
                                "date": datetime.now(ALMATY_TZ).date().isoformat(),
                                "source_file": campaign_summary_file.name,
                                "sent_at": datetime.now(ALMATY_TZ).isoformat(timespec="seconds"),
                            }
                        )
                else:
                    print("⚠️ Campaign daily summary skipped: could not build text.")
            except Exception as exc:
                print(f"⚠️ Campaign daily summary failed: {exc}")


if __name__ == "__main__":
    asyncio.run(main())
