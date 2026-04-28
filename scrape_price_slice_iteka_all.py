#!/usr/bin/env python3
from __future__ import annotations

import argparse
import math
import shutil
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from scrape_price_slice_kaspi_iteka import (
    DATA_START_ROW,
    extract_link,
    extract_iteka_total,
    find_column_by_header,
    iteka_page_url,
    normalize_iteka_url,
    parse_iteka_cards,
    request_with_retries,
    unavailable_note_from_text,
    HEADERS,
)


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE_DIR / "Срез цен_1_scraped_kaspi_iteka.xlsx"
SOURCE_SHEET = "рабочий"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape all i-teka pharmacy prices from the workbook.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--sheet", default=SOURCE_SHEET)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--workers", type=int, default=10)
    parser.add_argument("--max-pages", type=int, default=50)
    return parser.parse_args()


def base_row(ws, row_num: int, scraped_at: str) -> list[Any]:
    return [
        scraped_at,
        row_num,
        ws.cell(row=row_num, column=1).value,
        ws.cell(row=row_num, column=2).value,
        ws.cell(row=row_num, column=3).value,
        ws.cell(row=row_num, column=4).value,
        ws.cell(row=row_num, column=5).value,
        ws.cell(row=row_num, column=7).value,
        ws.cell(row=row_num, column=8).value,
        ws.cell(row=row_num, column=9).value,
        ws.cell(row=row_num, column=10).value,
    ]


def collect_source_rows(ws, limit: int) -> list[dict[str, Any]]:
    iteka_col = find_column_by_header(ws, "i-teka")
    stop_row = ws.max_row if limit <= 0 else min(ws.max_row, DATA_START_ROW + limit - 1)
    rows: list[dict[str, Any]] = []
    for row_num in range(DATA_START_ROW, stop_row + 1):
        url = normalize_iteka_url(extract_link(ws.cell(row=row_num, column=iteka_col)))
        rows.append({"row_num": row_num, "iteka_url": url})
    return rows


def fetch_all_iteka_prices(url: str, max_pages: int) -> dict[str, Any]:
    if not url:
        return {"url": url, "offers": [], "total_found": None, "pages_scraped": 0, "note": "i-teka URL missing"}

    seen: set[str] = set()
    offers: list[dict[str, Any]] = []
    total_found: int | None = None
    pages_scraped = 0
    note = ""
    first_page_text = ""

    for page in range(1, max(1, max_pages) + 1):
        page_url = iteka_page_url(url, page)
        try:
            response = request_with_retries("GET", page_url, headers=HEADERS)
        except Exception as exc:  # noqa: BLE001
            note = f"i-teka request failed: {exc}"
            break

        text = response.text
        if not first_page_text:
            first_page_text = text
        pages_scraped += 1
        if total_found is None:
            total_found = extract_iteka_total(text)

        cards = parse_iteka_cards(text, page_url)
        if not cards:
            break

        for card in cards:
            key = card.get("apteka_id") or f"{card.get('pharmacy')}|{card.get('address')}"
            if key in seen:
                continue
            seen.add(key)
            item = dict(card)
            item["site_order"] = len(offers) + 1
            item["source_page"] = page
            offers.append(item)

        if total_found:
            total_pages = max(1, math.ceil(total_found / 20))
            if page >= total_pages:
                break

    if not offers and not note:
        note = unavailable_note_from_text(first_page_text) or "No i-teka prices found"
    return {
        "url": url,
        "offers": offers,
        "total_found": total_found,
        "pages_scraped": pages_scraped,
        "note": note,
    }


def run_parallel(urls: list[str], workers: int, max_pages: int) -> dict[str, dict[str, Any]]:
    unique_urls = sorted({url for url in urls if url})
    results: dict[str, dict[str, Any]] = {}
    print(f"i-teka all: unique URLs {len(unique_urls)}", flush=True)
    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        future_map = {executor.submit(fetch_all_iteka_prices, url, max_pages): url for url in unique_urls}
        total = len(future_map)
        for idx, future in enumerate(as_completed(future_map), start=1):
            url = future_map[future]
            try:
                results[url] = future.result()
            except Exception as exc:  # noqa: BLE001
                results[url] = {"url": url, "offers": [], "total_found": None, "pages_scraped": 0, "note": str(exc)}
            if idx == 1 or idx % 20 == 0 or idx == total:
                print(f"i-teka all: {idx}/{total}", flush=True)
    return results


def delete_sheet_if_exists(wb, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]


def style_sheet(ws) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, min(ws.max_column, 55) + 1):
        letter = get_column_letter(column)
        max_len = 10
        for row in range(1, min(ws.max_row, 120) + 1):
            value = ws.cell(row=row, column=column).value
            if value is not None:
                max_len = max(max_len, min(65, len(str(value))))
        ws.column_dimensions[letter].width = max_len + 2


def write_outputs(wb, source_ws, source_rows: list[dict[str, Any]], results: dict[str, dict[str, Any]]) -> None:
    scraped_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    common_headers = [
        "scraped_at",
        "source_row",
        "Address Market",
        "Segment",
        "BrandEng",
        "BrandRus",
        "Corporation",
        "Fcc",
        "Fcc variant",
        "SKUEng",
        "SKUEng 2",
    ]
    summary_headers = common_headers + [
        "iteka_url",
        "total_found",
        "prices_scraped",
        "pages_scraped",
        "min_price",
        "min_pharmacy",
        "max_price",
        "max_pharmacy",
        "crown_count",
        "note",
    ]
    detail_headers = common_headers + [
        "site_order",
        "price_rank",
        "pharmacy",
        "price",
        "is_crown",
        "address",
        "updated",
        "tags",
        "apteka_id",
        "source_page",
        "pharmacy_url",
        "iteka_url",
    ]

    delete_sheet_if_exists(wb, "iteka all summary")
    delete_sheet_if_exists(wb, "iteka all prices")
    summary_ws = wb.create_sheet("iteka all summary")
    detail_ws = wb.create_sheet("iteka all prices")
    summary_ws.append(summary_headers)
    detail_ws.append(detail_headers)

    for item in source_rows:
        row_num = item["row_num"]
        url = item["iteka_url"]
        result = results.get(url) if url else None
        offers = list((result or {}).get("offers") or [])
        priced = [offer for offer in offers if offer.get("price_kzt") is not None]
        price_sorted = sorted(priced, key=lambda offer: (int(offer.get("price_kzt") or 10**18), offer.get("pharmacy") or ""))
        price_rank_by_key: dict[str, int] = {}
        for rank, offer in enumerate(price_sorted, start=1):
            key = offer.get("apteka_id") or f"{offer.get('pharmacy')}|{offer.get('address')}"
            price_rank_by_key[key] = rank

        min_offer = price_sorted[0] if price_sorted else {}
        max_offer = price_sorted[-1] if price_sorted else {}
        base = base_row(source_ws, row_num, scraped_at)
        summary_ws.append(
            base
            + [
                url,
                (result or {}).get("total_found") if result else None,
                len(priced),
                (result or {}).get("pages_scraped") if result else 0,
                min_offer.get("price_kzt"),
                min_offer.get("pharmacy"),
                max_offer.get("price_kzt"),
                max_offer.get("pharmacy"),
                sum(1 for offer in offers if offer.get("is_crown")),
                (result or {}).get("note") if result else ("i-teka URL missing" if not url else "i-teka result missing"),
            ]
        )

        for offer in offers:
            if offer.get("price_kzt") is None:
                continue
            key = offer.get("apteka_id") or f"{offer.get('pharmacy')}|{offer.get('address')}"
            detail_ws.append(
                base
                + [
                    offer.get("site_order"),
                    price_rank_by_key.get(key),
                    offer.get("pharmacy"),
                    offer.get("price_kzt"),
                    bool(offer.get("is_crown")),
                    offer.get("address"),
                    offer.get("updated"),
                    offer.get("tags"),
                    offer.get("apteka_id"),
                    offer.get("source_page"),
                    offer.get("pharmacy_url"),
                    url,
                ]
            )

    style_sheet(summary_ws)
    style_sheet(detail_ws)


def main() -> int:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(input_path)
    output_path = args.output.expanduser().resolve() if args.output else input_path
    if output_path != input_path:
        shutil.copy2(input_path, output_path)

    wb = load_workbook(output_path)
    if args.sheet not in wb.sheetnames:
        raise KeyError(f"Sheet {args.sheet!r} not found")
    source_ws = wb[args.sheet]
    source_rows = collect_source_rows(source_ws, args.limit)
    print(f"Rows to process: {len(source_rows)}", flush=True)
    results = run_parallel([row["iteka_url"] for row in source_rows], args.workers, args.max_pages)
    write_outputs(wb, source_ws, source_rows, results)
    wb.save(output_path)
    print(f"saved {output_path}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
