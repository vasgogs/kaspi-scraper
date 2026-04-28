#!/usr/bin/env python3
from __future__ import annotations

import argparse
import html
import json
import math
import re
import shutil
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlencode, urljoin, urlparse, urlunparse

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE_DIR / "Срез цен_1.xlsx"
CITY_ID = "750000000"
KASPI_ZONE_ID = "Magnum_ZONE1"
DATA_START_ROW = 5
HEADER_ROW = 4
PARENT_HEADER_ROW = 3
MAX_RETRIES = 3
TIMEOUT_SECONDS = 35
USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36"
)
HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept-Language": "ru,en;q=0.8",
}


@dataclass
class WorkRow:
    row_num: int
    address_market: Any
    segment: Any
    brand_eng: Any
    brand_rus: Any
    corporation: Any
    fcc: Any
    fcc_variant: Any
    sku_eng: Any
    sku_eng_2: Any
    kaspi_url: str
    iteka_url: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape Kaspi and i-teka links from Срез цен_1.xlsx.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--sheet", default="рабочий")
    parser.add_argument("--limit", type=int, default=0, help="Process first N rows for a smoke test.")
    parser.add_argument("--workers", type=int, default=18)
    return parser.parse_args()


def clean_text(value: Any) -> str:
    text = html.unescape(str(value or "")).replace("\xa0", " ")
    text = re.sub(r"<[^>]+>", " ", text)
    text = text.replace("&thinsp;", " ")
    return " ".join(text.split())


def unavailable_note_from_text(text: str) -> str:
    normalized = clean_text(text).lower()
    if not normalized:
        return ""
    phrases = [
        "нет в наличии",
        "нет наличии",
        "нет в продаже",
        "товар закончился",
        "товара нет",
        "отсутствует в продаже",
        "нет продавцов",
        "нет предложений",
        "найдено 0 аптек",
        "ничего не найдено",
        "товар не найден",
    ]
    return "Нет в наличии" if any(phrase in normalized for phrase in phrases) else ""


def extract_digits(value: Any) -> int | None:
    digits = re.sub(r"\D", "", str(value or ""))
    return int(digits) if digits else None


def extract_link(cell) -> str:
    if cell.hyperlink and cell.hyperlink.target:
        target = str(cell.hyperlink.target).strip()
        if target.startswith(("http://", "https://")):
            return target
    value = cell.value
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith(("http://", "https://")):
            return stripped
    return ""


def find_column_by_header(ws, wanted: str) -> int:
    wanted_norm = wanted.strip().lower()
    for column in range(1, ws.max_column + 1):
        for row in (HEADER_ROW, PARENT_HEADER_ROW):
            value = ws.cell(row=row, column=column).value
            if str(value or "").strip().lower() == wanted_norm:
                return column
    raise KeyError(f"Column with header {wanted!r} not found")


def collect_rows(ws, limit: int) -> list[WorkRow]:
    kaspi_col = find_column_by_header(ws, "Kaspi")
    iteka_col = find_column_by_header(ws, "i-teka")
    stop_row = ws.max_row if limit <= 0 else min(ws.max_row, DATA_START_ROW + limit - 1)
    rows: list[WorkRow] = []
    for row_num in range(DATA_START_ROW, stop_row + 1):
        rows.append(
            WorkRow(
                row_num=row_num,
                address_market=ws.cell(row=row_num, column=1).value,
                segment=ws.cell(row=row_num, column=2).value,
                brand_eng=ws.cell(row=row_num, column=3).value,
                brand_rus=ws.cell(row=row_num, column=4).value,
                corporation=ws.cell(row=row_num, column=5).value,
                fcc=ws.cell(row=row_num, column=7).value,
                fcc_variant=ws.cell(row=row_num, column=8).value,
                sku_eng=ws.cell(row=row_num, column=9).value,
                sku_eng_2=ws.cell(row=row_num, column=10).value,
                kaspi_url=normalize_kaspi_url(extract_link(ws.cell(row=row_num, column=kaspi_col))),
                iteka_url=normalize_iteka_url(extract_link(ws.cell(row=row_num, column=iteka_col))),
            )
        )
    return rows


def normalize_kaspi_url(url: str) -> str:
    value = (url or "").strip()
    if not value:
        return ""
    parsed = urlparse(value if "://" in value else f"https://{value}")
    query = parse_qs(parsed.query)
    query["c"] = [CITY_ID]
    new_query = urlencode(query, doseq=True)
    return parsed._replace(query=new_query, fragment="").geturl()


def normalize_iteka_url(url: str) -> str:
    value = (url or "").strip()
    if not value:
        return ""
    parsed = urlparse(value if "://" in value else f"https://{value}")
    parts = [part for part in parsed.path.split("/") if part]
    known_cities = {"almaty", "astana", "shymkent", "karaganda", "aktobe", "pavlodar", "taraz", "kostanay"}
    if parts and parts[0] in known_cities:
        parts[0] = "almaty"
    elif parts and parts[0] == "medicaments":
        parts.insert(0, "almaty")
    elif parts and "medicaments" not in parts:
        parts = ["almaty", "medicaments", *parts]
    elif not parts:
        parts = ["almaty"]
    path = "/" + "/".join(parts)
    query = {key: val for key, val in parse_qs(parsed.query, keep_blank_values=True).items() if key != "page"}
    return urlunparse((parsed.scheme, parsed.netloc.lower(), path, "", urlencode(query, doseq=True), ""))


def kaspi_product_code(url: str) -> str:
    match = re.search(r"/p/[^/]*?-(\d+)(?:/|\?|$)", url or "")
    return match.group(1) if match else ""


def request_with_retries(method: str, url: str, **kwargs) -> requests.Response:
    last_error: Exception | None = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = requests.request(method, url, timeout=TIMEOUT_SECONDS, **kwargs)
            if response.status_code in {403, 429, 500, 502, 503, 504}:
                raise RuntimeError(f"temporary HTTP {response.status_code}")
            response.raise_for_status()
            return response
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            if attempt < MAX_RETRIES:
                error_text = str(exc)
                delay = attempt * 5 if "HTTP 403" in error_text or "HTTP 429" in error_text else attempt
                time.sleep(delay)
    raise RuntimeError(str(last_error))


def normalize_seller_key(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ").strip()).lower()


def is_zerde_seller(value: Any) -> bool:
    key = normalize_seller_key(value)
    return "зерде" in key or "zerde" in key


def is_msp_seller(value: Any) -> bool:
    key = normalize_seller_key(value)
    return "msp" in key or "мсп" in key


def fetch_kaspi_offers(url: str) -> dict[str, Any]:
    product_id = kaspi_product_code(url)
    if not product_id:
        return {"url": url, "product_code": "", "offers": [], "offers_count": 0, "note": "Kaspi product code not found"}

    endpoint = f"https://kaspi.kz/yml/offer-view/offers/{product_id}"
    headers = dict(HEADERS)
    headers.update(
        {
            "Content-Type": "application/json",
            "Origin": "https://kaspi.kz",
            "Referer": url,
        }
    )
    all_offers: list[dict[str, Any]] = []
    offers_count = 0
    limit = 50
    note = ""
    for page in range(0, 10):
        payload = {
            "cityId": CITY_ID,
            "id": product_id,
            "merchantUID": [],
            "limit": limit,
            "page": page,
            "sortOption": "PRICE",
            "highRating": None,
            "searchText": None,
            "isExcellentMerchant": False,
            "zoneId": [KASPI_ZONE_ID],
            "installationId": "-1",
        }
        try:
            response = request_with_retries("POST", endpoint, headers=headers, json=payload)
            data = response.json()
        except Exception as exc:  # noqa: BLE001
            page_note = ""
            try:
                page_response = request_with_retries("GET", url, headers=HEADERS)
                page_note = unavailable_note_from_text(page_response.text)
            except Exception:
                pass
            return {
                "url": url,
                "product_code": product_id,
                "offers": all_offers,
                "offers_count": offers_count,
                "note": page_note or f"Kaspi API failed: {exc}",
            }

        offers = data.get("offers") or []
        offers_count = int(data.get("offersCount") or data.get("total") or offers_count or len(offers))
        all_offers.extend(offers)
        if not offers or len(all_offers) >= offers_count:
            break

    best_by_seller: dict[str, dict[str, Any]] = {}
    for offer in all_offers:
        price = offer.get("price")
        seller = clean_text(offer.get("merchantName") or "")
        if price is None or not seller:
            continue
        key = normalize_seller_key(seller)
        current = best_by_seller.get(key)
        if current is None or float(price) < float(current.get("price") or 10**18):
            normalized_offer = dict(offer)
            normalized_offer["merchantName"] = seller
            normalized_offer["price"] = int(float(price))
            best_by_seller[key] = normalized_offer

    deduped = sorted(best_by_seller.values(), key=lambda item: (int(item.get("price") or 10**18), item.get("merchantName") or ""))
    if not deduped and not note:
        try:
            page_response = request_with_retries("GET", url, headers=HEADERS)
            note = unavailable_note_from_text(page_response.text)
        except Exception as exc:  # noqa: BLE001
            note = f"Kaspi offers not found; page check failed: {exc}"
        if not note:
            note = "Kaspi offers not found"
    return {"url": url, "product_code": product_id, "offers": deduped, "offers_count": offers_count, "note": note}


def select_target_offer(offers: list[dict[str, Any]], matcher) -> dict[str, Any] | None:
    matched = [offer for offer in offers if matcher(offer.get("merchantName"))]
    if not matched:
        return None
    return sorted(matched, key=lambda item: (int(item.get("price") or 10**18), item.get("merchantName") or ""))[0]


def iteka_page_url(url: str, page: int) -> str:
    parsed = urlparse(url)
    query = parse_qs(parsed.query, keep_blank_values=True)
    if page > 1:
        query["page"] = [str(page)]
    else:
        query.pop("page", None)
    return parsed._replace(query=urlencode(query, doseq=True)).geturl()


def extract_iteka_total(text: str) -> int | None:
    match = re.search(r"Найдено\s+([0-9]+)\s+аптек", text)
    return int(match.group(1)) if match else None


def parse_iteka_cards(text: str, base_url: str) -> list[dict[str, Any]]:
    cards: list[dict[str, Any]] = []
    starts = [match.start() for match in re.finditer(r'<div class="list-item\s+gtm_block_apteka', text)]
    for index, start in enumerate(starts):
        end = starts[index + 1] if index + 1 < len(starts) else len(text)
        chunk = text[start:end]
        if "data-apteka-id" not in chunk:
            continue
        apteka_match = re.search(r'data-apteka-id="([^"]+)"', chunk)
        title_match = re.search(r'<a[^>]*href="([^"]+)"[^>]*class="title"[^>]*>(.*?)</a>', chunk, re.S)
        if not title_match:
            title_match = re.search(r'<a[^>]*class="title"[^>]*href="([^"]+)"[^>]*>(.*?)</a>', chunk, re.S)
        address_match = re.search(r'<div class="address">(.*?)</div>', chunk, re.S)
        price_match = re.search(r'<div class="price mt-2">\s*(.*?)\s*</div>', chunk, re.S)
        update_match = re.search(r"Обновлено:\s*([^<]+)", chunk)
        tag_matches = re.findall(r'<div class="tag-item[^"]*">(.*?)</div>', chunk, re.S)
        is_crown = 'class="crown"' in chunk or 'data-crown="1"' in chunk
        if not title_match:
            continue
        price = extract_digits(price_match.group(1) if price_match else "")
        cards.append(
            {
                "apteka_id": apteka_match.group(1) if apteka_match else "",
                "pharmacy": clean_text(title_match.group(2)),
                "pharmacy_url": urljoin(base_url, title_match.group(1)),
                "address": clean_text(address_match.group(1) if address_match else ""),
                "price_kzt": price,
                "updated": clean_text(update_match.group(1) if update_match else ""),
                "tags": "; ".join(clean_text(tag) for tag in tag_matches if clean_text(tag)),
                "is_crown": is_crown,
            }
        )
    return cards


def fetch_iteka_crown_offers(url: str) -> dict[str, Any]:
    if not url:
        return {"url": url, "total_found": None, "offers": [], "note": "i-teka URL missing"}
    selected: list[dict[str, Any]] = []
    total_found: int | None = None
    seen: set[str] = set()
    note = ""
    for page in range(1, 11):
        page_url = iteka_page_url(url, page)
        try:
            response = request_with_retries("GET", page_url, headers=HEADERS)
            text = response.text
        except Exception as exc:  # noqa: BLE001
            note = f"i-teka request failed: {exc}"
            break
        if total_found is None:
            total_found = extract_iteka_total(text)
        cards = parse_iteka_cards(text, page_url)
        for card in cards:
            key = card.get("apteka_id") or f"{card.get('pharmacy')}|{card.get('address')}"
            if key in seen:
                continue
            seen.add(key)
            if card.get("is_crown") and card.get("price_kzt") is not None:
                selected.append(card)
                if len(selected) >= 5:
                    break
        if len(selected) >= 5:
            break
        if total_found:
            total_pages = max(1, math.ceil(total_found / 20))
            if page >= total_pages:
                break
        if not cards:
            break
    if not selected and not note:
        note = "No crown i-teka prices found"
    return {"url": url, "total_found": total_found, "offers": selected[:5], "note": note}


def run_parallel(name: str, urls: list[str], worker_count: int, fn) -> dict[str, dict[str, Any]]:
    unique_urls = sorted({url for url in urls if url})
    results: dict[str, dict[str, Any]] = {}
    if not unique_urls:
        return results
    print(f"{name}: unique URLs {len(unique_urls)}", flush=True)
    with ThreadPoolExecutor(max_workers=max(1, worker_count)) as executor:
        future_map = {executor.submit(fn, url): url for url in unique_urls}
        total = len(future_map)
        for idx, future in enumerate(as_completed(future_map), start=1):
            url = future_map[future]
            try:
                results[url] = future.result()
            except Exception as exc:  # noqa: BLE001
                results[url] = {"url": url, "offers": [], "note": str(exc)}
            if idx == 1 or idx % 20 == 0 or idx == total:
                print(f"{name}: {idx}/{total}", flush=True)
    return results


def base_row(row: WorkRow, scraped_at: str) -> list[Any]:
    return [
        scraped_at,
        row.row_num,
        row.address_market,
        row.segment,
        row.brand_eng,
        row.brand_rus,
        row.corporation,
        row.fcc,
        row.fcc_variant,
        row.sku_eng,
        row.sku_eng_2,
    ]


def offer_value(offer: dict[str, Any] | None, key: str) -> Any:
    if not offer:
        return None
    return offer.get(key)


def selected_offer_row(row: WorkRow, scraped_at: str, offer_type: str, rank: int | None, offer: dict[str, Any], url: str, code: str) -> list[Any]:
    return base_row(row, scraped_at) + [
        code,
        offer_type,
        rank,
        offer.get("merchantName"),
        offer.get("price"),
        offer.get("merchantId"),
        offer.get("merchantRating"),
        offer.get("merchantReviewsQuantity"),
        offer.get("deliveryDuration"),
        url,
    ]


def delete_sheet_if_exists(wb, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]


def append_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def style_sheet(ws) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, min(ws.max_column, 45) + 1):
        letter = get_column_letter(column)
        max_len = 10
        for row in range(1, min(ws.max_row, 80) + 1):
            value = ws.cell(row=row, column=column).value
            if value is not None:
                max_len = max(max_len, min(55, len(str(value))))
        ws.column_dimensions[letter].width = max_len + 2


def build_workbook_output(wb, rows: list[WorkRow], kaspi_results: dict[str, dict[str, Any]], iteka_results: dict[str, dict[str, Any]]) -> None:
    scraped_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    common_headers = [
        "scraped_at",
        "source_row",
        "Address Market",
        "Segment",
        "BrandEng",
        "BrandRus",
        "Corporation",
        "Fcc",
        "Fcc variant",
        "SKUEng",
        "SKUEng 2",
    ]

    kaspi_summary_rows: list[list[Any]] = []
    kaspi_selected_rows: list[list[Any]] = []
    for row in rows:
        result = kaspi_results.get(row.kaspi_url) if row.kaspi_url else None
        offers = (result or {}).get("offers") or []
        top3 = offers[:3]
        zerde = select_target_offer(offers, is_zerde_seller)
        msp = select_target_offer(offers, is_msp_seller)
        code = (result or {}).get("product_code") or kaspi_product_code(row.kaspi_url)
        summary = base_row(row, scraped_at) + [
            row.kaspi_url,
            code,
            (result or {}).get("offers_count"),
        ]
        for idx in range(3):
            offer = top3[idx] if idx < len(top3) else None
            summary.extend([offer_value(offer, "merchantName"), offer_value(offer, "price")])
        summary.extend(
            [
                offer_value(zerde, "merchantName"),
                offer_value(zerde, "price"),
                offer_value(msp, "merchantName"),
                offer_value(msp, "price"),
                (result or {}).get("note") if result else ("Kaspi URL missing" if not row.kaspi_url else "Kaspi result missing"),
            ]
        )
        kaspi_summary_rows.append(summary)
        selected_seen: set[tuple[str, str]] = set()
        for rank, offer in enumerate(top3, start=1):
            key = ("top", str(offer.get("merchantName")))
            selected_seen.add(key)
            kaspi_selected_rows.append(selected_offer_row(row, scraped_at, f"top{rank}", rank, offer, row.kaspi_url, code))
        for offer_type, offer in (("zerde", zerde), ("msp", msp)):
            if not offer:
                continue
            key = (offer_type, str(offer.get("merchantName")))
            if key in selected_seen:
                continue
            kaspi_selected_rows.append(selected_offer_row(row, scraped_at, offer_type, None, offer, row.kaspi_url, code))

    iteka_summary_rows: list[list[Any]] = []
    iteka_selected_rows: list[list[Any]] = []
    for row in rows:
        result = iteka_results.get(row.iteka_url) if row.iteka_url else None
        offers = (result or {}).get("offers") or []
        summary = base_row(row, scraped_at) + [
            row.iteka_url,
            (result or {}).get("total_found"),
            len(offers),
        ]
        for idx in range(5):
            offer = offers[idx] if idx < len(offers) else {}
            summary.extend(
                [
                    offer.get("pharmacy"),
                    offer.get("price_kzt"),
                    offer.get("address"),
                    offer.get("updated"),
                    offer.get("tags"),
                ]
            )
        summary.append((result or {}).get("note") if result else ("i-teka URL missing" if not row.iteka_url else "i-teka result missing"))
        iteka_summary_rows.append(summary)
        for rank, offer in enumerate(offers, start=1):
            iteka_selected_rows.append(
                base_row(row, scraped_at)
                + [
                    rank,
                    offer.get("pharmacy"),
                    offer.get("price_kzt"),
                    offer.get("address"),
                    offer.get("updated"),
                    offer.get("tags"),
                    offer.get("apteka_id"),
                    offer.get("pharmacy_url"),
                    row.iteka_url,
                ]
            )

    for sheet_name in ("kaspi summary", "kaspi selected", "iteka summary", "iteka selected"):
        delete_sheet_if_exists(wb, sheet_name)

    kaspi_summary_headers = common_headers + [
        "kaspi_url",
        "product_code",
        "offers_count",
        "top1_seller",
        "top1_price",
        "top2_seller",
        "top2_price",
        "top3_seller",
        "top3_price",
        "zerde_seller",
        "zerde_price",
        "msp_seller",
        "msp_price",
        "note",
    ]
    append_rows(wb.create_sheet("kaspi summary"), kaspi_summary_headers, kaspi_summary_rows)

    kaspi_selected_headers = common_headers + [
        "product_code",
        "offer_type",
        "rank",
        "seller",
        "price",
        "merchant_id",
        "rating",
        "reviews",
        "delivery_duration",
        "kaspi_url",
    ]
    append_rows(wb.create_sheet("kaspi selected"), kaspi_selected_headers, kaspi_selected_rows)

    iteka_summary_headers = common_headers + ["iteka_url", "total_found", "crown_selected_count"]
    for rank in range(1, 6):
        iteka_summary_headers.extend(
            [
                f"crown{rank}_pharmacy",
                f"crown{rank}_price",
                f"crown{rank}_address",
                f"crown{rank}_updated",
                f"crown{rank}_tags",
            ]
        )
    iteka_summary_headers.append("note")
    append_rows(wb.create_sheet("iteka summary"), iteka_summary_headers, iteka_summary_rows)

    iteka_selected_headers = common_headers + [
        "rank",
        "pharmacy",
        "price",
        "address",
        "updated",
        "tags",
        "apteka_id",
        "pharmacy_url",
        "iteka_url",
    ]
    append_rows(wb.create_sheet("iteka selected"), iteka_selected_headers, iteka_selected_rows)


def main() -> int:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(input_path)
    output_path = (
        args.output.expanduser().resolve()
        if args.output
        else input_path.with_name(f"{input_path.stem}_scraped_kaspi_iteka.xlsx")
    )
    if input_path != output_path:
        shutil.copy2(input_path, output_path)

    wb = load_workbook(output_path)
    if args.sheet not in wb.sheetnames:
        raise KeyError(f"Sheet {args.sheet!r} not found")
    source = wb[args.sheet]
    rows = collect_rows(source, args.limit)
    print(f"Rows to process: {len(rows)}", flush=True)
    kaspi_results = run_parallel("Kaspi", [row.kaspi_url for row in rows], args.workers, fetch_kaspi_offers)
    iteka_results = run_parallel("i-teka", [row.iteka_url for row in rows], max(1, min(args.workers, 12)), fetch_iteka_crown_offers)
    build_workbook_output(wb, rows, kaspi_results, iteka_results)
    wb.save(output_path)
    print(json.dumps({"output": str(output_path), "rows": len(rows)}, ensure_ascii=False), flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
