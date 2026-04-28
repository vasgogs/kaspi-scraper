#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse, urlunparse

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from scrape_price_slice_kaspi_iteka import (
    DATA_START_ROW,
    HEADERS,
    clean_text,
    extract_digits,
    extract_link,
    find_column_by_header,
    request_with_retries,
    unavailable_note_from_text,
)
from scrape_workbook_prices import _rauza_resolver


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE_DIR / "Срез цен_1_scraped_kaspi_iteka.xlsx"
SOURCE_SHEET = "рабочий"
SITE_HEADERS = {
    "biosfera": "Биосфера",
    "rauza": "Рауза",
    "europharma": "Еврофарма",
}
SITE_SELLERS = {
    "biosfera": "Биосфера",
    "rauza": "Рауза",
    "europharma": "Еврофарма",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape Biosfera, Rauza and Europharma prices from workbook links.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--sheet", default=SOURCE_SHEET)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--workers", type=int, default=16)
    return parser.parse_args()


def normalize_url(url: str) -> str:
    value = (url or "").strip()
    if not value:
        return ""
    # Fix accidental pasted duplicate URLs, e.g.
    # https://europharma.kz/itemhttps:/europharma.kz/item
    for marker in ("https://", "http://", "https:/", "http:/"):
        pos = value.find(marker, 8)
        if pos != -1:
            value = value[:pos]
            break
    parsed = urlparse(value if "://" in value else f"https://{value}")
    path = parsed.path.rstrip("/")
    return urlunparse((parsed.scheme, parsed.netloc.lower(), path, "", "", ""))


def base_row(ws, row_num: int, scraped_at: str) -> list[Any]:
    return [
        scraped_at,
        row_num,
        ws.cell(row=row_num, column=1).value,
        ws.cell(row=row_num, column=2).value,
        ws.cell(row=row_num, column=3).value,
        ws.cell(row=row_num, column=4).value,
        ws.cell(row=row_num, column=5).value,
        ws.cell(row=row_num, column=7).value,
        ws.cell(row=row_num, column=8).value,
        ws.cell(row=row_num, column=9).value,
        ws.cell(row=row_num, column=10).value,
    ]


def collect_source_rows(ws, limit: int) -> list[dict[str, Any]]:
    site_columns = {site: find_column_by_header(ws, header) for site, header in SITE_HEADERS.items()}
    stop_row = ws.max_row if limit <= 0 else min(ws.max_row, DATA_START_ROW + limit - 1)
    rows: list[dict[str, Any]] = []
    for row_num in range(DATA_START_ROW, stop_row + 1):
        links = {
            site: normalize_url(extract_link(ws.cell(row=row_num, column=column)))
            for site, column in site_columns.items()
        }
        rows.append({"row_num": row_num, "links": links})
    return rows


def product_title_from_html(text: str) -> str:
    match = re.search(r"<title>(.*?)</title>", text, re.S | re.I)
    return clean_text(match.group(1)) if match else ""


def fetch_biosfera(url: str) -> dict[str, Any]:
    if not url:
        return {"url": url, "offers": [], "note": "Biosfera URL missing"}
    response = request_with_retries("GET", url, headers=HEADERS)
    text = response.text
    match = re.search(r'"priceCurrency":"KZT","price":([0-9]+(?:\.[0-9]+)?)', text)
    if not match:
        note = unavailable_note_from_text(text) or "Biosfera price not found"
        return {"url": url, "offers": [], "note": note, "product_name": product_title_from_html(text)}
    price = int(float(match.group(1)))
    return {
        "url": url,
        "product_name": product_title_from_html(text),
        "offers": [
            {
                "site_order": 1,
                "seller": SITE_SELLERS["biosfera"],
                "price": price,
                "address": "",
                "point_id": "",
                "quantity": None,
                "product_url": url,
            }
        ],
        "note": "",
    }


def fetch_europharma(url: str) -> dict[str, Any]:
    if not url:
        return {"url": url, "offers": [], "note": "Europharma URL missing"}
    response = request_with_retries("GET", url, headers=HEADERS)
    text = response.text
    match = re.search(r'<span class="product__price-value">\s*([0-9\s]+)\s*</span>', text, re.I)
    if not match:
        match = re.search(r'"priceCurrency":"KZT","price":([0-9]+(?:\.[0-9]+)?)', text)
    price = int(float(match.group(1))) if match and "." in match.group(1) else extract_digits(match.group(1) if match else "")
    if price is None:
        note = unavailable_note_from_text(text) or "Europharma price not found"
        return {"url": url, "offers": [], "note": note, "product_name": product_title_from_html(text)}
    return {
        "url": url,
        "product_name": product_title_from_html(text),
        "offers": [
            {
                "site_order": 1,
                "seller": SITE_SELLERS["europharma"],
                "price": price,
                "address": "",
                "point_id": "",
                "quantity": None,
                "product_url": url,
            }
        ],
        "note": "",
    }


def fetch_rauza(url: str) -> dict[str, Any]:
    if not url:
        return {"url": url, "offers": [], "note": "Rauza URL missing"}
    response = request_with_retries("GET", url, headers=HEADERS)
    text = response.text
    sku_match = re.search(r"/products/([^/?#]+)", url)
    if not sku_match:
        return {"url": url, "offers": [], "note": "Rauza SKU not found in URL"}
    sku = sku_match.group(1)
    payload_match = re.search(
        r'<script type="application/json" id="__NUXT_DATA__" data-ssr="true">(.*?)</script>',
        text,
        re.S,
    )
    if not payload_match:
        return {"url": url, "offers": [], "note": "Rauza __NUXT_DATA__ not found"}
    table = json.loads(payload_match.group(1))
    root = table[1]
    data_ref = root.get("data")
    if not isinstance(data_ref, int):
        return {"url": url, "offers": [], "note": "Rauza data ref not found"}
    raw_data = table[data_ref]
    product_ref = raw_data.get(f"getItemByID_{sku}") if isinstance(raw_data, dict) else None
    if not isinstance(product_ref, int):
        return {"url": url, "offers": [], "note": f"Rauza product ref missing for {sku}"}
    item = _rauza_resolver(table)(product_ref).get("item")
    if not isinstance(item, dict):
        return {"url": url, "offers": [], "note": f"Rauza item missing for {sku}"}
    product_name = clean_text(item.get("name") or item.get("publishName") or "")
    offers: list[dict[str, Any]] = []
    seen: set[tuple[Any, Any, Any]] = set()
    for position in item.get("positions") or []:
        if not isinstance(position, dict):
            continue
        price = position.get("price")
        if price is None:
            continue
        key = (position.get("point_id"), int(float(price)), position.get("qtty"))
        if key in seen:
            continue
        seen.add(key)
        offers.append(
            {
                "site_order": len(offers) + 1,
                "seller": SITE_SELLERS["rauza"],
                "price": int(float(price)),
                "address": "",
                "point_id": position.get("point_id") or "",
                "quantity": position.get("qtty"),
                "product_url": url,
            }
        )
    if not offers:
        meta = item.get("meta") if isinstance(item.get("meta"), dict) else {}
        fallback = meta.get("estPrice") or meta.get("maxPrice")
        if fallback is not None:
            offers.append(
                {
                    "site_order": 1,
                    "seller": SITE_SELLERS["rauza"],
                    "price": int(float(fallback)),
                    "address": "",
                    "point_id": "",
                    "quantity": None,
                    "product_url": url,
                }
            )
    return {
        "url": url,
        "product_name": product_name,
        "offers": offers,
        "note": "" if offers else (unavailable_note_from_text(text) or "Rauza price not found"),
    }


def fetch_site(site: str, url: str) -> dict[str, Any]:
    if site == "biosfera":
        return fetch_biosfera(url)
    if site == "rauza":
        return fetch_rauza(url)
    if site == "europharma":
        return fetch_europharma(url)
    return {"url": url, "offers": [], "note": f"Unsupported site {site}"}


def run_parallel(jobs: list[tuple[str, str]], workers: int) -> dict[tuple[str, str], dict[str, Any]]:
    unique_jobs = sorted({job for job in jobs if job[1]})
    results: dict[tuple[str, str], dict[str, Any]] = {}
    print(f"Sites: unique URLs {len(unique_jobs)}", flush=True)
    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        future_map = {executor.submit(fetch_site, site, url): (site, url) for site, url in unique_jobs}
        total = len(future_map)
        for idx, future in enumerate(as_completed(future_map), start=1):
            job = future_map[future]
            try:
                results[job] = future.result()
            except Exception as exc:  # noqa: BLE001
                results[job] = {"url": job[1], "offers": [], "note": str(exc)}
            if idx == 1 or idx % 25 == 0 or idx == total:
                print(f"Sites: {idx}/{total}", flush=True)
    return results


def delete_sheet_if_exists(wb, name: str) -> None:
    if name in wb.sheetnames:
        del wb[name]


def style_sheet(ws) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, min(ws.max_column, 55) + 1):
        letter = get_column_letter(column)
        max_len = 10
        for row in range(1, min(ws.max_row, 120) + 1):
            value = ws.cell(row=row, column=column).value
            if value is not None:
                max_len = max(max_len, min(65, len(str(value))))
        ws.column_dimensions[letter].width = max_len + 2


def write_outputs(wb, source_ws, source_rows: list[dict[str, Any]], results: dict[tuple[str, str], dict[str, Any]]) -> None:
    scraped_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    common_headers = [
        "scraped_at",
        "source_row",
        "Address Market",
        "Segment",
        "BrandEng",
        "BrandRus",
        "Corporation",
        "Fcc",
        "Fcc variant",
        "SKUEng",
        "SKUEng 2",
    ]
    summary_headers = common_headers + [
        "site",
        "site_url",
        "prices_scraped",
        "min_price",
        "min_seller",
        "max_price",
        "max_seller",
        "note",
    ]
    detail_headers = common_headers + [
        "site",
        "site_order",
        "price_rank",
        "seller",
        "price",
        "address",
        "point_id",
        "quantity",
        "product_url",
        "note",
    ]

    for sheet_name in ("sites summary", "sites prices"):
        delete_sheet_if_exists(wb, sheet_name)
    summary_ws = wb.create_sheet("sites summary")
    detail_ws = wb.create_sheet("sites prices")
    summary_ws.append(summary_headers)
    detail_ws.append(detail_headers)

    for item in source_rows:
        row_num = item["row_num"]
        base = base_row(source_ws, row_num, scraped_at)
        for site in ("biosfera", "rauza", "europharma"):
            url = item["links"].get(site) or ""
            result = results.get((site, url)) if url else None
            offers = list((result or {}).get("offers") or [])
            priced = [offer for offer in offers if offer.get("price") is not None]
            price_sorted = sorted(priced, key=lambda offer: (int(offer.get("price") or 10**18), offer.get("seller") or ""))
            rank_by_idx = {id(offer): rank for rank, offer in enumerate(price_sorted, start=1)}
            min_offer = price_sorted[0] if price_sorted else {}
            max_offer = price_sorted[-1] if price_sorted else {}
            note = (result or {}).get("note") if result else (f"{SITE_HEADERS[site]} URL missing" if not url else f"{SITE_HEADERS[site]} result missing")
            summary_ws.append(
                base
                + [
                    SITE_HEADERS[site],
                    url,
                    len(priced),
                    min_offer.get("price"),
                    min_offer.get("seller"),
                    max_offer.get("price"),
                    max_offer.get("seller"),
                    note,
                ]
            )
            for offer in offers:
                if offer.get("price") is None:
                    continue
                detail_ws.append(
                    base
                    + [
                        SITE_HEADERS[site],
                        offer.get("site_order"),
                        rank_by_idx.get(id(offer)),
                        offer.get("seller"),
                        offer.get("price"),
                        offer.get("address"),
                        offer.get("point_id"),
                        offer.get("quantity"),
                        offer.get("product_url") or url,
                        note,
                    ]
                )

    style_sheet(summary_ws)
    style_sheet(detail_ws)


def main() -> int:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(input_path)
    output_path = args.output.expanduser().resolve() if args.output else input_path
    if output_path != input_path:
        shutil.copy2(input_path, output_path)

    wb = load_workbook(output_path)
    if args.sheet not in wb.sheetnames:
        raise KeyError(f"Sheet {args.sheet!r} not found")
    source_ws = wb[args.sheet]
    source_rows = collect_source_rows(source_ws, args.limit)
    print(f"Rows to process: {len(source_rows)}", flush=True)
    jobs: list[tuple[str, str]] = []
    for item in source_rows:
        for site, url in item["links"].items():
            if url:
                jobs.append((site, url))
    results = run_parallel(jobs, args.workers)
    write_outputs(wb, source_ws, source_rows, results)
    wb.save(output_path)
    print(f"saved {output_path}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
