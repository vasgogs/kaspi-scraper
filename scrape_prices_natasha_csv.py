#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import shutil
import time
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from scrape_price_slice_kaspi_iteka import (
    fetch_kaspi_offers,
    is_msp_seller,
    is_zerde_seller,
    normalize_kaspi_url,
    select_target_offer,
)
from scrape_price_slice_iteka_all import fetch_all_iteka_prices
from scrape_price_slice_other_sites import SITE_HEADERS, fetch_site, normalize_url
from scrape_price_slice_kaspi_iteka import normalize_iteka_url


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = BASE_DIR / "prices_natasha.csv"
DEFAULT_OUTPUT = BASE_DIR / "prices_natasha_scraped.xlsx"
SITE_ORDER = ["kaspi", "biosfera", "rauza", "europharma", "iteka"]
CSV_HEADERS = {
    "kaspi": "Kaspi",
    "biosfera": "Биосфера",
    "rauza": "Рауза",
    "europharma": "Еврофарма",
    "iteka": "i-teka",
}
SITE_HOSTS = {
    "kaspi": "kaspi.kz",
    "biosfera": "biosfera.kz",
    "rauza": "rauza-ade.kz",
    "europharma": "europharma.kz",
    "iteka": "i-teka.kz",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape prices_natasha.csv into an Excel workbook.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--workers", type=int, default=12)
    parser.add_argument("--iteka-max-pages", type=int, default=50)
    return parser.parse_args()


def normalize_site_url(site: str, url: str) -> str:
    value = (url or "").strip()
    if not value:
        return ""
    if site == "kaspi":
        return normalize_kaspi_url(value)
    if site == "iteka":
        return normalize_iteka_url(value)
    return normalize_url(value)


def detect_site_from_url(url: str) -> str:
    host = urlparse(url if "://" in url else f"https://{url}").netloc.lower()
    for site, expected_host in SITE_HOSTS.items():
        if expected_host in host:
            return site
    return ""


def read_rows(path: Path, limit: int) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = list(csv.reader(fh))
    data_rows = reader[2:] if len(reader) >= 2 else reader
    if limit > 0:
        data_rows = data_rows[:limit]
    for idx, raw in enumerate(data_rows, start=3):
        padded = list(raw) + [""] * 6
        sku = str(padded[0] or "").strip()
        if not sku:
            continue
        links = {site: "" for site in SITE_ORDER}
        # Some input rows have a link pasted into the wrong site column
        # (for example, a Rauza link in the Biosfera column).  Route such URLs
        # by their real host instead of passing them to the wrong parser.
        for site, raw_url in zip(SITE_ORDER, padded[1:6], strict=False):
            normalized = normalize_site_url(site, raw_url)
            if not normalized:
                continue
            actual_site = detect_site_from_url(normalized)
            target_site = actual_site or site
            if target_site in links and not links[target_site]:
                links[target_site] = normalize_site_url(target_site, normalized)
        rows.append({"source_row": idx, "sku": sku, "links": links})
    return rows


def run_jobs(name: str, jobs: list[Any], workers: int, fn) -> dict[Any, Any]:
    unique_jobs = sorted(set(jobs))
    results: dict[Any, Any] = {}
    print(f"{name}: unique jobs {len(unique_jobs)}", flush=True)
    if not unique_jobs:
        return results
    with ThreadPoolExecutor(max_workers=max(1, workers)) as executor:
        future_map = {executor.submit(fn, job): job for job in unique_jobs}
        total = len(future_map)
        for idx, future in enumerate(as_completed(future_map), start=1):
            job = future_map[future]
            try:
                results[job] = future.result()
            except Exception as exc:  # noqa: BLE001
                url = job[1] if isinstance(job, tuple) and len(job) > 1 and isinstance(job[1], str) else job[0] if isinstance(job, tuple) else job
                results[job] = {"url": url, "offers": [], "note": f"Scrape failed: {exc}"}
            if idx == 1 or idx % 25 == 0 or idx == total:
                print(f"{name}: {idx}/{total}", flush=True)
    return results


def fetch_other_job(job: tuple[str, str]) -> dict[str, Any]:
    site, url = job
    return fetch_site(site, url)


def fetch_iteka_job(job: tuple[str, int]) -> dict[str, Any]:
    url, max_pages = job
    return fetch_all_iteka_prices(url, max_pages)


def style_sheet(ws) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column in range(1, min(ws.max_column, 55) + 1):
        letter = get_column_letter(column)
        max_len = 10
        for row_num in range(1, min(ws.max_row, 120) + 1):
            value = ws.cell(row=row_num, column=column).value
            if value is not None:
                max_len = max(max_len, min(65, len(str(value))))
        ws.column_dimensions[letter].width = max_len + 2


def append_rows(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_sheet(ws)


def base(row: dict[str, Any], scraped_at: str) -> list[Any]:
    return [scraped_at, row["source_row"], row["sku"]]


def offer_value(offer: dict[str, Any] | None, key: str) -> Any:
    return offer.get(key) if offer else None


def selected_kaspi_row(row: dict[str, Any], scraped_at: str, offer_type: str, rank: int | None, offer: dict[str, Any], url: str, code: str) -> list[Any]:
    return base(row, scraped_at) + [
        code,
        offer_type,
        rank,
        offer.get("merchantName"),
        offer.get("price"),
        offer.get("merchantId"),
        offer.get("merchantRating"),
        offer.get("merchantReviewsQuantity"),
        offer.get("deliveryDuration"),
        url,
    ]


def build_workbook(
    rows: list[dict[str, Any]],
    kaspi_results: dict[str, dict[str, Any]],
    iteka_results: dict[tuple[str, int], dict[str, Any]],
    other_results: dict[tuple[str, str], dict[str, Any]],
    output_path: Path,
    iteka_max_pages: int,
) -> None:
    scraped_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb = Workbook()
    wb.remove(wb.active)

    source_rows = []
    for row in rows:
        source_rows.append(
            [
                row["source_row"],
                row["sku"],
                row["links"].get("kaspi"),
                row["links"].get("biosfera"),
                row["links"].get("rauza"),
                row["links"].get("europharma"),
                row["links"].get("iteka"),
            ]
        )
    append_rows(wb.create_sheet("source"), ["source_row", "SKUEng", "Kaspi", "Биосфера", "Рауза", "Еврофарма", "i-teka"], source_rows)

    kaspi_summary_rows: list[list[Any]] = []
    kaspi_price_rows: list[list[Any]] = []
    for row in rows:
        url = row["links"].get("kaspi") or ""
        result = kaspi_results.get(url) if url else None
        offers = list((result or {}).get("offers") or [])
        top3 = offers[:3]
        zerde = select_target_offer(offers, is_zerde_seller)
        msp = select_target_offer(offers, is_msp_seller)
        code = (result or {}).get("product_code") or ""
        note = (result or {}).get("note") if result else ("Kaspi URL missing" if not url else "Kaspi result missing")
        summary = base(row, scraped_at) + [url, code, (result or {}).get("offers_count")]
        for idx in range(3):
            offer = top3[idx] if idx < len(top3) else None
            summary.extend([offer_value(offer, "merchantName"), offer_value(offer, "price")])
        summary.extend([offer_value(zerde, "merchantName"), offer_value(zerde, "price"), offer_value(msp, "merchantName"), offer_value(msp, "price"), note])
        kaspi_summary_rows.append(summary)
        for rank, offer in enumerate(top3, start=1):
            kaspi_price_rows.append(selected_kaspi_row(row, scraped_at, f"top{rank}", rank, offer, url, code))
        for offer_type, offer in (("zerde", zerde), ("msp", msp)):
            if offer:
                kaspi_price_rows.append(selected_kaspi_row(row, scraped_at, offer_type, None, offer, url, code))

    append_rows(
        wb.create_sheet("kaspi summary"),
        [
            "scraped_at",
            "source_row",
            "SKUEng",
            "kaspi_url",
            "product_code",
            "offers_count",
            "top1_seller",
            "top1_price",
            "top2_seller",
            "top2_price",
            "top3_seller",
            "top3_price",
            "zerde_seller",
            "zerde_price",
            "msp_seller",
            "msp_price",
            "note",
        ],
        kaspi_summary_rows,
    )
    append_rows(
        wb.create_sheet("kaspi prices"),
        [
            "scraped_at",
            "source_row",
            "SKUEng",
            "product_code",
            "offer_type",
            "rank",
            "seller",
            "price",
            "merchant_id",
            "rating",
            "reviews",
            "delivery_duration",
            "kaspi_url",
        ],
        kaspi_price_rows,
    )

    iteka_summary_rows: list[list[Any]] = []
    iteka_price_rows: list[list[Any]] = []
    for row in rows:
        url = row["links"].get("iteka") or ""
        key = (url, iteka_max_pages)
        result = iteka_results.get(key) if url else None
        offers = list((result or {}).get("offers") or [])
        priced = [offer for offer in offers if offer.get("price_kzt") is not None]
        price_sorted = sorted(priced, key=lambda offer: (int(offer.get("price_kzt") or 10**18), offer.get("pharmacy") or ""))
        rank_by_key = {}
        for rank, offer in enumerate(price_sorted, start=1):
            offer_key = offer.get("apteka_id") or f"{offer.get('pharmacy')}|{offer.get('address')}"
            rank_by_key[offer_key] = rank
        min_offer = price_sorted[0] if price_sorted else {}
        max_offer = price_sorted[-1] if price_sorted else {}
        note = (result or {}).get("note") if result else ("i-teka URL missing" if not url else "i-teka result missing")
        iteka_summary_rows.append(
            base(row, scraped_at)
            + [
                url,
                (result or {}).get("total_found"),
                len(priced),
                (result or {}).get("pages_scraped") if result else 0,
                min_offer.get("price_kzt"),
                min_offer.get("pharmacy"),
                max_offer.get("price_kzt"),
                max_offer.get("pharmacy"),
                sum(1 for offer in offers if offer.get("is_crown")),
                note,
            ]
        )
        for offer in priced:
            offer_key = offer.get("apteka_id") or f"{offer.get('pharmacy')}|{offer.get('address')}"
            iteka_price_rows.append(
                base(row, scraped_at)
                + [
                    offer.get("site_order"),
                    rank_by_key.get(offer_key),
                    offer.get("pharmacy"),
                    offer.get("price_kzt"),
                    bool(offer.get("is_crown")),
                    offer.get("address"),
                    offer.get("updated"),
                    offer.get("tags"),
                    offer.get("apteka_id"),
                    offer.get("source_page"),
                    offer.get("pharmacy_url"),
                    url,
                ]
            )

    append_rows(
        wb.create_sheet("iteka all summary"),
        [
            "scraped_at",
            "source_row",
            "SKUEng",
            "iteka_url",
            "total_found",
            "prices_scraped",
            "pages_scraped",
            "min_price",
            "min_pharmacy",
            "max_price",
            "max_pharmacy",
            "crown_count",
            "note",
        ],
        iteka_summary_rows,
    )
    append_rows(
        wb.create_sheet("iteka all prices"),
        [
            "scraped_at",
            "source_row",
            "SKUEng",
            "site_order",
            "price_rank",
            "pharmacy",
            "price",
            "is_crown",
            "address",
            "updated",
            "tags",
            "apteka_id",
            "source_page",
            "pharmacy_url",
            "iteka_url",
        ],
        iteka_price_rows,
    )

    site_summary_rows: list[list[Any]] = []
    site_price_rows: list[list[Any]] = []
    for row in rows:
        for site in ("biosfera", "rauza", "europharma"):
            url = row["links"].get(site) or ""
            result = other_results.get((site, url)) if url else None
            offers = list((result or {}).get("offers") or [])
            priced = [offer for offer in offers if offer.get("price") is not None]
            price_sorted = sorted(priced, key=lambda offer: (int(offer.get("price") or 10**18), offer.get("seller") or ""))
            rank_by_id = {id(offer): rank for rank, offer in enumerate(price_sorted, start=1)}
            min_offer = price_sorted[0] if price_sorted else {}
            max_offer = price_sorted[-1] if price_sorted else {}
            site_label = SITE_HEADERS[site]
            note = (result or {}).get("note") if result else (f"{site_label} URL missing" if not url else f"{site_label} result missing")
            site_summary_rows.append(
                base(row, scraped_at)
                + [
                    site_label,
                    url,
                    len(priced),
                    min_offer.get("price"),
                    min_offer.get("seller"),
                    max_offer.get("price"),
                    max_offer.get("seller"),
                    note,
                ]
            )
            for offer in priced:
                site_price_rows.append(
                    base(row, scraped_at)
                    + [
                        site_label,
                        offer.get("site_order"),
                        rank_by_id.get(id(offer)),
                        offer.get("seller"),
                        offer.get("price"),
                        offer.get("address"),
                        offer.get("point_id"),
                        offer.get("quantity"),
                        offer.get("product_url") or url,
                        note,
                    ]
                )

    append_rows(
        wb.create_sheet("sites summary"),
        [
            "scraped_at",
            "source_row",
            "SKUEng",
            "site",
            "site_url",
            "prices_scraped",
            "min_price",
            "min_seller",
            "max_price",
            "max_seller",
            "note",
        ],
        site_summary_rows,
    )
    append_rows(
        wb.create_sheet("sites prices"),
        [
            "scraped_at",
            "source_row",
            "SKUEng",
            "site",
            "site_order",
            "price_rank",
            "seller",
            "price",
            "address",
            "point_id",
            "quantity",
            "product_url",
            "note",
        ],
        site_price_rows,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> int:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    output_path = args.output.expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(input_path)
    rows = read_rows(input_path, args.limit)
    print(f"Rows to process: {len(rows)}", flush=True)
    kaspi_urls = [row["links"]["kaspi"] for row in rows if row["links"].get("kaspi")]
    iteka_jobs = [(row["links"]["iteka"], args.iteka_max_pages) for row in rows if row["links"].get("iteka")]
    other_jobs = [
        (site, row["links"][site])
        for row in rows
        for site in ("biosfera", "rauza", "europharma")
        if row["links"].get(site)
    ]
    kaspi_results = run_jobs("Kaspi", kaspi_urls, max(1, min(args.workers, 10)), fetch_kaspi_offers)
    kaspi_retry_urls = [
        url
        for url, result in kaspi_results.items()
        if "temporary HTTP 403" in str((result or {}).get("note") or "")
    ]
    if kaspi_retry_urls:
        print(f"Kaspi retry after HTTP 403: {len(kaspi_retry_urls)}", flush=True)
        time.sleep(20)
        kaspi_retry_results = run_jobs("Kaspi retry", kaspi_retry_urls, 2, fetch_kaspi_offers)
        for url, retry_result in kaspi_retry_results.items():
            # Keep the retry if it found offers or at least no longer failed with 403.
            retry_note = str((retry_result or {}).get("note") or "")
            if (retry_result or {}).get("offers") or "temporary HTTP 403" not in retry_note:
                kaspi_results[url] = retry_result
    iteka_results = run_jobs("i-teka all", iteka_jobs, max(1, min(args.workers, 14)), fetch_iteka_job)
    other_results = run_jobs("Other sites", other_jobs, args.workers, fetch_other_job)
    build_workbook(rows, kaspi_results, iteka_results, other_results, output_path, args.iteka_max_pages)
    print(f"saved {output_path}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
