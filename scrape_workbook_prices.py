#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/130.0.0.0 Safari/537.36"
)
HEADERS = {
    "User-Agent": USER_AGENT,
    "Accept-Language": "ru,en;q=0.8",
}
TIMEOUT_SECONDS = 40
MAX_RETRIES = 3
DATA_START_ROW = 5
SITE_COLUMNS = {
    "kaspi": 13,
    "biosfera": 14,
    "rauza": 15,
    "europharma": 16,
    "iteka": 17,
}
OUTPUT_COLUMNS = [
    "Kaspi price",
    "Biosfera price",
    "Rauza price",
    "Europharma price",
    "i-teka avg low 3",
    "i-teka low 1",
    "i-teka low 2",
    "i-teka low 3",
    "Scrape note",
]


@dataclass
class SiteResult:
    price: float | int | None = None
    low_prices: list[int] | None = None
    note: str = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Scrape competitor prices for the 'рабочий' sheet and write a new workbook "
            "with appended price columns."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("/home/vas/kaspi-scraper/Срез цен.xlsx"),
        help="Input workbook path.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output workbook path. Defaults to '<input stem>_scraped_prices.xlsx'.",
    )
    parser.add_argument(
        "--sheet",
        default="рабочий",
        help="Sheet to copy and enrich.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Only process the first N data rows for testing.",
    )
    parser.add_argument(
        "--workers",
        type=int,
        default=15,
        help="Parallel scraper workers.",
    )
    return parser.parse_args()


def extract_digits(value: str) -> int | None:
    digits = re.sub(r"\D", "", value or "")
    return int(digits) if digits else None


def normalize_url(url: str) -> str:
    value = (url or "").strip()
    if not value:
        return ""
    parsed = urlparse(value)
    host = parsed.netloc.lower()
    path = parsed.path.rstrip("/")
    if "kaspi.kz" in host:
        query = dict(parse_qsl(parsed.query, keep_blank_values=True))
        kept = {}
        if "c" in query:
            kept["c"] = query["c"]
        return urlunparse((parsed.scheme, host, path, "", urlencode(kept), ""))
    if "biosfera.kz" in host or "rauza-ade.kz" in host or "europharma.kz" in host:
        return urlunparse((parsed.scheme, host, path, "", "", ""))
    if "i-teka.kz" in host:
        query = {
            key: value
            for key, value in parse_qsl(parsed.query, keep_blank_values=True)
            if key != "page"
        }
        return urlunparse((parsed.scheme, host, path, "", urlencode(query), ""))
    return value


def extract_link(cell) -> str:
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
    value = cell.value
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.startswith(("http://", "https://")):
            return stripped
    return ""


def set_query_param(url: str, key: str, value: str | int) -> str:
    parsed = urlparse(url)
    query = dict(parse_qsl(parsed.query, keep_blank_values=True))
    query[str(key)] = str(value)
    return urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", urlencode(query), ""))


def fetch_text(url: str) -> str:
    last_error: Exception | None = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            response = requests.get(url, headers=HEADERS, timeout=TIMEOUT_SECONDS)
            if response.status_code in {429, 500, 502, 503, 504}:
                raise RuntimeError(f"temporary HTTP {response.status_code}")
            response.raise_for_status()
            return response.text
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            if attempt < MAX_RETRIES:
                time.sleep(attempt)
    raise RuntimeError(f"request failed for {url}: {last_error}")


def scrape_kaspi(url: str) -> SiteResult:
    text = fetch_text(url)
    match = re.search(
        r'property="product:price:amount"\s+content="([0-9]+(?:\.[0-9]+)?)"',
        text,
        re.I,
    )
    if not match:
        match = re.search(r'"price"\s*:\s*"?([0-9]+(?:\.[0-9]+)?)"?', text, re.I)
    if not match:
        return SiteResult(note="Kaspi price not found")
    return SiteResult(price=float(match.group(1)))


def scrape_biosfera(url: str) -> SiteResult:
    text = fetch_text(url)
    match = re.search(r'"priceCurrency":"KZT","price":([0-9]+(?:\.[0-9]+)?)', text)
    if not match:
        return SiteResult(note="Biosfera price not found")
    return SiteResult(price=float(match.group(1)))


def scrape_europharma(url: str) -> SiteResult:
    text = fetch_text(url)
    match = re.search(
        r'<span class="product__price-value">\s*([0-9\s]+)\s*</span>',
        text,
        re.I,
    )
    if not match:
        return SiteResult(note="Europharma price not found")
    price = extract_digits(match.group(1))
    return SiteResult(price=price) if price is not None else SiteResult(note="Europharma price parse failed")


def _rauza_resolver(table: list[object]):
    limit = len(table)

    @lru_cache(maxsize=None)
    def resolve_index(index: int):
        node = table[index]
        if isinstance(node, dict):
            return {key: resolve_ref(value) for key, value in node.items()}
        if isinstance(node, list):
            if node and node[0] == "Reactive":
                return resolve_ref(node[1])
            return [resolve_ref(item) for item in node]
        return node

    def resolve_ref(value):
        if isinstance(value, int):
            if value <= 1:
                return value
            if 0 <= value < limit:
                return resolve_index(value)
            return value
        if isinstance(value, dict):
            return {key: resolve_ref(item) for key, item in value.items()}
        if isinstance(value, list):
            if value and value[0] == "Reactive":
                return resolve_ref(value[1])
            return [resolve_ref(item) for item in value]
        return value

    return resolve_index


def scrape_rauza(url: str) -> SiteResult:
    text = fetch_text(url)
    sku_match = re.search(r"/products/([^/?#]+)", url)
    if not sku_match:
        return SiteResult(note="Rauza SKU not found in URL")
    sku = sku_match.group(1)
    payload_match = re.search(
        r'<script type="application/json" id="__NUXT_DATA__" data-ssr="true">(.*?)</script>',
        text,
        re.S,
    )
    if not payload_match:
        return SiteResult(note="Rauza __NUXT_DATA__ not found")
    table = json.loads(payload_match.group(1))
    root = table[1]
    data_ref = root.get("data")
    if not isinstance(data_ref, int):
        return SiteResult(note="Rauza data ref not found")
    raw_data = table[data_ref]
    if not isinstance(raw_data, dict):
        return SiteResult(note="Rauza raw data malformed")
    product_ref = raw_data.get(f"getItemByID_{sku}")
    if not isinstance(product_ref, int):
        return SiteResult(note=f"Rauza product ref missing for {sku}")
    resolve_index = _rauza_resolver(table)
    product_wrapper = resolve_index(product_ref)
    item = product_wrapper.get("item") if isinstance(product_wrapper, dict) else None
    if not isinstance(item, dict):
        return SiteResult(note=f"Rauza item missing for {sku}")
    meta = item.get("meta") if isinstance(item.get("meta"), dict) else {}
    positions = item.get("positions") if isinstance(item.get("positions"), list) else []
    price = meta.get("estPrice") or meta.get("maxPrice")
    if price is None:
        position_prices = sorted(
            {position.get("price") for position in positions if isinstance(position, dict) and position.get("price")}
        )
        if position_prices:
            price = position_prices[0]
    if price is None:
        return SiteResult(note="Rauza price not found")
    return SiteResult(price=price)


def extract_iteka_page_prices(text: str) -> list[int]:
    matches = re.findall(r'<div class="price mt-2">\s*([0-9&thinsp;\s]+)\s*тг\.', text, re.I)
    prices: list[int] = []
    for match in matches:
        price = extract_digits(match.replace("&thinsp;", " "))
        if price is not None:
            prices.append(price)
    return prices


def extract_iteka_total(text: str) -> int | None:
    match = re.search(r"Найдено\s+([0-9]+)\s+аптек", text)
    return int(match.group(1)) if match else None


def scrape_iteka(url: str) -> SiteResult:
    first_page = fetch_text(url)
    prices = extract_iteka_page_prices(first_page)
    total = extract_iteka_total(first_page)
    total_pages = 1
    if total:
        total_pages = max(1, math.ceil(total / 20))
    for page in range(2, total_pages + 1):
        page_url = set_query_param(url, "page", page)
        page_text = fetch_text(page_url)
        page_prices = extract_iteka_page_prices(page_text)
        if not page_prices and page > 2:
            break
        prices.extend(page_prices)
    unique_prices = sorted(price for price in prices if price > 0)
    if not unique_prices:
        return SiteResult(note="i-teka prices not found")
    low_prices = unique_prices[:3]
    average = round(sum(low_prices) / len(low_prices), 2)
    return SiteResult(price=average, low_prices=low_prices)


def scrape_site(site: str, url: str) -> SiteResult:
    if site == "kaspi":
        return scrape_kaspi(url)
    if site == "biosfera":
        return scrape_biosfera(url)
    if site == "rauza":
        return scrape_rauza(url)
    if site == "europharma":
        return scrape_europharma(url)
    if site == "iteka":
        return scrape_iteka(url)
    return SiteResult(note=f"unsupported site {site}")


def build_jobs(worksheet, max_row: int) -> dict[tuple[str, str], str]:
    jobs: dict[tuple[str, str], str] = {}
    for row in range(DATA_START_ROW, max_row + 1):
        for site, column in SITE_COLUMNS.items():
            url = normalize_url(extract_link(worksheet.cell(row=row, column=column)))
            if url:
                jobs[(site, url)] = url
    return jobs


def iter_data_rows(worksheet, limit: int = 0) -> Iterable[int]:
    max_row = worksheet.max_row
    stop_row = max_row
    if limit > 0:
        stop_row = min(max_row, DATA_START_ROW + limit - 1)
    for row in range(DATA_START_ROW, stop_row + 1):
        yield row


def prepare_output_workbook(input_path: Path, output_path: Path, sheet_name: str):
    shutil.copy2(input_path, output_path)
    workbook = load_workbook(output_path)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found")
    source_sheet = workbook[sheet_name]
    target_sheet = workbook.copy_worksheet(source_sheet)
    target_sheet.title = f"{sheet_name} цены"
    return workbook, target_sheet


def write_headers(worksheet, start_column: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for offset, label in enumerate(OUTPUT_COLUMNS):
        cell = worksheet.cell(row=4, column=start_column + offset)
        cell.value = label
        cell.fill = header_fill
    worksheet.cell(row=3, column=start_column).value = "Скрейп цен"
    worksheet.cell(row=3, column=start_column + 4).value = "i-teka lowest 3"


def apply_result(worksheet, row: int, start_column: int, row_results: dict[str, SiteResult]) -> None:
    result_order = ["kaspi", "biosfera", "rauza", "europharma", "iteka"]
    for index, site in enumerate(result_order):
        result = row_results.get(site, SiteResult())
        worksheet.cell(row=row, column=start_column + index).value = result.price
    iteka = row_results.get("iteka", SiteResult())
    lows = (iteka.low_prices or [])[:3]
    for offset in range(3):
        value = lows[offset] if offset < len(lows) else None
        worksheet.cell(row=row, column=start_column + 5 + offset).value = value
    notes = []
    for site in result_order:
        note = row_results.get(site, SiteResult()).note
        if note:
            notes.append(f"{site}: {note}")
    worksheet.cell(row=row, column=start_column + 8).value = "; ".join(notes)


def main() -> int:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    if args.output is None:
        output_path = input_path.with_name(f"{input_path.stem}_scraped_prices.xlsx")
    else:
        output_path = args.output.expanduser().resolve()

    workbook, target_sheet = prepare_output_workbook(input_path, output_path, args.sheet)
    source_sheet = workbook[args.sheet]
    data_rows = list(iter_data_rows(source_sheet, args.limit))
    if not data_rows:
        raise RuntimeError("No data rows found")
    max_row = data_rows[-1]
    jobs = build_jobs(source_sheet, max_row)
    print(f"Rows to process: {len(data_rows)}", flush=True)
    print(f"Unique URLs to scrape: {len(jobs)}", flush=True)

    results: dict[tuple[str, str], SiteResult] = {}
    with ThreadPoolExecutor(max_workers=max(1, args.workers)) as executor:
        future_map = {
            executor.submit(scrape_site, site, url): (site, url)
            for site, url in jobs
        }
        completed = 0
        total = len(future_map)
        for future in as_completed(future_map):
            site, url = future_map[future]
            try:
                results[(site, url)] = future.result()
            except Exception as exc:  # noqa: BLE001
                results[(site, url)] = SiteResult(note=str(exc))
            completed += 1
            if completed == 1 or completed % 20 == 0 or completed == total:
                print(f"Scraped {completed}/{total}", flush=True)

    start_column = 18
    target_sheet.insert_cols(start_column, len(OUTPUT_COLUMNS))
    write_headers(target_sheet, start_column)

    for row in data_rows:
        row_results: dict[str, SiteResult] = {}
        for site, column in SITE_COLUMNS.items():
            normalized = normalize_url(extract_link(source_sheet.cell(row=row, column=column)))
            if normalized:
                row_results[site] = results.get((site, normalized), SiteResult(note="missing cached result"))
        apply_result(target_sheet, row, start_column, row_results)

    for row in data_rows:
        target_sheet.cell(row=row, column=start_column + 4).number_format = "0.00"

    workbook.save(output_path)
    print(f"Saved: {output_path}", flush=True)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
