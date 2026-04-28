import csv
import tempfile
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any
from uuid import uuid4

from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
RESULTS_DIR = BASE_DIR.parent / "RESULTS"
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
router = APIRouter()

ALGATOP_MASTER_PATTERNS = ("algatop_market_masterfile_summary_*.xlsx",)
ALGATOP_COMPARE_PATTERNS = ("portfolio_vs_competitors_category_compare_*.xlsx",)
ALGATOP_CATEGORY_PATTERNS = ("algatop_category_penetration_*.xlsx",)
ALGATOP_MY_MONTHLY_PATTERNS = ("algatop_niche_monthly_*_my_products.xlsx",)
ALGATOP_COMPETITOR_MONTHLY_PATTERNS = (
    "algatop_niche_monthly_*_competitors_data_sku.xlsx",
    "algatop_niche_monthly_*_competitors_only.xlsx",
)
ALGATOP_FEATURED_CATEGORY_CODES = ("02807", "02863", "04550")
ALGATOP_EXPORT_DIR = Path(tempfile.gettempdir()) / "algatop_dashboard_exports"
ALGATOP_EXPORT_DIR.mkdir(parents=True, exist_ok=True)
ALGATOP_CACHE: dict[str, Any] = {"key": None, "payload": None}


def _invalidate_algatop_cache() -> None:
    ALGATOP_CACHE["key"] = None
    ALGATOP_CACHE["payload"] = None


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _number(value: Any) -> float:
    if value in ("", None):
        return 0.0
    try:
        number = float(value)
    except (TypeError, ValueError):
        return 0.0
    if number.is_integer():
        return int(number)
    return number


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return _text(value).lower() in {"1", "true", "yes", "y"}


def _pct(part: float, whole: float) -> float:
    if not whole:
        return 0.0
    return round((part / whole) * 100, 4)


def _month_key(month: str) -> int:
    return int(_text(month).replace("-", "") or 0)


def _date_key(value: str) -> int:
    return int(_text(value).replace("-", "") or 0)


def _sort_months(months: list[str]) -> list[str]:
    return sorted({month for month in months if month}, key=_month_key)


def _month_bounds(month: str) -> tuple[str, str]:
    year, month_num = [int(part) for part in _text(month).split("-")]
    last_day = monthrange(year, month_num)[1]
    return f"{year:04d}-{month_num:02d}-01", f"{year:04d}-{month_num:02d}-{last_day:02d}"


def _latest_report(patterns: tuple[str, ...]) -> Path | None:
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(RESULTS_DIR.glob(pattern))
    if not candidates:
        return None
    candidates.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return candidates[0]


def _sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        worksheet = workbook[sheet_name]
        row_iter = worksheet.iter_rows(values_only=True)
        headers = [_text(value) for value in next(row_iter, ())]
        rows: list[dict[str, Any]] = []
        for row in row_iter:
            item: dict[str, Any] = {}
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = row[index] if index < len(row) else ""
                if isinstance(value, (datetime, date)):
                    value = value.isoformat()
                item[header] = value
            if item:
                rows.append(item)
        return rows
    finally:
        workbook.close()


def _standardize_location_rows(rows: list[dict[str, Any]], cohort: str) -> list[dict[str, Any]]:
    prepared: list[dict[str, Any]] = []
    for row in rows:
        sku = _text(row.get("sku"))
        if not sku:
            continue
        prepared.append(
            {
                "cohort": cohort,
                "sku": sku,
                "found": _bool(row.get("found")),
                "product_name": _text(row.get("product_name")),
                "brand_name": _text(row.get("brand_name")),
                "category_name": _text(row.get("category_name")),
                "category_path": _text(row.get("category_path")),
                "level_1_code": _text(row.get("level_1_code")),
                "level_1_name": _text(row.get("level_1_name")),
                "level_2_code": _text(row.get("level_2_code")),
                "level_2_name": _text(row.get("level_2_name")),
                "level_3_code": _text(row.get("level_3_code")),
                "level_3_name": _text(row.get("level_3_name")),
                "product_url": _text(row.get("product_url")),
                "error": _text(row.get("error")),
            }
        )
    return prepared


def _standardize_category_compare_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    numeric_fields = {
        "level",
        "category_sale_qty",
        "category_sale_amount",
        "category_order_qty",
        "category_sale_product_qty",
        "category_sale_seller_qty",
        "category_sale_brand_qty",
        "my_total_sku_qty",
        "my_active_sku_qty",
        "my_sale_qty",
        "my_sale_amount",
        "my_share_qty_pct",
        "my_share_amount_pct",
        "competitor_total_sku_qty",
        "competitor_active_sku_qty",
        "competitor_sale_qty",
        "competitor_sale_amount",
        "competitor_share_qty_pct",
        "competitor_share_amount_pct",
        "combined_share_amount_pct",
        "share_gap_amount_pct",
        "my_vs_competitor_amount_ratio",
    }
    prepared: list[dict[str, Any]] = []
    for row in rows:
        item = {key: _text(value) for key, value in row.items()}
        for field in numeric_fields:
            item[field] = _number(row.get(field))
        prepared.append(item)
    return prepared


def _standardize_price_summary_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    numeric_fields = {
        "price_days_qty",
        "price_first_value",
        "price_last_value",
        "price_min_value",
        "price_max_value",
        "price_avg_value",
        "price_change_abs",
        "price_change_pct",
    }
    prepared: list[dict[str, Any]] = []
    for row in rows:
        item = {key: _text(value) for key, value in row.items()}
        item["cohort"] = _text(row.get("cohort"))
        item["sku"] = _text(row.get("sku"))
        item["price_first_date"] = _text(row.get("price_first_date"))
        item["price_last_date"] = _text(row.get("price_last_date"))
        for field in numeric_fields:
            item[field] = _number(row.get(field))
        if item["cohort"] and item["sku"]:
            prepared.append(item)
    return prepared


def _standardize_price_daily_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    numeric_fields = {
        "sale_price",
        "merchant_count",
        "product_rate",
        "review_qty",
        "sale_qty",
        "sale_amount",
    }
    prepared: list[dict[str, Any]] = []
    for row in rows:
        item = {key: _text(value) for key, value in row.items()}
        item["cohort"] = _text(row.get("cohort"))
        item["sku"] = _text(row.get("sku"))
        item["product_name"] = _text(row.get("product_name"))
        item["event_date"] = _text(row.get("event_date"))
        item["is_statistic_exists"] = _bool(row.get("is_statistic_exists"))
        for field in numeric_fields:
            item[field] = _number(row.get(field))
        if item["cohort"] and item["sku"] and item["event_date"]:
            prepared.append(item)
    return prepared


def _build_location_maps(
    my_rows: list[dict[str, Any]],
    competitor_rows: list[dict[str, Any]],
) -> dict[str, dict[str, dict[str, Any]]]:
    return {
        "my": {row["sku"]: row for row in my_rows},
        "competitor": {row["sku"]: row for row in competitor_rows},
    }


def _build_raw_rows(
    monthly_rows: list[dict[str, Any]],
    cohort: str,
    location_map: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    prepared: list[dict[str, Any]] = []
    for row in monthly_rows:
        sku = _text(row.get("sku"))
        if not sku:
            continue
        location = location_map.get(sku, {})
        if not location or not _bool(location.get("found")):
            continue
        sale_qty = _number(row.get("sale_qty"))
        sale_amount = _number(row.get("sale_amount"))
        prepared.append(
            {
                "cohort": cohort,
                "sku": sku,
                "month": _text(row.get("month")),
                "product_name": _text(location.get("product_name") or row.get("product_name")),
                "brand_name": _text(location.get("brand_name")),
                "category_path": _text(location.get("category_path")),
                "level_1_code": _text(location.get("level_1_code")),
                "level_1_name": _text(location.get("level_1_name")),
                "level_2_code": _text(location.get("level_2_code")),
                "level_2_name": _text(location.get("level_2_name")),
                "level_3_code": _text(location.get("level_3_code")),
                "level_3_name": _text(location.get("level_3_name")),
                "sale_qty": sale_qty,
                "sale_amount": sale_amount,
                "sale_price": _number(row.get("sale_price")),
                "merchant_count": _number(row.get("merchant_count")),
                "product_rate": _number(row.get("product_rate")),
                "review_qty": _number(row.get("review_qty")),
                "is_statistic_exists": _bool(row.get("is_statistic_exists")),
                "product_url": _text(location.get("product_url")),
                "found": _bool(location.get("found")),
            }
        )
    return prepared


def _build_filter_options(location_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, str]]]:
    level1: dict[str, dict[str, str]] = {}
    level2: dict[str, dict[str, str]] = {}
    level3: dict[str, dict[str, str]] = {}

    for row in location_rows:
        if row["level_1_code"]:
            level1[row["level_1_code"]] = {
                "code": row["level_1_code"],
                "name": row["level_1_name"],
            }
        if row["level_2_code"]:
            level2[row["level_2_code"]] = {
                "code": row["level_2_code"],
                "name": row["level_2_name"],
                "parent_code": row["level_1_code"],
                "path": " / ".join(part for part in [row["level_1_name"], row["level_2_name"]] if part),
            }
        if row["level_3_code"]:
            level3[row["level_3_code"]] = {
                "code": row["level_3_code"],
                "name": row["level_3_name"],
                "parent_code": row["level_2_code"],
                "path": " / ".join(
                    part for part in [row["level_1_name"], row["level_2_name"], row["level_3_name"]] if part
                ),
            }

    return {
        "level1_options": sorted(level1.values(), key=lambda row: row["name"]),
        "level2_options": sorted(level2.values(), key=lambda row: row["path"]),
        "level3_options": sorted(level3.values(), key=lambda row: row["path"]),
    }


def _build_sku_options(
    location_rows: list[dict[str, Any]],
    raw_rows: list[dict[str, Any]],
    price_summary_map: dict[str, dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    revenue_map: dict[str, float] = defaultdict(float)
    qty_map: dict[str, float] = defaultdict(float)
    latest_month_map: dict[str, str] = {}

    for row in raw_rows:
        key = f"{row['cohort']}::{row['sku']}"
        revenue_map[key] += _number(row.get("sale_amount"))
        qty_map[key] += _number(row.get("sale_qty"))
        month = _text(row.get("month"))
        if month and _month_key(month) >= _month_key(latest_month_map.get(key, "")):
            latest_month_map[key] = month

    options: dict[str, list[dict[str, Any]]] = {"my": [], "competitor": [], "all": []}
    for row in location_rows:
        if not row.get("found"):
            continue
        key = f"{row['cohort']}::{row['sku']}"
        price = price_summary_map.get(key, {})
        option = {
            "cohort": row["cohort"],
            "sku": row["sku"],
            "product_name": row["product_name"],
            "brand_name": row["brand_name"],
            "level_2_name": row["level_2_name"],
            "level_3_name": row["level_3_name"],
            "category_path": row["category_path"],
            "revenue": round(revenue_map.get(key, 0.0), 2),
            "sale_qty": round(qty_map.get(key, 0.0), 2),
            "latest_month": latest_month_map.get(key, ""),
            "latest_price": _number(price.get("price_last_value")),
            "label": f"{row['sku']} - {row['product_name']} [{row['brand_name']}]",
            "display": f"{row['sku']} - {row['product_name']}",
        }
        options[row["cohort"]].append(option)
        options["all"].append(option)

    for key in options:
        options[key].sort(
            key=lambda row: (
                -_number(row["revenue"]),
                -_number(row["sale_qty"]),
                row["product_name"].lower(),
                row["sku"],
            )
        )

    return options


def _load_algatop_store() -> dict[str, Any]:
    master_report_path = _latest_report(ALGATOP_MASTER_PATTERNS)
    compare_path = _latest_report(ALGATOP_COMPARE_PATTERNS)
    category_path = _latest_report(ALGATOP_CATEGORY_PATTERNS)
    my_monthly_path = _latest_report(ALGATOP_MY_MONTHLY_PATTERNS)
    competitor_monthly_path = _latest_report(ALGATOP_COMPETITOR_MONTHLY_PATTERNS)

    missing = [
        str(path)
        for path in [master_report_path, compare_path, category_path, my_monthly_path, competitor_monthly_path]
        if path is None
    ]
    if missing:
        raise HTTPException(status_code=404, detail="Не нашёл файлы выгрузки AlgaTop для dashboard")

    cache_key = tuple(
        f"{path}:{path.stat().st_mtime_ns}"
        for path in [master_report_path, compare_path, category_path, my_monthly_path, competitor_monthly_path]
    )
    if ALGATOP_CACHE["key"] == cache_key and ALGATOP_CACHE["payload"]:
        return ALGATOP_CACHE["payload"]

    latest_compare_rows = _standardize_category_compare_rows(_sheet_rows(compare_path, "latest_compare"))
    category_compare_rows = _standardize_category_compare_rows(_sheet_rows(compare_path, "category_compare_monthly"))
    top_brands_rows = _sheet_rows(category_path, "top_brands_monthly")
    price_summary_rows = _standardize_price_summary_rows(_sheet_rows(master_report_path, "price_summary"))
    price_daily_rows = _standardize_price_daily_rows(_sheet_rows(master_report_path, "price_daily"))
    my_location_rows = _standardize_location_rows(_sheet_rows(compare_path, "my_sku_locations"), "my")
    competitor_location_rows = _standardize_location_rows(
        _sheet_rows(compare_path, "competitor_sku_locations"), "competitor"
    )
    price_summary_map = {
        f"{row['cohort']}::{row['sku']}": row
        for row in price_summary_rows
    }
    price_daily_map = _build_price_daily_map(price_daily_rows)
    location_maps = _build_location_maps(my_location_rows, competitor_location_rows)
    my_monthly_rows = _sheet_rows(my_monthly_path, "monthly")
    competitor_monthly_rows = _sheet_rows(competitor_monthly_path, "monthly")
    raw_rows = _build_raw_rows(my_monthly_rows, "my", location_maps["my"])
    raw_rows.extend(_build_raw_rows(competitor_monthly_rows, "competitor", location_maps["competitor"]))

    all_location_rows = my_location_rows + competitor_location_rows
    filter_options = _build_filter_options(all_location_rows)
    sku_options = _build_sku_options(all_location_rows, raw_rows, price_summary_map)
    months = _sort_months([row["month"] for row in raw_rows])

    payload = {
        "master_path": compare_path,
        "master_report_path": master_report_path,
        "compare_path": compare_path,
        "category_path": category_path,
        "my_monthly_path": my_monthly_path,
        "competitor_monthly_path": competitor_monthly_path,
        "latest_compare_rows": latest_compare_rows,
        "category_compare_rows": category_compare_rows,
        "top_brands_rows": top_brands_rows,
        "price_summary_rows": price_summary_rows,
        "price_summary_map": price_summary_map,
        "price_daily_rows": price_daily_rows,
        "price_daily_map": price_daily_map,
        "sku_options": sku_options,
        "location_rows": {
            "my": my_location_rows,
            "competitor": competitor_location_rows,
        },
        "raw_rows": raw_rows,
        "months": months,
        "filters": filter_options,
        "updated_at": datetime.fromtimestamp(
            max(
                path.stat().st_mtime
                for path in [master_report_path, compare_path, category_path, my_monthly_path, competitor_monthly_path]
            )
        ).isoformat(timespec="seconds"),
    }
    ALGATOP_CACHE.update({"key": cache_key, "payload": payload})
    return payload


def _resolve_month_range(store: dict[str, Any], start_month: str, end_month: str) -> tuple[str, str, list[str]]:
    months = store["months"]
    if not months:
        raise HTTPException(status_code=404, detail="В выгрузке нет помесячных данных")
    start = start_month if start_month in months else months[0]
    end = end_month if end_month in months else months[-1]
    if _month_key(start) > _month_key(end):
        start, end = end, start
    selected_months = [month for month in months if _month_key(start) <= _month_key(month) <= _month_key(end)]
    return start, end, selected_months


def _filtered_raw_rows(
    store: dict[str, Any],
    start_month: str,
    end_month: str,
    cohort: str,
    level1_code: str,
    level2_code: str,
    level3_code: str,
    brand_query: str,
    search_query: str,
) -> list[dict[str, Any]]:
    _, _, selected_months = _resolve_month_range(store, start_month, end_month)
    selected_set = set(selected_months)
    brand_query_lc = _text(brand_query).lower()
    search_query_lc = _text(search_query).lower()
    rows: list[dict[str, Any]] = []

    for row in store["raw_rows"]:
        if row["month"] not in selected_set:
            continue
        if cohort in {"my", "competitor"} and row["cohort"] != cohort:
            continue
        if level1_code and row["level_1_code"] != level1_code:
            continue
        if level2_code and row["level_2_code"] != level2_code:
            continue
        if level3_code and row["level_3_code"] != level3_code:
            continue
        if brand_query_lc and brand_query_lc not in row["brand_name"].lower():
            continue
        if search_query_lc:
            haystack = " ".join(
                [
                    row["sku"],
                    row["product_name"],
                    row["brand_name"],
                    row["category_path"],
                ]
            ).lower()
            if search_query_lc not in haystack:
                continue
        rows.append(row)

    return rows


def _extract_sku_query(value: str) -> str:
    text_value = _text(value)
    match = re.search(r"\d{6,}", text_value)
    return match.group(0) if match else ""


def _build_price_daily_map(price_daily_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    rows_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in price_daily_rows:
        key = f"{row['cohort']}::{row['sku']}"
        rows_map[key].append(row)
    for key in rows_map:
        rows_map[key].sort(key=lambda row: _date_key(row["event_date"]))
    return rows_map


def _default_sku_pair(store: dict[str, Any]) -> tuple[str, str]:
    my_default = store["sku_options"]["my"][0]["sku"] if store["sku_options"]["my"] else ""
    competitor_default = store["sku_options"]["competitor"][0]["sku"] if store["sku_options"]["competitor"] else ""
    return my_default, competitor_default


def _selected_sku_option(store: dict[str, Any], cohort: str, sku_query: str) -> dict[str, Any] | None:
    sku = _extract_sku_query(sku_query)
    if not sku:
        return None
    options = store["sku_options"].get(cohort, [])
    for option in options:
        if option["sku"] == sku:
            return option
    return None


def _price_rows_for_period(
    store: dict[str, Any],
    cohort: str,
    sku: str,
    start_month: str,
    end_month: str,
) -> list[dict[str, Any]]:
    start_date, _ = _month_bounds(start_month)
    _, end_date = _month_bounds(end_month)
    rows = store["price_daily_map"].get(f"{cohort}::{sku}", [])
    return [
        row
        for row in rows
        if _date_key(start_date) <= _date_key(row["event_date"]) <= _date_key(end_date)
    ]


def _monthly_series_for_sku(
    store: dict[str, Any],
    cohort: str,
    sku: str,
    selected_months: list[str],
) -> list[dict[str, Any]]:
    rows_map = {
        row["month"]: row
        for row in store["raw_rows"]
        if row["cohort"] == cohort and row["sku"] == sku
    }
    return [
        {
            "month": month,
            "sale_amount": _number(rows_map.get(month, {}).get("sale_amount")),
            "sale_qty": _number(rows_map.get(month, {}).get("sale_qty")),
            "sale_price": _number(rows_map.get(month, {}).get("sale_price")),
            "merchant_count": _number(rows_map.get(month, {}).get("merchant_count")),
            "review_qty": _number(rows_map.get(month, {}).get("review_qty")),
        }
        for month in selected_months
    ]


def _sku_summary_card(
    option: dict[str, Any] | None,
    monthly_rows: list[dict[str, Any]],
    price_rows: list[dict[str, Any]],
    price_summary: dict[str, Any] | None,
) -> dict[str, Any]:
    if not option:
        return {}
    sale_amount = sum(_number(row["sale_amount"]) for row in monthly_rows)
    sale_qty = sum(_number(row["sale_qty"]) for row in monthly_rows)
    latest_month_row = monthly_rows[-1] if monthly_rows else {}
    price_summary = price_summary or {}
    return {
        "cohort": option["cohort"],
        "sku": option["sku"],
        "product_name": option["product_name"],
        "brand_name": option["brand_name"],
        "level_2_name": option["level_2_name"],
        "level_3_name": option["level_3_name"],
        "category_path": option["category_path"],
        "period_sale_amount": sale_amount,
        "period_sale_qty": sale_qty,
        "latest_month_sale_amount": _number(latest_month_row.get("sale_amount")),
        "latest_month_sale_qty": _number(latest_month_row.get("sale_qty")),
        "latest_month_sale_price": _number(latest_month_row.get("sale_price")),
        "latest_month_merchant_count": _number(latest_month_row.get("merchant_count")),
        "price_days_qty": _number(price_summary.get("price_days_qty")),
        "price_first_date": _text(price_summary.get("price_first_date")),
        "price_first_value": _number(price_summary.get("price_first_value")),
        "price_last_date": _text(price_summary.get("price_last_date")),
        "price_last_value": _number(price_summary.get("price_last_value")),
        "price_min_value": _number(price_summary.get("price_min_value")),
        "price_max_value": _number(price_summary.get("price_max_value")),
        "price_avg_value": _number(price_summary.get("price_avg_value")),
        "price_change_abs": _number(price_summary.get("price_change_abs")),
        "price_change_pct": _number(price_summary.get("price_change_pct")),
        "active_price_days": len([row for row in price_rows if _number(row["sale_price"]) > 0]),
    }


def _build_sku_analysis_payload(
    store: dict[str, Any],
    start_month: str,
    end_month: str,
    my_sku_query: str,
    compare_sku_query: str,
) -> dict[str, Any]:
    start, end, selected_months = _resolve_month_range(store, start_month, end_month)
    default_my_sku, default_compare_sku = _default_sku_pair(store)
    my_option = _selected_sku_option(store, "my", my_sku_query or default_my_sku)
    compare_option = _selected_sku_option(store, "competitor", compare_sku_query or default_compare_sku)
    if not my_option and default_my_sku:
        my_option = _selected_sku_option(store, "my", default_my_sku)
    if not compare_option and default_compare_sku:
        compare_option = _selected_sku_option(store, "competitor", default_compare_sku)

    my_monthly = _monthly_series_for_sku(store, "my", my_option["sku"], selected_months) if my_option else []
    compare_monthly = (
        _monthly_series_for_sku(store, "competitor", compare_option["sku"], selected_months) if compare_option else []
    )
    my_price_rows = _price_rows_for_period(store, "my", my_option["sku"], start, end) if my_option else []
    compare_price_rows = (
        _price_rows_for_period(store, "competitor", compare_option["sku"], start, end) if compare_option else []
    )

    my_price_summary = store["price_summary_map"].get(f"my::{my_option['sku']}") if my_option else None
    compare_price_summary = (
        store["price_summary_map"].get(f"competitor::{compare_option['sku']}") if compare_option else None
    )

    summary_cards = {
        "my": _sku_summary_card(my_option, my_monthly, my_price_rows, my_price_summary),
        "compare": _sku_summary_card(compare_option, compare_monthly, compare_price_rows, compare_price_summary),
    }

    monthly_compare_rows = []
    for index, month in enumerate(selected_months):
        my_row = my_monthly[index] if index < len(my_monthly) else {}
        compare_row = compare_monthly[index] if index < len(compare_monthly) else {}
        monthly_compare_rows.append(
            {
                "month": month,
                "my_sale_amount": _number(my_row.get("sale_amount")),
                "compare_sale_amount": _number(compare_row.get("sale_amount")),
                "my_sale_qty": _number(my_row.get("sale_qty")),
                "compare_sale_qty": _number(compare_row.get("sale_qty")),
                "my_sale_price": _number(my_row.get("sale_price")),
                "compare_sale_price": _number(compare_row.get("sale_price")),
                "my_merchant_count": _number(my_row.get("merchant_count")),
                "compare_merchant_count": _number(compare_row.get("merchant_count")),
                "revenue_gap": _number(my_row.get("sale_amount")) - _number(compare_row.get("sale_amount")),
                "price_gap": _number(my_row.get("sale_price")) - _number(compare_row.get("sale_price")),
            }
        )

    price_dates = sorted(
        {
            row["event_date"]
            for row in [*my_price_rows, *compare_price_rows]
        },
        key=_date_key,
    )
    my_price_map = {row["event_date"]: row for row in my_price_rows}
    compare_price_map = {row["event_date"]: row for row in compare_price_rows}
    price_trend = [
        {
            "event_date": event_date,
            "my_sale_price": _number(my_price_map.get(event_date, {}).get("sale_price")),
            "my_sale_amount": _number(my_price_map.get(event_date, {}).get("sale_amount")),
            "compare_sale_price": _number(compare_price_map.get(event_date, {}).get("sale_price")),
            "compare_sale_amount": _number(compare_price_map.get(event_date, {}).get("sale_amount")),
            "my_sale_qty": _number(my_price_map.get(event_date, {}).get("sale_qty")),
            "compare_sale_qty": _number(compare_price_map.get(event_date, {}).get("sale_qty")),
            "my_merchant_count": _number(my_price_map.get(event_date, {}).get("merchant_count")),
            "compare_merchant_count": _number(compare_price_map.get(event_date, {}).get("merchant_count")),
            "my_review_qty": _number(my_price_map.get(event_date, {}).get("review_qty")),
            "compare_review_qty": _number(compare_price_map.get(event_date, {}).get("review_qty")),
        }
        for event_date in price_dates
    ]

    compare_summary = {}
    if summary_cards["my"] and summary_cards["compare"]:
        my_latest_price = _number(summary_cards["my"].get("price_last_value"))
        compare_latest_price = _number(summary_cards["compare"].get("price_last_value"))
        my_period_revenue = _number(summary_cards["my"].get("period_sale_amount"))
        compare_period_revenue = _number(summary_cards["compare"].get("period_sale_amount"))
        compare_summary = {
            "price_gap_abs": my_latest_price - compare_latest_price,
            "price_gap_pct": round(((my_latest_price - compare_latest_price) / compare_latest_price) * 100, 4)
            if compare_latest_price
            else 0.0,
            "revenue_gap_abs": my_period_revenue - compare_period_revenue,
            "revenue_ratio": round(my_period_revenue / compare_period_revenue, 4) if compare_period_revenue else 0.0,
        }

    return {
        "period": {
            "start_month": start,
            "end_month": end,
            "selected_months": selected_months,
        },
        "selected": {
            "my_sku": my_option["sku"] if my_option else "",
            "compare_sku": compare_option["sku"] if compare_option else "",
        },
        "summary_cards": summary_cards,
        "compare_summary": compare_summary,
        "monthly_compare_rows": monthly_compare_rows,
        "monthly_trend": {
            "months": selected_months,
            "my": my_monthly,
            "compare": compare_monthly,
        },
        "price_trend": price_trend,
    }


def _paginate(rows: list[dict[str, Any]], page: int, page_size: int) -> tuple[list[dict[str, Any]], dict[str, int]]:
    safe_page_size = max(1, min(page_size, 500))
    safe_page = max(1, page)
    total = len(rows)
    start = (safe_page - 1) * safe_page_size
    end = start + safe_page_size
    return rows[start:end], {
        "page": safe_page,
        "page_size": safe_page_size,
        "total": total,
        "total_pages": max(1, (total + safe_page_size - 1) // safe_page_size),
    }


def _category_compare_featured(store: dict[str, Any], selected_months: list[str]) -> list[dict[str, Any]]:
    month_set = set(selected_months)
    by_code_month = {
        (row["category_code"], row["month"]): row
        for row in store["category_compare_rows"]
        if row["month"] in month_set
    }
    featured: list[dict[str, Any]] = []
    latest_month = selected_months[-1]
    latest_map = {
        row["category_code"]: row
        for row in store["category_compare_rows"]
        if row["month"] == latest_month
    }
    for code in ALGATOP_FEATURED_CATEGORY_CODES:
        latest_row = latest_map.get(code)
        if not latest_row:
            continue
        trend = []
        for month in selected_months:
            row = by_code_month.get((code, month))
            if not row:
                continue
            trend.append(
                {
                    "month": month,
                    "my_share_amount_pct": row["my_share_amount_pct"],
                    "competitor_share_amount_pct": row["competitor_share_amount_pct"],
                    "my_sale_amount": row["my_sale_amount"],
                    "competitor_sale_amount": row["competitor_sale_amount"],
                    "category_sale_amount": row["category_sale_amount"],
                }
            )
        featured.append(
            {
                "category_code": code,
                "category_name": latest_row["category_name"],
                "latest_month": latest_month,
                "latest_market_revenue": latest_row["category_sale_amount"],
                "latest_market_orders": latest_row["category_order_qty"],
                "latest_market_products": latest_row["category_sale_product_qty"],
                "latest_market_sellers": latest_row["category_sale_seller_qty"],
                "latest_market_brands": latest_row["category_sale_brand_qty"],
                "my_sale_amount": latest_row["my_sale_amount"],
                "my_share_amount_pct": latest_row["my_share_amount_pct"],
                "competitor_sale_amount": latest_row["competitor_sale_amount"],
                "competitor_share_amount_pct": latest_row["competitor_share_amount_pct"],
                "share_gap_amount_pct": latest_row["share_gap_amount_pct"],
                "trend": trend,
            }
        )
    return featured


def _top_groups(rows: list[dict[str, Any]], field: str, label_field: str | None = None, limit: int = 8) -> list[dict[str, Any]]:
    groups: dict[str, dict[str, Any]] = {}
    for row in rows:
        label = _text(row.get(field) or "(Без категории)")
        if label == "(Без категории)" and field == "level_3_name":
            continue
        if label not in groups:
            groups[label] = {
                "label": _text(row.get(label_field)) if label_field else label,
                "sale_amount": 0.0,
                "sale_qty": 0.0,
                "sku_set": set(),
            }
        groups[label]["sale_amount"] += _number(row.get("sale_amount"))
        groups[label]["sale_qty"] += _number(row.get("sale_qty"))
        groups[label]["sku_set"].add(f"{row['cohort']}::{row['sku']}")

    summary = [
        {
            "label": item["label"] or label,
            "sale_amount": item["sale_amount"],
            "sale_qty": item["sale_qty"],
            "sku_count": len(item["sku_set"]),
        }
        for label, item in groups.items()
        if item["sale_amount"] or item["sale_qty"]
    ]
    summary.sort(key=lambda row: (-row["sale_amount"], -row["sale_qty"], row["label"]))
    return summary[:limit]


def _build_overview_payload(
    store: dict[str, Any],
    start_month: str,
    end_month: str,
    cohort: str,
    level1_code: str,
    level2_code: str,
    level3_code: str,
    brand_query: str,
    search_query: str,
) -> dict[str, Any]:
    start, end, selected_months = _resolve_month_range(store, start_month, end_month)
    rows = _filtered_raw_rows(
        store,
        start,
        end,
        cohort,
        level1_code,
        level2_code,
        level3_code,
        brand_query,
        search_query,
    )
    latest_month = selected_months[-1]
    total_sale_amount = sum(_number(row["sale_amount"]) for row in rows)
    total_sale_qty = sum(_number(row["sale_qty"]) for row in rows)
    unique_skus = len({f"{row['cohort']}::{row['sku']}" for row in rows})
    active_skus = len(
        {
            f"{row['cohort']}::{row['sku']}"
            for row in rows
            if _number(row["sale_amount"]) > 0 or _number(row["sale_qty"]) > 0
        }
    )
    my_sale_amount = sum(_number(row["sale_amount"]) for row in rows if row["cohort"] == "my")
    competitor_sale_amount = sum(_number(row["sale_amount"]) for row in rows if row["cohort"] == "competitor")

    monthly_trend = []
    for month in selected_months:
        month_rows = [row for row in rows if row["month"] == month]
        monthly_trend.append(
            {
                "month": month,
                "my_sale_amount": sum(_number(row["sale_amount"]) for row in month_rows if row["cohort"] == "my"),
                "competitor_sale_amount": sum(
                    _number(row["sale_amount"]) for row in month_rows if row["cohort"] == "competitor"
                ),
                "total_sale_amount": sum(_number(row["sale_amount"]) for row in month_rows),
                "total_sale_qty": sum(_number(row["sale_qty"]) for row in month_rows),
            }
        )

    latest_rows = [row for row in rows if row["month"] == latest_month]
    latest_compare_rows = [
        row for row in store["category_compare_rows"] if row["month"] == latest_month
    ]
    best_vs_competitors = [
        row
        for row in sorted(
            [
                row
                for row in latest_compare_rows
                if _number(row["level"]) == 3 and _text(row["parent_category_code"]) == "02863"
            ],
            key=lambda row: (-_number(row["share_gap_amount_pct"]), -_number(row["my_sale_amount"])),
        )[:5]
    ]
    weak_vs_competitors = [
        row
        for row in sorted(
            [
                row
                for row in latest_compare_rows
                if _number(row["level"]) == 3 and _text(row["parent_category_code"]) == "02863"
            ],
            key=lambda row: (_number(row["share_gap_amount_pct"]), -_number(row["competitor_sale_amount"])),
        )[:5]
    ]

    return {
        "period": {
            "start_month": start,
            "end_month": end,
            "selected_months": selected_months,
            "latest_month": latest_month,
        },
        "filters": {
            "cohort": cohort,
            "level1_code": level1_code,
            "level2_code": level2_code,
            "level3_code": level3_code,
            "brand_query": brand_query,
            "search_query": search_query,
        },
        "metrics": [
            {"label": "Revenue", "value": total_sale_amount, "tone": "primary"},
            {"label": "Sales qty", "value": total_sale_qty, "tone": "neutral"},
            {"label": "Unique SKU", "value": unique_skus, "tone": "neutral"},
            {"label": "Active SKU", "value": active_skus, "tone": "neutral"},
            {
                "label": "My revenue",
                "value": my_sale_amount,
                "note": f"{_pct(my_sale_amount, total_sale_amount)}% of selected data",
                "tone": "success",
            },
            {
                "label": "Competitor revenue",
                "value": competitor_sale_amount,
                "note": f"{_pct(competitor_sale_amount, total_sale_amount)}% of selected data",
                "tone": "warning",
            },
        ],
        "monthly_trend": monthly_trend,
        "top_level2": _top_groups(rows, "level_2_name"),
        "top_level3": _top_groups(rows, "level_3_name"),
        "top_brands": _top_groups(rows, "brand_name"),
        "featured_categories": _category_compare_featured(store, selected_months),
        "best_vs_competitors": best_vs_competitors,
        "weak_vs_competitors": weak_vs_competitors,
        "latest_rows_count": len(latest_rows),
    }


def _build_masterfile_rows(
    store: dict[str, Any],
    start_month: str,
    end_month: str,
    cohort: str,
    level1_code: str,
    level2_code: str,
    level3_code: str,
    brand_query: str,
    search_query: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    start, end, selected_months = _resolve_month_range(store, start_month, end_month)
    rows = _filtered_raw_rows(
        store,
        start,
        end,
        cohort,
        level1_code,
        level2_code,
        level3_code,
        brand_query,
        search_query,
    )
    latest_month = selected_months[-1]
    groups: dict[str, dict[str, Any]] = {}
    total_sale_amount = sum(_number(row["sale_amount"]) for row in rows)

    for row in rows:
        key = f"{row['cohort']}::{row['sku']}"
        if key not in groups:
            groups[key] = {
                "cohort": row["cohort"],
                "sku": row["sku"],
                "product_name": row["product_name"],
                "brand_name": row["brand_name"],
                "category_path": row["category_path"],
                "level_1_name": row["level_1_name"],
                "level_2_name": row["level_2_name"],
                "level_3_name": row["level_3_name"],
                "selected_period_sale_qty": 0.0,
                "selected_period_sale_amount": 0.0,
                "latest_month_sale_qty": 0.0,
                "latest_month_sale_amount": 0.0,
                "latest_month_sale_price": 0.0,
                "latest_month_merchant_count": 0.0,
                "latest_month_product_rate": 0.0,
                "latest_month_review_qty": 0.0,
                "price_days_qty": 0.0,
                "price_first_date": "",
                "price_first_value": 0.0,
                "price_last_date": "",
                "price_last_value": 0.0,
                "price_min_value": 0.0,
                "price_max_value": 0.0,
                "price_avg_value": 0.0,
                "price_change_abs": 0.0,
                "price_change_pct": 0.0,
                "product_url": row["product_url"],
                "_active_months": set(),
            }
        entry = groups[key]
        entry["selected_period_sale_qty"] += _number(row["sale_qty"])
        entry["selected_period_sale_amount"] += _number(row["sale_amount"])
        if _number(row["sale_amount"]) > 0 or _number(row["sale_qty"]) > 0:
            entry["_active_months"].add(row["month"])
        if row["month"] == latest_month:
            entry["latest_month_sale_qty"] = _number(row["sale_qty"])
            entry["latest_month_sale_amount"] = _number(row["sale_amount"])
            entry["latest_month_sale_price"] = _number(row["sale_price"])
            entry["latest_month_merchant_count"] = _number(row["merchant_count"])
            entry["latest_month_product_rate"] = _number(row["product_rate"])
            entry["latest_month_review_qty"] = _number(row["review_qty"])

    result: list[dict[str, Any]] = []
    for entry in groups.values():
        active_months = sorted(entry.pop("_active_months"), key=_month_key)
        active_months_qty = len(active_months)
        entry["active_months_qty"] = active_months_qty
        entry["first_active_month"] = active_months[0] if active_months else ""
        entry["latest_active_month"] = active_months[-1] if active_months else ""
        entry["avg_sale_amount_per_active_month"] = round(
            entry["selected_period_sale_amount"] / active_months_qty, 2
        ) if active_months_qty else 0.0
        entry["selected_share_pct"] = _pct(entry["selected_period_sale_amount"], total_sale_amount)
        price_summary = store["price_summary_map"].get(f"{entry['cohort']}::{entry['sku']}")
        if price_summary:
            entry.update(price_summary)
        result.append(entry)

    meta = {
        "start_month": start,
        "end_month": end,
        "selected_months": selected_months,
        "latest_month": latest_month,
        "total_sale_amount": total_sale_amount,
        "total_sale_qty": sum(_number(row["sale_qty"]) for row in rows),
        "unique_sku_count": len(result),
    }
    return result, meta


def _sort_masterfile_rows(rows: list[dict[str, Any]], sort: str) -> list[dict[str, Any]]:
    if sort == "qty_desc":
        return sorted(rows, key=lambda row: (-row["selected_period_sale_qty"], -row["selected_period_sale_amount"]))
    if sort == "latest_desc":
        return sorted(rows, key=lambda row: (-row["latest_month_sale_amount"], -row["selected_period_sale_amount"]))
    if sort == "reviews_desc":
        return sorted(rows, key=lambda row: (-row["latest_month_review_qty"], -row["selected_period_sale_amount"]))
    if sort == "name_asc":
        return sorted(rows, key=lambda row: (row["product_name"].lower(), row["sku"]))
    return sorted(rows, key=lambda row: (-row["selected_period_sale_amount"], -row["selected_period_sale_qty"]))


def _build_category_compare_rows(
    store: dict[str, Any],
    start_month: str,
    end_month: str,
    level: int,
    search_query: str,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    start, end, selected_months = _resolve_month_range(store, start_month, end_month)
    month_set = set(selected_months)
    search_lc = _text(search_query).lower()
    filtered_rows = [
        row
        for row in store["category_compare_rows"]
        if row["month"] in month_set
        and int(_number(row["level"])) == level
        and (not search_lc or search_lc in row["category_name"].lower() or search_lc in row["full_path"].lower())
    ]
    groups: dict[str, dict[str, Any]] = {}

    for row in filtered_rows:
        code = row["category_code"]
        if code not in groups:
            groups[code] = {
                "category_code": code,
                "category_name": row["category_name"],
                "full_path": row["full_path"],
                "level": row["level"],
                "selected_period_market_revenue": 0.0,
                "selected_period_market_qty": 0.0,
                "selected_period_my_revenue": 0.0,
                "selected_period_competitor_revenue": 0.0,
                "selected_period_my_qty": 0.0,
                "selected_period_competitor_qty": 0.0,
            }
        entry = groups[code]
        entry["selected_period_market_revenue"] += row["category_sale_amount"]
        entry["selected_period_market_qty"] += row["category_sale_qty"]
        entry["selected_period_my_revenue"] += row["my_sale_amount"]
        entry["selected_period_competitor_revenue"] += row["competitor_sale_amount"]
        entry["selected_period_my_qty"] += row["my_sale_qty"]
        entry["selected_period_competitor_qty"] += row["competitor_sale_qty"]

    result: list[dict[str, Any]] = []
    for entry in groups.values():
        entry["selected_period_my_share_pct"] = _pct(
            entry["selected_period_my_revenue"], entry["selected_period_market_revenue"]
        )
        entry["selected_period_competitor_share_pct"] = _pct(
            entry["selected_period_competitor_revenue"], entry["selected_period_market_revenue"]
        )
        entry["share_gap_amount_pct"] = round(
            entry["selected_period_my_share_pct"] - entry["selected_period_competitor_share_pct"], 4
        )
        result.append(entry)

    latest_month = selected_months[-1]
    featured_trend = []
    for code in ALGATOP_FEATURED_CATEGORY_CODES:
        trend_rows = [
            row
            for row in store["category_compare_rows"]
            if row["category_code"] == code and row["month"] in month_set
        ]
        if not trend_rows:
            continue
        trend_rows.sort(key=lambda row: _month_key(row["month"]))
        featured_trend.append(
            {
                "category_code": code,
                "category_name": trend_rows[-1]["category_name"],
                "series": [
                    {
                        "month": row["month"],
                        "my_share_amount_pct": row["my_share_amount_pct"],
                        "competitor_share_amount_pct": row["competitor_share_amount_pct"],
                        "my_sale_amount": row["my_sale_amount"],
                        "competitor_sale_amount": row["competitor_sale_amount"],
                        "category_sale_amount": row["category_sale_amount"],
                    }
                    for row in trend_rows
                ],
            }
        )

    meta = {
        "start_month": start,
        "end_month": end,
        "selected_months": selected_months,
        "latest_month": latest_month,
    }
    return result, {"meta": meta, "featured_trend": featured_trend}


def _sort_category_compare_rows(rows: list[dict[str, Any]], sort: str) -> list[dict[str, Any]]:
    if sort == "my_share_desc":
        return sorted(rows, key=lambda row: (-row["selected_period_my_share_pct"], -row["selected_period_my_revenue"]))
    if sort == "competitor_share_desc":
        return sorted(
            rows,
            key=lambda row: (-row["selected_period_competitor_share_pct"], -row["selected_period_competitor_revenue"]),
        )
    if sort == "gap_desc":
        return sorted(rows, key=lambda row: (-row["share_gap_amount_pct"], -row["selected_period_my_revenue"]))
    if sort == "gap_asc":
        return sorted(rows, key=lambda row: (row["share_gap_amount_pct"], -row["selected_period_competitor_revenue"]))
    return sorted(rows, key=lambda row: (-row["selected_period_market_revenue"], row["category_name"]))


def _sort_raw_rows(rows: list[dict[str, Any]], sort: str) -> list[dict[str, Any]]:
    if sort == "amount_asc":
        return sorted(rows, key=lambda row: (_month_key(row["month"]) * -1, row["sale_amount"], row["sku"]))
    if sort == "qty_desc":
        return sorted(rows, key=lambda row: (-_month_key(row["month"]), -row["sale_qty"], -row["sale_amount"]))
    if sort == "name_asc":
        return sorted(rows, key=lambda row: (-_month_key(row["month"]), row["product_name"].lower(), row["sku"]))
    return sorted(rows, key=lambda row: (-_month_key(row["month"]), -row["sale_amount"], -row["sale_qty"]))


def _overview_export_rows(payload: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in payload["metrics"]:
        rows.append(
            {
                "section": "metrics",
                "metric": item["label"],
                "value": item["value"],
                "note": item.get("note", ""),
            }
        )
    for item in payload["featured_categories"]:
        rows.append(
            {
                "section": "featured_categories",
                "metric": item["category_name"],
                "value": item["my_share_amount_pct"],
                "note": f"my share vs competitor {item['competitor_share_amount_pct']}%",
            }
        )
    return rows


def _write_export(rows: list[dict[str, Any]], filename: str, file_format: str) -> Path:
    safe_ext = "xlsx" if file_format == "xlsx" else "csv"
    export_path = ALGATOP_EXPORT_DIR / f"{filename}_{uuid4().hex[:8]}.{safe_ext}"

    if safe_ext == "csv":
        headers = list(rows[0].keys()) if rows else ["empty"]
        with export_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers)
            writer.writeheader()
            for row in rows:
                writer.writerow(row)
        return export_path

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "data"
    headers = list(rows[0].keys()) if rows else ["empty"]
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])
    workbook.save(export_path)
    return export_path


@router.get("/algatop", response_class=HTMLResponse)
async def algatop_home(request: Request):
    store = _load_algatop_store()
    return templates.TemplateResponse(
        "algatop_dashboard.html",
        {
            "request": request,
            "updated_at": store["updated_at"],
            "master_filename": store["master_report_path"].name,
        },
    )


@router.get("/algatop/api/filters")
async def algatop_filters():
    store = _load_algatop_store()
    default_my_sku, default_compare_sku = _default_sku_pair(store)
    payload = {
        "months": store["months"],
        "defaults": {
            "start_month": store["months"][0],
            "end_month": store["months"][-1],
            "cohort": "all",
            "level1_code": "",
            "level2_code": "",
            "level3_code": "",
            "brand_query": "",
            "search_query": "",
            "master_sort": "revenue_desc",
            "raw_sort": "amount_desc",
            "compare_level": 2,
            "compare_sort": "market_desc",
            "my_sku": default_my_sku,
            "compare_sku": default_compare_sku,
        },
        "cohorts": [
            {"value": "all", "label": "All"},
            {"value": "my", "label": "My portfolio"},
            {"value": "competitor", "label": "Competitors"},
        ],
        "download_datasets": [
            {"value": "overview", "label": "Overview summary"},
            {"value": "masterfile", "label": "Masterfile"},
            {"value": "raw", "label": "Raw monthly data"},
            {"value": "category_compare", "label": "Category compare"},
        ],
        "source_files": {
            "master": store["master_report_path"].name,
            "compare": store["compare_path"].name,
            "category": store["category_path"].name,
            "my_monthly": store["my_monthly_path"].name,
            "competitor_monthly": store["competitor_monthly_path"].name,
        },
        "updated_at": store["updated_at"],
        "sku_options": store["sku_options"],
        **store["filters"],
    }
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@router.post("/algatop/api/reload")
async def algatop_reload():
    _invalidate_algatop_cache()
    store = _load_algatop_store()
    payload = {
        "status": "ok",
        "updated_at": store["updated_at"],
        "source_files": {
            "master": store["master_report_path"].name,
            "compare": store["compare_path"].name,
            "category": store["category_path"].name,
            "my_monthly": store["my_monthly_path"].name,
            "competitor_monthly": store["competitor_monthly_path"].name,
        },
    }
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@router.get("/algatop/api/overview")
async def algatop_overview(
    start_month: str = "",
    end_month: str = "",
    cohort: str = "all",
    level1_code: str = "",
    level2_code: str = "",
    level3_code: str = "",
    brand_query: str = "",
    search_query: str = "",
):
    store = _load_algatop_store()
    payload = _build_overview_payload(
        store,
        start_month,
        end_month,
        cohort,
        level1_code,
        level2_code,
        level3_code,
        brand_query,
        search_query,
    )
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@router.get("/algatop/api/masterfile")
async def algatop_masterfile(
    start_month: str = "",
    end_month: str = "",
    cohort: str = "all",
    level1_code: str = "",
    level2_code: str = "",
    level3_code: str = "",
    brand_query: str = "",
    search_query: str = "",
    sort: str = "revenue_desc",
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=200),
):
    store = _load_algatop_store()
    rows, meta = _build_masterfile_rows(
        store,
        start_month,
        end_month,
        cohort,
        level1_code,
        level2_code,
        level3_code,
        brand_query,
        search_query,
    )
    sorted_rows = _sort_masterfile_rows(rows, sort)
    page_rows, pagination = _paginate(sorted_rows, page, page_size)
    return JSONResponse(
        {"rows": page_rows, "pagination": pagination, "meta": meta},
        headers={"Cache-Control": "no-store"},
    )


@router.get("/algatop/api/raw")
async def algatop_raw(
    start_month: str = "",
    end_month: str = "",
    cohort: str = "all",
    level1_code: str = "",
    level2_code: str = "",
    level3_code: str = "",
    brand_query: str = "",
    search_query: str = "",
    sort: str = "amount_desc",
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=300),
):
    store = _load_algatop_store()
    rows = _filtered_raw_rows(
        store,
        start_month,
        end_month,
        cohort,
        level1_code,
        level2_code,
        level3_code,
        brand_query,
        search_query,
    )
    sorted_rows = _sort_raw_rows(rows, sort)
    page_rows, pagination = _paginate(sorted_rows, page, page_size)
    meta = {
        "total_sale_amount": sum(_number(row["sale_amount"]) for row in rows),
        "total_sale_qty": sum(_number(row["sale_qty"]) for row in rows),
    }
    return JSONResponse(
        {"rows": page_rows, "pagination": pagination, "meta": meta},
        headers={"Cache-Control": "no-store"},
    )


@router.get("/algatop/api/category-compare")
async def algatop_category_compare(
    start_month: str = "",
    end_month: str = "",
    level: int = Query(2, ge=1, le=3),
    search_query: str = "",
    sort: str = "market_desc",
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=200),
):
    store = _load_algatop_store()
    rows, extra = _build_category_compare_rows(store, start_month, end_month, level, search_query)
    sorted_rows = _sort_category_compare_rows(rows, sort)
    page_rows, pagination = _paginate(sorted_rows, page, page_size)
    return JSONResponse(
        {
            "rows": page_rows,
            "pagination": pagination,
            "meta": extra["meta"],
            "featured_trend": extra["featured_trend"],
        },
        headers={"Cache-Control": "no-store"},
    )


@router.get("/algatop/api/sku-analysis")
async def algatop_sku_analysis(
    start_month: str = "",
    end_month: str = "",
    my_sku: str = "",
    compare_sku: str = "",
):
    store = _load_algatop_store()
    payload = _build_sku_analysis_payload(store, start_month, end_month, my_sku, compare_sku)
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@router.get("/algatop/download")
async def algatop_download(
    dataset: str = Query("masterfile"),
    file_format: str = Query("xlsx"),
    start_month: str = "",
    end_month: str = "",
    cohort: str = "all",
    level1_code: str = "",
    level2_code: str = "",
    level3_code: str = "",
    brand_query: str = "",
    search_query: str = "",
    master_sort: str = "revenue_desc",
    raw_sort: str = "amount_desc",
    compare_level: int = Query(2, ge=1, le=3),
    compare_search_query: str = "",
    compare_sort: str = "market_desc",
):
    store = _load_algatop_store()
    if file_format not in {"csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="Поддерживаются только csv и xlsx")

    if dataset == "overview":
        overview = _build_overview_payload(
            store,
            start_month,
            end_month,
            cohort,
            level1_code,
            level2_code,
            level3_code,
            brand_query,
            search_query,
        )
        rows = _overview_export_rows(overview)
    elif dataset == "masterfile":
        rows, _ = _build_masterfile_rows(
            store,
            start_month,
            end_month,
            cohort,
            level1_code,
            level2_code,
            level3_code,
            brand_query,
            search_query,
        )
        rows = _sort_masterfile_rows(rows, master_sort)
    elif dataset == "raw":
        rows = _filtered_raw_rows(
            store,
            start_month,
            end_month,
            cohort,
            level1_code,
            level2_code,
            level3_code,
            brand_query,
            search_query,
        )
        rows = _sort_raw_rows(rows, raw_sort)
    elif dataset == "category_compare":
        rows, _ = _build_category_compare_rows(store, start_month, end_month, compare_level, compare_search_query)
        rows = _sort_category_compare_rows(rows, compare_sort)
    else:
        raise HTTPException(status_code=400, detail="Неизвестный dataset для скачивания")

    start, end, _ = _resolve_month_range(store, start_month, end_month)
    filename = f"algatop_{dataset}_{start}_{end}"
    export_path = _write_export(rows, filename, file_format)
    media_type = "text/csv" if file_format == "csv" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return FileResponse(export_path, filename=export_path.name, media_type=media_type)
