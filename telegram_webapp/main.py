import json
import os
import sys
import threading
import csv
import statistics
import math
import re
import html
from collections import Counter, defaultdict
from datetime import datetime, date, timedelta
from pathlib import Path

from fastapi import FastAPI, HTTPException, Request, Query
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import load_workbook
from zoneinfo import ZoneInfo
import pandas as pd
import sqlite3

BASE_DIR = Path(__file__).resolve().parent

app = FastAPI(title="LilPonk Telegram WebApp")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
ROOT_DIR = BASE_DIR.parent
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from Scraper_Kaspi import run_kaspi_scrape, run_keyword_search_report, run_reviews_scrape, normalize_seller_name  # noqa: E402
from algatop_dashboard import router as algatop_router  # noqa: E402

RESULTS_DIR = BASE_DIR.parent / "RESULTS"
BUNDLE_CACHE_PATH = ROOT_DIR / "state" / "bundle_price_cache.json"
WOLT_RESULTS_DIR = Path(
    os.environ.get("WOLT_RESULTS_DIR", str(BASE_DIR.parent / "wolt_project" / "RESULTS"))
)
GLOVO_RESULTS_DIR = Path(
    os.environ.get("GLOVO_RESULTS_DIR", str(BASE_DIR.parent / "glovo_project" / "RESULTS"))
)
REVIEWS_DB_PATH = Path(os.environ.get("REVIEWS_DB_PATH", str(BASE_DIR.parent / "state" / "reviews.db")))
MISSION_FILE_PREFIX = os.environ.get("MISSION_FILE_PREFIX") or "mission_april"
MAIN_FILE_PREFIX = os.environ.get("MAIN_FILE_PREFIX") or "kaspi_prices"
COMPETITOR_FILE_PREFIX = os.environ.get("COMPETITOR_FILE_PREFIX") or "kaspi_competitors_almaty"
KEYWORD_FILE_PREFIX = os.environ.get("KEYWORD_FILE_PREFIX") or "kaspi_search"
MISSION_VIEW_LIMIT = int(os.environ.get("MISSION_VIEW_LIMIT", "1000"))
MISSION_REQUIRED_COLUMNS = {
    "region",
    "seller",
    "input_product",
    "product",
    "price_kzt",
    "actual_price_kzt",
    "status",
    "best_price_kzt",
    "better_price_analysis",
    "second_price_analysis",
    "scraped_at",
    "review_count",
    "review_progress",
}
MISSION_CAPTURE_COLUMNS = MISSION_REQUIRED_COLUMNS | {
    "product_url",
    "review_count",
    "review_progress",
    "bundle_of",
    "bundle_qty",
    "bundle_1_sku_price_kzt",
    "bundle_price_kzt",
    "bundle_unit_price_kzt",
    "bundle_same_sku_total_kzt",
    "bundle_savings_vs_same_sku_total_kzt",
    "bundle_vs_single_kzt",
    "bundle_best_seller",
    "1_sku_best_price_kzt",
    "bundle_best_price_kzt",
    "bundle_savings_vs_best_total_kzt",
    "bundle_vs_best_kzt",
    "bundle_discount_pct",
    "bundle_note",
}
CACHE_VERSION = 2
MISSION_CACHE: dict = {"path": None, "mtime": 0.0, "payload": None, "version": CACHE_VERSION}
KEYWORD_LOCK = threading.Lock()
KEYWORD_STATE: dict = {"status": "idle", "last_run": "", "last_error": "", "last_file": ""}
REVIEWS_LOCK = threading.Lock()
REVIEWS_STATE: dict = {"status": "idle", "last_run": "", "last_error": "", "last_file": ""}
COMPETITOR_LOCK = threading.Lock()
COMPETITOR_STATE: dict = {"status": "idle", "last_run": "", "last_error": "", "last_file": ""}
ALMATY_TZ = ZoneInfo("Asia/Almaty")
KEYWORD_ROWS_LIMIT = int(os.environ.get("KEYWORD_ROWS_LIMIT", "120"))
WOLT_ITEMS_PATTERN = re.compile(
    r"^wolt_brand_(?P<brand>[a-z0-9_-]+)_(?P<stamp>\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})_items\.csv$",
    re.IGNORECASE,
)
WOLT_PHARMACIES_PATTERN = re.compile(
    r"^wolt_almaty_pharmacies_(?P<stamp>\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.csv$",
    re.IGNORECASE,
)
GLOVO_ITEMS_PATTERN = re.compile(
    r"^glovo_brand_(?P<brand>[a-z0-9_-]+)_(?P<stamp>\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})_items\.csv$",
    re.IGNORECASE,
)
GLOVO_PHARMACIES_PATTERN = re.compile(
    r"^glovo_almaty_pharmacies_(?P<stamp>\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.csv$",
    re.IGNORECASE,
)
GLOVO_GRID_PATTERN = re.compile(
    r"^glovo_almaty_address_grid_(?P<stamp>\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.csv$",
    re.IGNORECASE,
)
GLOVO_VISIBILITY_PATTERN = re.compile(
    r"^glovo_almaty_store_visibility_(?P<stamp>\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.csv$",
    re.IGNORECASE,
)
GLOVO_COMPARE_PATTERN = re.compile(
    r"^glovo_almaty_vs_wolt_(?P<stamp>\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.csv$",
    re.IGNORECASE,
)
GLOVO_BRANCH_CONTEXTS_PATTERN = re.compile(
    r"^glovo_almaty_branch_contexts_(?P<stamp>\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.csv$",
    re.IGNORECASE,
)
WOLT_ASSORTMENT_GAP_PATTERN = re.compile(
    r"^wolt_almaty_assortment_gap_(?P<stamp>\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})\.json$",
    re.IGNORECASE,
)
CAMPAIGN_QUALITY_SUMMARY_PATTERN = re.compile(
    r"^mission_campaign_quality_(?P<campaign_start>\d{4}-\d{2}-\d{2})_(?P<campaign_end>\d{4}-\d{2}-\d{2})_(?P<stamp>\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})_summary\.json$",
    re.IGNORECASE,
)
GLOVO_SLUG_NAME_OVERRIDES = {
    "ak-zhol-pharm-ala": "Ак Жол Pharm",
    "alma-ala": "Алма",
    "apteka-so-sklada-ala": "Аптека со склада",
    "dobraya-apteka-ala": "Добрая аптека",
    "el-lab-pcr-abylaikhan-ala": "El Lab PCR",
    "europharma-ala": "Europharma",
    "pharmacom-ala": "Pharmacom",
    "sadyhan-ala": "Садыхан",
    "sadykhan-ala": "Садыхан",
    "yves-rocher-ala-1": "Yves Rocher",
}
MSP_SELLER_PATTERN = re.compile(r"^аптека\s*msp(?:\s+(?:алматы|астана|шымкент))?$", re.IGNORECASE)
WOLT_MAX_REPORTS = int(os.environ.get("WOLT_MAX_REPORTS", "30"))
WOLT_MAX_ITEMS = int(os.environ.get("WOLT_MAX_ITEMS", "2000"))
GLOVO_MAX_REPORTS = int(os.environ.get("GLOVO_MAX_REPORTS", "30"))
GLOVO_MAX_ITEMS = int(os.environ.get("GLOVO_MAX_ITEMS", "2000"))
WOLT_GEO_CACHE: dict = {"path": None, "mtime": 0.0, "payload": None}
WOLT_GAP_CACHE: dict = {"path": None, "mtime": 0.0, "payload": None}
GLOVO_CACHE: dict = {"reports_at": 0.0, "summary_at": 0.0, "summary_brand": "", "path": None, "mtime": 0.0, "payload": None}
GLOVO_PHARMACY_CACHE: dict = {"path": None, "mtime": 0.0, "payload": None}
GLOVO_COVERAGE_CACHE: dict = {"key": None, "payload": None}
_price_leaders_live_json_raw = os.environ.get("PRICE_LEADERS_LIVE_JSON", "").strip()
if _price_leaders_live_json_raw:
    PRICE_LEADERS_LIVE_JSON = Path(_price_leaders_live_json_raw)
    if not PRICE_LEADERS_LIVE_JSON.is_absolute():
        PRICE_LEADERS_LIVE_JSON = (BASE_DIR.parent / PRICE_LEADERS_LIVE_JSON).resolve()
else:
    PRICE_LEADERS_LIVE_JSON = RESULTS_DIR / "kaspi_price_leaders_live_latest.json"
PRICE_LEADERS_CACHE: dict = {"path": None, "mtime": 0.0, "payload": None}
PRICE_LEADERS_ROWS_CACHE: dict = {"path": None, "mtime": 0.0, "rows": None}
PRICE_LEADERS_ALL_TIME_CACHE: dict = {"key": None, "rows": None, "meta": None}
BUNDLE_CACHE: dict = {"path": None, "mtime": 0.0, "rows": None}
COMPETITOR_CACHE: dict = {"path": None, "mtime": 0.0, "city": "", "limit": 0, "payload": None}
CAMPAIGN_QUALITY_CACHE: dict = {"path": None, "mtime": 0.0, "payload": None}
PRICE_DYNAMICS_OPTIONS_CACHE: dict = {"key": None, "payload": None}
PRICE_DYNAMICS_SUMMARY_CACHE: dict = {"key": None, "payload": None}

app.include_router(algatop_router)


@app.middleware("http")
async def disable_cache(request: Request, call_next):
    response = await call_next(request)
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


def _mission_prefixes() -> list[str]:
    raw = [
        MISSION_FILE_PREFIX,
        "mission",
        "mission_april",
        "mission_april_temp",
        "mission_march",
        "mission_march_temp",
        "mission_february",
        "mission_february_temp",
        "mission_january",
        "mission_january_temp",
    ]
    seen: set[str] = set()
    prefixes: list[str] = []
    for prefix in raw:
        norm = (prefix or "").strip()
        if norm and norm not in seen:
            seen.add(norm)
            prefixes.append(norm)
    return prefixes


def _latest_report(prefixes: list[str]) -> Path | None:
    if not RESULTS_DIR.exists():
        return None
    candidates: list[Path] = []
    for prefix in prefixes:
        norm = (prefix or "").strip()
        if not norm:
            continue
        candidates.extend(RESULTS_DIR.glob(f"{norm}_*.xlsx"))
    if not candidates:
        return None
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates[0]


def _keyword_latest_report() -> Path | None:
    return _latest_report([KEYWORD_FILE_PREFIX, "kaspi_search"])


def _main_prefixes() -> list[str]:
    raw = [MAIN_FILE_PREFIX, "kaspi_prices"]
    seen: set[str] = set()
    prefixes: list[str] = []
    for prefix in raw:
        norm = (prefix or "").strip()
        if norm and norm not in seen:
            seen.add(norm)
            prefixes.append(norm)
    return prefixes


def _competitor_prefixes() -> list[str]:
    raw = [COMPETITOR_FILE_PREFIX, "kaspi_competitors_almaty"]
    seen: set[str] = set()
    prefixes: list[str] = []
    for prefix in raw:
        norm = (prefix or "").strip()
        if norm and norm not in seen:
            seen.add(norm)
            prefixes.append(norm)
    return prefixes


def _competitor_report_candidates() -> list[Path]:
    if not RESULTS_DIR.exists():
        return []
    candidates: list[Path] = []
    for prefix in _competitor_prefixes():
        candidates.extend(RESULTS_DIR.glob(f"{prefix}_*.xlsx"))
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates


def _extract_product_code_from_url(url: str) -> str:
    text = str(url or "")
    if not text:
        return ""
    matches = re.findall(r"(\d{6,})", text)
    if not matches:
        return ""
    return matches[-1]


def _canonical_seller_name(value: str) -> str:
    text = " ".join(str(value or "").replace("\u00a0", " ").split()).strip()
    if not text:
        return ""
    if MSP_SELLER_PATTERN.match(text):
        return "Аптека MSP"
    return text


def _load_bundle_cache_rows() -> dict[str, dict]:
    path = BUNDLE_CACHE_PATH
    if not path.exists():
        return {}
    mtime = path.stat().st_mtime
    if BUNDLE_CACHE.get("path") == path and BUNDLE_CACHE.get("mtime") == mtime and BUNDLE_CACHE.get("rows") is not None:
        return BUNDLE_CACHE["rows"]
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        BUNDLE_CACHE.update({"path": path, "mtime": mtime, "rows": {}})
        return {}
    rows = payload.get("rows") if isinstance(payload, dict) else {}
    rows = rows if isinstance(rows, dict) else {}
    BUNDLE_CACHE.update({"path": path, "mtime": mtime, "rows": rows})
    return rows


def _bundle_cache_note(cache_entry: dict | None) -> str:
    if not cache_entry:
        return ""
    stamp = str(cache_entry.get("cached_at") or "").strip()
    if not stamp:
        return "Показаны последние успешные данные по 1 SKU"
    try:
        dt = datetime.fromisoformat(stamp)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ALMATY_TZ)
        else:
            dt = dt.astimezone(ALMATY_TZ)
        return f"Показаны последние успешные данные по 1 SKU от {dt.strftime('%d.%m %H:%M')}"
    except Exception:
        return "Показаны последние успешные данные по 1 SKU"


def _enrich_bundle_item(record: dict, cache_rows: dict[str, dict]) -> dict:
    bundle_of = str(record.get("bundle_of") or "").strip()
    bundle_qty = _int_or_none(record.get("bundle_qty"))
    if not bundle_of or not bundle_qty:
        return record

    bundle_price = _int_or_none(record.get("bundle_price_kzt"))
    if bundle_price is None:
        bundle_price = _int_or_none(record.get("price_kzt"))
    record["bundle_price_kzt"] = bundle_price
    if record.get("bundle_unit_price_kzt") is None and bundle_price is not None and bundle_qty:
        record["bundle_unit_price_kzt"] = int(math.ceil(bundle_price / bundle_qty))

    cache_key = "||".join(
        [
            str(record.get("region") or "").strip(),
            normalize_seller_name(str(record.get("seller_raw") or record.get("seller") or "").strip()),
            bundle_of,
        ]
    )
    cache_entry = cache_rows.get(cache_key) or {}
    cache_used = False

    if record.get("bundle_1_sku_price_kzt") is None:
        cached_base = _int_or_none(cache_entry.get("bundle_1_sku_price_kzt"))
        if cached_base is not None:
            record["bundle_1_sku_price_kzt"] = cached_base
            cache_used = True
    if record.get("1_sku_best_price_kzt") is None:
        cached_best = _int_or_none(cache_entry.get("1_sku_best_price_kzt"))
        if cached_best is not None:
            record["1_sku_best_price_kzt"] = cached_best
            cache_used = True
    if not (record.get("bundle_best_seller") or "").strip():
        cached_seller = str(cache_entry.get("bundle_best_seller") or "").strip()
        if cached_seller:
            record["bundle_best_seller"] = cached_seller
            cache_used = True

    base_price = _int_or_none(record.get("bundle_1_sku_price_kzt"))
    best_price = _int_or_none(record.get("1_sku_best_price_kzt"))
    if record.get("bundle_same_sku_total_kzt") is None and base_price is not None:
        record["bundle_same_sku_total_kzt"] = base_price * bundle_qty
    if record.get("bundle_savings_vs_same_sku_total_kzt") is None and record.get("bundle_same_sku_total_kzt") is not None and bundle_price is not None:
        record["bundle_savings_vs_same_sku_total_kzt"] = _int_or_none(record.get("bundle_same_sku_total_kzt")) - bundle_price
    if record.get("bundle_best_price_kzt") is None and best_price is not None:
        record["bundle_best_price_kzt"] = best_price * bundle_qty
    if record.get("bundle_savings_vs_best_total_kzt") is None and record.get("bundle_best_price_kzt") is not None and bundle_price is not None:
        record["bundle_savings_vs_best_total_kzt"] = _int_or_none(record.get("bundle_best_price_kzt")) - bundle_price
    if record.get("bundle_vs_single_kzt") is None and record.get("bundle_unit_price_kzt") is not None and base_price is not None:
        record["bundle_vs_single_kzt"] = _int_or_none(record.get("bundle_unit_price_kzt")) - base_price
    if record.get("bundle_vs_best_kzt") is None and record.get("bundle_best_price_kzt") is not None and bundle_price is not None:
        record["bundle_vs_best_kzt"] = bundle_price - _int_or_none(record.get("bundle_best_price_kzt"))
    if not str(record.get("bundle_discount_pct") or "").strip() and record.get("bundle_best_price_kzt") and bundle_price is not None:
        best_total = _int_or_none(record.get("bundle_best_price_kzt"))
        if best_total:
            record["bundle_discount_pct"] = f"{round((best_total - bundle_price) / best_total * 100, 2):.2f}%"

    if cache_used and not str(record.get("bundle_note") or "").strip():
        record["bundle_note"] = _bundle_cache_note(cache_entry)
    elif not str(record.get("bundle_note") or "").strip() and record.get("bundle_1_sku_price_kzt") is None and record.get("1_sku_best_price_kzt") is None:
        record["bundle_note"] = "1 SKU будет показан после следующего bundle-пересчёта"
    return record


def _read_competitor_product_map(path: Path, city: str | None = "Алматы") -> dict[str, dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        header = next(rows_iter, None)
        if not header:
            return {}
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in header]
        idx = {name: i for i, name in enumerate(headers)}
        product_idx = idx.get("product")
        input_product_idx = idx.get("input_product")
        url_idx = idx.get("product_url")
        code_idx = idx.get("product_code")
        seller_idx = idx.get("seller")
        region_idx = idx.get("region")
        price_idx = idx.get("price_kzt")
        review_idx = idx.get("review_count")
        scraped_idx = idx.get("scraped_date")

        city_norm = str(city or "").strip().lower()
        by_product: dict[str, dict] = {}

        for row in rows_iter:
            region_value = str(row[region_idx] or "").strip() if region_idx is not None and region_idx < len(row) else ""
            if city_norm and region_value and region_value.lower() != city_norm:
                continue

            product_name = ""
            if product_idx is not None and product_idx < len(row):
                product_name = str(row[product_idx] or "").strip()
            if not product_name and input_product_idx is not None and input_product_idx < len(row):
                product_name = str(row[input_product_idx] or "").strip()

            product_url = str(row[url_idx] or "").strip() if url_idx is not None and url_idx < len(row) else ""
            product_code = str(row[code_idx] or "").strip() if code_idx is not None and code_idx < len(row) else ""
            if not product_code:
                product_code = _extract_product_code_from_url(product_url)

            key = product_code or product_url or _normalize_search_text(product_name)
            if not key:
                continue

            seller_name = str(row[seller_idx] or "").strip() if seller_idx is not None and seller_idx < len(row) else ""
            price_value = _int_or_none(row[price_idx]) if price_idx is not None and price_idx < len(row) else None
            review_value = _int_or_none(row[review_idx]) if review_idx is not None and review_idx < len(row) else None
            scraped_value = str(row[scraped_idx] or "").strip() if scraped_idx is not None and scraped_idx < len(row) else ""

            slot = by_product.setdefault(
                key,
                {
                    "product_key": key,
                    "product_code": product_code,
                    "product": product_name,
                    "product_url": product_url,
                    "region": region_value,
                    "best_price_kzt": None,
                    "best_seller": "",
                    "review_count": None,
                    "offers_count": 0,
                    "sellers": set(),
                    "scraped_date": scraped_value,
                },
            )
            if product_name and not slot.get("product"):
                slot["product"] = product_name
            if product_url and not slot.get("product_url"):
                slot["product_url"] = product_url
            if product_code and not slot.get("product_code"):
                slot["product_code"] = product_code
            if region_value and not slot.get("region"):
                slot["region"] = region_value
            if scraped_value and not slot.get("scraped_date"):
                slot["scraped_date"] = scraped_value

            slot["offers_count"] += 1
            if seller_name:
                slot["sellers"].add(seller_name)
            if price_value is not None and (slot.get("best_price_kzt") is None or price_value < slot["best_price_kzt"]):
                slot["best_price_kzt"] = price_value
                slot["best_seller"] = seller_name
            if review_value is not None and (slot.get("review_count") is None or review_value > slot["review_count"]):
                slot["review_count"] = review_value

        return by_product
    finally:
        wb.close()


def _competitors_summary_payload(city: str = "Алматы", limit: int = 260) -> dict | None:
    candidates = _competitor_report_candidates()
    if not candidates:
        return None

    path = candidates[0]
    try:
        mtime = path.stat().st_mtime
    except OSError:
        return None

    cached_payload = COMPETITOR_CACHE.get("payload")
    if (
        cached_payload
        and COMPETITOR_CACHE.get("path") == str(path)
        and COMPETITOR_CACHE.get("mtime") == mtime
        and COMPETITOR_CACHE.get("city") == city
        and COMPETITOR_CACHE.get("limit") == int(limit)
    ):
        return cached_payload

    current = _read_competitor_product_map(path, city=city)
    if not current:
        payload = {
            "rows": [],
            "total": 0,
            "priced_products": 0,
            "reviewed_products": 0,
            "offers_count": 0,
            "city": city,
            "updated_at": datetime.fromtimestamp(mtime, tz=ALMATY_TZ).isoformat(timespec="seconds"),
            "updated_human": datetime.fromtimestamp(mtime, tz=ALMATY_TZ).strftime("%d.%m %H:%M"),
            "source_file": path.name,
            "previous_file": candidates[1].name if len(candidates) > 1 else "",
        }
        COMPETITOR_CACHE.update({"path": str(path), "mtime": mtime, "city": city, "limit": int(limit), "payload": payload})
        return payload

    previous = _read_competitor_product_map(candidates[1], city=city) if len(candidates) > 1 else {}

    rows: list[dict] = []
    for key, slot in current.items():
        prev_slot = previous.get(key, {})
        best_price = slot.get("best_price_kzt")
        prev_price = prev_slot.get("best_price_kzt")
        review_count = slot.get("review_count")
        prev_reviews = prev_slot.get("review_count")
        rows.append(
            {
                "product_key": key,
                "product_code": slot.get("product_code") or "",
                "product": slot.get("product") or "",
                "product_url": slot.get("product_url") or "",
                "region": slot.get("region") or city,
                "best_price_kzt": best_price,
                "best_seller": slot.get("best_seller") or "",
                "prev_best_price_kzt": prev_price,
                "price_delta_kzt": (best_price - prev_price) if best_price is not None and prev_price is not None else None,
                "review_count": review_count,
                "prev_review_count": prev_reviews,
                "review_delta": (review_count - prev_reviews) if review_count is not None and prev_reviews is not None else None,
                "offers_count": int(slot.get("offers_count") or 0),
                "sellers_count": len(slot.get("sellers") or []),
                "scraped_date": slot.get("scraped_date") or "",
            }
        )

    rows.sort(
        key=lambda row: (
            row.get("best_price_kzt") is None,
            row.get("best_price_kzt") if row.get("best_price_kzt") is not None else 10**12,
            -(row.get("review_count") or -1),
            str(row.get("product") or "").lower(),
        )
    )
    total_products = len(rows)
    visible_rows = rows[: max(1, limit)]
    for row in visible_rows:
        row.pop("product_key", None)

    payload = {
        "rows": visible_rows,
        "total": total_products,
        "priced_products": sum(1 for row in rows if row.get("best_price_kzt") is not None),
        "reviewed_products": sum(1 for row in rows if row.get("review_count") is not None),
        "offers_count": sum(int(row.get("offers_count") or 0) for row in rows),
        "city": city,
        "updated_at": datetime.fromtimestamp(mtime, tz=ALMATY_TZ).isoformat(timespec="seconds"),
        "updated_human": datetime.fromtimestamp(mtime, tz=ALMATY_TZ).strftime("%d.%m %H:%M"),
        "source_file": path.name,
        "previous_file": candidates[1].name if len(candidates) > 1 else "",
    }
    COMPETITOR_CACHE.update({"path": str(path), "mtime": mtime, "city": city, "limit": int(limit), "payload": payload})
    return payload


def _competitors_status_payload() -> dict:
    payload = {"status": "idle", "last_run": "", "last_error": "", "last_file": ""}
    with COMPETITOR_LOCK:
        payload.update(COMPETITOR_STATE)
    latest = _latest_report(_competitor_prefixes())
    if latest:
        payload["latest_file"] = latest.name
        payload["latest_mtime"] = datetime.fromtimestamp(latest.stat().st_mtime, tz=ALMATY_TZ).isoformat(timespec="seconds")
    return payload


def _wolt_results_dirs() -> list[Path]:
    dirs: list[Path] = []
    seen: set[str] = set()
    for path in [WOLT_RESULTS_DIR, RESULTS_DIR]:
        key = str(path)
        if key in seen:
            continue
        seen.add(key)
        dirs.append(path)
    return dirs


def _wolt_stamp_sort_key(path: Path, pattern: re.Pattern[str]) -> tuple[str, float]:
    match = pattern.match(path.name)
    stamp = match.group("stamp") if match else ""
    try:
        mtime = path.stat().st_mtime
    except OSError:
        mtime = 0.0
    return stamp, mtime


def _wolt_report_candidates() -> list[Path]:
    by_name: dict[str, Path] = {}
    for base_dir in _wolt_results_dirs():
        if not base_dir.exists():
            continue
        for path in base_dir.glob("wolt_brand_*_items.csv"):
            if not path.is_file():
                continue
            prev = by_name.get(path.name)
            if prev is None or _wolt_stamp_sort_key(path, WOLT_ITEMS_PATTERN) > _wolt_stamp_sort_key(prev, WOLT_ITEMS_PATTERN):
                by_name[path.name] = path
    candidates = list(by_name.values())
    candidates.sort(key=lambda p: _wolt_stamp_sort_key(p, WOLT_ITEMS_PATTERN), reverse=True)
    return candidates


def _wolt_latest_report() -> Path | None:
    candidates = _wolt_report_candidates()
    return candidates[0] if candidates else None


def _wolt_pharmacies_candidates() -> list[Path]:
    by_name: dict[str, Path] = {}
    for base_dir in _wolt_results_dirs():
        if not base_dir.exists():
            continue
        for path in base_dir.glob("wolt_almaty_pharmacies_*.csv"):
            if not path.is_file() or not WOLT_PHARMACIES_PATTERN.match(path.name):
                continue
            prev = by_name.get(path.name)
            if prev is None or _wolt_stamp_sort_key(path, WOLT_PHARMACIES_PATTERN) > _wolt_stamp_sort_key(prev, WOLT_PHARMACIES_PATTERN):
                by_name[path.name] = path
    candidates = list(by_name.values())
    candidates.sort(key=lambda p: _wolt_stamp_sort_key(p, WOLT_PHARMACIES_PATTERN), reverse=True)
    return candidates


def _wolt_assortment_gap_candidates() -> list[Path]:
    by_name: dict[str, Path] = {}
    for base_dir in _wolt_results_dirs():
        if not base_dir.exists():
            continue
        for path in base_dir.glob("wolt_almaty_assortment_gap_*.json"):
            if not path.is_file() or not WOLT_ASSORTMENT_GAP_PATTERN.match(path.name):
                continue
            prev = by_name.get(path.name)
            if prev is None or _wolt_stamp_sort_key(path, WOLT_ASSORTMENT_GAP_PATTERN) > _wolt_stamp_sort_key(prev, WOLT_ASSORTMENT_GAP_PATTERN):
                by_name[path.name] = path
    candidates = list(by_name.values())
    candidates.sort(key=lambda p: _wolt_stamp_sort_key(p, WOLT_ASSORTMENT_GAP_PATTERN), reverse=True)
    return candidates


def _wolt_latest_assortment_gap() -> Path | None:
    candidates = _wolt_assortment_gap_candidates()
    return candidates[0] if candidates else None


def _safe_int(value):
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def _safe_float(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except Exception:
        return None


def _campaign_quality_summary_candidates() -> list[Path]:
    if not RESULTS_DIR.exists():
        return []
    candidates = [
        path
        for path in RESULTS_DIR.glob("mission_campaign_quality_*_summary.json")
        if path.is_file() and CAMPAIGN_QUALITY_SUMMARY_PATTERN.match(path.name)
    ]
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    return candidates


def _campaign_quality_related_path(summary_path: Path, suffix: str) -> Path:
    stem = summary_path.name[: -len("_summary.json")]
    return summary_path.with_name(f"{stem}_{suffix}")


def _campaign_quality_base_meta(summary_path: Path) -> tuple[float, dict] | tuple[None, None]:
    if not summary_path.exists() or not summary_path.is_file():
        return (None, None)
    try:
        mtime = summary_path.stat().st_mtime
    except OSError:
        return (None, None)
    try:
        with open(summary_path, "r", encoding="utf-8") as fh:
            summary = json.load(fh)
    except Exception:
        return (None, None)
    if not isinstance(summary, dict):
        return (None, None)
    return (mtime, summary)


def _campaign_quality_rows_payload(path: Path) -> list[dict]:
    rows: list[dict] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            seller = _canonical_seller_name(row.get("seller") or "")
            region = str(row.get("region") or "").strip()
            input_product = html.unescape(str(row.get("input_product") or row.get("product") or "").strip())
            product = html.unescape(str(row.get("product") or input_product).strip())
            product_url = str(row.get("product_url") or "").strip()
            product_code = str(row.get("product_code") or "").strip()
            if not product_code:
                product_code = _extract_product_code_from_url(product_url)
            rows.append(
                {
                    "row_key": str(row.get("row_key") or "").strip(),
                    "region": region,
                    "seller": seller,
                    "input_product": input_product,
                    "product": product,
                    "product_url": product_url,
                    "product_code": product_code,
                    "price_kzt": _safe_int(row.get("price_kzt")),
                    "observed_hours": _safe_float(row.get("observed_hours")) or 0.0,
                    "live_hours": _safe_float(row.get("live_hours")) or 0.0,
                    "market_hours": _safe_float(row.get("market_hours")) or 0.0,
                    "correct_price_hours": _safe_float(row.get("correct_price_hours")) or 0.0,
                    "best_price_hours": _safe_float(row.get("best_price_hours")) or 0.0,
                    "attention_hours": _safe_float(row.get("attention_hours")) or 0.0,
                    "problem_hours": _safe_float(row.get("problem_hours")) or 0.0,
                    "snapshots": int(_safe_int(row.get("snapshots")) or 0),
                    "live_snapshots": int(_safe_int(row.get("live_snapshots")) or 0),
                    "market_snapshots": int(_safe_int(row.get("market_snapshots")) or 0),
                    "correct_price_snapshots": int(_safe_int(row.get("correct_price_snapshots")) or 0),
                    "best_price_snapshots": int(_safe_int(row.get("best_price_snapshots")) or 0),
                    "availability_pct_observed": _safe_float(row.get("availability_pct_observed")) or 0.0,
                    "correct_price_pct_observed": _safe_float(row.get("correct_price_pct_observed")) or 0.0,
                    "correct_price_pct_live": _safe_float(row.get("correct_price_pct_live")) or 0.0,
                    "best_price_pct_observed": _safe_float(row.get("best_price_pct_observed")) or 0.0,
                    "best_price_pct_live": _safe_float(row.get("best_price_pct_live")) or 0.0,
                    "best_price_pct_market": _safe_float(row.get("best_price_pct_market")) or 0.0,
                    "attention_pct_observed": _safe_float(row.get("attention_pct_observed")) or 0.0,
                    "problem_pct_observed": _safe_float(row.get("problem_pct_observed")) or 0.0,
                    "first_seen_snapshot": str(row.get("first_seen_snapshot") or "").strip(),
                    "last_seen_snapshot": str(row.get("last_seen_snapshot") or "").strip(),
                }
            )
    return rows


def _campaign_quality_payload() -> dict | None:
    candidates = _campaign_quality_summary_candidates()
    for path in candidates[:20]:
        mtime, summary = _campaign_quality_base_meta(path)
        if mtime is None or summary is None:
            continue
        cached = CAMPAIGN_QUALITY_CACHE.get("payload")
        if CAMPAIGN_QUALITY_CACHE.get("path") == str(path) and CAMPAIGN_QUALITY_CACHE.get("mtime") == mtime and cached:
            return cached

        per_row_path = _campaign_quality_related_path(path, "per_row.csv")
        if not per_row_path.exists() or not per_row_path.is_file():
            continue
        try:
            rows = _campaign_quality_rows_payload(per_row_path)
        except Exception:
            continue

        payload = {
            "summary": summary,
            "rows": rows,
            "total_rows": len(rows),
            "source_file": path.name,
            "updated_at": datetime.fromtimestamp(mtime).isoformat(timespec="seconds"),
            "updated_human": datetime.fromtimestamp(mtime).strftime("%d.%m %H:%M"),
            "files": {
                "summary": path.name,
                "per_row": per_row_path.name,
                "per_seller": _campaign_quality_related_path(path, "per_seller.csv").name,
                "per_region": _campaign_quality_related_path(path, "per_region.csv").name,
                "per_seller_region": _campaign_quality_related_path(path, "per_seller_region.csv").name,
                "per_sku": _campaign_quality_related_path(path, "per_sku.csv").name,
            },
        }
        CAMPAIGN_QUALITY_CACHE.update({"path": str(path), "mtime": mtime, "payload": payload})
        return payload
    return None


def _price_leaders_summary_payload() -> dict | None:
    path = PRICE_LEADERS_LIVE_JSON
    if not path.exists() or not path.is_file():
        return None
    try:
        mtime = path.stat().st_mtime
    except OSError:
        return None
    cached = PRICE_LEADERS_CACHE.get("payload")
    if PRICE_LEADERS_CACHE.get("path") == str(path) and PRICE_LEADERS_CACHE.get("mtime") == mtime and cached:
        return cached
    try:
        with open(path, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    payload["source_file"] = path.name
    PRICE_LEADERS_CACHE.update({"path": str(path), "mtime": mtime, "payload": payload})
    return payload


def _parse_bool(value) -> bool:
    raw = str(value or "").strip().lower()
    return raw in {"1", "true", "yes", "y", "on"}


def _normalize_search_text(value: str) -> str:
    return re.sub(r"[^a-z0-9а-яё]+", " ", str(value or "").lower()).strip()


def _extract_price_leaders_brand(product_name: str) -> str:
    text = " ".join(str(product_name or "").strip().split())
    if not text:
        return ""
    tokens = text.split()
    tokens_lower = [tok.lower() for tok in tokens]
    if len(tokens_lower) >= 2 and tokens_lower[0] == "new" and tokens_lower[1] == "nordic":
        return "NEW NORDIC"
    if tokens_lower and tokens_lower[0] == "stada" and len(tokens) >= 2:
        return tokens[1]
    return tokens[0]


def _price_leaders_match_filter(value: str, filter_text: str, *, exact: bool = False) -> bool:
    if not filter_text:
        return True
    value_norm = _normalize_search_text(value)
    filter_norm = _normalize_search_text(filter_text)
    if not filter_norm:
        return True
    if exact:
        return value_norm == filter_norm
    return filter_norm in value_norm


def _price_leaders_row_match(
    row: dict,
    *,
    pharmacy: str,
    brand: str,
    search: str,
) -> bool:
    if pharmacy and not _price_leaders_match_filter(str(row.get("seller") or ""), pharmacy, exact=True):
        return False
    if brand:
        brand_text = str(row.get("brand") or "")
        product_text = str(row.get("product_name") or "")
        if not (
            _price_leaders_match_filter(brand_text, brand)
            or _price_leaders_match_filter(product_text, brand)
        ):
            return False
    if search:
        hay = " ".join(
            [
                str(row.get("product_name") or ""),
                str(row.get("product_code") or ""),
                str(row.get("seller") or ""),
                str(row.get("brand") or ""),
            ]
        )
        if not _price_leaders_match_filter(hay, search):
            return False
    return True


def _price_leaders_csv_path(payload: dict | None = None) -> Path | None:
    payload = payload or _price_leaders_summary_payload()
    if not payload:
        return None
    candidates: list[Path] = []
    for key in ("details_file_latest", "details_file"):
        raw_name = str(payload.get(key) or "").strip()
        if raw_name:
            candidates.append(RESULTS_DIR / Path(raw_name).name)
    candidates.append(RESULTS_DIR / "kaspi_price_leaders_live_latest.csv")
    for path in candidates:
        if path.exists() and path.is_file():
            return path
    return None


def _price_leaders_rows_payload() -> list[dict]:
    summary = _price_leaders_summary_payload()
    path = _price_leaders_csv_path(summary)
    if not path:
        return []
    try:
        mtime = path.stat().st_mtime
    except OSError:
        return []
    cached_rows = PRICE_LEADERS_ROWS_CACHE.get("rows")
    if (
        PRICE_LEADERS_ROWS_CACHE.get("path") == str(path)
        and PRICE_LEADERS_ROWS_CACHE.get("mtime") == mtime
        and isinstance(cached_rows, list)
    ):
        return cached_rows

    rows: list[dict] = []
    try:
        with open(path, "r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                month = str(row.get("month") or "").strip()
                seller = str(row.get("seller") or "").strip()
                if not month or not seller:
                    continue
                rows.append(
                    {
                        "month": month,
                        "region": str(row.get("region") or "").strip(),
                        "product_code": str(row.get("product_code") or "").strip(),
                        "product_name": str(row.get("product_name") or "").strip(),
                        "seller": seller,
                        "wins": int(_safe_int(row.get("wins")) or 0),
                        "product_snapshots": int(_safe_int(row.get("product_snapshots")) or 0),
                        "win_share_pct": float(_safe_float(row.get("win_share_pct")) or 0.0),
                        "seller_presence_snapshots": int(_safe_int(row.get("seller_presence_snapshots")) or 0),
                        "seller_presence_share_pct": float(_safe_float(row.get("seller_presence_share_pct")) or 0.0),
                        "avg_sellers_per_snapshot": float(_safe_float(row.get("avg_sellers_per_snapshot")) or 0.0),
                        "seller_min_price_kzt": _safe_int(row.get("seller_min_price_kzt")),
                        "seller_max_price_kzt": _safe_int(row.get("seller_max_price_kzt")),
                        "is_qualified": _parse_bool(row.get("is_qualified")),
                    }
                )
                rows[-1]["brand"] = _extract_price_leaders_brand(rows[-1]["product_name"])
    except Exception:
        return []

    PRICE_LEADERS_ROWS_CACHE.update({"path": str(path), "mtime": mtime, "rows": rows})
    return rows


def _price_leaders_sort_key(row: dict) -> tuple:
    return (
        -float(row.get("win_share_pct") or 0.0),
        -int(row.get("wins") or 0),
        -int(row.get("product_snapshots") or 0),
        -float(row.get("avg_sellers_per_snapshot") or 0.0),
        str(row.get("seller") or ""),
        str(row.get("product_name") or ""),
    )


def _price_leaders_all_time_rows(summary: dict, rows: list[dict]) -> tuple[list[dict], dict]:
    cache_key = (str(summary.get("generated_at") or ""), len(rows))
    cached_rows = PRICE_LEADERS_ALL_TIME_CACHE.get("rows")
    cached_meta = PRICE_LEADERS_ALL_TIME_CACHE.get("meta")
    if PRICE_LEADERS_ALL_TIME_CACHE.get("key") == cache_key and isinstance(cached_rows, list) and isinstance(cached_meta, dict):
        return cached_rows, cached_meta

    filters = summary.get("filters") or {}
    min_product_snapshots = int(_safe_int(filters.get("min_product_snapshots")) or 0)
    min_avg_sellers = float(_safe_float(filters.get("min_avg_sellers")) or 0.0)
    grouped: dict[tuple[str, str], dict] = {}
    product_keys: set[str] = set()

    for row in rows:
        product_code = str(row.get("product_code") or "").strip()
        product_name = str(row.get("product_name") or "").strip()
        seller = str(row.get("seller") or "").strip()
        if not seller or (not product_code and not product_name):
            continue
        product_key = product_code or f"name::{_normalize_search_text(product_name)}"
        key = (product_key, seller)
        slot = grouped.setdefault(
            key,
            {
                "month": "all",
                "region": str(row.get("region") or summary.get("region") or "").strip(),
                "product_code": product_code,
                "product_name": product_name,
                "seller": seller,
                "wins": 0,
                "product_snapshots": 0,
                "seller_presence_snapshots": 0,
                "avg_sellers_weighted_sum": 0.0,
                "avg_sellers_weighted_den": 0.0,
                "seller_min_price_kzt": None,
                "seller_max_price_kzt": None,
                "brand": str(row.get("brand") or "").strip(),
            },
        )
        slot["wins"] += int(row.get("wins") or 0)
        slot["product_snapshots"] += int(row.get("product_snapshots") or 0)
        slot["seller_presence_snapshots"] += int(row.get("seller_presence_snapshots") or 0)
        row_snapshots = int(row.get("product_snapshots") or 0)
        slot["avg_sellers_weighted_sum"] += float(row.get("avg_sellers_per_snapshot") or 0.0) * row_snapshots
        slot["avg_sellers_weighted_den"] += row_snapshots

        row_min = _safe_int(row.get("seller_min_price_kzt"))
        row_max = _safe_int(row.get("seller_max_price_kzt"))
        if row_min is not None:
            slot["seller_min_price_kzt"] = row_min if slot["seller_min_price_kzt"] is None else min(int(slot["seller_min_price_kzt"]), row_min)
        if row_max is not None:
            slot["seller_max_price_kzt"] = row_max if slot["seller_max_price_kzt"] is None else max(int(slot["seller_max_price_kzt"]), row_max)
        if product_code and not slot["product_code"]:
            slot["product_code"] = product_code
        if product_name and not slot["product_name"]:
            slot["product_name"] = product_name
        if not slot["brand"]:
            slot["brand"] = str(row.get("brand") or "").strip()
        product_keys.add(product_key)

    all_rows: list[dict] = []
    for slot in grouped.values():
        product_snapshots = int(slot.get("product_snapshots") or 0)
        if product_snapshots <= 0:
            continue
        wins = int(slot.get("wins") or 0)
        presence = int(slot.get("seller_presence_snapshots") or 0)
        avg_den = float(slot.get("avg_sellers_weighted_den") or 0.0)
        avg_sellers = round((float(slot.get("avg_sellers_weighted_sum") or 0.0) / avg_den), 2) if avg_den else 0.0
        all_rows.append(
            {
                "month": "all",
                "region": slot.get("region") or "",
                "product_code": slot.get("product_code") or "",
                "product_name": slot.get("product_name") or "",
                "seller": slot.get("seller") or "",
                "wins": wins,
                "product_snapshots": product_snapshots,
                "win_share_pct": round((wins / product_snapshots) * 100, 2),
                "seller_presence_snapshots": presence,
                "seller_presence_share_pct": round((presence / product_snapshots) * 100, 2),
                "avg_sellers_per_snapshot": avg_sellers,
                "seller_min_price_kzt": _safe_int(slot.get("seller_min_price_kzt")),
                "seller_max_price_kzt": _safe_int(slot.get("seller_max_price_kzt")),
                "is_qualified": bool(product_snapshots >= min_product_snapshots and avg_sellers >= min_avg_sellers),
                "brand": slot.get("brand") or "",
            }
        )

    all_rows.sort(key=_price_leaders_sort_key)
    month_meta = {
        "month": "all",
        "snapshot_files": int(_safe_int(summary.get("files_total")) or 0),
        "products_count": len(product_keys),
        "pairs_count": len(all_rows),
        "qualified_pairs_count": sum(1 for row in all_rows if row.get("is_qualified")),
        "latest_snapshot_dt": str(summary.get("latest_snapshot_dt") or ""),
    }
    PRICE_LEADERS_ALL_TIME_CACHE.update({"key": cache_key, "rows": all_rows, "meta": month_meta})
    return all_rows, month_meta


def _price_leaders_month_context(
    month: str | None,
) -> tuple[dict | None, list[dict], str, list[str], dict]:
    summary = _price_leaders_summary_payload()
    rows = _price_leaders_rows_payload()
    if not summary or not rows:
        return None, [], "", [], {}

    months = [str(item.get("month") or "").strip() for item in (summary.get("months") or []) if item.get("month")]
    months = [item for item in months if item]
    if not months:
        months = sorted({str(row.get("month") or "").strip() for row in rows if row.get("month")}, reverse=True)
    months_with_all = ["all", *months] if months else ["all"]

    selected_month = str(month or summary.get("selected_month") or "").strip()
    if not selected_month:
        selected_month = "all"
    if selected_month == "all":
        month_rows_all, month_meta = _price_leaders_all_time_rows(summary, rows)
        month_rows_qualified = [row for row in month_rows_all if row.get("is_qualified")]
        month_rows = month_rows_qualified or month_rows_all
        return summary, month_rows, "all", months_with_all, month_meta
    if selected_month not in months and months:
        selected_month = months[0]

    month_meta = next(
        (item for item in (summary.get("months") or []) if str(item.get("month") or "").strip() == selected_month),
        None,
    )

    month_rows_all = [dict(row) for row in rows if str(row.get("month") or "").strip() == selected_month]
    month_rows_qualified = [row for row in month_rows_all if row.get("is_qualified")]
    month_rows = month_rows_qualified or month_rows_all
    return summary, month_rows, selected_month, months_with_all, month_meta or {}


def _price_leaders_filters_payload(
    month: str | None,
    *,
    pharmacy: str | None = None,
    brand: str | None = None,
    search: str | None = None,
    sku_query: str | None = None,
    q: str | None = None,
    limit: int = 20,
) -> dict | None:
    summary, month_rows, selected_month, months, month_meta = _price_leaders_month_context(month)
    if not summary:
        return None

    pharmacy = str(pharmacy or "").strip()
    brand = str(brand or "").strip()
    search = str(search or "").strip()
    sku_query = str(sku_query or "").strip()
    sku_query_norm = _normalize_search_text(sku_query)
    q = str(q or "").strip()
    q_norm = _normalize_search_text(q)

    rows_for_sellers = [
        row for row in month_rows if _price_leaders_row_match(row, pharmacy="", brand=brand, search=search)
    ]
    rows_for_brands = [
        row for row in month_rows if _price_leaders_row_match(row, pharmacy=pharmacy, brand="", search=search)
    ]
    rows_for_queries = [
        row for row in month_rows if _price_leaders_row_match(row, pharmacy=pharmacy, brand=brand, search="")
    ]

    sellers_counter: Counter[str] = Counter()
    for row in rows_for_sellers:
        seller = str(row.get("seller") or "").strip()
        if seller:
            sellers_counter[seller] += 1
    sellers = [
        {"value": name, "count": int(cnt)}
        for name, cnt in sellers_counter.most_common(max(1, limit))
    ]

    brands_counter: Counter[str] = Counter()
    for row in rows_for_brands:
        brand_name = str(row.get("brand") or "").strip()
        if brand_name:
            brands_counter[brand_name] += 1
    brands = [
        {"value": name, "count": int(cnt)}
        for name, cnt in brands_counter.most_common(max(1, limit))
    ]

    sku_counter: Counter[str] = Counter()
    sku_name_by_code: dict[str, str] = {}
    for row in rows_for_queries:
        product_code = str(row.get("product_code") or "").strip()
        if not product_code:
            continue
        code_norm = _normalize_search_text(product_code)
        if sku_query_norm and sku_query_norm not in code_norm:
            continue
        sku_counter[product_code] += 1
        if product_code not in sku_name_by_code:
            sku_name_by_code[product_code] = str(row.get("product_name") or "").strip()
    sku_suggestions = [
        {"value": code, "product_name": sku_name_by_code.get(code, ""), "count": int(cnt)}
        for code, cnt in sku_counter.most_common(max(1, limit * 2))
    ]

    query_items: list[dict] = []
    for row in rows_for_queries:
        product_code = str(row.get("product_code") or "").strip()
        product_name = str(row.get("product_name") or "").strip()
        seller = str(row.get("seller") or "").strip()
        brand_name = str(row.get("brand") or "").strip()
        for value, kind in (
            (product_code, "sku"),
            (product_name, "product"),
            (seller, "seller"),
            (brand_name, "brand"),
        ):
            if not value:
                continue
            norm = _normalize_search_text(value)
            if q_norm and q_norm not in norm:
                continue
            query_items.append({"value": value, "kind": kind})

    query_counter: Counter[tuple[str, str]] = Counter((item["value"], item["kind"]) for item in query_items)
    query_suggestions = [
        {"value": value, "kind": kind, "count": int(cnt)}
        for (value, kind), cnt in query_counter.most_common(max(1, limit * 2))
    ]

    return {
        "generated_at": summary.get("generated_at") or "",
        "generated_at_human": summary.get("generated_at_human") or "",
        "region": summary.get("region") or "",
        "month": selected_month,
        "months": months,
        "month_meta": month_meta,
        "filters": {
            "pharmacy": pharmacy,
            "brand": brand,
            "search": search,
            "sku_query": sku_query,
            "q": q,
        },
        "sellers": sellers,
        "brands": brands,
        "sku_suggestions": sku_suggestions,
        "query_suggestions": query_suggestions,
    }


def _price_leaders_ranking_payload(
    month: str | None,
    level: str,
    query: str | None,
    pharmacy: str | None,
    brand: str | None,
    search: str | None,
    sku_query: str | None,
    top_n: int,
) -> dict | None:
    summary, month_rows, selected_month, months, month_meta = _price_leaders_month_context(month)
    if not summary:
        return None

    raw_level = str(level or "month").strip().lower()
    if raw_level not in {"month", "sku", "brand"}:
        raw_level = "month"
    raw_query = str(query or "").strip()
    query_norm = _normalize_search_text(raw_query)
    raw_pharmacy = str(pharmacy or "").strip()
    raw_brand = str(brand or "").strip()
    raw_search = str(search or "").strip()
    raw_sku_query = str(sku_query or "").strip()
    filtered_rows = [
        row
        for row in month_rows
        if _price_leaders_row_match(
            row,
            pharmacy=raw_pharmacy,
            brand=raw_brand,
            search=raw_search,
        )
    ]

    payload: dict = {
        "generated_at": summary.get("generated_at") or "",
        "generated_at_human": summary.get("generated_at_human") or "",
        "region": summary.get("region") or "",
        "month": selected_month,
        "months": months,
        "level": raw_level,
        "query": raw_query,
        "top_n": int(top_n),
        "rows": [],
        "total_rows": 0,
        "month_meta": month_meta or {},
        "meta": {},
        "filters": {
            "pharmacy": raw_pharmacy,
            "brand": raw_brand,
            "search": raw_search,
            "sku_query": raw_sku_query,
        },
    }

    if raw_level == "month":
        rows_for_month = filtered_rows
        if raw_sku_query:
            rows_for_month = [
                row
                for row in rows_for_month
                if _price_leaders_match_filter(str(row.get("product_code") or ""), raw_sku_query)
            ]
        grouped: dict[str, dict] = {}
        for row in rows_for_month:
            product_code = str(row.get("product_code") or "").strip()
            product_name = str(row.get("product_name") or "").strip()
            if not product_code and not product_name:
                continue
            key = product_code or f"name::{_normalize_search_text(product_name)}"
            slot = grouped.setdefault(
                key,
                {
                    "product_code": product_code,
                    "product_name": product_name,
                    "brand": str(row.get("brand") or "").strip(),
                    "product_snapshots": 0,
                    "avg_sellers_per_snapshot": 0.0,
                    "_sellers": {},
                },
            )
            if product_code and not slot["product_code"]:
                slot["product_code"] = product_code
            if product_name and not slot["product_name"]:
                slot["product_name"] = product_name
            if not slot["brand"]:
                slot["brand"] = str(row.get("brand") or "").strip()
            slot["product_snapshots"] = max(int(slot["product_snapshots"] or 0), int(row.get("product_snapshots") or 0))
            slot["avg_sellers_per_snapshot"] = max(
                float(slot["avg_sellers_per_snapshot"] or 0.0),
                float(row.get("avg_sellers_per_snapshot") or 0.0),
            )

            seller = str(row.get("seller") or "").strip()
            if not seller:
                continue
            seller_slot = slot["_sellers"].setdefault(
                seller,
                {
                    "seller": seller,
                    "wins": 0,
                    "product_snapshots": 0,
                    "win_share_pct": 0.0,
                    "avg_sellers_per_snapshot": 0.0,
                    "seller_presence_snapshots": 0,
                    "seller_presence_share_pct": 0.0,
                },
            )
            seller_slot["wins"] += int(row.get("wins") or 0)
            seller_slot["product_snapshots"] = max(
                int(seller_slot["product_snapshots"] or 0),
                int(row.get("product_snapshots") or 0),
            )
            seller_slot["seller_presence_snapshots"] = max(
                int(seller_slot["seller_presence_snapshots"] or 0),
                int(row.get("seller_presence_snapshots") or 0),
            )
            seller_slot["avg_sellers_per_snapshot"] = max(
                float(seller_slot["avg_sellers_per_snapshot"] or 0.0),
                float(row.get("avg_sellers_per_snapshot") or 0.0),
            )
            snapshots = int(seller_slot["product_snapshots"] or 0)
            wins = int(seller_slot["wins"] or 0)
            seller_slot["win_share_pct"] = round((wins / snapshots) * 100, 2) if snapshots else float(
                row.get("win_share_pct") or 0.0
            )
            presence = int(seller_slot["seller_presence_snapshots"] or 0)
            seller_slot["seller_presence_share_pct"] = round((presence / snapshots) * 100, 2) if snapshots else float(
                row.get("seller_presence_share_pct") or 0.0
            )

        sku_rows: list[dict] = []
        for slot in grouped.values():
            sellers = sorted(
                slot["_sellers"].values(),
                key=lambda item: (
                    -float(item.get("win_share_pct") or 0.0),
                    -int(item.get("wins") or 0),
                    -int(item.get("product_snapshots") or 0),
                    str(item.get("seller") or ""),
                ),
            )
            if not sellers:
                continue
            top_seller = sellers[0]
            sku_rows.append(
                {
                    "product_code": slot["product_code"],
                    "product_name": slot["product_name"],
                    "brand": slot["brand"],
                    "product_snapshots": int(slot["product_snapshots"] or 0),
                    "avg_sellers_per_snapshot": round(float(slot["avg_sellers_per_snapshot"] or 0.0), 2),
                    "top_seller": top_seller.get("seller") or "",
                    "top_win_share_pct": float(top_seller.get("win_share_pct") or 0.0),
                    "top_wins": int(top_seller.get("wins") or 0),
                    "sellers_count": len(sellers),
                    "sellers": sellers,
                }
            )

        sku_rows.sort(
            key=lambda item: (
                -float(item.get("top_win_share_pct") or 0.0),
                -int(item.get("top_wins") or 0),
                -int(item.get("product_snapshots") or 0),
                str(item.get("product_name") or ""),
            )
        )
        payload["rows"] = sku_rows[:top_n]
        payload["total_rows"] = len(sku_rows)
        payload["meta"] = {
            "mode": "month",
            "filtered_rows": len(rows_for_month),
            "base_rows": len(month_rows),
            "sku_rows": len(sku_rows),
            "seller_rows": sum(len(item.get("sellers") or []) for item in sku_rows),
        }
        return payload

    if raw_level == "sku" and not raw_query:
        payload["meta"] = {"error": "query_required"}
        return payload

    if raw_level == "sku":
        products: dict[str, dict] = {}
        raw_query_lc = raw_query.lower()
        for row in filtered_rows:
            product_code = str(row.get("product_code") or "").strip()
            if not product_code:
                continue
            product_name = str(row.get("product_name") or "").strip()
            slot = products.setdefault(
                product_code,
                {"product_code": product_code, "product_name": product_name, "score": 0.0, "rows": 0},
            )
            slot["rows"] += 1
            if product_name and not slot["product_name"]:
                slot["product_name"] = product_name
            product_name_norm = _normalize_search_text(product_name)
            product_code_norm = _normalize_search_text(product_code)

            if raw_query_lc == product_code.lower():
                slot["score"] += 2200
            if query_norm and query_norm == product_name_norm:
                slot["score"] += 1400
            if query_norm and query_norm in product_code_norm:
                slot["score"] += 900
            if query_norm and query_norm in product_name_norm:
                slot["score"] += 500
            slot["score"] += float(row.get("win_share_pct") or 0.0) * 0.01

        ranked_products = sorted(
            products.values(),
            key=lambda item: (-float(item.get("score") or 0.0), -int(item.get("rows") or 0), str(item.get("product_name") or "")),
        )
        if not ranked_products:
            payload["meta"] = {"mode": "sku", "matched_products": []}
            return payload

        selected = ranked_products[0]
        selected_code = str(selected.get("product_code") or "").strip()
        selected_rows = [row for row in filtered_rows if str(row.get("product_code") or "").strip() == selected_code]
        ranked = sorted(selected_rows, key=_price_leaders_sort_key)

        payload["rows"] = ranked[:top_n]
        payload["total_rows"] = len(ranked)
        payload["meta"] = {
            "mode": "sku",
            "selected_product_code": selected_code,
            "selected_product_name": str(selected.get("product_name") or "").strip(),
            "matched_products": [
                {"product_code": item.get("product_code"), "product_name": item.get("product_name"), "rows": item.get("rows")}
                for item in ranked_products[:30]
            ],
        }
        return payload

    # brand mode
    brand_query = raw_query or raw_brand
    brand_query_norm = _normalize_search_text(brand_query)
    if not brand_query_norm:
        payload["meta"] = {"error": "query_required"}
        return payload
    matched_rows = [
        row
        for row in filtered_rows
        if brand_query_norm in _normalize_search_text(str(row.get("product_name") or ""))
        or brand_query_norm in _normalize_search_text(str(row.get("brand") or ""))
    ]

    grouped: dict[str, dict] = {}
    for row in matched_rows:
        seller = str(row.get("seller") or "").strip()
        if not seller:
            continue
        slot = grouped.setdefault(
            seller,
            {
                "seller": seller,
                "wins": 0,
                "product_snapshots": 0,
                "seller_presence_snapshots": 0,
                "avg_sellers_weighted_sum": 0.0,
                "avg_sellers_weighted_den": 0.0,
                "sku_codes": set(),
                "sample_products": [],
            },
        )
        wins = int(row.get("wins") or 0)
        snapshots = int(row.get("product_snapshots") or 0)
        presence = int(row.get("seller_presence_snapshots") or 0)
        avg_sellers = float(row.get("avg_sellers_per_snapshot") or 0.0)

        slot["wins"] += wins
        slot["product_snapshots"] += snapshots
        slot["seller_presence_snapshots"] += presence
        slot["avg_sellers_weighted_sum"] += avg_sellers * snapshots
        slot["avg_sellers_weighted_den"] += snapshots

        product_code = str(row.get("product_code") or "").strip()
        if product_code:
            slot["sku_codes"].add(product_code)
        product_name = str(row.get("product_name") or "").strip()
        if product_name and product_name not in slot["sample_products"] and len(slot["sample_products"]) < 3:
            slot["sample_products"].append(product_name)

    brand_rows: list[dict] = []
    for slot in grouped.values():
        snapshots = int(slot.get("product_snapshots") or 0)
        if snapshots <= 0:
            continue
        wins = int(slot.get("wins") or 0)
        presence = int(slot.get("seller_presence_snapshots") or 0)
        avg_weighted_den = float(slot.get("avg_sellers_weighted_den") or 0.0)
        avg_sellers = (float(slot.get("avg_sellers_weighted_sum") or 0.0) / avg_weighted_den) if avg_weighted_den else 0.0
        brand_rows.append(
            {
                "month": selected_month,
                "region": summary.get("region") or "",
                "product_code": "",
                "product_name": f"Бренд: {raw_query}",
                "seller": slot["seller"],
                "wins": wins,
                "product_snapshots": snapshots,
                "win_share_pct": round((wins / snapshots) * 100, 2),
                "seller_presence_snapshots": presence,
                "seller_presence_share_pct": round((presence / snapshots) * 100, 2),
                "avg_sellers_per_snapshot": round(avg_sellers, 2),
                "seller_min_price_kzt": None,
                "seller_max_price_kzt": None,
                "is_qualified": True,
                "sku_count": len(slot["sku_codes"]),
                "sample_products": slot["sample_products"],
            }
        )

    brand_rows.sort(
        key=lambda row: (
            -float(row.get("win_share_pct") or 0.0),
            -int(row.get("wins") or 0),
            -int(row.get("sku_count") or 0),
            str(row.get("seller") or ""),
        )
    )
    payload["rows"] = brand_rows[:top_n]
    payload["total_rows"] = len(brand_rows)
    payload["meta"] = {
        "mode": "brand",
        "matched_rows": len(matched_rows),
        "matched_skus": len({str(row.get("product_code") or "").strip() for row in matched_rows if row.get("product_code")}),
        "matched_sellers": len(grouped),
    }
    return payload


def _normalize_text(value: str) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _humanize_brand_slug(slug: str) -> str:
    parts = [part for part in re.split(r"[_\s-]+", slug) if part]
    if not parts:
        return slug
    return " ".join(part.capitalize() for part in parts)


def _is_bad_glovo_name(value: str, slug: str = "") -> bool:
    text = str(value or "").strip()
    norm = _normalize_text(text)
    slug_norm = _normalize_text(str(slug or "").replace("-", " "))
    if not text or text in {"--", "-", "—"}:
        return True
    if norm in {"", slug_norm}:
        return True
    if re.fullmatch(r"\d+\s*%?", text):
        return True
    return False


def _glovo_fallback_name_from_slug(slug: str) -> str:
    slug_key = str(slug or "").strip().lower()
    if slug_key in GLOVO_SLUG_NAME_OVERRIDES:
        return GLOVO_SLUG_NAME_OVERRIDES[slug_key]
    base = re.sub(r"-ala(?:-\d+)?$", "", slug_key)
    parts = [part for part in re.split(r"[_\s-]+", base) if part]
    if not parts:
        return slug_key
    return " ".join(part.capitalize() for part in parts)


def _glovo_best_store_name(slug: str, *candidates: str) -> str:
    for candidate in candidates:
        text = str(candidate or "").strip()
        if not _is_bad_glovo_name(text, slug):
            return text
    return _glovo_fallback_name_from_slug(slug)


def _wolt_load_pharmacy_geo(path: Path) -> dict:
    by_slug: dict[str, dict] = {}
    by_name: dict[str, dict] = {}
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            slug = str(row.get("slug") or "").strip().lower()
            name = str(row.get("name") or "").strip()
            lat = _safe_float(row.get("lat"))
            lon = _safe_float(row.get("lon"))
            if lat is None or lon is None:
                continue
            payload = {
                "lat": lat,
                "lon": lon,
                "name": name,
                "slug": slug,
                "address": str(row.get("address") or "").strip(),
                "city": str(row.get("city") or "").strip(),
                "country": str(row.get("country") or "").strip(),
                "venue_url": str(row.get("venue_url") or "").strip(),
            }
            if slug:
                by_slug[slug] = payload
            name_key = _normalize_text(name)
            if name_key:
                by_name[name_key] = payload
    return {
        "by_slug": by_slug,
        "by_name": by_name,
        "source_file": path.name,
    }


def _wolt_geo_index() -> dict:
    for path in _wolt_pharmacies_candidates():
        try:
            mtime = path.stat().st_mtime
        except OSError:
            continue
        cached = WOLT_GEO_CACHE.get("payload")
        if WOLT_GEO_CACHE.get("path") == str(path) and WOLT_GEO_CACHE.get("mtime") == mtime and cached:
            if cached.get("by_slug"):
                return cached
        try:
            payload = _wolt_load_pharmacy_geo(path)
        except Exception:
            continue
        if payload.get("by_slug"):
            WOLT_GEO_CACHE["path"] = str(path)
            WOLT_GEO_CACHE["mtime"] = mtime
            WOLT_GEO_CACHE["payload"] = payload
            return payload
    return {"by_slug": {}, "by_name": {}, "source_file": ""}


def _parse_wolt_items_report(path: Path, limit_rows: int = WOLT_MAX_ITEMS) -> dict:
    geo_index = _wolt_geo_index()
    geo_by_slug = geo_index.get("by_slug") or {}
    geo_by_name = geo_index.get("by_name") or {}

    rows: list[dict] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for idx, row in enumerate(reader):
            if idx >= limit_rows:
                break
            status = str(row.get("status") or "").strip().lower()
            price_minor = _safe_int(row.get("price_minor"))
            price = _safe_float(row.get("price"))
            pharmacy = str(row.get("pharmacy") or "").strip()
            venue_slug = str(row.get("venue_slug") or "").strip()
            venue_slug_key = venue_slug.lower()
            geo = geo_by_slug.get(venue_slug_key) or geo_by_name.get(_normalize_text(pharmacy))
            rows.append(
                {
                    "checked_at": str(row.get("checked_at") or "").strip(),
                    "brand": str(row.get("brand") or "").strip(),
                    "query": str(row.get("query") or "").strip(),
                    "pharmacy": pharmacy,
                    "venue_slug": venue_slug,
                    "venue_url": str(row.get("venue_url") or "").strip(),
                    "item_id": str(row.get("item_id") or "").strip(),
                    "item_name": str(row.get("item_name") or "").strip(),
                    "status": status,
                    "disable_text": str(row.get("disable_text") or "").strip(),
                    "price_minor": price_minor,
                    "price": price,
                    "product_link": str(row.get("product_link") or "").strip(),
                    "search_link": str(row.get("search_link") or "").strip(),
                    "lat": geo.get("lat") if geo else None,
                    "lon": geo.get("lon") if geo else None,
                    "pharmacy_address": geo.get("address") if geo else "",
                    "pharmacy_city": geo.get("city") if geo else "",
                    "pharmacy_country": geo.get("country") if geo else "",
                }
            )

    if not rows:
        return {
            "rows": [],
            "brand": "",
            "query": "",
            "checked_at": "",
            "status_counts": {},
            "pharmacies_count": 0,
            "item_ids_count": 0,
            "price_stats": {"min": None, "median": None, "max": None},
            "top_offers": [],
            "pharmacy_stats": [],
            "geo_source_file": geo_index.get("source_file") or "",
        }

    status_counts: dict[str, int] = {}
    pharmacy_stats: dict[str, dict] = {}
    unique_items = set()
    in_stock_prices: list[float] = []
    best_offers_map: dict[str, dict] = {}

    for row in rows:
        status = row.get("status") or "unknown"
        status_counts[status] = status_counts.get(status, 0) + 1
        unique_items.add(row.get("item_id") or "")

        pharmacy_key = (row.get("venue_slug") or row.get("pharmacy") or "").strip().lower()
        bucket = pharmacy_stats.setdefault(
            pharmacy_key,
            {
                "pharmacy": row.get("pharmacy") or "",
                "venue_slug": row.get("venue_slug") or "",
                "venue_url": row.get("venue_url") or "",
                "total": 0,
                "in_stock": 0,
                "out_of_stock": 0,
                "unavailable": 0,
                "not_found": 0,
                "best_price": None,
                "lat": row.get("lat"),
                "lon": row.get("lon"),
                "pharmacy_address": row.get("pharmacy_address") or "",
                "pharmacy_city": row.get("pharmacy_city") or "",
                "pharmacy_country": row.get("pharmacy_country") or "",
            },
        )
        bucket["total"] += 1
        if status in bucket:
            bucket[status] += 1
        if bucket.get("lat") is None and row.get("lat") is not None:
            bucket["lat"] = row.get("lat")
            bucket["lon"] = row.get("lon")
            bucket["pharmacy_address"] = row.get("pharmacy_address") or bucket.get("pharmacy_address") or ""
            bucket["pharmacy_city"] = row.get("pharmacy_city") or bucket.get("pharmacy_city") or ""
            bucket["pharmacy_country"] = row.get("pharmacy_country") or bucket.get("pharmacy_country") or ""

        price_value = row.get("price")
        if status == "in_stock" and price_value is not None:
            in_stock_prices.append(price_value)
            if bucket["best_price"] is None or price_value < bucket["best_price"]:
                bucket["best_price"] = price_value

            item_key = _normalize_text(row.get("item_name") or row.get("item_id") or "")
            prev = best_offers_map.get(item_key)
            if prev is None or (prev.get("price") is None) or price_value < prev.get("price"):
                best_offers_map[item_key] = row

    top_offers = sorted(
        best_offers_map.values(),
        key=lambda x: (x.get("price") if x.get("price") is not None else 10**12, _normalize_text(x.get("item_name") or "")),
    )[:24]
    pharmacy_rows = sorted(
        pharmacy_stats.values(),
        key=lambda x: (x.get("in_stock", 0), -(x.get("out_of_stock", 0))),
        reverse=True,
    )[:80]

    median_price = round(statistics.median(in_stock_prices), 2) if in_stock_prices else None
    return {
        "rows": rows,
        "brand": rows[0].get("brand") or "",
        "query": rows[0].get("query") or "",
        "checked_at": rows[0].get("checked_at") or "",
        "status_counts": status_counts,
        "pharmacies_count": len({(r.get("venue_slug") or r.get("pharmacy") or "").strip().lower() for r in rows if (r.get("venue_slug") or r.get("pharmacy"))}),
        "item_ids_count": len({i for i in unique_items if i}),
        "price_stats": {
            "min": round(min(in_stock_prices), 2) if in_stock_prices else None,
            "median": median_price,
            "max": round(max(in_stock_prices), 2) if in_stock_prices else None,
        },
        "top_offers": top_offers,
        "pharmacy_stats": pharmacy_rows,
        "geo_source_file": geo_index.get("source_file") or "",
    }


def _wolt_assortment_gap_payload(path: Path | None = None) -> dict | None:
    target = path or _wolt_latest_assortment_gap()
    if not target or not target.exists() or not target.is_file():
        return None
    try:
        mtime = target.stat().st_mtime
    except OSError:
        return None
    if (
        WOLT_GAP_CACHE.get("path") == str(target)
        and WOLT_GAP_CACHE.get("mtime") == mtime
        and WOLT_GAP_CACHE.get("payload") is not None
    ):
        return WOLT_GAP_CACHE["payload"]
    try:
        payload = json.loads(target.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(payload, dict):
        return None
    payload["report_file"] = target.name
    WOLT_GAP_CACHE["path"] = str(target)
    WOLT_GAP_CACHE["mtime"] = mtime
    WOLT_GAP_CACHE["payload"] = payload
    return payload


def _wolt_reports_payload() -> dict:
    reports = []
    per_brand: dict[str, dict] = {}
    for path in _wolt_report_candidates()[: max(1, WOLT_MAX_REPORTS)]:
        mtime = datetime.fromtimestamp(path.stat().st_mtime, tz=ALMATY_TZ).isoformat(timespec="seconds")
        match = WOLT_ITEMS_PATTERN.match(path.name)
        if match:
            brand_slug = match.group("brand")
            stamp = match.group("stamp")
            brand = _humanize_brand_slug(brand_slug)
        else:
            brand_slug = "unknown"
            brand = "Unknown"
            stamp = ""
        row = {
            "file": path.name,
            "brand": brand,
            "brand_slug": brand_slug,
            "stamp": stamp,
            "mtime": mtime,
        }
        reports.append(row)
        slot = per_brand.get(brand_slug)
        if slot is None:
            per_brand[brand_slug] = {
                "brand": brand,
                "brand_slug": brand_slug,
                "latest_file": path.name,
                "latest_stamp": stamp,
                "latest_mtime": mtime,
                "reports_count": 1,
            }
        else:
            slot["reports_count"] = int(slot.get("reports_count") or 0) + 1

    brands = sorted(
        per_brand.values(),
        key=lambda x: str(x.get("latest_mtime") or ""),
        reverse=True,
    )
    return {"reports": reports, "brands": brands}


def _resolve_wolt_report(report: str | None, brand: str | None = None) -> Path | None:
    if report:
        safe_name = Path(report).name
        if not WOLT_ITEMS_PATTERN.match(safe_name):
            return None
        matches: list[Path] = []
        for base_dir in _wolt_results_dirs():
            path = base_dir / safe_name
            if path.exists():
                matches.append(path)
        if matches:
            matches.sort(key=lambda p: _wolt_stamp_sort_key(p, WOLT_ITEMS_PATTERN), reverse=True)
            return matches[0]
        return None

    if brand:
        brand_raw = str(brand or "").strip()
        brand_slug_query = re.sub(r"[^a-z0-9_-]+", "_", brand_raw.lower()).strip("_")
        brand_norm_query = _normalize_text(brand_raw.replace("_", " ").replace("-", " "))
        for path in _wolt_report_candidates():
            match = WOLT_ITEMS_PATTERN.match(path.name)
            if not match:
                continue
            brand_slug = str(match.group("brand") or "").strip().lower()
            brand_human = _humanize_brand_slug(brand_slug)
            brand_norm = _normalize_text(brand_human)
            brand_slug_norm = _normalize_text(brand_slug.replace("_", " ").replace("-", " "))
            if (
                brand_slug == brand_slug_query
                or brand_norm == brand_norm_query
                or brand_slug_norm == brand_norm_query
            ):
                return path
        return None

    return _wolt_latest_report()


def _glovo_results_dirs() -> list[Path]:
    dirs: list[Path] = []
    seen: set[str] = set()
    for path in [GLOVO_RESULTS_DIR, RESULTS_DIR]:
        key = str(path)
        if key in seen:
            continue
        seen.add(key)
        dirs.append(path)
    return dirs


def _glovo_report_candidates() -> list[Path]:
    by_name: dict[str, Path] = {}
    for base_dir in _glovo_results_dirs():
        if not base_dir.exists():
            continue
        for path in base_dir.glob("glovo_brand_*_items.csv"):
            if not path.is_file() or not GLOVO_ITEMS_PATTERN.match(path.name):
                continue
            prev = by_name.get(path.name)
            if prev is None or _wolt_stamp_sort_key(path, GLOVO_ITEMS_PATTERN) > _wolt_stamp_sort_key(prev, GLOVO_ITEMS_PATTERN):
                by_name[path.name] = path
    candidates = list(by_name.values())
    candidates.sort(key=lambda p: _wolt_stamp_sort_key(p, GLOVO_ITEMS_PATTERN), reverse=True)
    return candidates


def _glovo_latest_report() -> Path | None:
    candidates = _glovo_report_candidates()
    return candidates[0] if candidates else None


def _glovo_pharmacies_candidates() -> list[Path]:
    by_name: dict[str, Path] = {}
    for base_dir in _glovo_results_dirs():
        if not base_dir.exists():
            continue
        for path in base_dir.glob("glovo_almaty_pharmacies_*.csv"):
            if not path.is_file() or not GLOVO_PHARMACIES_PATTERN.match(path.name):
                continue
            prev = by_name.get(path.name)
            if prev is None or _wolt_stamp_sort_key(path, GLOVO_PHARMACIES_PATTERN) > _wolt_stamp_sort_key(prev, GLOVO_PHARMACIES_PATTERN):
                by_name[path.name] = path
    candidates = list(by_name.values())
    candidates.sort(key=lambda p: _wolt_stamp_sort_key(p, GLOVO_PHARMACIES_PATTERN), reverse=True)
    return candidates


def _glovo_load_pharmacy_index(path: Path) -> dict:
    by_slug: dict[str, dict] = {}
    by_name: dict[str, dict] = {}
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            slug = str(row.get("slug") or "").strip().lower()
            name = str(row.get("name") or "").strip()
            address = str(row.get("address") or "").strip()
            payload = {
                "slug": slug,
                "name": name,
                "address": address,
                "store_url": str(row.get("store_url") or "").strip(),
            }
            if slug:
                by_slug[slug] = payload
            if name:
                by_name[_normalize_text(name)] = payload
    return {"by_slug": by_slug, "by_name": by_name, "source_file": path.name}


def _glovo_pharmacy_index() -> dict:
    for path in _glovo_pharmacies_candidates():
        try:
            mtime = path.stat().st_mtime
        except OSError:
            continue
        cached = GLOVO_PHARMACY_CACHE.get("payload")
        if (
            GLOVO_PHARMACY_CACHE.get("path") == str(path)
            and GLOVO_PHARMACY_CACHE.get("mtime") == mtime
            and cached is not None
        ):
            return cached
        try:
            payload = _glovo_load_pharmacy_index(path)
        except Exception:
            continue
        if payload.get("by_slug") or payload.get("by_name"):
            GLOVO_PHARMACY_CACHE["path"] = str(path)
            GLOVO_PHARMACY_CACHE["mtime"] = mtime
            GLOVO_PHARMACY_CACHE["payload"] = payload
            return payload
    return {"by_slug": {}, "by_name": {}, "source_file": ""}


def _glovo_discovery_candidates(glob_pattern: str, pattern: re.Pattern[str]) -> list[Path]:
    by_name: dict[str, Path] = {}
    for base_dir in _glovo_results_dirs():
        if not base_dir.exists():
            continue
        for path in base_dir.glob(glob_pattern):
            if not path.is_file() or not pattern.match(path.name):
                continue
            prev = by_name.get(path.name)
            if prev is None or _wolt_stamp_sort_key(path, pattern) > _wolt_stamp_sort_key(prev, pattern):
                by_name[path.name] = path
    candidates = list(by_name.values())
    candidates.sort(key=lambda p: _wolt_stamp_sort_key(p, pattern), reverse=True)
    return candidates


def _glovo_latest_discovery_path(glob_pattern: str, pattern: re.Pattern[str]) -> Path | None:
    candidates = _glovo_discovery_candidates(glob_pattern, pattern)
    return candidates[0] if candidates else None


def _glovo_stamp_from_path(path: Path | None, pattern: re.Pattern[str]) -> str:
    if not path:
        return ""
    match = pattern.match(path.name)
    return str(match.group("stamp") or "") if match else ""


def _glovo_find_discovery_file(file_name: str) -> Path | None:
    safe_name = Path(file_name).name
    for base_dir in _glovo_results_dirs():
        path = base_dir / safe_name
        if path.exists() and path.is_file():
            return path
    return None


def _glovo_split_multi(value: str) -> list[str]:
    raw = str(value or "").strip()
    if not raw:
        return []
    return [part.strip() for part in raw.split(" || ") if part.strip()]


def _glovo_load_pharmacies_rows(path: Path | None) -> list[dict]:
    if not path or not path.exists() or not path.is_file():
        return []
    rows: list[dict] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            rows.append(
                {
                    "slug": str(row.get("slug") or "").strip().lower(),
                    "name": str(row.get("name") or "").strip(),
                    "address": str(row.get("address") or "").strip(),
                    "store_url": str(row.get("store_url") or "").strip(),
                    "store_id": str(row.get("store_id") or "").strip(),
                    "address_id": str(row.get("address_id") or "").strip(),
                    "store_geo_lat": _safe_float(row.get("store_geo_lat")),
                    "store_geo_lon": _safe_float(row.get("store_geo_lon")),
                    "store_geo_place_id": str(row.get("store_geo_place_id") or "").strip(),
                    "store_geo_address": str(row.get("store_geo_address") or "").strip(),
                    "grid_probe_hits": _safe_int(row.get("grid_probe_hits")) or 0,
                    "sample_contexts": _glovo_split_multi(str(row.get("sample_contexts") or "")),
                    "sample_probe_addresses": _glovo_split_multi(str(row.get("sample_probe_addresses") or "")),
                    "is_pharmacy": str(row.get("is_pharmacy") or "").strip(),
                    "store_type": str(row.get("store_type") or "").strip(),
                    "skip_reason": str(row.get("skip_reason") or "").strip(),
                }
            )
    return rows


def _glovo_load_grid_rows(path: Path | None) -> list[dict]:
    if not path or not path.exists() or not path.is_file():
        return []
    rows: list[dict] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            rows.append(
                {
                    "probe_label": str(row.get("probe_label") or "").strip(),
                    "probe_type": str(row.get("probe_type") or "").strip(),
                    "input_address": str(row.get("input_address") or "").strip(),
                    "resolved_title": str(row.get("resolved_title") or "").strip(),
                    "resolved_subtitle": str(row.get("resolved_subtitle") or "").strip(),
                    "resolved_full_address": str(row.get("resolved_full_address") or "").strip(),
                    "place_id": str(row.get("place_id") or "").strip(),
                    "latitude": _safe_float(row.get("latitude")),
                    "longitude": _safe_float(row.get("longitude")),
                    "store_count": _safe_int(row.get("store_count")) or 0,
                    "store_slugs": _glovo_split_multi(str(row.get("store_slugs") or "")),
                    "status": str(row.get("status") or "").strip().lower(),
                    "error": str(row.get("error") or "").strip(),
                }
            )
    return rows


def _glovo_load_visibility_rows(path: Path | None) -> list[dict]:
    if not path or not path.exists() or not path.is_file():
        return []
    rows: list[dict] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            rows.append(
                {
                    "slug": str(row.get("slug") or "").strip().lower(),
                    "store_url": str(row.get("store_url") or "").strip(),
                    "display_names": _glovo_split_multi(str(row.get("display_names") or "")),
                    "discovery_methods": _glovo_split_multi(str(row.get("discovery_methods") or "")),
                    "grid_probe_hits": _safe_int(row.get("grid_probe_hits")) or 0,
                    "source_urls": _glovo_split_multi(str(row.get("source_urls") or "")),
                    "probe_labels": _glovo_split_multi(str(row.get("probe_labels") or "")),
                    "sample_contexts": _glovo_split_multi(str(row.get("sample_contexts") or "")),
                    "sample_probe_addresses": _glovo_split_multi(str(row.get("sample_probe_addresses") or "")),
                }
            )
    return rows


def _glovo_load_compare_rows(path: Path | None) -> list[dict]:
    if not path or not path.exists() or not path.is_file():
        return []
    rows: list[dict] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            rows.append(
                {
                    "glovo_name": str(row.get("glovo_name") or "").strip(),
                    "glovo_slug": str(row.get("glovo_slug") or "").strip().lower(),
                    "glovo_address": str(row.get("glovo_address") or "").strip(),
                    "glovo_chain": str(row.get("glovo_chain") or "").strip(),
                    "glovo_store_url": str(row.get("glovo_store_url") or "").strip(),
                    "glovo_lat": _safe_float(row.get("glovo_lat")),
                    "glovo_lon": _safe_float(row.get("glovo_lon")),
                    "match_status": str(row.get("match_status") or "").strip(),
                    "distance_m": _safe_float(row.get("distance_m")),
                    "wolt_name": str(row.get("wolt_name") or "").strip(),
                    "wolt_slug": str(row.get("wolt_slug") or "").strip(),
                    "wolt_address": str(row.get("wolt_address") or "").strip(),
                    "wolt_lat": _safe_float(row.get("wolt_lat")),
                    "wolt_lon": _safe_float(row.get("wolt_lon")),
                }
            )
    return rows


def _glovo_load_branch_context_payload(path: Path | None) -> dict:
    if not path or not path.exists() or not path.is_file():
        return {"rows": [], "slug_summaries": []}

    raw_rows: list[dict] = []
    slug_groups: dict[str, dict] = {}
    branch_groups: dict[tuple[str, str], dict] = {}
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            slug = str(row.get("slug") or "").strip().lower()
            branch_key = str(row.get("branch_key") or "").strip()
            fetch_status = str(row.get("fetch_status") or "").strip().lower()
            item = {
                "probe_label": str(row.get("probe_label") or "").strip(),
                "probe_type": str(row.get("probe_type") or "").strip(),
                "probe_input_address": str(row.get("probe_input_address") or "").strip(),
                "probe_resolved_title": str(row.get("probe_resolved_title") or "").strip(),
                "probe_resolved_subtitle": str(row.get("probe_resolved_subtitle") or "").strip(),
                "probe_resolved_full_address": str(row.get("probe_resolved_full_address") or "").strip(),
                "probe_place_id": str(row.get("probe_place_id") or "").strip(),
                "probe_latitude": _safe_float(row.get("probe_latitude")),
                "probe_longitude": _safe_float(row.get("probe_longitude")),
                "slug": slug,
                "store_url": str(row.get("store_url") or "").strip(),
                "display_name": str(row.get("display_name") or "").strip(),
                "context_store_name": str(row.get("context_store_name") or "").strip(),
                "context_store_id": str(row.get("context_store_id") or "").strip(),
                "context_address_id": str(row.get("context_address_id") or "").strip(),
                "context_address": str(row.get("context_address") or "").strip(),
                "context_geo_lat": _safe_float(row.get("context_geo_lat")),
                "context_geo_lon": _safe_float(row.get("context_geo_lon")),
                "context_geo_place_id": str(row.get("context_geo_place_id") or "").strip(),
                "context_geo_address": str(row.get("context_geo_address") or "").strip(),
                "branch_key": branch_key,
                "fetch_status": fetch_status,
                "error": str(row.get("error") or "").strip(),
            }
            raw_rows.append(item)
            if not slug:
                continue
            slug_bucket = slug_groups.setdefault(
                slug,
                {
                    "slug": slug,
                    "pharmacy": item["display_name"] or item["context_store_name"] or slug,
                    "store_url": item["store_url"] or "",
                    "probe_hits": 0,
                    "ok_hits": 0,
                    "error_hits": 0,
                    "probe_labels": set(),
                    "areas": set(),
                    "branches": [],
                },
            )
            slug_bucket["probe_hits"] += 1
            if item["probe_label"]:
                slug_bucket["probe_labels"].add(item["probe_label"])
            if item["probe_resolved_title"]:
                slug_bucket["areas"].add(item["probe_resolved_title"])
            if fetch_status == "ok":
                slug_bucket["ok_hits"] += 1
            else:
                slug_bucket["error_hits"] += 1

            if not branch_key:
                continue
            bucket = branch_groups.setdefault(
                (slug, branch_key),
                {
                    "slug": slug,
                    "pharmacy": item["display_name"] or item["context_store_name"] or slug,
                    "store_url": item["store_url"] or "",
                    "branch_key": branch_key,
                    "context_store_id": item["context_store_id"] or "",
                    "context_address_id": item["context_address_id"] or "",
                    "address": item["context_address"] or "",
                    "geo_lat": item["context_geo_lat"],
                    "geo_lon": item["context_geo_lon"],
                    "geo_address": item["context_geo_address"] or "",
                    "probe_hits": 0,
                    "ok_hits": 0,
                    "error_hits": 0,
                    "probe_labels": set(),
                    "areas": set(),
                    "sample_contexts": set(),
                },
            )
            bucket["probe_hits"] += 1
            if item["probe_label"]:
                bucket["probe_labels"].add(item["probe_label"])
            if item["probe_resolved_title"]:
                bucket["areas"].add(item["probe_resolved_title"])
            if item["probe_resolved_full_address"]:
                bucket["sample_contexts"].add(item["probe_resolved_full_address"])
            if fetch_status == "ok":
                bucket["ok_hits"] += 1
            else:
                bucket["error_hits"] += 1
            if not bucket.get("address") and item["context_address"]:
                bucket["address"] = item["context_address"]
            if bucket.get("geo_lat") is None and item["context_geo_lat"] is not None:
                bucket["geo_lat"] = item["context_geo_lat"]
            if bucket.get("geo_lon") is None and item["context_geo_lon"] is not None:
                bucket["geo_lon"] = item["context_geo_lon"]
            if not bucket.get("geo_address") and item["context_geo_address"]:
                bucket["geo_address"] = item["context_geo_address"]

    for branch in branch_groups.values():
        slug_bucket = slug_groups.get(branch["slug"])
        if not slug_bucket:
            continue
        slug_bucket["branches"].append(
            {
                "branch_key": branch["branch_key"],
                "context_store_id": branch["context_store_id"],
                "context_address_id": branch["context_address_id"],
                "address": branch["address"],
                "geo_lat": branch["geo_lat"],
                "geo_lon": branch["geo_lon"],
                "geo_address": branch["geo_address"],
                "probe_hits": branch["probe_hits"],
                "ok_hits": branch["ok_hits"],
                "error_hits": branch["error_hits"],
                "probe_labels": sorted(branch["probe_labels"]),
                "areas": sorted(branch["areas"]),
                "sample_contexts": sorted(branch["sample_contexts"]),
            }
        )

    slug_summaries: list[dict] = []
    for slug, bucket in slug_groups.items():
        branches = sorted(
            bucket["branches"],
            key=lambda item: (-int(item.get("ok_hits") or 0), -int(item.get("probe_hits") or 0), str(item.get("address") or "")),
        )
        slug_summaries.append(
            {
                "slug": slug,
                "pharmacy": bucket["pharmacy"],
                "store_url": bucket["store_url"],
                "probe_hits": bucket["probe_hits"],
                "ok_hits": bucket["ok_hits"],
                "error_hits": bucket["error_hits"],
                "branch_count": len(branches),
                "probe_labels": sorted(bucket["probe_labels"]),
                "areas": sorted(bucket["areas"]),
                "branches": branches,
            }
        )
    slug_summaries.sort(
        key=lambda item: (-int(item.get("branch_count") or 0), -int(item.get("ok_hits") or 0), _normalize_text(item.get("pharmacy") or "")),
    )
    return {"rows": raw_rows, "slug_summaries": slug_summaries}


def _glovo_coverage_payload() -> dict:
    pharmacies_path = _glovo_latest_discovery_path("glovo_almaty_pharmacies_*.csv", GLOVO_PHARMACIES_PATTERN)
    stamp = _glovo_stamp_from_path(pharmacies_path, GLOVO_PHARMACIES_PATTERN)
    if not pharmacies_path or not stamp:
        return {
            "stamp": "",
            "generated_at": "",
            "files": {},
            "probes": [],
            "visibility": [],
            "pharmacies": [],
            "compare_with_wolt": [],
            "branch_rows": [],
            "slug_contexts": [],
        }

    expected_files = {
        "pharmacies": pharmacies_path,
        "grid": _glovo_find_discovery_file(f"glovo_almaty_address_grid_{stamp}.csv"),
        "visibility": _glovo_find_discovery_file(f"glovo_almaty_store_visibility_{stamp}.csv"),
        "compare_wolt": _glovo_find_discovery_file(f"glovo_almaty_vs_wolt_{stamp}.csv"),
        "branch_contexts": _glovo_find_discovery_file(f"glovo_almaty_branch_contexts_{stamp}.csv"),
    }
    cache_key = []
    for key in sorted(expected_files):
        path = expected_files[key]
        try:
            cache_key.append((key, str(path) if path else "", path.stat().st_mtime if path else 0.0))
        except OSError:
            cache_key.append((key, str(path) if path else "", 0.0))
    cache_key_tuple = tuple(cache_key)
    if GLOVO_COVERAGE_CACHE.get("key") == cache_key_tuple and GLOVO_COVERAGE_CACHE.get("payload") is not None:
        return GLOVO_COVERAGE_CACHE["payload"]

    pharmacies = _glovo_load_pharmacies_rows(expected_files["pharmacies"])
    probes = _glovo_load_grid_rows(expected_files["grid"])
    visibility = _glovo_load_visibility_rows(expected_files["visibility"])
    compare_rows = _glovo_load_compare_rows(expected_files["compare_wolt"])
    branch_payload = _glovo_load_branch_context_payload(expected_files["branch_contexts"])

    visibility_by_slug = {str(item.get("slug") or "").strip().lower(): item for item in visibility}
    compare_by_slug = {str(item.get("glovo_slug") or "").strip().lower(): item for item in compare_rows}
    slug_contexts = branch_payload.get("slug_summaries") or []
    pharmacy_by_slug = {str(item.get("slug") or "").strip().lower(): item for item in pharmacies}

    normalized_slug_contexts: list[dict] = []
    for entry in slug_contexts:
        slug = str(entry.get("slug") or "").strip().lower()
        visibility_meta = visibility_by_slug.get(slug) or {}
        pharmacy_meta = pharmacy_by_slug.get(slug) or {}
        compare = compare_by_slug.get(slug) or {}
        display_names = [
            name
            for name in (visibility_meta.get("display_names") or [])
            if not _is_bad_glovo_name(str(name or ""), slug)
        ]
        best_name = _glovo_best_store_name(
            slug,
            entry.get("pharmacy") or "",
            pharmacy_meta.get("name") or "",
            *(display_names or []),
            compare.get("glovo_name") or "",
        )
        normalized_entry = dict(entry)
        normalized_entry["pharmacy"] = best_name
        normalized_entry["branches"] = [
            {
                **dict(branch),
                "address": str(branch.get("address") or "").strip() or str(pharmacy_meta.get("address") or "").strip(),
            }
            for branch in (entry.get("branches") or [])
        ]
        normalized_slug_contexts.append(normalized_entry)
    slug_contexts = normalized_slug_contexts
    slug_contexts_by_slug = {str(item.get("slug") or "").strip().lower(): item for item in slug_contexts}

    enriched_pharmacies: list[dict] = []
    for row in pharmacies:
        slug = str(row.get("slug") or "").strip().lower()
        compare = compare_by_slug.get(slug) or {}
        contexts = slug_contexts_by_slug.get(slug) or {}
        visibility_meta = visibility_by_slug.get(slug) or {}
        display_names = [
            name
            for name in (visibility_meta.get("display_names") or [])
            if not _is_bad_glovo_name(str(name or ""), slug)
        ]
        resolved_name = _glovo_best_store_name(
            slug,
            row.get("name") or "",
            *(display_names or []),
            contexts.get("pharmacy") or "",
            compare.get("glovo_name") or "",
        )
        enriched = dict(row)
        enriched["name"] = resolved_name
        enriched["display_names"] = visibility_meta.get("display_names") or []
        enriched["probe_labels"] = visibility_meta.get("probe_labels") or []
        enriched["sample_contexts"] = visibility_meta.get("sample_contexts") or row.get("sample_contexts") or []
        enriched["sample_probe_addresses"] = visibility_meta.get("sample_probe_addresses") or row.get("sample_probe_addresses") or []
        enriched["branch_count"] = int(contexts.get("branch_count") or 0)
        enriched["wolt_match_status"] = compare.get("match_status") or ""
        enriched["wolt_name"] = compare.get("wolt_name") or ""
        enriched["wolt_address"] = compare.get("wolt_address") or ""
        enriched["distance_to_wolt_m"] = compare.get("distance_m")
        enriched_pharmacies.append(enriched)

    enriched_probes: list[dict] = []
    for probe in probes:
        stores: list[dict] = []
        for slug in probe.get("store_slugs") or []:
            slug_key = str(slug or "").strip().lower()
            pharmacy = pharmacy_by_slug.get(slug_key) or {}
            visibility_meta = visibility_by_slug.get(slug_key) or {}
            contexts = slug_contexts_by_slug.get(slug_key) or {}
            display_names = [
                name
                for name in (visibility_meta.get("display_names") or [])
                if not _is_bad_glovo_name(str(name or ""), slug_key)
            ]
            stores.append(
                {
                    "slug": slug_key,
                    "name": _glovo_best_store_name(
                        slug_key,
                        pharmacy.get("name") or "",
                        *(display_names or []),
                        contexts.get("pharmacy") or "",
                        (compare_by_slug.get(slug_key) or {}).get("glovo_name") or "",
                    ),
                    "store_url": pharmacy.get("store_url") or visibility_meta.get("store_url") or "",
                    "address": pharmacy.get("address") or "",
                    "branch_count": int(contexts.get("branch_count") or 0),
                }
            )
        item = dict(probe)
        item["stores"] = stores
        item["pharmacy_count"] = len(stores)
        enriched_probes.append(item)

    generated_at = ""
    try:
        generated_at = datetime.fromtimestamp(expected_files["pharmacies"].stat().st_mtime, tz=ALMATY_TZ).isoformat(
            timespec="seconds"
        )
    except OSError:
        generated_at = ""

    payload = {
        "stamp": stamp,
        "generated_at": generated_at,
        "files": {key: path.name if path else "" for key, path in expected_files.items()},
        "probes": enriched_probes,
        "visibility": visibility,
        "pharmacies": enriched_pharmacies,
        "compare_with_wolt": compare_rows,
        "branch_rows": branch_payload.get("rows") or [],
        "slug_contexts": slug_contexts,
    }
    GLOVO_COVERAGE_CACHE["key"] = cache_key_tuple
    GLOVO_COVERAGE_CACHE["payload"] = payload
    return payload


def _parse_glovo_items_report(path: Path, limit_rows: int = GLOVO_MAX_ITEMS) -> dict:
    pharmacy_index = _glovo_pharmacy_index()
    pharmacy_by_slug = pharmacy_index.get("by_slug") or {}
    pharmacy_by_name = pharmacy_index.get("by_name") or {}
    rows: list[dict] = []
    with open(path, "r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        for idx, row in enumerate(reader):
            if idx >= limit_rows:
                break
            status = str(row.get("status") or "").strip().lower()
            price_minor = _safe_int(row.get("price_minor"))
            price = _safe_float(row.get("price"))
            pharmacy = str(row.get("pharmacy") or "").strip()
            venue_slug = str(row.get("venue_slug") or "").strip()
            venue_slug_key = venue_slug.lower()
            pharmacy_meta = pharmacy_by_slug.get(venue_slug_key) or pharmacy_by_name.get(_normalize_text(pharmacy))
            pharmacy_address = str(row.get("pharmacy_address") or "").strip() or str((pharmacy_meta or {}).get("address") or "").strip()
            rows.append(
                {
                    "checked_at": str(row.get("checked_at") or "").strip(),
                    "brand": str(row.get("brand") or "").strip(),
                    "query": str(row.get("query") or "").strip(),
                    "pharmacy": pharmacy,
                    "pharmacy_address": pharmacy_address,
                    "venue_slug": venue_slug,
                    "venue_url": str(row.get("venue_url") or "").strip(),
                    "item_id": str(row.get("item_id") or "").strip(),
                    "item_name": str(row.get("item_name") or "").strip(),
                    "status": status,
                    "disable_text": str(row.get("disable_text") or "").strip(),
                    "price_minor": price_minor,
                    "price": price,
                    "product_link": str(row.get("product_link") or "").strip(),
                    "search_link": str(row.get("search_link") or "").strip(),
                }
            )

    if not rows:
        return {
            "rows": [],
            "brand": "",
            "query": "",
            "checked_at": "",
            "status_counts": {},
            "pharmacies_count": 0,
            "item_ids_count": 0,
            "price_stats": {"min": None, "median": None, "max": None},
            "top_offers": [],
            "pharmacy_stats": [],
        }

    status_counts: dict[str, int] = {}
    pharmacy_stats: dict[str, dict] = {}
    unique_items = set()
    in_stock_prices: list[float] = []
    best_offers_map: dict[str, dict] = {}

    for row in rows:
        status = row.get("status") or "unknown"
        status_counts[status] = status_counts.get(status, 0) + 1
        unique_items.add(row.get("item_id") or "")

        pharmacy_key = (row.get("venue_slug") or row.get("pharmacy") or "").strip().lower()
        bucket = pharmacy_stats.setdefault(
            pharmacy_key,
            {
                "pharmacy": row.get("pharmacy") or "",
                "venue_slug": row.get("venue_slug") or "",
                "venue_url": row.get("venue_url") or "",
                "pharmacy_address": row.get("pharmacy_address") or "",
                "total": 0,
                "in_stock": 0,
                "out_of_stock": 0,
                "unavailable": 0,
                "not_found": 0,
                "best_price": None,
            },
        )
        bucket["total"] += 1
        if status in bucket:
            bucket[status] += 1

        price_value = row.get("price")
        if status == "in_stock" and price_value is not None:
            in_stock_prices.append(price_value)
            if bucket["best_price"] is None or price_value < bucket["best_price"]:
                bucket["best_price"] = price_value

            item_key = _normalize_text(row.get("item_name") or row.get("item_id") or "")
            prev = best_offers_map.get(item_key)
            if prev is None or (prev.get("price") is None) or price_value < prev.get("price"):
                best_offers_map[item_key] = row

    top_offers = sorted(
        best_offers_map.values(),
        key=lambda x: (x.get("price") if x.get("price") is not None else 10**12, _normalize_text(x.get("item_name") or "")),
    )[:24]
    pharmacy_rows = sorted(
        pharmacy_stats.values(),
        key=lambda x: (x.get("in_stock", 0), -(x.get("unavailable", 0))),
        reverse=True,
    )[:80]

    median_price = round(statistics.median(in_stock_prices), 2) if in_stock_prices else None
    return {
        "rows": rows,
        "brand": rows[0].get("brand") or "",
        "query": rows[0].get("query") or "",
        "checked_at": rows[0].get("checked_at") or "",
        "status_counts": status_counts,
        "pharmacies_count": len(
            {
                (r.get("venue_slug") or r.get("pharmacy") or "").strip().lower()
                for r in rows
                if (r.get("venue_slug") or r.get("pharmacy"))
            }
        ),
        "item_ids_count": len({i for i in unique_items if i}),
        "price_stats": {
            "min": round(min(in_stock_prices), 2) if in_stock_prices else None,
            "median": median_price,
            "max": round(max(in_stock_prices), 2) if in_stock_prices else None,
        },
        "top_offers": top_offers,
        "pharmacy_stats": pharmacy_rows,
        "pharmacy_source_file": pharmacy_index.get("source_file") or "",
    }


def _glovo_reports_payload() -> dict:
    reports = []
    per_brand: dict[str, dict] = {}
    for path in _glovo_report_candidates()[: max(1, GLOVO_MAX_REPORTS)]:
        mtime = datetime.fromtimestamp(path.stat().st_mtime, tz=ALMATY_TZ).isoformat(timespec="seconds")
        match = GLOVO_ITEMS_PATTERN.match(path.name)
        if match:
            brand_slug = match.group("brand")
            stamp = match.group("stamp")
            brand = _humanize_brand_slug(brand_slug)
        else:
            brand_slug = "unknown"
            brand = "Unknown"
            stamp = ""
        row = {
            "file": path.name,
            "brand": brand,
            "brand_slug": brand_slug,
            "stamp": stamp,
            "mtime": mtime,
        }
        reports.append(row)
        slot = per_brand.get(brand_slug)
        if slot is None:
            per_brand[brand_slug] = {
                "brand": brand,
                "brand_slug": brand_slug,
                "latest_file": path.name,
                "latest_stamp": stamp,
                "latest_mtime": mtime,
                "reports_count": 1,
            }
        else:
            slot["reports_count"] = int(slot.get("reports_count") or 0) + 1

    brands = sorted(
        per_brand.values(),
        key=lambda x: str(x.get("latest_mtime") or ""),
        reverse=True,
    )
    return {"reports": reports, "brands": brands}


def _resolve_glovo_report(report: str | None, brand: str | None = None) -> Path | None:
    if report:
        safe_name = Path(report).name
        if not GLOVO_ITEMS_PATTERN.match(safe_name):
            return None
        matches: list[Path] = []
        for base_dir in _glovo_results_dirs():
            path = base_dir / safe_name
            if path.exists():
                matches.append(path)
        if matches:
            matches.sort(key=lambda p: _wolt_stamp_sort_key(p, GLOVO_ITEMS_PATTERN), reverse=True)
            return matches[0]
        return None

    if brand:
        brand_raw = str(brand or "").strip()
        brand_slug_query = re.sub(r"[^a-z0-9_-]+", "_", brand_raw.lower()).strip("_")
        brand_norm_query = _normalize_text(brand_raw.replace("_", " ").replace("-", " "))
        for path in _glovo_report_candidates():
            match = GLOVO_ITEMS_PATTERN.match(path.name)
            if not match:
                continue
            brand_slug = str(match.group("brand") or "").strip().lower()
            brand_human = _humanize_brand_slug(brand_slug)
            brand_norm = _normalize_text(brand_human)
            brand_slug_norm = _normalize_text(brand_slug.replace("_", " ").replace("-", " "))
            if (
                brand_slug == brand_slug_query
                or brand_norm == brand_norm_query
                or brand_slug_norm == brand_norm_query
            ):
                return path
        return None

    return _glovo_latest_report()


def _keyword_status_payload() -> dict:
    payload = {}
    with KEYWORD_LOCK:
        payload.update(KEYWORD_STATE)
    latest = _keyword_latest_report()
    if latest:
        payload["latest_file"] = latest.name
        payload["latest_mtime"] = datetime.fromtimestamp(latest.stat().st_mtime, tz=ALMATY_TZ).isoformat(timespec="seconds")
    return payload


def _reviews_status_payload() -> dict:
    payload = {}
    with REVIEWS_LOCK:
        payload.update(REVIEWS_STATE)
    latest = None
    if RESULTS_DIR.exists():
        candidates = list(RESULTS_DIR.glob("reviews_*.csv"))
        if candidates:
            candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            latest = candidates[0]
    if latest:
        payload["latest_file"] = latest.name
        payload["latest_mtime"] = datetime.fromtimestamp(latest.stat().st_mtime, tz=ALMATY_TZ).isoformat(timespec="seconds")
    return payload


def _keyword_queries_payload() -> dict:
    raw_path = os.environ.get("SEARCH_KEYWORDS_CSV", "search_keywords.csv")
    path = Path(raw_path)
    if not path.is_absolute():
        path = BASE_DIR.parent / path
    if not path.exists():
        return {"queries": []}
    try:
        df = pd.read_csv(path)
    except Exception:
        return {"queries": []}
    if df.empty:
        return {"queries": []}
    columns = {str(col).strip().lower(): col for col in df.columns}
    keyword_col = columns.get("keyword") or columns.get("query") or df.columns[0]
    url_col = columns.get("search_url") or columns.get("url")
    queries = []
    for _, row in df.iterrows():
        keyword = str(row.get(keyword_col) or "").strip()
        search_url = str(row.get(url_col) or "").strip() if url_col else ""
        if not keyword and not search_url:
            continue
        queries.append({"keyword": keyword, "search_url": search_url})
    return {"queries": queries}


def _keyword_rows_payload(status: str | None = None, pos_le: int | None = None) -> dict:
    latest = _keyword_latest_report()
    if not latest:
        return {}
    wb = load_workbook(latest, read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        header = next(rows_iter, None)
        if not header:
            return {}
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in header]
        idx = {name: i for i, name in enumerate(headers)}
        keyword_idx = idx.get("keyword")
        status_idx = idx.get("status")
        position_idx = idx.get("position")
        name_idx = idx.get("product_name")
        url_idx = idx.get("product_url")
        code_idx = idx.get("product_code")
        if keyword_idx is None or status_idx is None:
            return {}
        result_rows = []
        for row in rows_iter:
            if len(result_rows) >= KEYWORD_ROWS_LIMIT:
                break
            status_raw = row[status_idx] if status_idx < len(row) else ""
            status_norm = str(status_raw or "").strip().lower()
            if status and status != "all" and status_norm != status:
                continue
            pos_raw = row[position_idx] if position_idx is not None and position_idx < len(row) else None
            try:
                pos = int(pos_raw) if pos_raw not in (None, "") else None
            except Exception:
                pos = None
            if pos_le is not None and pos is not None and pos > pos_le:
                continue
            result_rows.append({
                "keyword": str(row[keyword_idx] or "").strip(),
                "status": status_norm or "",
                "position": pos,
                "product_name": str(row[name_idx] or "").strip() if name_idx is not None else "",
                "product_url": str(row[url_idx] or "").strip() if url_idx is not None else "",
                "product_code": str(row[code_idx] or "").strip() if code_idx is not None else "",
            })
        return {
            "rows": result_rows,
            "latest_file": latest.name,
            "latest_mtime": datetime.fromtimestamp(latest.stat().st_mtime, tz=ALMATY_TZ).isoformat(timespec="seconds"),
        }
    finally:
        wb.close()


def _normalize_region_key(value: str | None) -> str:
    return " ".join(str(value or "").replace("\u00a0", " ").split()).strip().lower()


def _safe_ratio_pct(numerator: int | float, denominator: int | float) -> float | None:
    if not denominator:
        return None
    try:
        return round((float(numerator) / float(denominator)) * 100.0, 2)
    except Exception:
        return None


def _iter_price_snapshot_files() -> list[tuple[datetime, Path]]:
    if not RESULTS_DIR.exists():
        return []
    stamp_pattern = re.compile(
        r"^(?P<date>\d{4}-\d{2}-\d{2})_(?P<time>\d{2}-\d{2}-\d{2})(?:_.+)?\.xlsx$",
        re.IGNORECASE,
    )
    rows: list[tuple[datetime, Path]] = []
    seen: set[Path] = set()
    for prefix in _main_prefixes():
        for path in RESULTS_DIR.glob(f"{prefix}_*.xlsx"):
            if path in seen:
                continue
            suffix = path.name[len(prefix) + 1 :]
            match = stamp_pattern.match(suffix)
            if not match:
                continue
            try:
                snap_dt = datetime.strptime(
                    f"{match.group('date')} {match.group('time')}",
                    "%Y-%m-%d %H-%M-%S",
                )
            except Exception:
                continue
            rows.append((snap_dt, path))
            seen.add(path)
    rows.sort(key=lambda item: item[0])
    return rows


def _price_dynamics_options_payload(region: str = "Алматы", q: str | None = None, limit: int = 40) -> dict:
    files = _iter_price_snapshot_files()
    if not files:
        return {
            "options": [],
            "regions": [],
            "latest_snapshot_file": "",
            "latest_snapshot_dt": "",
            "available_start": "",
            "available_end": "",
            "total_matches": 0,
        }

    latest_dt, latest_path = files[-1]
    earliest_dt, _ = files[0]
    region_key = _normalize_region_key(region)
    query = " ".join(str(q or "").strip().lower().split())
    cache_key = (
        latest_path.name,
        latest_path.stat().st_mtime,
        region_key,
        query,
        int(limit),
    )
    if PRICE_DYNAMICS_OPTIONS_CACHE.get("key") == cache_key and PRICE_DYNAMICS_OPTIONS_CACHE.get("payload") is not None:
        return PRICE_DYNAMICS_OPTIONS_CACHE["payload"]

    wb = load_workbook(latest_path, read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        header = next(rows_iter, None)
        if not header:
            payload = {
                "options": [],
                "regions": [],
                "latest_snapshot_file": latest_path.name,
                "latest_snapshot_dt": latest_dt.isoformat(timespec="seconds"),
                "available_start": earliest_dt.date().isoformat(),
                "available_end": latest_dt.date().isoformat(),
                "total_matches": 0,
            }
            PRICE_DYNAMICS_OPTIONS_CACHE.update({"key": cache_key, "payload": payload})
            return payload

        headers = [str(cell).strip().lower() if cell is not None else "" for cell in header]
        idx = {name: i for i, name in enumerate(headers)}
        region_idx = idx.get("region")
        code_idx = idx.get("product_code")
        product_idx = idx.get("product")
        input_product_idx = idx.get("input_product")
        url_idx = idx.get("product_url")

        options_map: dict[str, dict] = {}
        region_labels: dict[str, str] = {}
        for row in rows_iter:
            raw_region = str(row[region_idx] or "").strip() if region_idx is not None and region_idx < len(row) else ""
            norm_region = _normalize_region_key(raw_region)
            if raw_region and norm_region not in region_labels:
                region_labels[norm_region] = raw_region
            if region_key and norm_region != region_key:
                continue

            product_code = str(row[code_idx] or "").strip() if code_idx is not None and code_idx < len(row) else ""
            if not product_code and url_idx is not None and url_idx < len(row):
                product_code = _extract_product_code_from_url(str(row[url_idx] or "").strip())
            if not product_code:
                continue

            product_name = ""
            if product_idx is not None and product_idx < len(row):
                product_name = str(row[product_idx] or "").strip()
            if not product_name and input_product_idx is not None and input_product_idx < len(row):
                product_name = str(row[input_product_idx] or "").strip()

            haystack = " ".join(
                part for part in [product_code.lower(), product_name.lower()] if part
            ).strip()
            if query and query not in haystack:
                continue

            if product_code not in options_map:
                options_map[product_code] = {
                    "product_code": product_code,
                    "product_name": product_name,
                    "label": f"{product_code} • {product_name}" if product_name else product_code,
                }
            elif product_name and not options_map[product_code].get("product_name"):
                options_map[product_code]["product_name"] = product_name
                options_map[product_code]["label"] = f"{product_code} • {product_name}"

        options = sorted(
            options_map.values(),
            key=lambda item: (
                str(item.get("product_name") or "").lower(),
                str(item.get("product_code") or ""),
            ),
        )
        payload = {
            "options": options[: max(int(limit), 1)],
            "regions": sorted(region_labels.values(), key=lambda value: value.lower()),
            "latest_snapshot_file": latest_path.name,
            "latest_snapshot_dt": latest_dt.isoformat(timespec="seconds"),
            "available_start": earliest_dt.date().isoformat(),
            "available_end": latest_dt.date().isoformat(),
            "total_matches": len(options),
        }
        PRICE_DYNAMICS_OPTIONS_CACHE.update({"key": cache_key, "payload": payload})
        return payload
    finally:
        wb.close()


def _read_snapshot_sku_prices(path: Path, region: str, sku: str) -> tuple[dict[str, int], str, str]:
    target_region = _normalize_region_key(region)
    target_sku = str(sku or "").strip()
    if not target_sku:
        return {}, "", ""

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        header = next(rows_iter, None)
        if not header:
            return {}, "", ""

        headers = [str(cell).strip().lower() if cell is not None else "" for cell in header]
        idx = {name: i for i, name in enumerate(headers)}
        region_idx = idx.get("region")
        seller_idx = idx.get("seller")
        price_idx = idx.get("price_kzt")
        code_idx = idx.get("product_code")
        url_idx = idx.get("product_url")
        product_idx = idx.get("product")
        input_product_idx = idx.get("input_product")
        if seller_idx is None or price_idx is None:
            return {}, "", ""

        seller_prices: dict[str, int] = {}
        product_name = ""
        product_url = ""
        for row in rows_iter:
            if region_idx is not None and region_idx < len(row) and target_region:
                row_region = _normalize_region_key(row[region_idx])
                if row_region != target_region:
                    continue

            product_code = str(row[code_idx] or "").strip() if code_idx is not None and code_idx < len(row) else ""
            raw_url = str(row[url_idx] or "").strip() if url_idx is not None and url_idx < len(row) else ""
            if not product_code and raw_url:
                product_code = _extract_product_code_from_url(raw_url)
            if product_code != target_sku:
                continue

            seller = _canonical_seller_name(str(row[seller_idx] or "").strip())
            price = _int_or_none(row[price_idx] if price_idx < len(row) else None)
            if not seller or price is None:
                continue

            if not product_name and product_idx is not None and product_idx < len(row):
                product_name = str(row[product_idx] or "").strip()
            if not product_name and input_product_idx is not None and input_product_idx < len(row):
                product_name = str(row[input_product_idx] or "").strip()
            if raw_url and not product_url:
                product_url = raw_url

            prev_price = seller_prices.get(seller)
            if prev_price is None or price < prev_price:
                seller_prices[seller] = price

        return seller_prices, product_name, product_url
    finally:
        wb.close()


def _price_dynamics_summary_payload(region: str, sku: str, start: str, end: str) -> dict:
    start_date = _parse_date(start)
    end_date = _parse_date(end)
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="end before start")

    sku_value = str(sku or "").strip()
    if not sku_value:
        raise HTTPException(status_code=400, detail="sku required")

    files = [
        (snap_dt, path)
        for snap_dt, path in _iter_price_snapshot_files()
        if start_date <= snap_dt.date() <= end_date
    ]
    if not files:
        return {}

    files_signature = tuple((path.name, int(path.stat().st_mtime)) for _, path in files)
    cache_key = (
        _normalize_region_key(region),
        sku_value,
        start_date.isoformat(),
        end_date.isoformat(),
        files_signature,
    )
    if PRICE_DYNAMICS_SUMMARY_CACHE.get("key") == cache_key and PRICE_DYNAMICS_SUMMARY_CACHE.get("payload") is not None:
        return PRICE_DYNAMICS_SUMMARY_CACHE["payload"]

    timeline: list[dict] = []
    seller_presence: Counter[str] = Counter()
    seller_wins: Counter[str] = Counter()
    seller_prices: dict[str, list[int]] = defaultdict(list)
    seller_counts: list[int] = []
    min_prices: list[int] = []
    absolute_min_sellers: set[str] = set()
    product_name = ""
    product_url = ""

    for snap_dt, path in files:
        price_map, title, url = _read_snapshot_sku_prices(path=path, region=region, sku=sku_value)
        if title and not product_name:
            product_name = title
        if url and not product_url:
            product_url = url
        if not price_map:
            continue

        snapshot_min = min(price_map.values())
        snapshot_max = max(price_map.values())
        best_sellers = sorted([seller for seller, value in price_map.items() if value == snapshot_min], key=lambda text: text.lower())
        snapshot_label = snap_dt.strftime("%d.%m %H:%M")
        timeline.append(
            {
                "snapshot_ts": snap_dt.isoformat(timespec="seconds"),
                "snapshot_label": snapshot_label,
                "source_file": path.name,
                "min_price_kzt": snapshot_min,
                "max_price_kzt": snapshot_max,
                "spread_kzt": snapshot_max - snapshot_min,
                "sellers_count": len(price_map),
                "best_sellers": best_sellers,
            }
        )
        min_prices.append(snapshot_min)
        seller_counts.append(len(price_map))
        for seller, price in price_map.items():
            seller_presence[seller] += 1
            seller_prices[seller].append(price)
            if price == snapshot_min:
                seller_wins[seller] += 1

    if not timeline:
        return {}

    absolute_min = min(min_prices)
    absolute_min_hits = 0
    absolute_min_labels: list[str] = []
    for point in timeline:
        if int(point.get("min_price_kzt") or 0) == absolute_min:
            absolute_min_hits += 1
            absolute_min_labels.append(str(point.get("snapshot_label") or ""))
            for seller in point.get("best_sellers") or []:
                absolute_min_sellers.add(str(seller))

    priced_snapshots = len(timeline)
    unique_sellers = set(seller_prices.keys())
    avg_sellers = round(statistics.fmean(seller_counts), 2) if seller_counts else 0.0
    max_sellers = max(seller_counts) if seller_counts else 0
    avg_min_price = round(statistics.fmean(min_prices), 2) if min_prices else 0.0
    median_min_price = round(float(statistics.median(min_prices)), 2) if min_prices else 0.0
    variability_std = round(float(statistics.pstdev(min_prices)), 2) if len(min_prices) > 1 else 0.0
    variability_cv = _safe_ratio_pct(variability_std, avg_min_price) if avg_min_price else None

    seller_rows: list[dict] = []
    for seller, prices in seller_prices.items():
        presence = int(seller_presence.get(seller) or 0)
        wins = int(seller_wins.get(seller) or 0)
        avg_price = round(statistics.fmean(prices), 2) if prices else None
        price_std = round(float(statistics.pstdev(prices)), 2) if len(prices) > 1 else 0.0
        seller_rows.append(
            {
                "seller": seller,
                "wins": wins,
                "win_share_pct": _safe_ratio_pct(wins, priced_snapshots),
                "presence_snapshots": presence,
                "presence_share_pct": _safe_ratio_pct(presence, priced_snapshots),
                "min_price_kzt": min(prices) if prices else None,
                "max_price_kzt": max(prices) if prices else None,
                "avg_price_kzt": avg_price,
                "price_std_kzt": price_std,
                "price_cv_pct": _safe_ratio_pct(price_std, avg_price) if avg_price else None,
            }
        )
    seller_rows.sort(
        key=lambda row: (
            -int(row.get("wins") or 0),
            -float(row.get("win_share_pct") or 0),
            -int(row.get("presence_snapshots") or 0),
            float(row.get("avg_price_kzt") or 0),
            str(row.get("seller") or "").lower(),
        )
    )
    top_seller = seller_rows[0] if seller_rows else {}
    payload = {
        "region": region,
        "sku": sku_value,
        "product_name": product_name,
        "product_url": product_url,
        "start": start_date.isoformat(),
        "end": end_date.isoformat(),
        "available_snapshots": len(files),
        "priced_snapshots": priced_snapshots,
        "missing_snapshots": max(len(files) - priced_snapshots, 0),
        "latest_snapshot_file": files[-1][1].name,
        "latest_snapshot_dt": files[-1][0].isoformat(timespec="seconds"),
        "min_price_kzt": absolute_min,
        "min_price_hits": absolute_min_hits,
        "min_price_sellers": sorted(absolute_min_sellers, key=lambda text: text.lower()),
        "min_price_labels": [label for label in absolute_min_labels if label],
        "max_min_price_kzt": max(min_prices),
        "avg_min_price_kzt": avg_min_price,
        "median_min_price_kzt": median_min_price,
        "variability_std_kzt": variability_std,
        "variability_cv_pct": variability_cv,
        "avg_sellers_per_snapshot": avg_sellers,
        "max_sellers_per_snapshot": max_sellers,
        "unique_sellers_count": len(unique_sellers),
        "top_min_seller": top_seller.get("seller") or "",
        "top_min_seller_snapshots": int(top_seller.get("wins") or 0),
        "top_min_seller_share_pct": top_seller.get("win_share_pct"),
        "timeline": timeline,
        "sellers": seller_rows,
    }
    PRICE_DYNAMICS_SUMMARY_CACHE.update({"key": cache_key, "payload": payload})
    return payload


def _parse_date(value: str) -> date:
    try:
        return datetime.fromisoformat(value).date()
    except Exception as exc:
        raise HTTPException(status_code=400, detail="invalid date") from exc


def _reviews_db_exists() -> bool:
    return REVIEWS_DB_PATH.exists()


def _normalize_review_dates(conn: sqlite3.Connection):
    conn.execute(
        """
        UPDATE reviews
        SET review_date = substr(review_date, 7, 4) || '-' || substr(review_date, 4, 2) || '-' || substr(review_date, 1, 2)
        WHERE review_date LIKE '__.__.____'
        """
    )
    conn.execute(
        """
        UPDATE reviews
        SET review_date = substr(review_date, 7, 4) || '-' || substr(review_date, 4, 2) || '-' || substr(review_date, 1, 2)
        WHERE review_date LIKE '__/__/____'
        """
    )
    conn.commit()


def _reviews_timeline_payload(
    conn: sqlite3.Connection,
    start_date: date,
    end_exclusive: date,
    sku: str | None,
) -> dict:
    where_clause = (
        "comment IS NOT NULL AND TRIM(comment) != '' "
        "AND review_date IS NOT NULL AND review_date != '' "
        "AND review_date >= ? AND review_date < ?"
    )
    params: list = [start_date.isoformat(), end_exclusive.isoformat()]
    if sku:
        where_clause += " AND sku = ?"
        params.append(sku)

    cur = conn.cursor()
    rows = cur.execute(
        f"""
        SELECT review_date, COUNT(*)
        FROM reviews
        WHERE {where_clause}
        GROUP BY review_date
        ORDER BY review_date
        """,
        params,
    ).fetchall()
    daily_counts = {day: int(cnt or 0) for day, cnt in rows if day}
    all_days = [start_date + timedelta(days=offset) for offset in range((end_exclusive - start_date).days)]
    timeline: list[dict] = []
    cumulative = 0
    for day in all_days:
        day_iso = day.isoformat()
        count = int(daily_counts.get(day_iso, 0))
        cumulative += count
        timeline.append(
            {
                "date": day_iso,
                "count": count,
                "cumulative": cumulative,
            }
        )

    if sku:
        stacked = [{"date": row["date"], "total": row["count"], "parts": [{"sku": sku, "count": row["count"]}]} for row in timeline]
        return {"timeline": timeline, "stacked": stacked, "legend": [sku]}

    top_rows = cur.execute(
        f"""
        SELECT COALESCE(NULLIF(TRIM(sku), ''), 'Без SKU') AS sku_key, COUNT(*) AS cnt
        FROM reviews
        WHERE {where_clause}
        GROUP BY sku_key
        ORDER BY cnt DESC
        LIMIT 4
        """,
        params,
    ).fetchall()
    top_skus = [str(row[0]) for row in top_rows if row and row[0]]
    top_set = set(top_skus)

    by_day_rows = cur.execute(
        f"""
        SELECT review_date, COALESCE(NULLIF(TRIM(sku), ''), 'Без SKU') AS sku_key, COUNT(*) AS cnt
        FROM reviews
        WHERE {where_clause}
        GROUP BY review_date, sku_key
        ORDER BY review_date ASC
        """,
        params,
    ).fetchall()

    day_parts: dict[str, dict[str, int]] = {}
    for day, sku_key, cnt in by_day_rows:
        if not day:
            continue
        bucket = day_parts.setdefault(day, {})
        key = str(sku_key)
        if key in top_set:
            bucket[key] = bucket.get(key, 0) + int(cnt or 0)
        else:
            bucket["Другие"] = bucket.get("Другие", 0) + int(cnt or 0)

    legend = list(top_skus)
    if any("Другие" in values for values in day_parts.values()):
        legend.append("Другие")

    stacked: list[dict] = []
    for row in timeline:
        day = row["date"]
        parts_map = day_parts.get(day, {})
        parts = [{"sku": label, "count": int(parts_map.get(label, 0))} for label in legend if int(parts_map.get(label, 0)) > 0]
        stacked.append({"date": day, "total": row["count"], "parts": parts})

    return {"timeline": timeline, "stacked": stacked, "legend": legend}


def _reviews_summary_payload(start: str, end: str, sku: str | None = None) -> dict:
    def _safe_pct(change: int, base: int) -> float | None:
        if base == 0:
            return None
        return round((change / base) * 100.0, 2)

    if not _reviews_db_exists():
        return {
            "start": start,
            "end": end,
            "sku": sku or "",
            "period_count": 0,
            "total_before": 0,
            "total_at_end": 0,
            "delta": 0,
            "prev_start": "",
            "prev_end": "",
            "prev_count": 0,
            "delta_vs_prev": 0,
            "delta_pct": None,
            "delta_vs_prev_pct": None,
            "latest_scrape": "",
            "latest_review_date": "",
            "timeline": [],
            "timeline_stacked": [],
            "timeline_legend": [],
        }
    start_date = _parse_date(start)
    end_date = _parse_date(end)
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="end before start")
    period_days = (end_date - start_date).days + 1
    prev_start = start_date - timedelta(days=period_days)
    prev_end = start_date - timedelta(days=1)
    end_exclusive = end_date + timedelta(days=1)

    where_clause = (
        "comment IS NOT NULL AND TRIM(comment) != '' "
        "AND review_date IS NOT NULL AND review_date != ''"
    )
    params_base = []
    if sku:
        where_clause += " AND sku = ?"
        params_base.append(sku)

    def _count_between(date_from: date, date_to: date):
        sql = f"""
            SELECT COUNT(*)
            FROM reviews
            WHERE {where_clause}
              AND review_date >= ?
              AND review_date < ?
        """
        params = params_base + [date_from.isoformat(), date_to.isoformat()]
        return sql, params

    def _count_before(date_to: date):
        sql = f"""
            SELECT COUNT(*)
            FROM reviews
            WHERE {where_clause}
              AND review_date < ?
        """
        params = params_base + [date_to.isoformat()]
        return sql, params

    with sqlite3.connect(REVIEWS_DB_PATH) as conn:
        _normalize_review_dates(conn)
        cur = conn.cursor()
        sql, params = _count_between(start_date, end_exclusive)
        period_count = cur.execute(sql, params).fetchone()[0] or 0
        sql, params = _count_before(start_date)
        total_before = cur.execute(sql, params).fetchone()[0] or 0
        sql, params = _count_before(end_exclusive)
        total_at_end = cur.execute(sql, params).fetchone()[0] or 0
        sql, params = _count_between(prev_start, start_date)
        prev_count = cur.execute(sql, params).fetchone()[0] or 0
        latest_scrape = cur.execute("SELECT MAX(scraped_at) FROM reviews").fetchone()[0] or ""
        latest_review_date = cur.execute("SELECT MAX(review_date) FROM reviews").fetchone()[0] or ""
        timeline_payload = _reviews_timeline_payload(conn, start_date=start_date, end_exclusive=end_exclusive, sku=sku)

    delta = total_at_end - total_before
    delta_vs_prev = period_count - prev_count
    delta_pct = _safe_pct(int(delta), int(total_before))
    delta_vs_prev_pct = _safe_pct(int(delta_vs_prev), int(prev_count))

    return {
        "start": start_date.isoformat(),
        "end": end_date.isoformat(),
        "sku": sku or "",
        "period_count": int(period_count),
        "total_before": int(total_before),
        "total_at_end": int(total_at_end),
        "delta": int(delta),
        "prev_start": prev_start.isoformat(),
        "prev_end": prev_end.isoformat(),
        "prev_count": int(prev_count),
        "delta_vs_prev": int(delta_vs_prev),
        "delta_pct": delta_pct,
        "delta_vs_prev_pct": delta_vs_prev_pct,
        "latest_scrape": latest_scrape,
        "latest_review_date": latest_review_date,
        "timeline": timeline_payload["timeline"],
        "timeline_stacked": timeline_payload["stacked"],
        "timeline_legend": timeline_payload["legend"],
    }


def _reviews_timeline_api_payload(start: str, end: str, sku: str | None = None) -> dict:
    if not _reviews_db_exists():
        return {
            "start": start,
            "end": end,
            "sku": sku or "",
            "timeline": [],
            "timeline_stacked": [],
            "timeline_legend": [],
            "period_count": 0,
        }
    start_date = _parse_date(start)
    end_date = _parse_date(end)
    if end_date < start_date:
        raise HTTPException(status_code=400, detail="end before start")
    end_exclusive = end_date + timedelta(days=1)
    with sqlite3.connect(REVIEWS_DB_PATH) as conn:
        _normalize_review_dates(conn)
        timeline_payload = _reviews_timeline_payload(
            conn,
            start_date=start_date,
            end_exclusive=end_exclusive,
            sku=sku,
        )
    period_count = sum(int(item.get("count") or 0) for item in timeline_payload["timeline"])
    return {
        "start": start_date.isoformat(),
        "end": end_date.isoformat(),
        "sku": sku or "",
        "timeline": timeline_payload["timeline"],
        "timeline_stacked": timeline_payload["stacked"],
        "timeline_legend": timeline_payload["legend"],
        "period_count": int(period_count),
    }


def _reviews_skus_payload() -> dict:
    if not _reviews_db_exists():
        return {"skus": []}
    with sqlite3.connect(REVIEWS_DB_PATH) as conn:
        rows = conn.execute(
            "SELECT DISTINCT sku FROM reviews WHERE sku IS NOT NULL AND TRIM(sku) != '' ORDER BY sku"
        ).fetchall()
    return {"skus": [row[0] for row in rows]}


def _keyword_summary_payload() -> dict:
    latest = _keyword_latest_report()
    if not latest:
        return {}
    wb = load_workbook(latest, read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        header = next(rows_iter, None)
        if not header:
            return {}
        headers = [str(cell).strip().lower() if cell is not None else "" for cell in header]
        idx = {name: i for i, name in enumerate(headers)}
        keyword_idx = idx.get("keyword")
        status_idx = idx.get("status")
        position_idx = idx.get("position")
        if keyword_idx is None or status_idx is None:
            return {}

        total = 0
        ok = 0
        wrong = 0
        wrong_by_keyword: dict[str, dict] = {}
        for row in rows_iter:
            total += 1
            status_raw = row[status_idx] if status_idx < len(row) else ""
            status = str(status_raw or "").strip().lower()
            keyword_raw = row[keyword_idx] if keyword_idx < len(row) else ""
            keyword = str(keyword_raw or "").strip()
            pos_raw = row[position_idx] if position_idx is not None and position_idx < len(row) else None
            try:
                pos = int(pos_raw) if pos_raw not in (None, "") else None
            except Exception:
                pos = None
            if status == "ok":
                ok += 1
            else:
                wrong += 1
                bucket = wrong_by_keyword.setdefault(keyword or "Без ключа", {"count": 0, "best_pos": None})
                bucket["count"] += 1
                if pos is not None:
                    if bucket["best_pos"] is None or pos < bucket["best_pos"]:
                        bucket["best_pos"] = pos

        top_wrong = sorted(
            [
                {"keyword": key, "count": val["count"], "best_pos": val["best_pos"]}
                for key, val in wrong_by_keyword.items()
            ],
            key=lambda item: (-item["count"], item["best_pos"] or 9999),
        )[:8]

        return {
            "total": total,
            "ok": ok,
            "wrong": wrong,
            "top_wrong": top_wrong,
            "latest_file": latest.name,
            "latest_mtime": datetime.fromtimestamp(latest.stat().st_mtime, tz=ALMATY_TZ).isoformat(timespec="seconds"),
        }
    finally:
        wb.close()


def _keyword_history_payload(days: int = 10) -> dict:
    if not RESULTS_DIR.exists():
        return {"series": []}
    candidates = sorted(RESULTS_DIR.glob("kaspi_search_*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    series = []
    for path in candidates[: max(days, 1) * 3]:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = wb.active
            rows_iter = sheet.iter_rows(min_row=1, values_only=True)
            header = next(rows_iter, None)
            if not header:
                continue
            headers = [str(cell).strip().lower() if cell is not None else "" for cell in header]
            idx = {name: i for i, name in enumerate(headers)}
            status_idx = idx.get("status")
            if status_idx is None:
                continue
            total = 0
            ok = 0
            wrong = 0
            for row in rows_iter:
                total += 1
                status_norm = str(row[status_idx] or "").strip().lower()
                if status_norm == "ok":
                    ok += 1
                else:
                    wrong += 1
            mtime = datetime.fromtimestamp(path.stat().st_mtime, tz=ALMATY_TZ)
            series.append({
                "date": mtime.strftime("%m-%d %H:%M"),
                "ts": mtime.isoformat(timespec="seconds"),
                "ok": ok,
                "wrong": wrong,
                "total": total,
            })
            if len(series) >= days:
                break
        finally:
            wb.close()
    series.reverse()
    return {"series": series}


def _int_or_none(value):
    if value is None:
        return None
    try:
        return int(float(value))
    except Exception:
        return None


def _normalize_status(value) -> str:
    """Normalize status text for stable comparisons."""
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    # collapse duplicate whitespace to avoid surprises from Excel exports
    return " ".join(text.split()).lower()


def _is_issue(status_raw, price_kzt, actual_price_kzt) -> bool:
    """Problems are everything that is not explicitly OK in the status column."""
    status_norm = _normalize_status(status_raw)
    if status_norm.startswith("ok") or status_norm.startswith("ок"):
        return False
    return True


def _read_mission_rows(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = wb.active
        rows_iter = sheet.iter_rows(min_row=1, values_only=True)
        header = next(rows_iter, None)
        if not header:
            return None
        header_labels = [(str(cell).strip() if cell is not None else "") for cell in header]
        header_set = set(header_labels)
        if not MISSION_REQUIRED_COLUMNS.issubset(header_set):
            return None
        idx_map = {idx: col for idx, col in enumerate(header_labels) if col in MISSION_CAPTURE_COLUMNS}
        rows = []
        for raw in rows_iter:
            record = {}
            for idx, col in idx_map.items():
                record[col] = raw[idx] if idx < len(raw) else None
            rows.append(record)
        return rows
    finally:
        wb.close()


def load_latest_mission(limit: int | None = 250):
    if not RESULTS_DIR.exists():
        return None
    prefixes = _mission_prefixes()
    if not prefixes:
        return None
    candidates: list[Path] = []
    for prefix in prefixes:
        candidates.extend(RESULTS_DIR.glob(f"{prefix}_*.xlsx"))
    candidates = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
    for path in candidates[:80]:
        mtime = path.stat().st_mtime
        # отключаем кеш если структура могла поменяться
        rows = _read_mission_rows(path)
        if not rows:
            continue
        items = []
        total_issue_count = 0
        sellers: dict[str, str] = {}  # normalized -> original
        cities: set[str] = set()
        seller_sku_map: dict[str, set[str]] = {}
        bundle_cache_rows = _load_bundle_cache_rows()
        sliced_rows = rows if limit is None or limit <= 0 else rows[:limit]
        for raw in sliced_rows:
            status_raw = raw.get("status")
            price_kzt = _int_or_none(raw.get("price_kzt"))
            actual_price_kzt = _int_or_none(raw.get("actual_price_kzt"))
            is_issue = _is_issue(status_raw, price_kzt, actual_price_kzt)
            total_issue_count += 1 if is_issue else 0
            seller_name_raw = (raw.get("seller") or "").strip()
            seller_name = _canonical_seller_name(seller_name_raw)
            city_name = (raw.get("region") or "").strip()
            sku_label = (raw.get("input_product") or raw.get("product") or "").strip()
            product_url = raw.get("product_url") or ""
            product_code = _extract_product_code_from_url(product_url)
            seller_key = seller_name.lower()
            if seller_name:
                sellers[seller_key] = seller_name
                seller_sku_map.setdefault(seller_key, set()).add(sku_label or f"#{len(seller_sku_map.get(seller_key, []))+1}")
            if city_name:
                cities.add(city_name)
            item = {
                "region": city_name,
                "seller": seller_name,
                "seller_raw": seller_name_raw,
                "seller_key": seller_key,
                "product_code": product_code,
                "product": raw.get("product") or raw.get("input_product") or "",
                "input_product": raw.get("input_product") or "",
                "product_url": product_url,
                "review_count": _int_or_none(raw.get("review_count")),
                "review_progress": (str(raw.get("review_progress") or "").strip() if raw.get("review_progress") is not None else ""),
                "price_kzt": price_kzt,
                "actual_price_kzt": actual_price_kzt,
                "best_price_kzt": _int_or_none(raw.get("best_price_kzt")),
                "status": (status_raw or "").strip() or "—",
                "better_price_analysis": (raw.get("better_price_analysis") or "").strip(),
                "second_price_analysis": (raw.get("second_price_analysis") or "").strip(),
                "scraped_at": raw.get("scraped_at") or "",
                "bundle_of": (str(raw.get("bundle_of") or "").strip() if raw.get("bundle_of") is not None else ""),
                "bundle_qty": _int_or_none(raw.get("bundle_qty")),
                "bundle_1_sku_price_kzt": _int_or_none(raw.get("bundle_1_sku_price_kzt")),
                "bundle_price_kzt": _int_or_none(raw.get("bundle_price_kzt")),
                "bundle_unit_price_kzt": _int_or_none(raw.get("bundle_unit_price_kzt")),
                "bundle_same_sku_total_kzt": _int_or_none(raw.get("bundle_same_sku_total_kzt")),
                "bundle_savings_vs_same_sku_total_kzt": _int_or_none(raw.get("bundle_savings_vs_same_sku_total_kzt")),
                "bundle_vs_single_kzt": _int_or_none(raw.get("bundle_vs_single_kzt")),
                "bundle_best_seller": (raw.get("bundle_best_seller") or "").strip(),
                "1_sku_best_price_kzt": _int_or_none(raw.get("1_sku_best_price_kzt")),
                "bundle_best_price_kzt": _int_or_none(raw.get("bundle_best_price_kzt")),
                "bundle_savings_vs_best_total_kzt": _int_or_none(raw.get("bundle_savings_vs_best_total_kzt")),
                "bundle_vs_best_kzt": _int_or_none(raw.get("bundle_vs_best_kzt")),
                "bundle_discount_pct": (str(raw.get("bundle_discount_pct") or "").strip() if raw.get("bundle_discount_pct") is not None else ""),
                "bundle_note": (str(raw.get("bundle_note") or "").strip() if raw.get("bundle_note") is not None else ""),
                "is_issue": is_issue,
            }
            items.append(_enrich_bundle_item(item, bundle_cache_rows))

        seller_stats = []
        for key, sku_set in seller_sku_map.items():
            display_name = sellers.get(key) or key
            seller_rows = [row for row in items if (row.get("seller_key") or "") == key]
            ok_count = len([r for r in seller_rows if not r.get("is_issue")])
            seller_issue_count = len([r for r in seller_rows if r.get("is_issue")])
            seller_stats.append({
                "seller": display_name,
                "seller_key": key,
                "sku_count": len(sku_set),
                "ok": int(ok_count),
                "issues": int(seller_issue_count),
            })
        seller_stats = sorted(seller_stats, key=lambda x: x["sku_count"], reverse=True)
        # Recompute totals from the cleaned items to avoid any mismatch with display logic
        total_issue_count = sum(1 for item in items if item.get("is_issue"))
        ok_count = len(items) - total_issue_count
        payload = {
            "items": items,
            "total": len(items),
            "issues": total_issue_count,
            "ok": max(0, ok_count),
            "updated_at": datetime.fromtimestamp(mtime).isoformat(timespec="seconds"),
            "updated_human": datetime.fromtimestamp(mtime).strftime("%d.%m %H:%M"),
            "source_file": path.name,
            "sellers": sorted(sellers.values()),
            "sellers_count": len(sellers),
            "cities": sorted(cities),
            "cities_count": len(cities),
            "seller_stats": seller_stats,
        }
        MISSION_CACHE.update({"path": path, "mtime": mtime, "payload": payload})
        return payload
    return None


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    mission_payload = load_latest_mission(limit=MISSION_VIEW_LIMIT)
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "tg_bot_name": "LilPonk Bot",
            "mission_data": mission_payload or {},
            "mission_data_json": json.dumps(mission_payload or {}, ensure_ascii=False),
        },
    )


@app.get("/health", response_class=HTMLResponse)
async def healthcheck():
    return "ok"


@app.get("/mission/data")
async def mission_data():
    payload = load_latest_mission(limit=MISSION_VIEW_LIMIT)
    if not payload:
        raise HTTPException(status_code=404, detail="Нет свежего отчёта по миссии")
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/mission/campaign-quality")
async def mission_campaign_quality():
    payload = _campaign_quality_payload()
    if not payload:
        raise HTTPException(status_code=404, detail="Нет свежей аналитики по кампании")
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/download/mission")
async def download_mission():
    path = _latest_report(_mission_prefixes())
    if not path:
        raise HTTPException(status_code=404, detail="Нет свежего отчёта по миссии")
    return FileResponse(
        path,
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/download/main")
async def download_main():
    path = _latest_report([MAIN_FILE_PREFIX, "kaspi_prices"])
    if not path:
        raise HTTPException(status_code=404, detail="Нет свежего основного отчёта")
    return FileResponse(
        path,
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/download/competitors")
async def download_competitors():
    path = _latest_report(_competitor_prefixes())
    if not path:
        raise HTTPException(status_code=404, detail="Нет свежего отчёта по конкурентам")
    return FileResponse(
        path,
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/download/keywords")
async def download_keywords():
    path = _keyword_latest_report()
    if not path:
        raise HTTPException(status_code=404, detail="Нет свежего отчёта по ключам")
    return FileResponse(
        path,
        filename=path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/competitors/summary")
async def competitors_summary(
    city: str = Query(default="Алматы"),
    limit: int = Query(default=260, ge=20, le=1200),
):
    payload = _competitors_summary_payload(city=city, limit=limit)
    if not payload:
        raise HTTPException(status_code=404, detail="Нет свежего отчёта по конкурентам")
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/competitors/status")
async def competitors_status():
    payload = _competitors_status_payload()
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.post("/competitors/run")
async def competitors_run(city: str = Query(default="Алматы")):
    with COMPETITOR_LOCK:
        if COMPETITOR_STATE.get("status") == "running":
            return JSONResponse({"status": "running"}, headers={"Cache-Control": "no-store"})
        COMPETITOR_STATE.update({"status": "running", "last_error": ""})

    def _job():
        try:
            csv_arg = os.environ.get("COMPETITOR_PRODUCTS_CSV") or "competitor_products.csv"
            run_kaspi_scrape(
                cities=[city],
                extra_csv_paths=csv_arg,
                include_base=False,
                output_prefix=COMPETITOR_FILE_PREFIX,
                priority="secondary",
            )
            latest = _latest_report(_competitor_prefixes())
            with COMPETITOR_LOCK:
                COMPETITOR_STATE["status"] = "idle"
                COMPETITOR_STATE["last_run"] = datetime.now().isoformat(timespec="seconds")
                COMPETITOR_STATE["last_file"] = latest.name if latest else ""
                if not latest:
                    COMPETITOR_STATE["last_error"] = "no_data"
        except Exception as exc:
            with COMPETITOR_LOCK:
                COMPETITOR_STATE["status"] = "idle"
                COMPETITOR_STATE["last_run"] = datetime.now().isoformat(timespec="seconds")
                COMPETITOR_STATE["last_error"] = str(exc)
        finally:
            COMPETITOR_CACHE.update({"path": None, "mtime": 0.0, "city": "", "limit": 0, "payload": None})

    threading.Thread(target=_job, daemon=True).start()
    return JSONResponse({"status": "started"}, headers={"Cache-Control": "no-store"})


@app.get("/price-leaders/summary")
async def price_leaders_summary():
    payload = _price_leaders_summary_payload()
    if not payload:
        raise HTTPException(status_code=404, detail="Нет live-аналитики по лучшим ценам")
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/price-leaders/ranking")
async def price_leaders_ranking(
    month: str | None = Query(default=None),
    level: str = Query(default="month"),
    query: str | None = Query(default=None),
    pharmacy: str | None = Query(default=None),
    brand: str | None = Query(default=None),
    search: str | None = Query(default=None),
    sku_query: str | None = Query(default=None),
    top_n: int = Query(default=50, ge=1, le=500),
):
    payload = _price_leaders_ranking_payload(
        month=month,
        level=level,
        query=query,
        pharmacy=pharmacy,
        brand=brand,
        search=search,
        sku_query=sku_query,
        top_n=top_n,
    )
    if not payload:
        raise HTTPException(status_code=404, detail="Нет live-аналитики по лучшим ценам")
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/price-leaders/filters")
async def price_leaders_filters(
    month: str | None = Query(default=None),
    pharmacy: str | None = Query(default=None),
    brand: str | None = Query(default=None),
    search: str | None = Query(default=None),
    sku_query: str | None = Query(default=None),
    q: str | None = Query(default=None),
    limit: int = Query(default=20, ge=5, le=100),
):
    payload = _price_leaders_filters_payload(
        month=month,
        pharmacy=pharmacy,
        brand=brand,
        search=search,
        sku_query=sku_query,
        q=q,
        limit=limit,
    )
    if not payload:
        raise HTTPException(status_code=404, detail="Нет live-аналитики по лучшим ценам")
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/price-dynamics/options")
async def price_dynamics_options(
    region: str = Query(default="Алматы"),
    q: str | None = Query(default=None),
    limit: int = Query(default=40, ge=5, le=120),
):
    payload = _price_dynamics_options_payload(region=region, q=q, limit=limit)
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/price-dynamics/summary")
async def price_dynamics_summary(
    region: str = Query(default="Алматы"),
    sku: str = Query(...),
    start: str = Query(...),
    end: str = Query(...),
):
    payload = _price_dynamics_summary_payload(region=region, sku=sku, start=start, end=end)
    if not payload:
        raise HTTPException(status_code=404, detail="Нет данных по SKU за выбранный период")
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/download/price-leaders")
async def download_price_leaders():
    payload = _price_leaders_summary_payload()
    if not payload:
        raise HTTPException(status_code=404, detail="Нет live-аналитики по лучшим ценам")
    for key in ("details_file_latest", "details_file"):
        raw_name = str(payload.get(key) or "").strip()
        if not raw_name:
            continue
        safe_name = Path(raw_name).name
        path = RESULTS_DIR / safe_name
        if path.exists() and path.is_file():
            return FileResponse(path, filename=path.name, media_type="text/csv")
    raise HTTPException(status_code=404, detail="CSV файл live-аналитики не найден")


@app.get("/wolt/reports")
async def wolt_reports():
    payload = _wolt_reports_payload()
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/wolt/summary")
async def wolt_summary(
    report: str | None = Query(default=None),
    brand: str | None = Query(default=None),
):
    path = _resolve_wolt_report(report=report, brand=brand)
    if not path:
        raise HTTPException(status_code=404, detail="Нет свежего отчёта по Wolt")
    payload = _parse_wolt_items_report(path)
    payload["report_file"] = path.name
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/wolt/assortment-gap")
async def wolt_assortment_gap():
    payload = _wolt_assortment_gap_payload()
    if not payload:
        raise HTTPException(status_code=404, detail="Нет свежего gap-отчёта по Wolt")
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/glovo/reports")
async def glovo_reports():
    payload = _glovo_reports_payload()
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/glovo/summary")
async def glovo_summary(
    report: str | None = Query(default=None),
    brand: str | None = Query(default=None),
):
    path = _resolve_glovo_report(report=report, brand=brand)
    if not path:
        raise HTTPException(status_code=404, detail="Нет свежего отчёта по Glovo")
    payload = _parse_glovo_items_report(path)
    payload["report_file"] = path.name
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/glovo/coverage")
async def glovo_coverage():
    payload = _glovo_coverage_payload()
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/keywords/status")
async def keyword_status():
    payload = _keyword_status_payload()
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/keywords/queries")
async def keyword_queries():
    payload = _keyword_queries_payload()
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/keywords/summary")
async def keyword_summary():
    payload = _keyword_summary_payload()
    if not payload:
        raise HTTPException(status_code=404, detail="Нет свежего отчёта по ключам")
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/keywords/rows")
async def keyword_rows(
    status: str | None = Query(default=None),
    pos_le: int | None = Query(default=None),
):
    payload = _keyword_rows_payload(status=status, pos_le=pos_le)
    if not payload:
        raise HTTPException(status_code=404, detail="Нет свежего отчёта по ключам")
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/keywords/history")
async def keyword_history():
    payload = _keyword_history_payload()
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/reviews/summary")
async def reviews_summary(
    start: str = Query(...),
    end: str = Query(...),
    sku: str | None = Query(default=None),
):
    payload = _reviews_summary_payload(start=start, end=end, sku=sku)
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/reviews/timeline")
async def reviews_timeline(
    start: str = Query(...),
    end: str = Query(...),
    sku: str | None = Query(default=None),
):
    payload = _reviews_timeline_api_payload(start=start, end=end, sku=sku)
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/reviews/skus")
async def reviews_skus():
    payload = _reviews_skus_payload()
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.get("/reviews/status")
async def reviews_status():
    payload = _reviews_status_payload()
    return JSONResponse(payload, headers={"Cache-Control": "no-store"})


@app.post("/reviews/run")
async def reviews_run():
    with REVIEWS_LOCK:
        if REVIEWS_STATE.get("status") == "running":
            return JSONResponse({"status": "running"}, headers={"Cache-Control": "no-store"})
        REVIEWS_STATE.update({"status": "running", "last_error": ""})

    def _job():
        try:
            path = run_reviews_scrape(reviews_csv=BASE_DIR.parent / "reviews.csv")
            with REVIEWS_LOCK:
                REVIEWS_STATE["status"] = "idle"
                REVIEWS_STATE["last_run"] = datetime.now().isoformat(timespec="seconds")
                REVIEWS_STATE["last_file"] = path.name if path else ""
                if not path:
                    REVIEWS_STATE["last_error"] = "no_data"
        except Exception as exc:
            with REVIEWS_LOCK:
                REVIEWS_STATE["status"] = "idle"
                REVIEWS_STATE["last_run"] = datetime.now().isoformat(timespec="seconds")
                REVIEWS_STATE["last_error"] = str(exc)

    threading.Thread(target=_job, daemon=True).start()
    return JSONResponse({"status": "started"}, headers={"Cache-Control": "no-store"})


@app.post("/keywords/run")
async def run_keywords():
    with KEYWORD_LOCK:
        if KEYWORD_STATE.get("status") == "running":
            return JSONResponse({"status": "running"}, headers={"Cache-Control": "no-store"})
        KEYWORD_STATE.update({"status": "running", "last_error": ""})

    def _job():
        try:
            path = run_keyword_search_report(include_base=True, chat_id=None)
            with KEYWORD_LOCK:
                KEYWORD_STATE["status"] = "idle"
                KEYWORD_STATE["last_run"] = datetime.now().isoformat(timespec="seconds")
                KEYWORD_STATE["last_file"] = path.name if path else ""
                if not path:
                    KEYWORD_STATE["last_error"] = "no_data"
        except Exception as exc:
            with KEYWORD_LOCK:
                KEYWORD_STATE["status"] = "idle"
                KEYWORD_STATE["last_run"] = datetime.now().isoformat(timespec="seconds")
                KEYWORD_STATE["last_error"] = str(exc)

    threading.Thread(target=_job, daemon=True).start()
    return JSONResponse({"status": "started"}, headers={"Cache-Control": "no-store"})
